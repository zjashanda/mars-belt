from __future__ import annotations

import argparse
import hashlib
import json
import os
import platform
import re
import shlex
import shutil
import subprocess
import sys
import time
import zipfile
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Optional, Sequence

# Detect current OS for smart defaults
IS_WINDOWS = platform.system() == "Windows"
IS_LINUX = platform.system() == "Linux"

# Smart defaults based on OS
DEFAULT_PORT = "/dev/ttyACM0" if IS_LINUX else "COM14"
DEFAULT_CTRL_PORT = "/dev/ttyACM4" if IS_LINUX else "COM15"

INTERNAL_PY_ROOT = Path(__file__).resolve().parent / "py"
if str(INTERNAL_PY_ROOT) not in sys.path:
    sys.path.insert(0, str(INTERNAL_PY_ROOT))

from listenai_audio_skill_bootstrap import ensure_audio_skill_set, ensure_laid_helper
from listenai_task_support import (
    BURN_LOG_ROOT,
    LOCAL_TOOLS_MD,
    PACKAGE_CACHE_ROOT,
    STATE_ROOT,
    TASKS_ROOT,
    append_error_entry,
    ensure_runtime_dir,
    load_local_listenai_token,
    resolve_listenai_token,
    resolve_user_path,
    runtime_dir_for_task,
    task_dir_for_runtime,
)


SCRIPT_ROOT = Path(__file__).resolve().parent
SKILL_ROOT = SCRIPT_ROOT.parent
PY_ROOT = SCRIPT_ROOT / "py"
BURN_ROOT = SCRIPT_ROOT / "burn"
RESULT_ROOT = TASKS_ROOT
PACKAGE_ROOT = PACKAGE_CACHE_ROOT
WAV_ROOT = SCRIPT_ROOT / "wavSource"
RUNTIME_ROOT = ensure_runtime_dir()
TOOLS_MD = LOCAL_TOOLS_MD


def ensure_workspace() -> None:
    for path in (
        RESULT_ROOT,
        PACKAGE_ROOT,
        RUNTIME_ROOT,
        BURN_LOG_ROOT,
        STATE_ROOT,
    ):
        path.mkdir(parents=True, exist_ok=True)


def load_token_from_tools_md() -> str:
    return load_local_listenai_token()


def set_token(token: str, allow_missing: bool = False) -> None:
    try:
        resolve_listenai_token(token, allow_missing=allow_missing, persist=True)
    except RuntimeError as exc:
        raise SystemExit(str(exc)) from exc


def add_arg(args: List[str], name: str, value: Any, allow_empty: bool = False) -> None:
    if value is None:
        return
    text = str(value)
    if not allow_empty and text == "":
        return
    args.extend([name, text])


def add_path_arg(args: List[str], name: str, value: Any, allow_empty: bool = False) -> None:
    if value is None:
        return
    text = str(value)
    if not allow_empty and text == "":
        return
    args.extend([name, resolve_user_path(text)])


def split_repeat(values: Optional[Sequence[str]], *, split_comma: bool = True) -> List[str]:
    result: List[str] = []
    for value in values or []:
        if value is None:
            continue
        items = str(value).split(",") if split_comma else [str(value)]
        for item in items:
            text = item.strip()
            if text:
                result.append(text)
    return result


def runtime_file(*parts: str) -> Path:
    if not parts:
        raise ValueError("runtime_file requires at least one path segment")
    if len(parts) == 1:
        path = RUNTIME_ROOT / parts[0]
    else:
        path = ensure_runtime_dir(*parts[:-1]) / parts[-1]
    path.parent.mkdir(parents=True, exist_ok=True)
    return path


def resolve_task_runtime_dir(task_dir: Path) -> Path:
    base_dir = runtime_dir_for_task(task_dir)
    candidates = [base_dir]
    candidates.extend(
        sorted(
            item.parent
            for item in base_dir.glob("*/summary.json")
            if item.is_file()
        )
    )
    for candidate in candidates:
        if (candidate / "summary.json").exists():
            return candidate
    return base_dir


def infer_task_dir(package_zip: str = "", suite_dir: str = "") -> Optional[Path]:
    if package_zip:
        return Path(resolve_user_path(package_zip)).resolve().parent
    if suite_dir:
        suite_path = Path(resolve_user_path(suite_dir)).resolve()
        for parent in [suite_path, *suite_path.parents]:
            if parent.parent == RESULT_ROOT:
                return parent
    return None


def latest_result_subdir(result_root: Path) -> Optional[Path]:
    if not result_root.exists():
        return None
    candidates = [item for item in result_root.iterdir() if item.is_dir()]
    if not candidates:
        return None
    candidates.sort(key=lambda item: item.stat().st_mtime, reverse=True)
    return candidates[0]


def copy_if_exists(source: Path, target: Path) -> None:
    if not source.exists():
        return
    target.parent.mkdir(parents=True, exist_ok=True)
    shutil.copy2(source, target)


def stage_validation_artifacts(task_dir: Optional[Path], suite_dir: Path) -> None:
    if task_dir is None:
        print("[mars-belt] stage_validation_artifacts: task_dir is None, skipping")
        return
    result_dir = latest_result_subdir(suite_dir / "result")
    if result_dir is None:
        print(f"[mars-belt] stage_validation_artifacts: no result dir found in {suite_dir / 'result'}")
        return
    
    print(f"[mars-belt] stage_validation_artifacts: copying from {result_dir} to {task_dir}")
    
    copy_if_exists(result_dir / "testResult.xlsx", task_dir / "testResult.xlsx")
    copy_if_exists(result_dir / "serial_raw.log", task_dir / "serial_raw.log")
    test_logs = sorted(result_dir.glob("test_*.log"), key=lambda item: item.stat().st_mtime, reverse=True)
    if test_logs:
        copy_if_exists(test_logs[0], task_dir / "test_tool.log")
    
    # Generate HTML email report and zip file
    generate_email_report(task_dir, result_dir)
    
    print(f"[mars-belt] stage_validation_artifacts: done")


def generate_email_report(task_dir: Path, result_dir: Path) -> None:
    """Generate HTML email report and zip file with attachments."""
    import csv
    from html import escape

    print("[mars-belt] generate_email_report: starting...")

    task_dir_abs = task_dir.resolve() if not task_dir.is_absolute() else task_dir
    summary_path = resolve_task_runtime_dir(task_dir_abs) / "summary.json"
    if not summary_path.exists():
        print(f"[mars-belt] generate_email_report: summary.json not found at {summary_path}")
        return

    runtime_dir = summary_path.parent
    suite_dir = runtime_dir / "suite"
    serial_path = task_dir / "serial_raw.log"
    xlsx_path = task_dir / "testResult.xlsx"
    test_log_candidates = sorted(result_dir.glob("test_*.log"), key=lambda item: item.stat().st_mtime)

    with summary_path.open("r", encoding="utf-8") as fp:
        summary = json.load(fp)

    device_info = {}
    for device_info_path in (
        result_dir / "deviceInfo_generated.json",
        suite_dir / "deviceInfo_generated.json",
        task_dir / "deviceInfo_generated.json",
    ):
        if device_info_path.exists():
            with device_info_path.open("r", encoding="utf-8") as fp:
                device_info = json.load(fp)
            break

    test_cases: List[Dict[str, str]] = []
    for test_cases_csv in (
        suite_dir / "testCases.csv",
        result_dir.parent / "testCases.csv",
        task_dir / "testCases.csv",
    ):
        if test_cases_csv.exists():
            with test_cases_csv.open("r", encoding="utf-8-sig", newline="") as fp:
                test_cases = list(csv.DictReader(fp))
            break
    test_cases_by_id = {row.get("用例编号", "").strip(): row for row in test_cases if row.get("用例编号")}

    xlsx_rows: Dict[str, Dict[str, str]] = {}
    if xlsx_path.exists():
        try:
            from openpyxl import load_workbook

            wb = load_workbook(xlsx_path, read_only=True, data_only=True)
            ws = wb.active
            headers = [str(cell) if cell is not None else "" for cell in next(ws.iter_rows(min_row=1, max_row=1, values_only=True))]
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not row:
                    continue
                data = {
                    headers[index]: "" if index >= len(row) or row[index] is None else str(row[index])
                    for index in range(len(headers))
                    if headers[index]
                }
                case_id = data.get("用例编号", "").strip()
                if case_id:
                    xlsx_rows[case_id] = data
        except Exception as exc:
            print(f"[mars-belt] generate_email_report: error reading {xlsx_path}: {exc}")

    serial_content = ""
    if serial_path.exists():
        serial_content = serial_path.read_text(encoding="utf-8", errors="replace")

    selected = summary.get("selected") or {}
    package_summary = summary.get("packageSummary") or {}
    final_release = summary.get("finalReleaseSubset") or {}
    product_name = package_summary.get("productName") or selected.get("productLabel") or "未知产品"
    module = selected.get("moduleMark") or selected.get("moduleBoard") or "未知模组"
    version = final_release.get("version") or package_summary.get("releaseVersion") or "V-未知"
    language = selected.get("language") or "中文"
    flash = selected.get("flash") or "-"
    sram = selected.get("sram") or "-"
    power = selected.get("powerSupply") or "-"
    firmware_timeout = final_release.get("timeout", "-")
    firmware_vol_level = final_release.get("volLevel", "-")
    firmware_default_vol = final_release.get("defaultVol", "-")
    wake_word = device_info.get("wakeupWord") or "小聆小聆"

    test_time = datetime.fromtimestamp(result_dir.stat().st_mtime).strftime("%Y-%m-%d %H:%M:%S")

    case_order = [row["用例编号"].strip() for row in test_cases if row.get("用例编号")]
    if not case_order:
        case_order = list(xlsx_rows.keys())

    def case_kind(case_id: str) -> str:
        if "-VERSION-" in case_id:
            return "version"
        if "-TIMEOUT-" in case_id:
            return "timeout"
        if "-DEFAULTVOL-" in case_id:
            return "defaultvol"
        if "-VOLUME-" in case_id:
            return "volume"
        if "-WAKE-" in case_id:
            return "wake"
        if "-CMD-" in case_id:
            return "cmd"
        return "other"

    def display_type(case_id: str) -> str:
        mapping = {
            "version": "固件版本校验",
            "timeout": "超时退出",
            "volume": "音量档位验证",
            "defaultvol": "默认音量验证",
            "wake": "唤醒识别",
            "cmd": "功能验证",
        }
        kind = case_kind(case_id)
        return mapping.get(kind) or test_cases_by_id.get(case_id, {}).get("测试类型", "") or xlsx_rows.get(case_id, {}).get("测试类型", "")

    def display_command(case_id: str) -> str:
        kind = case_kind(case_id)
        if kind in {"version", "timeout", "volume", "defaultvol"}:
            return "-"
        if kind == "wake":
            return wake_word
        return test_cases_by_id.get(case_id, {}).get("命令词", "") or xlsx_rows.get(case_id, {}).get("命令词", "") or "-"

    def parse_version_actual(row: Dict[str, str]) -> str:
        detail = row.get("识别结果", "")
        device_resp = row.get("设备响应列表", "")
        for source in (detail, device_resp, serial_content):
            match = re.search(r"device=([^;|]+)", source)
            if match:
                return match.group(1).strip()
            match = re.search(r"config version:\s*(\S+)", source)
            if match:
                return match.group(1).strip()
        return "-"

    def parse_timeout_actual(row: Dict[str, str]) -> str:
        detail = row.get("识别结果", "")
        match = re.search(r"actual=([0-9.]+)s", detail)
        if match:
            return f"设备在{match.group(1)}s退出唤醒态"
        return detail or "-"

    def parse_volume_actual(row: Dict[str, str]) -> str:
        detail = row.get("识别结果", "")
        match = re.search(r"实际检测=(\d+)", detail)
        if match:
            return f"音量设置 最小→最大，实测档位={match.group(1)}"
        values = [int(item) for item in re.findall(r"set vol:\s*\d+\s*->\s*(\d+)", serial_content)]
        if values:
            return f"音量设置 最小→最大，实测档位={len(set(values)) + 1}"
        return detail or "-"

    def infer_default_slot(raw_value: str) -> Optional[int]:
        if not raw_value:
            return None
        try:
            observed = int(raw_value)
        except ValueError:
            return None
        levels_path = runtime_dir / "web_config.json"
        levels: List[int] = []
        if levels_path.exists():
            try:
                web_cfg = json.loads(levels_path.read_text(encoding="utf-8"))
                levels = [
                    int(item)
                    for item in ((web_cfg.get("firmware") or {}).get("volume_config") or {}).get("level") or []
                    if isinstance(item, int)
                ]
            except Exception:
                levels = []
        if observed in levels:
            return levels.index(observed) + 1
        if levels and 0 <= observed < len(levels):
            return observed + 1
        return observed + 1 if observed >= 0 else None

    def parse_default_actual(row: Dict[str, str]) -> str:
        detail = row.get("识别结果", "")
        match = re.search(r"inferred_default=(\d+)", detail)
        if match:
            return f"断电重启后默认第{match.group(1)}档"
        match = re.search(r"boot_volume=(\d+)", detail)
        if match:
            inferred = infer_default_slot(match.group(1))
            if inferred is not None:
                return f"断电重启后默认第{inferred}档"
        device_resp = row.get("设备响应列表", "")
        match = re.search(r"volume\s*:\s*(\d+)", device_resp)
        if match:
            inferred = infer_default_slot(match.group(1))
            if inferred is not None:
                return f"断电重启后默认第{inferred}档"
        return detail or "-"

    def parse_wake_actual(row: Dict[str, str]) -> str:
        raw = row.get("识别原始结果", "").strip()
        recognized = row.get("识别结果", "").strip()
        if raw and recognized and raw != recognized:
            return f"{raw} → {recognized}"
        return recognized or raw or "-"

    def expected_value(case_id: str, row: Dict[str, str]) -> Tuple[str, bool]:
        kind = case_kind(case_id)
        if kind == "version":
            return str(version), False
        if kind == "timeout":
            return f"timeout = {firmware_timeout}s", False
        if kind == "volume":
            return f"总档位 = {firmware_vol_level}档", False
        if kind == "defaultvol":
            return f"默认第{firmware_default_vol}档", False
        if kind == "wake":
            return "唤醒成功", False
        protocol = row.get("期望协议", "").strip()
        return (protocol or "-", bool(protocol))

    def actual_value(case_id: str, row: Dict[str, str]) -> Tuple[str, bool]:
        kind = case_kind(case_id)
        if kind == "version":
            return parse_version_actual(row), False
        if kind == "timeout":
            return parse_timeout_actual(row), False
        if kind == "volume":
            return parse_volume_actual(row), False
        if kind == "defaultvol":
            return parse_default_actual(row), False
        if kind == "wake":
            return parse_wake_actual(row), False
        protocol = row.get("实际发送协议", "").strip()
        return (protocol or "-", bool(protocol))

    def render_value(value: str, *, proto: bool = False) -> str:
        text = escape(value or "-")
        if proto and value and value != "-":
            return f'<span class="proto">{text}</span>'
        return text

    pass_count = 0
    rows_html: List[str] = []
    notes: List[str] = []
    failed_case_lines: List[str] = []
    for case_id in case_order:
        csv_row = test_cases_by_id.get(case_id, {})
        xlsx_row = xlsx_rows.get(case_id, {})
        verdict = xlsx_row.get("识别判定", "")
        if verdict == "OK":
            pass_count += 1
        else:
            expected_text, _ = expected_value(case_id, csv_row)
            actual_text, _ = actual_value(case_id, xlsx_row)
            failed_case_lines.append(
                f"{case_id} | {display_command(case_id) or display_type(case_id)} | 期望={expected_text or '-'} | 实测={actual_text or '-'}"
            )
        status_class = "status-ok" if verdict == "OK" else "status-fail"
        status_text = "OK" if verdict == "OK" else ("Fail" if verdict else "-")
        expected_text, expected_proto = expected_value(case_id, csv_row)
        actual_text, actual_proto = actual_value(case_id, xlsx_row)
        rows_html.append(
            "                <tr>\n"
            f"                    <td>{escape(case_id)}</td>\n"
            f"                    <td>{escape(display_type(case_id))}</td>\n"
            f"                    <td>{escape(display_command(case_id))}</td>\n"
            f"                    <td>{render_value(expected_text, proto=expected_proto)}</td>\n"
            f"                    <td>{render_value(actual_text, proto=actual_proto)}</td>\n"
            f"                    <td><span class=\"{status_class}\">{status_text}</span></td>\n"
            "                </tr>"
        )
        if case_kind(case_id) == "version" and verdict != "OK" and actual_text.endswith(".."):
            notes.append(f"{case_id} 启动日志中的固件版本被截断为 {actual_text}，未获取到完整版本号，按失败判定。")
        if case_kind(case_id) == "cmd" and verdict == "OK":
            retry_match = re.search(r"attempt=(\d+)/(\d+)", xlsx_row.get("设备响应列表", ""))
            if retry_match and int(retry_match.group(1)) > 1:
                notes.append(f"{case_id} 首轮协议日志存在截断或未完整捕获，重试后通过，不影响功能判定。")
        if case_kind(case_id) == "volume" and verdict == "OK" and "ConfigFail" in xlsx_row.get("设备响应列表", ""):
            notes.append(f"{case_id} 音量档位探测过程中出现协议日志截断，但实测档位数与配置一致，按功能通过判定。")

    total_count = len(case_order)
    fail_count = max(total_count - pass_count, 0)
    pass_rate = int(pass_count / total_count * 100) if total_count else 0
    conclusion = "PASS" if total_count and fail_count == 0 else "FAIL"
    result_text = "全部通过" if conclusion == "PASS" else "存在失败项"
    notes = list(dict.fromkeys(notes))

    config_rows = [
        ("唤醒超时", f"{firmware_timeout}s", parse_timeout_actual(xlsx_rows.get(next((cid for cid in case_order if case_kind(cid) == 'timeout'), ""), {}))),
        ("音量档位", f"{firmware_vol_level}档", parse_volume_actual(xlsx_rows.get(next((cid for cid in case_order if case_kind(cid) == 'volume'), ""), {}))),
        ("默认音量", f"第{firmware_default_vol}档", parse_default_actual(xlsx_rows.get(next((cid for cid in case_order if case_kind(cid) == 'defaultvol'), ""), {}))),
        ("唤醒词", wake_word, wake_word),
        ("Flash / SRAM", f"{flash} / {sram}", f"{flash} / {sram}"),
        ("供电电压", power, power),
    ]
    config_rows_html = "\n".join(
        "                <tr>"
        f"<td>{escape(label)}</td>"
        f"<td>{escape(expected)}</td>"
        f"<td>{escape(actual)}</td>"
        "</tr>"
        for label, expected, actual in config_rows
    )

    note_html = ""
    if notes:
        note_items = "".join(f"<li>{escape(item)}</li>" for item in notes)
        note_html = (
            '<div class="attn-box">'
            "<strong>异常备注</strong>"
            f"<ul>{note_items}</ul>"
            "</div>"
        )

    html_body = f"""<!DOCTYPE html>
<html>
<head>
    <meta charset="UTF-8">
    <style>
        body {{ font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif; background-color: #f5f7fa; margin: 0; padding: 20px; }}
        .container {{ max-width: 760px; margin: 0 auto; background: #ffffff; border-radius: 12px; overflow: hidden; box-shadow: 0 4px 20px rgba(0,0,0,0.08); }}
        .header {{ background: linear-gradient(135deg, #667eea 0%, #764ba2 100%); color: white; padding: 28px 30px; }}
        .header h1 {{ margin: 0; font-size: 22px; font-weight: 600; }}
        .header .subtitle {{ margin-top: 6px; opacity: 0.9; font-size: 13px; }}
        .content {{ padding: 28px 30px; }}
        .info-grid {{ display: grid; grid-template-columns: repeat(3, 1fr); gap: 12px; margin-bottom: 20px; }}
        .info-item {{ background: #f8f9fa; padding: 12px 14px; border-radius: 8px; border-left: 4px solid #667eea; }}
        .info-item label {{ display: block; color: #6c757d; font-size: 11px; margin-bottom: 4px; text-transform: uppercase; letter-spacing: 0.5px; }}
        .info-item span {{ color: #2d3748; font-weight: 600; font-size: 13px; }}
        .section-title {{ color: #2d3748; font-size: 15px; font-weight: 700; margin: 20px 0 12px; padding-bottom: 8px; border-bottom: 2px solid #667eea; display: inline-block; }}
        .config-table, table {{ width: 100%; border-collapse: collapse; margin-top: 12px; }}
        .config-table th, th {{ background: #667eea; color: white; padding: 10px 12px; text-align: left; font-size: 12px; }}
        .config-table td, td {{ padding: 9px 12px; border-bottom: 1px solid #e9ecef; font-size: 12px; color: #4a5568; }}
        .config-table td:first-child, td:first-child {{ font-weight: 600; color: #2d3748; }}
        .result-box {{ background: linear-gradient(135deg, #11998e 0%, #38ef7d 100%); color: white; padding: 22px; border-radius: 10px; text-align: center; margin-bottom: 20px; }}
        .result-box.fail {{ background: linear-gradient(135deg, #eb4d4b 0%, #f093fb 100%); }}
        .pass-rate {{ font-size: 40px; font-weight: 700; }}
        .result-text {{ font-size: 16px; margin-top: 4px; opacity: 0.95; }}
        .status-ok {{ color: #218838; font-weight: 600; background: #e6fffa; padding: 3px 9px; border-radius: 20px; font-size: 11px; }}
        .status-fail {{ color: #c53030; font-weight: 600; background: #fff5f5; padding: 3px 9px; border-radius: 20px; font-size: 11px; }}
        .proto {{ font-family: 'Courier New', monospace; font-size: 11px; color: #805ad5; background: #faf5ff; padding: 2px 5px; border-radius: 4px; }}
        .attn-box {{ background: #fff8e6; padding: 13px 15px; border-radius: 8px; border-left: 4px solid #d69e2e; font-size: 12px; color: #4a5568; margin-top: 14px; }}
        .attn-box ul {{ margin: 8px 0 0 18px; padding: 0; }}
        .footer {{ background: #f8f9fa; padding: 18px; text-align: center; color: #6c757d; font-size: 11px; }}
    </style>
</head>
<body>
    <div class="container">
        <div class="header">
            <h1>固件测试报告</h1>
            <div class="subtitle">{escape(product_name)}</div>
        </div>
        <div class="content">
            <div class="info-grid">
                <div class="info-item"><label>产品名称</label><span>{escape(product_name)}</span></div>
                <div class="info-item"><label>芯片型号</label><span>{escape(module)}</span></div>
                <div class="info-item"><label>固件版本</label><span>{escape(str(version))}</span></div>
                <div class="info-item"><label>语言</label><span>{escape(language)}</span></div>
                <div class="info-item"><label>测试时间</label><span>{escape(test_time)}</span></div>
                <div class="info-item"><label>串口配置</label><span>/dev/ttyACM4 + /dev/ttyACM0</span></div>
            </div>
            <div class="result-box{' fail' if conclusion != 'PASS' else ''}">
                <div class="pass-rate">{pass_rate}%</div>
                <div class="result-text">{escape(result_text)} ({pass_count}/{total_count})</div>
            </div>
            <div class="section-title">固件配置参数</div>
            <table class="config-table">
                <thead>
                    <tr><th>配置项</th><th>配置期望值</th><th>实测校验值</th></tr>
                </thead>
                <tbody>
{config_rows_html}
                </tbody>
            </table>
            <div class="section-title">测试用例详情</div>
            <table>
                <thead>
                    <tr><th>用例ID</th><th>测试类型</th><th>命令词</th><th>配置期望值</th><th>实测校验值</th><th>结果</th></tr>
                </thead>
                <tbody>
{chr(10).join(rows_html)}
                </tbody>
            </table>
            {note_html}
            <div class="section-title">附件说明</div>
            <div style="background: #f8f9fa; padding: 13px 15px; border-radius: 8px; border-left: 4px solid #667eea; font-size: 12px; color: #4a5568;">
                <p style="margin: 4px 0;"><strong>result.zip</strong> 包含：</p>
                <ul style="margin: 4px 0; padding-left: 18px;">
                    <li>testResult.xlsx - 测试结果Excel</li>
                    <li>serial_raw.log - 串口原始日志</li>
                    <li>test_*.log - 测试工具日志</li>
                    <li>testCases.csv - 本次测试用例</li>
                    <li>固件zip - 本次测试的固件包（{escape(str(version))}）</li>
                </ul>
            </div>
        </div>
        <div class="footer">此邮件由自动测试系统发送 | 测试结论：<strong>{escape(conclusion)}</strong></div>
    </div>
</body>
</html>
"""

    html_path = task_dir / "test_report.html"
    html_path.write_text(html_body, encoding="utf-8")
    print(f"[mars-belt] generate_email_report: HTML saved to {html_path}")

    compact_config_items = [
        f"{label}: {expected}"
        for label, expected, _actual in config_rows[:4]
    ]
    failed_items_html = "".join(
        f"<li>{escape(line)}</li>"
        for line in failed_case_lines[:6]
    ) or "<li>无异常用例</li>"
    note_items_html = "".join(f"<li>{escape(item)}</li>" for item in notes[:4]) or "<li>无额外备注</li>"
    mail_summary_body = f"""<!DOCTYPE html>
<html>
<head>
    <meta charset="UTF-8">
    <style>
        body {{ margin: 0; padding: 0; background: #08111f; color: #d8e3f0; font-family: 'Segoe UI', 'PingFang SC', 'Microsoft YaHei', sans-serif; }}
        .shell {{ padding: 24px 16px; background:
            radial-gradient(circle at top right, rgba(34,211,238,0.16), transparent 26%),
            radial-gradient(circle at top left, rgba(99,102,241,0.14), transparent 24%),
            #08111f; }}
        .panel {{ max-width: 780px; margin: 0 auto; background: rgba(10, 18, 34, 0.94); border: 1px solid rgba(148, 163, 184, 0.18); border-radius: 18px; overflow: hidden; }}
        .hero {{ padding: 24px 26px 18px; border-bottom: 1px solid rgba(148, 163, 184, 0.14); }}
        .hero small {{ color: #67e8f9; letter-spacing: 1px; text-transform: uppercase; }}
        .hero h1 {{ margin: 10px 0 8px; color: #f8fafc; font-size: 26px; }}
        .hero p {{ margin: 0; color: #94a3b8; font-size: 13px; }}
        .metrics {{ display: grid; grid-template-columns: repeat(3, 1fr); gap: 12px; margin-top: 18px; }}
        .metric {{ padding: 14px 16px; border-radius: 14px; background: rgba(15, 23, 42, 0.76); border: 1px solid rgba(148, 163, 184, 0.14); }}
        .metric label {{ display: block; color: #94a3b8; font-size: 11px; text-transform: uppercase; margin-bottom: 6px; }}
        .metric strong {{ color: #f8fafc; font-size: 22px; }}
        .metric span {{ display: block; margin-top: 6px; color: #cbd5e1; font-size: 12px; }}
        .content {{ padding: 22px 26px 28px; }}
        .section {{ margin-top: 22px; }}
        .section:first-child {{ margin-top: 0; }}
        .section h2 {{ margin: 0 0 10px; color: #f8fafc; font-size: 16px; }}
        .chips {{ margin-top: 10px; }}
        .chip {{ display: inline-block; margin: 0 8px 8px 0; padding: 6px 10px; border-radius: 999px; background: rgba(15, 23, 42, 0.9); border: 1px solid rgba(148, 163, 184, 0.14); color: #dbeafe; font-size: 12px; }}
        .card {{ padding: 14px 16px; border-radius: 14px; background: rgba(15, 23, 42, 0.78); border: 1px solid rgba(148, 163, 184, 0.14); }}
        .card p {{ margin: 0; color: #cbd5e1; font-size: 12px; line-height: 1.7; }}
        .card ul {{ margin: 8px 0 0 18px; color: #e2e8f0; font-size: 12px; line-height: 1.7; }}
        .footer {{ margin-top: 16px; padding: 12px 14px; border-radius: 12px; background: rgba(15, 23, 42, 0.82); border: 1px solid rgba(148, 163, 184, 0.14); color: #cbd5e1; font-size: 12px; }}
        @media (max-width: 720px) {{ .metrics {{ grid-template-columns: 1fr; }} }}
    </style>
</head>
<body>
    <div class="shell">
        <div class="panel">
            <div class="hero">
                <small>ListenAI Validation Mail Summary</small>
                <h1>Mars-Belt Test Result</h1>
                <p>{escape(product_name)} | {escape(module)} | {escape(language)} | {escape(str(version))}</p>
                <div class="metrics">
                    <div class="metric"><label>总用例</label><strong>{total_count}</strong><span>本轮执行总数</span></div>
                    <div class="metric"><label>通过率</label><strong>{pass_rate}%</strong><span>{escape(result_text)} ({pass_count}/{total_count})</span></div>
                    <div class="metric"><label>失败数</label><strong>{fail_count}</strong><span>{escape(test_time)}</span></div>
                </div>
            </div>
            <div class="content">
                <div class="section">
                    <h2>本轮配置与测试范围</h2>
                    <div class="chips">
                        {''.join(f'<span class="chip">{escape(item)}</span>' for item in compact_config_items)}
                    </div>
                    <div class="card"><p>正文只展示关键配置、结果和失败原因；完整用例明细、串口日志和 Excel 结果继续保留在附件中。</p></div>
                </div>
                <div class="section">
                    <h2>重点结果</h2>
                    <div class="card">
                        <ul>{failed_items_html}</ul>
                    </div>
                </div>
                <div class="section">
                    <h2>补充备注</h2>
                    <div class="card">
                        <ul>{note_items_html}</ul>
                    </div>
                </div>
                <div class="footer"><strong>附件说明</strong><br>`result.zip` 中仍保留 `testResult.xlsx`、`serial_raw.log`、`test_*.log`、`testCases.csv` 与固件包本体。</div>
            </div>
        </div>
    </div>
</body>
</html>
"""
    mail_summary_path = task_dir / "mail_summary.html"
    mail_summary_path.write_text(mail_summary_body, encoding="utf-8")
    print(f"[mars-belt] generate_email_report: Mail summary saved to {mail_summary_path}")

    zip_path = task_dir / "result.zip"
    with zipfile.ZipFile(zip_path, "w", zipfile.ZIP_DEFLATED) as zipf:
        for file_name in ("testResult.xlsx", "serial_raw.log", "test_report.html"):
            candidate = task_dir / file_name
            if candidate.exists():
                zipf.write(candidate, candidate.name)

        if test_cases:
            source_test_cases = suite_dir / "testCases.csv"
            if source_test_cases.exists():
                zipf.write(source_test_cases, "testCases.csv")

        if test_log_candidates:
            for log_path in test_log_candidates:
                zipf.write(log_path, log_path.name)
        elif (task_dir / "test_tool.log").exists():
            zipf.write(task_dir / "test_tool.log", "test_tool.log")

        for firmware_zip in sorted(task_dir.glob("*.zip")):
            if firmware_zip.name == "result.zip":
                continue
            zipf.write(firmware_zip, firmware_zip.name)

    print(f"[mars-belt] generate_email_report: Zip created at {zip_path}")
    print("[mars-belt] generate_email_report: done")


def add_repeat_arg(
    args: List[str],
    name: str,
    values: Optional[Sequence[str]],
    *,
    split_comma: bool = True,
) -> None:
    for value in split_repeat(values, split_comma=split_comma):
        args.extend([name, value])


def add_flag(args: List[str], name: str, enabled: bool) -> None:
    if enabled:
        args.append(name)


def run_python_script(
    script_name: str,
    arguments: Sequence[str],
    *,
    context: Optional[Dict[str, Any]] = None,
    command_name: str = "",
    use_dialout: bool = False,
) -> int:
    ensure_workspace()
    script_path = PY_ROOT / script_name
    if not script_path.exists():
        raise SystemExit(f"Missing bundled Python script: {script_path}")
    invoke_args = [sys.executable, "-X", "utf8", str(script_path), *arguments]
    
    # Use sg dialout for scripts that need serial port and audio access
    if use_dialout:
        cmd_str = " ".join(invoke_args)
        completed = subprocess.run(
            f'sg dialout -c "{cmd_str}"',
            cwd=str(SCRIPT_ROOT),
            check=False,
            shell=True,
            capture_output=True,
            text=True,
            encoding="utf-8",
            errors="replace",
        )
    else:
        completed = subprocess.run(
            invoke_args,
            cwd=str(SCRIPT_ROOT),
            check=False,
            capture_output=True,
            text=True,
            encoding="utf-8",
            errors="replace",
        )
    if completed.stdout:
        print(completed.stdout, end="")
    if completed.stderr:
        print(completed.stderr, end="", file=sys.stderr)
    if completed.returncode != 0:
        append_error_entry(
            command_name=command_name or script_name,
            cli_args=context,
            phenomenon=f"{script_name} exited with code {completed.returncode}",
            stdout_text=completed.stdout,
            stderr_text=completed.stderr,
            work_dir=str(SCRIPT_ROOT),
        )
    return int(completed.returncode)


def copy_directory_contents(source_dir: Path, destination_dir: Path) -> None:
    if not source_dir.exists():
        return
    destination_dir.mkdir(parents=True, exist_ok=True)
    for item in source_dir.iterdir():
        if item.is_symlink() and not item.exists():
            continue
        target = destination_dir / item.name
        if item.is_dir():
            shutil.copytree(item, target, dirs_exist_ok=True)
        elif item.is_file():
            shutil.copy2(item, target)


def sync_shared_audio_from_suite(suite_dir: Path) -> None:
    copy_directory_contents(suite_dir / "wavSource", WAV_ROOT)


def sync_shared_audio_to_suite(suite_dir: Path) -> None:
    copy_directory_contents(WAV_ROOT, suite_dir / "wavSource")


def text_value(value: Any) -> bool:
    return value is not None and str(value).strip() != ""


def module_hint(text: str) -> str:
    match = re.search(r"(?<!\d)(\d{4})(?!\d)", text or "")
    return match.group(1) if match else ""


def suite_runtime_state(suite_dir: Path) -> Dict[str, Any]:
    suite_dir = suite_dir.resolve()
    info_path = suite_dir / "deviceInfo_generated.json"
    wav_dir = suite_dir / "wavSource"
    state: Dict[str, Any] = {
        "suiteDir": suite_dir,
        "infoPath": info_path,
        "projectInfo": "",
        "hasDeviceInfo": False,
        "hasTts": False,
        "hasPorts": False,
        "hasWav": False,
    }
    if info_path.exists():
        state["hasDeviceInfo"] = True
        raw = info_path.read_text(encoding="utf-8", errors="ignore")
        try:
            payload = json.loads(raw)
            state["projectInfo"] = str(payload.get("projectInfo") or "")
            tts = payload.get("ttsConfig") or {}
            device = payload.get("deviceListInfo") or {}
            csk = device.get("cskApLog") or {}
            pretest = payload.get("pretestConfig") or {}
            state["hasTts"] = text_value(tts.get("app_id")) and text_value(tts.get("api_key"))
            state["hasPorts"] = text_value(csk.get("port")) and text_value(pretest.get("ctrlPort"))
        except Exception:
            project_match = re.search(r'"projectInfo"\s*:\s*"([^"]*)"', raw)
            app_id_match = re.search(r'"app_id"\s*:\s*"([^"]*)"', raw)
            api_key_match = re.search(r'"api_key"\s*:\s*"([^"]*)"', raw)
            port_match = re.search(r'"port"\s*:\s*"([^"]*)"', raw)
            ctrl_port_match = re.search(r'"ctrlPort"\s*:\s*"([^"]*)"', raw)
            state["projectInfo"] = project_match.group(1) if project_match else ""
            state["hasTts"] = text_value(app_id_match.group(1) if app_id_match else "") and text_value(api_key_match.group(1) if api_key_match else "")
            state["hasPorts"] = text_value(port_match.group(1) if port_match else "") and text_value(ctrl_port_match.group(1) if ctrl_port_match else "")
    if wav_dir.exists():
        state["hasWav"] = any(item.is_file() for item in wav_dir.iterdir())
    return state


def find_suite_runtime_donor(target_suite: Path) -> Optional[Dict[str, Any]]:
    target_state = suite_runtime_state(target_suite)
    project_info = target_state["projectInfo"]
    if not project_info or not RESULT_ROOT.exists():
        return None
    candidates = sorted(RESULT_ROOT.rglob("deviceInfo_generated.json"), key=lambda item: item.stat().st_mtime, reverse=True)
    hint = module_hint(str(target_state["suiteDir"]))
    fallback: Optional[Dict[str, Any]] = None
    for candidate in candidates:
        if candidate.resolve() == Path(target_state["infoPath"]).resolve():
            continue
        candidate_state = suite_runtime_state(candidate.parent)
        if (
            candidate_state["projectInfo"] == project_info
            and candidate_state["hasTts"]
            and candidate_state["hasPorts"]
            and candidate_state["hasWav"]
        ):
            if hint and re.search(rf"(?<!\d){re.escape(hint)}(?!\d)", str(candidate_state["suiteDir"])):
                return candidate_state
            if fallback is None:
                fallback = candidate_state
    return fallback


def ensure_suite_runtime_assets(suite_dir: Path) -> Optional[Dict[str, Any]]:
    target_state = suite_runtime_state(suite_dir)
    if not target_state["hasWav"]:
        sync_shared_audio_to_suite(suite_dir)
        target_state = suite_runtime_state(suite_dir)
    if target_state["hasTts"] and target_state["hasPorts"] and target_state["hasWav"]:
        return None
    donor = find_suite_runtime_donor(suite_dir)
    if donor is None:
        return None
    shutil.copy2(str(donor["infoPath"]), str(target_state["infoPath"]))
    target_wav_dir = Path(target_state["suiteDir"]) / "wavSource"
    target_wav_dir.mkdir(parents=True, exist_ok=True)
    copy_directory_contents(Path(donor["suiteDir"]) / "wavSource", target_wav_dir)
    return donor


def set_playback_volume(percent: int) -> int:
    target = max(0, min(100, int(percent)))
    try:
        from ctypes import POINTER, cast
        from comtypes import CLSCTX_ALL
        from pycaw.pycaw import AudioUtilities, IAudioEndpointVolume
    except Exception as exc:
        # Linux has no pycaw/comtypes - skip volume setting (not critical for burn/test)
        print(f"[mars-belt] volume setting skipped (Linux: {exc})")
        return 0

    devices = AudioUtilities.GetSpeakers()
    interface = devices.Activate(IAudioEndpointVolume._iid_, CLSCTX_ALL, None)
    volume = cast(interface, POINTER(IAudioEndpointVolume))
    volume.SetMute(0, None)
    volume.SetMasterVolumeLevelScalar(target / 100.0, None)
    print(round(volume.GetMasterVolumeLevelScalar() * 100))
    return 0


def burn_log_dir() -> Path:
    path = BURN_LOG_ROOT
    path.mkdir(parents=True, exist_ok=True)
    return path


def write_burn_log(log_file: Path, message: str) -> None:
    line = f"[{datetime.now().strftime('%H:%M:%S')}] {message}"
    print(line)
    with log_file.open("a", encoding="utf-8") as handle:
        handle.write(line + "\n")


def serial_module():
    try:
        import serial
        import serial.tools.list_ports
    except Exception as exc:
        raise SystemExit(f"burn requires pyserial: {exc}") from exc
    return serial


def available_ports() -> List[str]:
    serial = serial_module()
    return [item.device for item in serial.tools.list_ports.comports()]


def wait_port(port_name: str, attempts: int, delay_ms: int, log_file: Path) -> None:
    for _ in range(attempts):
        if port_name in available_ports():
            write_burn_log(log_file, f"Detected port {port_name}")
            return
        time.sleep(delay_ms / 1000.0)
    raise RuntimeError(f"Timeout waiting for port {port_name}")


def log_port_state(port_name: str, log_file: Path, label: str) -> None:
    state = "present" if port_name in available_ports() else "absent"
    write_burn_log(log_file, f"{label}: port {port_name} is {state}")


def open_serial_port(port_name: str, baud_rate: int, read_timeout: float = 0.5, write_timeout: float = 0.5):
    serial = serial_module()
    return serial.Serial(port_name, baud_rate, timeout=read_timeout, write_timeout=write_timeout)


def invoke_ctrl_sequence(
    commands: Sequence[str],
    ctrl_port: str,
    ctrl_baud: int,
    cmd_delay_ms: int,
    log_file: Path,
    prompt_timeout_ms: int = 2000,
) -> None:
    wait_port(ctrl_port, attempts=20, delay_ms=500, log_file=log_file)
    # Use sg dialout for serial port access (preserves audio environment)
    ctrl_script = BURN_ROOT / "sudo_ctrl.py"
    cmd_args = [
        "python3",
        str(ctrl_script),
        "send",
        ctrl_port,
        str(ctrl_baud),
        "--delay-ms",
        str(int(cmd_delay_ms or 0)),
        "--prompt-timeout-ms",
        str(int(prompt_timeout_ms or 0)),
        *list(commands),
    ]
    cmd_str = " ".join(shlex.quote(arg) for arg in cmd_args)
    write_burn_log(log_file, f"Invoke ctrl sequence via sg dialout: commands={commands} delayMs={cmd_delay_ms}")
    completed = subprocess.run(
        ["sg", "dialout", "-c", cmd_str],
        cwd=str(BURN_ROOT),
        stdout=subprocess.PIPE,
        stderr=subprocess.STDOUT,
        text=True,
        encoding="utf-8",
        errors="ignore",
        check=False,
    )
    output = completed.stdout or ""
    if output:
        for line in output.strip().splitlines():
            write_burn_log(log_file, line)
    if completed.returncode != 0:
        write_burn_log(log_file, f"Ctrl sequence failed with code {completed.returncode}")
        raise RuntimeError(f"Ctrl sequence failed: {output}")


def invoke_ctrl_command(
    command: str,
    ctrl_port: str,
    ctrl_baud: int,
    cmd_delay_ms: int,
    log_file: Path,
    prompt_timeout_ms: int = 2000,
) -> None:
    invoke_ctrl_sequence(
        [command],
        ctrl_port=ctrl_port,
        ctrl_baud=ctrl_baud,
        cmd_delay_ms=cmd_delay_ms,
        log_file=log_file,
        prompt_timeout_ms=prompt_timeout_ms,
    )


def wait_ctrl_prompt_ready(
    ctrl_port: str,
    ctrl_baud: int,
    log_file: Path,
    *,
    timeout_seconds: float = 15.0,
) -> None:
    deadline = time.time() + max(timeout_seconds, 1.0)
    while time.time() < deadline:
        try:
            wait_port(ctrl_port, attempts=1, delay_ms=200, log_file=log_file)
        except RuntimeError:
            time.sleep(0.2)
            continue
        try:
            with open_serial_port(ctrl_port, ctrl_baud, read_timeout=0.2, write_timeout=0.2) as port:
                time.sleep(0.2)
                try:
                    port.reset_input_buffer()
                except Exception:
                    pass
                try:
                    port.write(b"\r\n")
                    port.flush()
                except Exception:
                    pass
                captured: list[str] = []
                window_end = time.time() + 1.5
                while time.time() < window_end:
                    try:
                        waiting = port.in_waiting
                    except Exception:
                        waiting = 0
                    data = port.read(waiting or 256)
                    if data:
                        text = data.decode("utf-8", errors="ignore")
                        captured.append(text)
                        joined = "".join(captured)
                        if "root:/$" in joined or "root:/" in joined:
                            write_burn_log(log_file, f"Ctrl prompt detected on {ctrl_port}")
                            return
                    time.sleep(0.05)
        except Exception as exc:
            write_burn_log(log_file, f"Ctrl prompt wait retry on {ctrl_port}: {exc}")
        time.sleep(0.2)
    raise RuntimeError(f"Timeout waiting for ctrl prompt on {ctrl_port}")


def reboot_device_before_burn(
    ctrl_port: str,
    ctrl_baud: int,
    log_file: Path,
    *,
    prompt_timeout_seconds: float = 15.0,
) -> None:
    write_burn_log(log_file, "Pre-burn reboot via ctrl port")
    wait_port(ctrl_port, attempts=20, delay_ms=500, log_file=log_file)
    try:
        with open_serial_port(ctrl_port, ctrl_baud, read_timeout=0.2, write_timeout=0.2) as port:
            time.sleep(0.2)
            try:
                port.reset_input_buffer()
            except Exception:
                pass
            port.write(b"reboot\r\n")
            port.flush()
            write_burn_log(log_file, "CTRL SEND reboot")
            time.sleep(0.2)
            try:
                waiting = port.in_waiting
            except Exception:
                waiting = 0
            if waiting:
                raw = port.read(waiting).decode("utf-8", errors="ignore")
                if raw:
                    write_burn_log(log_file, f"CTRL RAW reboot: {raw!r}")
    except Exception as exc:
        raise RuntimeError(f"Failed to send reboot on ctrl port {ctrl_port}: {exc}") from exc
    time.sleep(1.0)
    wait_ctrl_prompt_ready(
        ctrl_port,
        ctrl_baud,
        log_file,
        timeout_seconds=prompt_timeout_seconds,
    )


def enter_burn_mode(
    ctrl_port: str,
    ctrl_baud: int,
    burn_port: str,
    cmd_delay_ms: int,
    burn_mode_wait_ms: int,
    log_file: Path,
    *,
    pre_burn_reboot: bool = False,
) -> None:
    write_burn_log(log_file, "Enter burn mode")
    if pre_burn_reboot:
        reboot_device_before_burn(ctrl_port, ctrl_baud, log_file)
    else:
        write_burn_log(log_file, "Skip pre-burn reboot; use stable 4-step burn entry flow")
    wait_port(burn_port, attempts=20, delay_ms=500, log_file=log_file)
    write_burn_log(
        log_file,
        "Burn mode 4-step sequence: uut-switch1.off -> uut-switch2.on -> uut-switch1.on -> uut-switch2.off",
    )
    invoke_ctrl_sequence(
        ["uut-switch1.off", "uut-switch2.on", "uut-switch1.on", "uut-switch2.off"],
        ctrl_port=ctrl_port,
        ctrl_baud=ctrl_baud,
        cmd_delay_ms=cmd_delay_ms,
        log_file=log_file,
    )
    log_port_state(burn_port, log_file, "After burn-entry sequence")
    wait_port(burn_port, attempts=20, delay_ms=300, log_file=log_file)
    time.sleep(burn_mode_wait_ms / 1000.0)


def exit_burn_mode(
    ctrl_port: str,
    ctrl_baud: int,
    burn_port: str,
    cmd_delay_ms: int,
    boot_wait_seconds: int,
    log_file: Path,
) -> None:
    write_burn_log(log_file, "Exit burn mode and restore power")
    write_burn_log(
        log_file,
        "Restore 3-step sequence: uut-switch2.off -> uut-switch1.off -> uut-switch1.on",
    )
    invoke_ctrl_sequence(
        ["uut-switch2.off", "uut-switch1.off", "uut-switch1.on"],
        ctrl_port=ctrl_port,
        ctrl_baud=ctrl_baud,
        cmd_delay_ms=cmd_delay_ms,
        log_file=log_file,
    )
    log_port_state(burn_port, log_file, "After restore sequence")
    wait_port(burn_port, attempts=20, delay_ms=300, log_file=log_file)
    time.sleep(boot_wait_seconds)


def compute_file_sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        while True:
            chunk = handle.read(1024 * 1024)
            if not chunk:
                break
            digest.update(chunk)
    return digest.hexdigest()


def validate_firmware_file(path: Path, log_file: Path, *, origin: str, package_entry: str = "") -> Path:
    resolved = path.resolve()
    if not resolved.exists():
        raise RuntimeError(f"{origin} firmware file does not exist: {resolved}")
    size = resolved.stat().st_size
    if size <= 0:
        raise RuntimeError(f"{origin} firmware file is empty: {resolved}")
    sha256 = compute_file_sha256(resolved)
    extra = f" entry={package_entry}" if package_entry else ""
    write_burn_log(log_file, f"Firmware preflight OK: origin={origin}{extra} path={resolved} size={size} sha256={sha256}")
    return resolved


def clear_staged_firmware(staged_fw: Path, log_file: Path) -> None:
    if staged_fw.exists():
        staged_fw.unlink()
        write_burn_log(log_file, f"Removed existing staged firmware: {staged_fw}")


def stage_explicit_firmware(source_path: Path, staged_fw: Path, log_file: Path) -> Path:
    resolved_source = validate_firmware_file(source_path, log_file, origin="explicit-source")
    resolved_staged = staged_fw.resolve()
    if resolved_source == resolved_staged:
        write_burn_log(log_file, f"Explicit firmware already staged as app.bin: {resolved_staged}")
        return validate_firmware_file(resolved_staged, log_file, origin="explicit-staged")
    clear_staged_firmware(staged_fw, log_file)
    shutil.copyfile(resolved_source, staged_fw)
    staged_path = validate_firmware_file(staged_fw, log_file, origin="explicit-staged")
    write_burn_log(log_file, f"Copied explicit firmware to staged app.bin: source={resolved_source} target={staged_path}")
    return staged_path


def extract_firmware_from_package(zip_path: Path, staged_fw: Path, log_file: Path) -> bool:
    if not zip_path.exists():
        raise RuntimeError(f"Package zip does not exist: {zip_path}")
    write_burn_log(log_file, f"Extract fw.bin from package: {zip_path}")
    clear_staged_firmware(staged_fw, log_file)
    with zipfile.ZipFile(zip_path) as archive:
        entry_name = next((name for name in archive.namelist() if re.search(r"(^|/)Standard_product/fw\.bin$", name)), "")
        if not entry_name:
            raise RuntimeError("Standard_product/fw.bin was not found in package")
        entry_info = archive.getinfo(entry_name)
        if entry_info.file_size <= 0:
            raise RuntimeError(f"Package firmware entry is empty: {entry_name}")
        with archive.open(entry_name) as source, staged_fw.open("wb") as target:
            shutil.copyfileobj(source, target)
    validate_firmware_file(staged_fw, log_file, origin="package", package_entry=entry_name)
    write_burn_log(log_file, f"Prepared staged firmware: {staged_fw}")
    return True


def resolve_firmware_path(package_zip: str, firmware_bin: str, staged_fw: Path, log_file: Path) -> tuple[Path, bool]:
    if firmware_bin:
        resolved = stage_explicit_firmware(Path(firmware_bin), staged_fw, log_file)
        write_burn_log(log_file, f"Use staged firmware: {resolved}")
        return resolved, False
    if package_zip:
        extract_firmware_from_package(Path(package_zip).resolve(), staged_fw, log_file)
        return staged_fw, True
    if staged_fw.exists():
        resolved = validate_firmware_file(staged_fw, log_file, origin="default-staged")
        write_burn_log(log_file, f"Use default staged firmware: {resolved}")
        return resolved, False
    raise RuntimeError("No --package-zip or --firmware-bin was provided and burn/app.bin does not exist")


def diagnose_burn_failure(output_path: Path, exit_code: int) -> tuple[bool, str, str]:
    markers = [
        "CONNECT ROM AND DOWNLOAD RAM LOADER SUCCESS",
        "SEND MD5 COMMAND WITH RAM SUCCESS",
        "SEND END COMMAND SUCCESS",
    ]
    content = output_path.read_text(encoding="utf-8", errors="ignore") if output_path.exists() else ""
    if markers[0] not in content and "RECEIVE OVERTIME" in content:
        detail = "ROM handshake failed before RAM loader download (missing CONNECT ROM..., saw RECEIVE OVERTIME)"
        if exit_code != 0:
            detail = f"{detail}; burn tool exit code: {exit_code}"
        return False, detail, content
    if exit_code != 0:
        return False, f"Burn tool returned non-zero exit code: {exit_code}", content
    if not content.strip():
        return False, "Burn tool produced no output", content
    for marker in markers:
        if marker not in content:
            if marker == markers[0] and "RECEIVE OVERTIME" in content:
                return False, "ROM handshake failed before RAM loader download (missing CONNECT ROM..., saw RECEIVE OVERTIME)", content
            return False, f"Missing success marker: {marker}", content
    return True, "", content


def log_burn_output_tail(content: str, log_file: Path, *, limit: int = 20) -> None:
    lines = [line for line in content.splitlines() if line.strip()]
    if not lines:
        return
    for line in lines[-limit:]:
        write_burn_log(log_file, f"[burn-tail] {line}")


def invoke_burn_tool(tool_path: Path, fw_path: Path, burn_port: str, baud: int, log_file: Path) -> None:
    wait_port(burn_port, attempts=20, delay_ms=500, log_file=log_file)
    tool_output = burn_log_dir() / f"burn_tool_{datetime.now().strftime('%Y%m%d_%H%M%S')}.log"
    staged_fw = (BURN_ROOT / "app.bin").resolve()
    if fw_path.resolve() != staged_fw:
        raise RuntimeError(f"Refuse to burn non-staged firmware: {fw_path}; expected {staged_fw}")
    args = [
        f"./{tool_path.name}",
        "-b",
        str(baud),
        "-p",
        burn_port,
        "-f",
        "app.bin",
        "-m",
        "-d",
        "-a",
        "0x0",
        "-i",
        "adaptive-duplex",
        "-s",
    ]
    write_burn_log(log_file, "Run burn tool: " + " ".join(args))
    completed = subprocess.run(
        args,
        cwd=str(BURN_ROOT),
        stdout=subprocess.PIPE,
        stderr=subprocess.STDOUT,
        text=True,
        encoding="utf-8",
        errors="ignore",
        check=False,
    )
    output = completed.stdout or ""
    if output:
        print(output, end="" if output.endswith("\n") else "\n")
        tool_output.write_text(output, encoding="utf-8")
        with log_file.open("a", encoding="utf-8") as handle:
            handle.write(output)
            if not output.endswith("\n"):
                handle.write("\n")
    ok, failure_reason, content = diagnose_burn_failure(tool_output, completed.returncode)
    if not ok:
        write_burn_log(log_file, f"Burn tool failure reason: {failure_reason}")
        log_burn_output_tail(content, log_file)
        raise RuntimeError(f"{failure_reason}. Check {tool_output}")
    write_burn_log(log_file, "Burn tool reported success")


def set_device_loglevel(runtime_log_port: str, runtime_log_baud: int, skip_loglevel: bool, log_file: Path) -> None:
    if skip_loglevel:
        write_burn_log(log_file, "Skip setting loglevel 4")
        return
    # Use sg dialout for serial port access (preserves audio environment)
    ctrl_script = BURN_ROOT / "sudo_ctrl.py"
    cmd_args = ["python3", str(ctrl_script), "setloglevel", runtime_log_port, str(int(runtime_log_baud or 115200))]
    cmd_str = " ".join(cmd_args)
    write_burn_log(log_file, f"Set device loglevel via sg dialout port={runtime_log_port} baud={runtime_log_baud}")
    completed = subprocess.run(
        f'sg dialout -c "{cmd_str}"',
        cwd=str(BURN_ROOT),
        shell=True,
        stdout=subprocess.PIPE,
        stderr=subprocess.STDOUT,
        text=True,
        encoding="utf-8",
        errors="ignore",
        check=False,
    )
    output = completed.stdout or ""
    if output:
        for line in output.strip().split("\n"):
            write_burn_log(log_file, line)


def extract_version_from_package(package_zip: Path) -> Optional[str]:
    """从 package zip 的 web_config.json 中提取 version 字段"""
    try:
        with zipfile.ZipFile(package_zip) as archive:
            entry_name = next((name for name in archive.namelist()
                               if re.search(r"Standard_product/web_config\.json$", name)), "")
            if not entry_name:
                return None
            with archive.open(entry_name) as f:
                web_cfg = json.load(f)
                return web_cfg.get("version")
    except Exception:
        return None


def verify_firmware_version(runtime_log_port: str, runtime_log_baud: int, expected_version: str, log_file: Path) -> tuple[bool, str]:
    """
    烧录完成后校验固件版本：
    1. 捕获设备上电后的启动日志
    2. 从日志中提取 config version
    3. 与 expected_version 比对（设备端显示截断至分钟级，比对前缀）
    返回 (成功标志, 详情描述)
    """
    if not expected_version:
        return True, "无 expected_version，跳过校验"

    wait_port(runtime_log_port, attempts=20, delay_ms=500, log_file=log_file)
    port = None
    try:
        port = open_serial_port(runtime_log_port, int(runtime_log_baud or 115200), read_timeout=0.2, write_timeout=0.2)
        time.sleep(1)

        # 清空缓冲区
        port.read_all()

        # 等待 boot log 出现（最长15秒）
        boot_output = []
        start = time.time()
        while time.time() - start < 15:
            data = port.read(port.in_waiting)
            if data:
                try:
                    text = data.decode("utf-8", errors="replace")
                    boot_output.append(text)
                    # 一旦看到 "config version:" 说明版本行已输出
                    if "config version:" in text:
                        break
                except Exception:
                    pass
            time.sleep(0.1)

        full_log = "".join(boot_output)

        # 从 boot log 中提取 config version
        # 格式示例: "config version: V-2026.03.30_17.."
        match = re.search(r"config version:\s*(V-\d+\.\d+\.\d+_\d+)", full_log)
        if not match:
            # 尝试备用匹配（更宽松的格式）
            match = re.search(r"config version:\s*(\S+)", full_log)

        if not match:
            write_burn_log(log_file, f"版本校验：未在启动日志中找到 config version，整行：{full_log[:200]}")
            return False, f"未找到 config version in boot log (前200字: {full_log[:200]})"

        device_version = match.group(1).strip()
        write_burn_log(log_file, f"版本校验：设备 config version = {device_version}")
        write_burn_log(log_file, f"版本校验：期望 version = {expected_version}")

        # 设备显示截断规则：V-YYYY.MM.DD_HH.MM.SS → V-YYYY.MM.DD_HH.MM..
        # 因此取 expected_version 的前17字符作为比对前缀
        # V-2026.03.30_17.29.33 → 前17位 → "V-2026.03.30_17"
        expected_prefix = expected_version[:17]
        if device_version.startswith(expected_prefix):
            write_burn_log(log_file, f"版本校验：✅ PASS (前缀匹配 {expected_prefix})")
            return True, f"PASS (设备: {device_version}, 期望: {expected_version}, 前缀: {expected_prefix})"
        else:
            write_burn_log(log_file, f"版本校验：❌ FAIL (设备: {device_version}, 期望前缀: {expected_prefix})")
            return False, f"FAIL (设备: {device_version}, 期望: {expected_version})"

    except Exception as exc:
        write_burn_log(log_file, f"版本校验异常: port={runtime_log_port} baud={runtime_log_baud} exc={exc}")
        return False, str(exc)
    finally:
        if port is not None and port.is_open:
            port.close()


def run_burn(args: argparse.Namespace) -> int:
    ensure_workspace()
    args.max_retry = max(args.max_retry, 1)
    tool_path = BURN_ROOT / "Uart_Burn_Tool"
    staged_fw = BURN_ROOT / "app.bin"
    task_dir = infer_task_dir(package_zip=args.package_zip)
    if not tool_path.exists():
        raise SystemExit(f"Missing burn tool: {tool_path}")
    log_file = burn_log_dir() / f"burn_{datetime.now().strftime('%Y%m%d_%H%M%S')}.log"
    fw_path, extracted = resolve_firmware_path(args.package_zip, args.firmware_bin, staged_fw, log_file)

    # ========== 烧录后版本号校验（阶段3改进）==========
    # 从 package zip 提取期望版本号
    expected_version: Optional[str] = None
    if args.package_zip:
        pkg_path = Path(args.package_zip).resolve()
        expected_version = extract_version_from_package(pkg_path)
        if expected_version:
            write_burn_log(log_file, f"期望固件版本: {expected_version}")
        else:
            write_burn_log(log_file, "未从 package zip 中提取到 version，跳过版本校验")
    # =================================================

    runtime_log_port = str(getattr(args, "runtime_log_port", "") or args.burn_port).strip()
    runtime_log_baud = int(getattr(args, "runtime_log_baud", 115200) or 115200)

    write_burn_log(log_file, "========== Burn flow start ==========")
    write_burn_log(
        log_file,
        f"CtrlPort={args.ctrl_port} BurnPort={args.burn_port} Baud={args.baud} RuntimeLogPort={runtime_log_port} RuntimeLogBaud={runtime_log_baud} Firmware={fw_path}",
    )
    try:
        for attempt in range(1, args.max_retry + 1):
            try:
                write_burn_log(log_file, f"Burn attempt {attempt}/{args.max_retry}")
                enter_burn_mode(
                    args.ctrl_port,
                    args.ctrl_baud,
                    args.burn_port,
                    args.cmd_delay_ms,
                    args.burn_mode_wait_ms,
                    log_file,
                    pre_burn_reboot=args.pre_burn_reboot,
                )
                invoke_burn_tool(tool_path, fw_path, args.burn_port, args.baud, log_file)
                exit_burn_mode(
                    args.ctrl_port,
                    args.ctrl_baud,
                    args.burn_port,
                    args.cmd_delay_ms,
                    args.boot_wait_seconds,
                    log_file,
                )
                set_device_loglevel(runtime_log_port, runtime_log_baud, args.skip_loglevel, log_file)

                # ========== 版本号校验（接在 set_device_loglevel 后）==========
                if expected_version:
                    write_burn_log(log_file, "开始校验固件版本...")
                    ver_ok, ver_detail = verify_firmware_version(runtime_log_port, runtime_log_baud, expected_version, log_file)
                    if not ver_ok:
                        write_burn_log(log_file, f"❌ 版本校验失败: {ver_detail}")
                        if task_dir is not None:
                            copy_if_exists(log_file, task_dir / "burn.log")
                        return 2  # 版本校验失败专属退出码
                # ==============================================================

                write_burn_log(log_file, "Burn flow completed")
                if task_dir is not None:
                    copy_if_exists(log_file, task_dir / "burn.log")
                return 0
            except Exception as exc:
                write_burn_log(log_file, f"Attempt failed: {exc}")
                if attempt >= args.max_retry:
                    raise
                write_burn_log(log_file, "Retry after power restore")
                try:
                    exit_burn_mode(
                        args.ctrl_port,
                        args.ctrl_baud,
                        args.burn_port,
                        args.cmd_delay_ms,
                        args.boot_wait_seconds,
                        log_file,
                    )
                except Exception as restore_exc:
                    write_burn_log(log_file, f"Power restore also failed: {restore_exc}")
                time.sleep(2)
    except Exception as exc:
        write_burn_log(log_file, f"Burn flow failed: {exc}")
        return 1
    finally:
        if extracted and not args.keep_extracted and staged_fw.exists():
            staged_fw.unlink()
            write_burn_log(log_file, f"Removed staged firmware: {staged_fw}")


def copy_relative_item(source_root: Path, relative_path: str, destination_root: Path) -> None:
    source = source_root / relative_path
    if not source.exists():
        return
    destination = destination_root / relative_path
    destination.parent.mkdir(parents=True, exist_ok=True)
    if source.is_dir():
        shutil.copytree(source, destination, dirs_exist_ok=True)
    else:
        shutil.copy2(source, destination)


def copy_absolute_item(source: Path, destination: Path) -> None:
    if not source.exists():
        return
    destination.parent.mkdir(parents=True, exist_ok=True)
    if source.is_dir():
        shutil.copytree(source, destination, dirs_exist_ok=True)
    else:
        shutil.copy2(source, destination)


def backup_state_once(root: Path, memory_root: Path) -> None:
    timestamp = datetime.now().strftime("%Y%m%d-%H%M%S")
    backup_root = memory_root / "backups"
    latest_root = backup_root / "latest"
    snapshot_root = backup_root / timestamp
    state_source = None
    for candidate in (
        RESULT_ROOT / "listenai_firmware_packaging_state.md",
        SCRIPT_ROOT / "result" / "listenai_firmware_packaging_state.md",
        root / "scripts" / "result" / "listenai_firmware_packaging_state.md",
    ):
        if candidate.exists():
            state_source = candidate
            break
    memory_state = memory_root / "listenai-firmware-packaging-state.md"
    memory_latest = memory_root / "listenai-firmware-packaging-latest.json"
    manifest_path = backup_root / "latest-backup.json"
    include_items = [
        (root / "SKILL.md", "SKILL.md"),
        (SCRIPT_ROOT / "py", "scripts/py"),
        (BURN_ROOT, "scripts/burn"),
        (RESULT_ROOT, "artifacts/tasks"),
        (PACKAGE_ROOT, "artifacts/package"),
        (RUNTIME_ROOT, "artifacts/runtime"),
        (WAV_ROOT, "scripts/wavSource"),
    ]

    backup_root.mkdir(parents=True, exist_ok=True)
    snapshot_root.mkdir(parents=True, exist_ok=True)
    memory_root.mkdir(parents=True, exist_ok=True)
    if latest_root.exists():
        shutil.rmtree(latest_root)
    latest_root.mkdir(parents=True, exist_ok=True)

    for source, relative_path in include_items:
        copy_absolute_item(source, snapshot_root / relative_path)
        copy_absolute_item(source, latest_root / relative_path)

    if state_source and state_source.exists():
        shutil.copy2(state_source, memory_state)

    manifest = {
        "timestamp": datetime.now().isoformat(timespec="seconds"),
        "root": str(root),
        "snapshot": str(snapshot_root),
        "latest": str(latest_root),
        "memoryState": str(memory_state),
        "includeCount": len(include_items),
    }
    manifest_json = json.dumps(manifest, ensure_ascii=False, indent=2)
    manifest_path.write_text(manifest_json, encoding="utf-8")
    memory_latest.write_text(manifest_json, encoding="utf-8")
    print(f"backup ok: {snapshot_root}")


def forward_list_catalog(args: argparse.Namespace) -> int:
    set_token(args.token)
    json_out = args.json_out or str(runtime_file("catalog", "listenai_product_options.json"))
    command: List[str] = []
    add_repeat_arg(command, "--product", args.product)
    add_repeat_arg(command, "--scene", args.scene)
    add_path_arg(command, "--json-out", json_out)
    return run_python_script("listenai_product_options.py", command, context=vars(args), command_name="list-catalog")


def forward_resolve(args: argparse.Namespace) -> int:
    if args.action == "package":
        set_token(args.token)
        command: List[str] = []
        add_arg(command, "--product", args.product)
        add_arg(command, "--module", args.module)
        add_arg(command, "--language", args.language)
        add_arg(command, "--version", args.version)
        add_arg(command, "--scene", args.scene)
        add_arg(command, "--source-release-id", args.source_release_id)
        add_arg(command, "--product-name", args.product_name)
        add_arg(command, "--timeout-sec", args.timeout_sec)
        add_flag(command, "--refresh-live", args.refresh_live)
        return run_python_script("listenai_custom_package.py", command, context=vars(args), command_name="resolve(package)")
    if args.action in {"prepare", "package"} or args.refresh_live:
        set_token(args.token)
    elif args.token:
        set_token(args.token, allow_missing=True)
    command: List[str] = []
    resolve_dir = ensure_runtime_dir("resolve", "catalog")
    add_arg(command, "--action", args.action)
    add_arg(command, "--product", args.product)
    add_arg(command, "--module", args.module)
    add_arg(command, "--language", args.language)
    add_arg(command, "--version", args.version)
    add_arg(command, "--scene", args.scene)
    add_arg(command, "--product-name", args.product_name)
    add_arg(command, "--source-release-id", args.source_release_id)
    add_arg(command, "--timeout-sec", args.timeout_sec)
    add_flag(command, "--refresh-live", args.refresh_live)
    add_path_arg(command, "--json-out", args.json_out or str(resolve_dir / "listenai_product_options.json"))
    add_path_arg(command, "--products-csv-out", args.products_csv_out or str(resolve_dir / "listenai_product_catalog.csv"))
    add_path_arg(command, "--modules-csv-out", args.modules_csv_out or str(resolve_dir / "listenai_module_catalog.csv"))
    add_path_arg(command, "--matrix-csv-out", args.matrix_csv_out or str(resolve_dir / "listenai_product_options_matrix.csv"))
    add_path_arg(command, "--duplicates-csv-out", args.duplicates_csv_out or str(resolve_dir / "listenai_version_defid_duplicates.csv"))
    add_path_arg(command, "--matrix-md-out", args.matrix_md_out or str(resolve_dir / "listenai_product_options_matrix.md"))
    add_path_arg(command, "--resolution-out", args.resolution_out or str(resolve_dir / "listenai_resolved_product.json"))
    add_path_arg(command, "--summary-out", args.summary_out or str(resolve_dir / "listenai_selected_package_summary.json"))
    return run_python_script("listenai_resolve_and_package.py", command, context=vars(args), command_name="resolve")


def _contains_algo_overrides(overrides: Sequence[str]) -> bool:
    for item in overrides or []:
        text = str(item or "")
        if text.startswith("releaseAlgoList") or text.startswith("releaseDepthList"):
            return True
    return False


def forward_package_custom(args: argparse.Namespace) -> int:
    set_token(args.token)
    command: List[str] = []
    add_arg(command, "--product", args.product)
    add_arg(command, "--module", args.module)
    add_arg(command, "--language", args.language)
    add_arg(command, "--version", args.version)
    add_arg(command, "--scene", args.scene)
    add_path_arg(command, "--catalog-json", args.catalog_json)
    add_arg(command, "--source-release-id", args.source_release_id)
    add_arg(command, "--product-name", args.product_name)
    add_arg(command, "--run-id", args.run_id)
    add_path_arg(command, "--result-root", args.result_root)
    add_path_arg(command, "--package-root", args.package_root)
    add_arg(command, "--timeout-sec", args.timeout_sec)
    add_arg(command, "--retry-limit", args.retry_limit)

    # Keep base packaging untouched unless the task explicitly asks for algorithm word changes.
    if getattr(args, "enable_algo_words", False) and _contains_algo_overrides(args.override):
        add_repeat_arg(command, "--override", args.override, split_comma=False)
    elif _contains_algo_overrides(args.override):
        print("[mars-belt] Skip releaseAlgoList/releaseDepthList overrides because --enable-algo-words was not set.")
        safe_overrides = [item for item in (args.override or []) if not str(item or "").startswith(("releaseAlgoList", "releaseDepthList"))]
        add_repeat_arg(command, "--override", safe_overrides, split_comma=False)
    else:
        add_repeat_arg(command, "--override", args.override, split_comma=False)

    add_repeat_arg(command, "--require-feature", args.require_feature)
    add_arg(command, "--comments", args.comments)
    add_flag(command, "--refresh-live", args.refresh_live)
    add_flag(command, "--dry-run", args.dry_run)
    return run_python_script("listenai_custom_package.py", command, context=vars(args), command_name="package-custom")


def forward_generate_algo_words(args: argparse.Namespace) -> int:
    command: List[str] = []
    add_path_arg(command, "--base-json", args.base_json)
    add_path_arg(command, "--output", args.output)
    add_arg(command, "--append-count", args.append_count)
    add_arg(command, "--proto-start", args.proto_start)
    return run_python_script("listenai_generate_algo_words.py", command, context=vars(args), command_name="generate-algo-words")


def forward_package_voice_reg(args: argparse.Namespace) -> int:
    set_token(args.token)
    command: List[str] = []
    add_arg(command, "--product", args.product)
    add_arg(command, "--module", args.module)
    add_arg(command, "--language", args.language)
    add_arg(command, "--version", args.version)
    add_arg(command, "--scene", args.scene)
    add_path_arg(command, "--catalog-json", args.catalog_json)
    add_arg(command, "--source-release-id", args.source_release_id)
    add_arg(command, "--product-name", args.product_name)
    add_arg(command, "--run-id", args.run_id)
    add_path_arg(command, "--result-root", args.result_root)
    add_path_arg(command, "--package-root", args.package_root)
    add_arg(command, "--timeout-sec", args.timeout_sec)
    add_arg(command, "--retry-limit", args.retry_limit)
    add_arg(command, "--timeout-value", args.timeout_value)
    add_arg(command, "--vol-level", args.vol_level)
    add_arg(command, "--algo-view-mode", args.algo_view_mode)
    add_repeat_arg(command, "--study-reg-command", args.study_reg_command, split_comma=False)
    add_repeat_arg(command, "--require-feature", args.require_feature)
    add_arg(command, "--comments", args.comments)
    add_flag(command, "--refresh-live", args.refresh_live)
    add_flag(command, "--dry-run", args.dry_run)
    return run_python_script("listenai_custom_voice_reg_package.py", command, context=vars(args), command_name="package-voice-reg")


def forward_package_weekly(args: argparse.Namespace) -> int:
    set_token(args.token)
    command: List[str] = []
    add_arg(command, "--product", args.product)
    add_arg(command, "--module", args.module)
    add_arg(command, "--language", args.language)
    add_arg(command, "--version", args.version)
    add_arg(command, "--scene", args.scene)
    add_arg(command, "--source-release-id", args.source_release_id)
    add_arg(command, "--timeout-sec", args.timeout_sec)
    add_arg(command, "--log-port", args.log_port)
    add_arg(command, "--protocol-port", getattr(args, "protocol_port", ""))
    add_arg(command, "--uart0-port", getattr(args, "uart0_port", ""))
    add_arg(command, "--uart1-port", getattr(args, "uart1_port", ""))
    add_arg(command, "--ctrl-port", args.ctrl_port)
    add_arg(command, "--burn-port", getattr(args, "burn_port", ""))
    add_arg(command, "--task-dir", getattr(args, "task_dir", ""))
    add_arg(command, "--variants", getattr(args, "variants", ""))
    add_flag(command, "--refresh-live", args.refresh_live)
    add_flag(command, "--update-audio-skills", getattr(args, "update_audio_skills", False))
    add_flag(command, "--skip-device", args.skip_device)
    return run_python_script("listenai_weekly_validation_runner.py", command, context=vars(args), command_name="package-weekly")


def forward_package_batch(args: argparse.Namespace) -> int:
    set_token(args.token)
    command: List[str] = []
    add_path_arg(command, "--catalog-json", args.catalog_json)
    add_path_arg(command, "--test-catalog-json", args.test_catalog_json)
    add_path_arg(command, "--batch-root", args.batch_root)
    add_path_arg(command, "--package-root", args.package_root)
    add_arg(command, "--batch-id", args.batch_id)
    add_arg(command, "--timeout-sec", args.timeout_sec)
    add_flag(command, "--dry-run", args.dry_run)
    return run_python_script("listenai_batch_package_parameters.py", command, context=vars(args), command_name="package-batch")


def forward_package_grouped(args: argparse.Namespace) -> int:
    set_token(args.token)
    command: List[str] = []
    add_path_arg(command, "--catalog-json", args.catalog_json)
    add_path_arg(command, "--test-catalog-json", args.test_catalog_json)
    add_path_arg(command, "--batch-root", args.batch_root)
    add_path_arg(command, "--package-root", args.package_root)
    add_arg(command, "--batch-id", args.batch_id)
    add_arg(command, "--timeout-sec", args.timeout_sec)
    add_flag(command, "--dry-run", args.dry_run)
    return run_python_script("listenai_grouped_product_package.py", command, context=vars(args), command_name="package-grouped")


def forward_build_case_catalog(args: argparse.Namespace) -> int:
    task_dir = infer_task_dir(package_zip=args.package_zip)
    default_dir = runtime_dir_for_task(task_dir, "catalog") if task_dir is not None else ensure_runtime_dir("catalog", "adhoc")
    if args.package_zip or args.web_config or args.profile != "parameter":
        command: List[str] = ["--catalog-only"]
        add_path_arg(command, "--package-zip", args.package_zip)
        add_path_arg(command, "--web-config", args.web_config)
        add_arg(command, "--profile", "auto" if args.profile == "parameter" else args.profile)
        add_path_arg(command, "--json-out", args.json_out or str(default_dir / "test_case_catalog.json"))
        add_path_arg(command, "--md-out", args.md_out or str(default_dir / "test_case_catalog.md"))
        add_repeat_arg(command, "--override", args.override, split_comma=False)
        return run_python_script("listenai_profile_suite.py", command, context=vars(args), command_name="build-case-catalog(profile)")

    set_token(args.token)
    command = []
    add_arg(command, "--product", args.product)
    add_arg(command, "--module", args.module)
    add_arg(command, "--language", args.language)
    add_arg(command, "--version", args.version)
    add_arg(command, "--scene", args.scene)
    add_arg(command, "--source-release-id", args.source_release_id)
    add_flag(command, "--refresh-live", args.refresh_live)
    add_path_arg(command, "--md-out", args.md_out or str(default_dir / "listenai_parameter_catalog.md"))
    add_path_arg(command, "--json-out", args.json_out or str(default_dir / "listenai_parameter_catalog.json"))
    add_path_arg(command, "--json-out-catalog", args.json_out_catalog or str(default_dir / "listenai_product_options.json"))
    add_path_arg(command, "--products-csv-out", args.products_csv_out or str(default_dir / "listenai_product_catalog.csv"))
    add_path_arg(command, "--modules-csv-out", args.modules_csv_out or str(default_dir / "listenai_module_catalog.csv"))
    add_path_arg(command, "--matrix-csv-out", args.matrix_csv_out or str(default_dir / "listenai_product_options_matrix.csv"))
    add_path_arg(command, "--duplicates-csv-out", args.duplicates_csv_out or str(default_dir / "listenai_version_defid_duplicates.csv"))
    add_path_arg(command, "--matrix-md-out", args.matrix_md_out or str(default_dir / "listenai_product_options_matrix.md"))
    add_path_arg(command, "--resolution-out", args.resolution_out or str(default_dir / "listenai_resolved_product.json"))
    add_path_arg(command, "--summary-out", args.summary_out or str(default_dir / "listenai_selected_package_summary.json"))
    return run_python_script("listenai_test_case_catalog.py", command, context=vars(args), command_name="build-case-catalog")


def forward_generate_suite(args: argparse.Namespace) -> int:
    task_dir = infer_task_dir(package_zip=args.package_zip)
    default_out_dir = runtime_dir_for_task(task_dir, "suite_cli") if task_dir is not None else ensure_runtime_dir("suite_cli", "adhoc")
    if args.profile != "parameter" or not args.test_catalog_json:
        command: List[str] = []
        add_path_arg(command, "--package-zip", args.package_zip)
        add_path_arg(command, "--web-config", args.web_config)
        add_arg(command, "--profile", "auto" if args.profile == "parameter" else args.profile)
        add_path_arg(command, "--out-dir", args.out_dir or str(default_out_dir))
        add_arg(command, "--port", args.port)
        add_arg(command, "--ctrl-port", args.ctrl_port)
        add_repeat_arg(command, "--override", args.override, split_comma=False)
        return run_python_script("listenai_profile_suite.py", command, context=vars(args), command_name="generate-suite(profile)")

    command = []
    add_path_arg(command, "--test-catalog-json", args.test_catalog_json)
    add_path_arg(command, "--package-zip", args.package_zip)
    add_path_arg(command, "--web-config", args.web_config)
    add_path_arg(command, "--out-dir", args.out_dir or str(default_out_dir))
    add_arg(command, "--port", args.port)
    add_arg(command, "--ctrl-port", args.ctrl_port)
    return run_python_script("listenai_executable_case_suite.py", command, context=vars(args), command_name="generate-suite")


def forward_validate(args: argparse.Namespace) -> int:
    suite_dir = Path(args.suite_dir).resolve()

    # Try to get task_dir from package_zip first (most reliable)
    task_dir = infer_task_dir(package_zip=args.package_zip)

    # If no task_dir from package_zip, try to derive it from the unified runtime tree.
    if task_dir is None and suite_dir:
        derived_task_dir = task_dir_for_runtime(suite_dir.parent if suite_dir.name == "suite" else suite_dir)
        if derived_task_dir is not None:
            task_dir = derived_task_dir

    if task_dir is None:
        # Fallback to original logic
        task_dir = infer_task_dir(package_zip=args.package_zip, suite_dir=str(suite_dir))

    # Debug: show what task_dir was derived
    print(f"[mars-belt] Suite dir: {suite_dir}")
    print(f"[mars-belt] Derived task dir: {task_dir}")

    if not args.config_only:
        set_playback_volume(30)
        sync_shared_audio_to_suite(suite_dir)
        donor = ensure_suite_runtime_assets(suite_dir)
        if donor is not None:
            print(f"[mars-belt] reused suite runtime assets from: {donor['suiteDir']}")
    command: List[str] = []
    add_arg(command, "--suite-dir", str(suite_dir))
    add_path_arg(command, "--package-zip", args.package_zip)
    add_path_arg(command, "--web-config", args.web_config)
    add_path_arg(command, "--base-script", args.base_script)
    add_flag(command, "--config-only", args.config_only)
    file_path = Path(args.file)
    add_arg(command, "-f", str(file_path.resolve()) if file_path.is_absolute() else args.file)
    add_arg(command, "-r", args.run_times)
    add_arg(command, "-l", args.label)
    add_arg(command, "-p", args.port)
    add_arg(command, "--ctrl-port", args.ctrl_port)
    add_arg(command, "--protocol-port", getattr(args, "protocol_port", ""))
    add_flag(command, "--pretest", args.pretest)
    add_flag(command, "--skip-pretest", getattr(args, 'skip_pretest', False))
    add_flag(command, "--update-audio-skills", getattr(args, "update_audio_skills", False))
    exit_code = run_python_script("listenai_voice_test_lite.py", command, context=vars(args), command_name="validate", use_dialout=True)
    if not args.config_only:
        sync_shared_audio_from_suite(suite_dir)
        stage_validation_artifacts(task_dir, suite_dir)
    return exit_code


def run_backup_state(args: argparse.Namespace) -> int:
    root = Path(args.root).resolve() if args.root else SKILL_ROOT
    memory_root = Path(args.memory_root).resolve() if args.memory_root else STATE_ROOT
    if args.loop:
        while True:
            backup_state_once(root, memory_root)
            time.sleep(max(args.interval_seconds, 1))
    else:
        backup_state_once(root, memory_root)
    return 0


def run_ensure_audio_skills(args: argparse.Namespace) -> int:
    report = ensure_audio_skill_set(update=getattr(args, "update_audio_skills", False))
    if getattr(args, "ensure_laid", False):
        report["laid"] = ensure_laid_helper(
            update=False,
            force=getattr(args, "force_laid", False),
        )
    print(json.dumps(report, ensure_ascii=False, indent=2))
    return 0


def _copy_sheet_cells(src_ws, dst_ws) -> None:
    from copy import copy

    for row in src_ws.iter_rows():
        for cell in row:
            new_cell = dst_ws.cell(row=cell.row, column=cell.column, value=cell.value)
            if cell.has_style:
                new_cell.font = copy(cell.font)
                new_cell.fill = copy(cell.fill)
                new_cell.border = copy(cell.border)
                new_cell.alignment = copy(cell.alignment)
                new_cell.number_format = copy(cell.number_format)
                new_cell.protection = copy(cell.protection)
            if cell.hyperlink:
                new_cell._hyperlink = copy(cell.hyperlink)
            if cell.comment:
                new_cell.comment = copy(cell.comment)
    for key, dimension in src_ws.column_dimensions.items():
        dst_ws.column_dimensions[key].width = dimension.width
        dst_ws.column_dimensions[key].hidden = dimension.hidden
    for key, dimension in src_ws.row_dimensions.items():
        dst_ws.row_dimensions[key].height = dimension.height
        dst_ws.row_dimensions[key].hidden = dimension.hidden
    for merged in src_ws.merged_cells.ranges:
        dst_ws.merge_cells(str(merged))
    dst_ws.freeze_panes = src_ws.freeze_panes
    dst_ws.sheet_view.zoomScale = src_ws.sheet_view.zoomScale


def _safe_sheet_name(name: str, used: set[str]) -> str:
    cleaned = re.sub(r"[\\/*?:\[\]]+", "_", name).strip() or "Sheet"
    cleaned = cleaned[:31]
    base = cleaned
    index = 2
    while cleaned in used:
        suffix = f"_{index}"
        cleaned = (base[: 31 - len(suffix)] + suffix) if len(base) + len(suffix) > 31 else base + suffix
        index += 1
    used.add(cleaned)
    return cleaned


def _write_table_sheet(ws, title: str, rows: List[Dict[str, Any]]) -> None:
    from openpyxl.styles import Font

    ws["A1"] = title
    ws["A1"].font = Font(bold=True, size=13)
    if not rows:
        ws["A3"] = "无数据"
        return
    headers = list(rows[0].keys())
    for col, header in enumerate(headers, start=1):
        ws.cell(row=3, column=col, value=header)
        ws.cell(row=3, column=col).font = Font(bold=True)
    for row_idx, row in enumerate(rows, start=4):
        for col_idx, header in enumerate(headers, start=1):
            ws.cell(row=row_idx, column=col_idx, value=row.get(header, ""))


def run_merge_report(args: argparse.Namespace) -> int:
    try:
        from openpyxl import Workbook, load_workbook
    except Exception as exc:
        raise SystemExit(f"merge-report requires openpyxl: {exc}") from exc

    output = Path(args.output).resolve()
    output.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    summary_ws = wb.active
    summary_ws.title = "任务汇总"
    used = {"任务汇总"}
    row = 1
    summary_ws.cell(row=row, column=1, value="任务汇总")
    row += 2

    if args.summary_json:
        payload = json.loads(Path(args.summary_json).read_text(encoding="utf-8"))
        selected = payload.get("selected") or {}
        final_release = payload.get("finalReleaseSubset") or {}
        summary_rows = [
            ("生成时间", payload.get("generatedAt", "")),
            ("运行ID", payload.get("runId", "")),
            ("产品", selected.get("productLabel", "")),
            ("模组", selected.get("moduleBoard", "")),
            ("语言", selected.get("language", "")),
            ("场景", selected.get("sceneLabel", "")),
            ("版本", selected.get("versionLabel", "")),
            ("releaseId", (payload.get("packageSummary") or {}).get("releaseId", "")),
            ("releaseVersion", (payload.get("packageSummary") or {}).get("releaseVersion", "")),
            ("timeout", final_release.get("timeout", "")),
            ("volLevel", final_release.get("volLevel", "")),
            ("releaseVerifyOk", payload.get("releaseVerifyOk", "")),
            ("zipVerifyOk", payload.get("zipVerifyOk", "")),
            ("packageZip", ((payload.get("artifacts") or {}).get("packageZip") or "")),
            ("webConfigJson", ((payload.get("artifacts") or {}).get("webConfigJson") or "")),
        ]
        for key, value in summary_rows:
            summary_ws.cell(row=row, column=1, value=key)
            summary_ws.cell(row=row, column=2, value=value)
            row += 1

    if args.config_json:
        config_payload = json.loads(Path(args.config_json).read_text(encoding="utf-8"))
        cfg_ws = wb.create_sheet(_safe_sheet_name("配置校验", used))
        summary = config_payload.get("summary") or {}
        rows = config_payload.get("rows") or []
        cfg_ws["A1"] = "配置校验汇总"
        cfg_ws["A3"] = "总数"
        cfg_ws["B3"] = summary.get("total", "")
        counters = summary.get("counters") or {}
        current_row = 5
        for key, value in counters.items():
            cfg_ws.cell(row=current_row, column=1, value=key)
            cfg_ws.cell(row=current_row, column=2, value=value)
            current_row += 1
        if rows:
            headers = list(rows[0].keys())
            start_row = current_row + 2
            for col_idx, header in enumerate(headers, start=1):
                cfg_ws.cell(row=start_row, column=col_idx, value=header)
            for row_idx, item in enumerate(rows, start=start_row + 1):
                for col_idx, header in enumerate(headers, start=1):
                    cfg_ws.cell(row=row_idx, column=col_idx, value=item.get(header, ""))

    for spec in args.xlsx:
        if "=" in spec:
            sheet_title, raw_path = spec.split("=", 1)
        else:
            raw_path = spec
            sheet_title = Path(raw_path).stem
        src_path = Path(raw_path).resolve()
        src_wb = load_workbook(src_path)
        for src_ws in src_wb.worksheets:
            title = _safe_sheet_name(f"{sheet_title}_{src_ws.title}" if len(src_wb.worksheets) > 1 else sheet_title, used)
            dst_ws = wb.create_sheet(title)
            _copy_sheet_cells(src_ws, dst_ws)

    wb.save(output)
    print(str(output))
    return 0


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description="Mars Belt direct Python entrypoint")
    subparsers = parser.add_subparsers(dest="command", required=True)

    list_catalog = subparsers.add_parser("list-catalog", help="Export the current live selectable matrix")
    list_catalog.add_argument("--token", default="")
    list_catalog.add_argument("--product", action="append", default=[])
    list_catalog.add_argument("--scene", action="append", default=[])
    list_catalog.add_argument("--json-out", default="")
    list_catalog.set_defaults(func=forward_list_catalog)

    resolve = subparsers.add_parser("resolve", help="Resolve one target or package one base firmware")
    resolve.add_argument("--token", default="")
    resolve.add_argument("--action", choices=["resolve", "prepare", "package"], default="resolve")
    resolve.add_argument("--product", required=True)
    resolve.add_argument("--module", required=True)
    resolve.add_argument("--language", required=True)
    resolve.add_argument("--version", required=True)
    resolve.add_argument("--scene", default="")
    resolve.add_argument("--product-name", default="")
    resolve.add_argument("--source-release-id", default="")
    resolve.add_argument("--timeout-sec", type=int, default=600)
    resolve.add_argument("--refresh-live", action="store_true")
    resolve.add_argument("--json-out", default="")
    resolve.add_argument("--products-csv-out", default="")
    resolve.add_argument("--modules-csv-out", default="")
    resolve.add_argument("--matrix-csv-out", default="")
    resolve.add_argument("--duplicates-csv-out", default="")
    resolve.add_argument("--matrix-md-out", default="")
    resolve.add_argument("--resolution-out", default="")
    resolve.add_argument("--summary-out", default="")
    resolve.set_defaults(func=forward_resolve)

    package_custom = subparsers.add_parser("package-custom", help="Package one firmware; no override=base mode, with override=specified-parameter mode")
    package_custom.add_argument("--token", default="")
    package_custom.add_argument("--product", required=True)
    package_custom.add_argument("--module", required=True)
    package_custom.add_argument("--language", required=True)
    package_custom.add_argument("--version", required=True)
    package_custom.add_argument("--scene", default="")
    package_custom.add_argument("--catalog-json", default="")
    package_custom.add_argument("--source-release-id", default="")
    package_custom.add_argument("--product-name", default="")
    package_custom.add_argument("--run-id", default="")
    package_custom.add_argument("--result-root", default="")
    package_custom.add_argument("--package-root", default="")
    package_custom.add_argument("--timeout-sec", type=int, default=1800)
    package_custom.add_argument("--retry-limit", type=int, default=2)
    package_custom.add_argument("--override", action="append", default=[])
    package_custom.add_argument("--enable-algo-words", action="store_true", help="Only when explicitly requested by the task, allow releaseAlgoList/releaseDepthList overrides for adding algorithm words.")
    package_custom.add_argument("--require-feature", action="append", default=[])
    package_custom.add_argument("--comments", default="")
    package_custom.add_argument("--refresh-live", action="store_true")
    package_custom.add_argument("--dry-run", action="store_true")
    package_custom.set_defaults(func=forward_package_custom)

    generate_algo_words = subparsers.add_parser("generate-algo-words", help="Generate a releaseAlgoList JSON with unique Chinese command words and protocols")
    generate_algo_words.add_argument("--base-json", required=True)
    generate_algo_words.add_argument("--output", required=True)
    generate_algo_words.add_argument("--append-count", type=int, required=True)
    generate_algo_words.add_argument("--proto-start", default="0x0083")
    generate_algo_words.set_defaults(func=forward_generate_algo_words)

    package_voice_reg = subparsers.add_parser("package-voice-reg", help="Package one specific-learning voice-registration firmware")
    package_voice_reg.add_argument("--token", default="")
    package_voice_reg.add_argument("--product", required=True)
    package_voice_reg.add_argument("--module", required=True)
    package_voice_reg.add_argument("--language", required=True)
    package_voice_reg.add_argument("--version", required=True)
    package_voice_reg.add_argument("--scene", default="")
    package_voice_reg.add_argument("--catalog-json", default="")
    package_voice_reg.add_argument("--source-release-id", default="")
    package_voice_reg.add_argument("--product-name", default="")
    package_voice_reg.add_argument("--run-id", default="")
    package_voice_reg.add_argument("--result-root", default="")
    package_voice_reg.add_argument("--package-root", default="")
    package_voice_reg.add_argument("--timeout-sec", type=int, default=1800)
    package_voice_reg.add_argument("--retry-limit", type=int, default=2)
    package_voice_reg.add_argument("--timeout-value", type=int, default=30)
    package_voice_reg.add_argument("--vol-level", type=int, default=5)
    package_voice_reg.add_argument("--algo-view-mode", default="full")
    package_voice_reg.add_argument("--study-reg-command", action="append", default=[])
    package_voice_reg.add_argument("--require-feature", action="append", default=[])
    package_voice_reg.add_argument("--comments", default="")
    package_voice_reg.add_argument("--refresh-live", action="store_true")
    package_voice_reg.add_argument("--dry-run", action="store_true")
    package_voice_reg.set_defaults(func=forward_package_voice_reg)

    package_weekly = subparsers.add_parser("package-weekly", help="Run base + timeout/volLevel/defaultVol low-mid-high multi-package validation")
    package_weekly.add_argument("--token", default="")
    package_weekly.add_argument("--product", required=True)
    package_weekly.add_argument("--module", required=True)
    package_weekly.add_argument("--language", required=True)
    package_weekly.add_argument("--version", required=True)
    package_weekly.add_argument("--scene", default="")
    package_weekly.add_argument("--source-release-id", default="")
    package_weekly.add_argument("--timeout-sec", type=int, default=1800)
    package_weekly.add_argument("--log-port", default=DEFAULT_PORT)
    package_weekly.add_argument("--protocol-port", default="/dev/ttyACM1" if IS_LINUX else "COM13")
    package_weekly.add_argument("--uart0-port", default="")
    package_weekly.add_argument("--uart1-port", default="")
    package_weekly.add_argument("--ctrl-port", default=DEFAULT_CTRL_PORT)
    package_weekly.add_argument("--burn-port", default=DEFAULT_PORT)
    package_weekly.add_argument("--task-dir", default="")
    package_weekly.add_argument("--variants", default="")
    package_weekly.add_argument("--refresh-live", action="store_true")
    package_weekly.add_argument("--update-audio-skills", action="store_true")
    package_weekly.add_argument("--skip-device", action="store_true")
    package_weekly.set_defaults(func=forward_package_weekly)

    package_batch = subparsers.add_parser("package-batch", help="Package one representative firmware per direct parameter case")
    package_batch.add_argument("--token", default="")
    package_batch.add_argument("--catalog-json", required=True)
    package_batch.add_argument("--test-catalog-json", required=True)
    package_batch.add_argument("--batch-root", default="")
    package_batch.add_argument("--package-root", default="")
    package_batch.add_argument("--batch-id", default="")
    package_batch.add_argument("--timeout-sec", type=int, default=1800)
    package_batch.add_argument("--dry-run", action="store_true")
    package_batch.set_defaults(func=forward_package_batch)

    package_grouped = subparsers.add_parser("package-grouped", help="Package grouped multi-parameter bundles")
    package_grouped.add_argument("--token", default="")
    package_grouped.add_argument("--catalog-json", required=True)
    package_grouped.add_argument("--test-catalog-json", required=True)
    package_grouped.add_argument("--batch-root", default="")
    package_grouped.add_argument("--package-root", default="")
    package_grouped.add_argument("--batch-id", default="")
    package_grouped.add_argument("--timeout-sec", type=int, default=1800)
    package_grouped.add_argument("--dry-run", action="store_true")
    package_grouped.set_defaults(func=forward_package_grouped)

    build_case_catalog = subparsers.add_parser("build-case-catalog", help="Export the selected target into a parameter test-case catalog")
    build_case_catalog.add_argument("--token", default="")
    build_case_catalog.add_argument("--product", default="")
    build_case_catalog.add_argument("--module", default="")
    build_case_catalog.add_argument("--language", default="")
    build_case_catalog.add_argument("--version", default="")
    build_case_catalog.add_argument("--scene", default="")
    build_case_catalog.add_argument("--source-release-id", default="")
    build_case_catalog.add_argument("--refresh-live", action="store_true")
    build_case_catalog.add_argument("--package-zip", default="")
    build_case_catalog.add_argument("--web-config", default="")
    build_case_catalog.add_argument("--profile", choices=["parameter", "auto", "base", "changed", "voice-reg", "multi-wke"], default="parameter")
    build_case_catalog.add_argument("--override", action="append", default=[])
    build_case_catalog.add_argument("--md-out", default="")
    build_case_catalog.add_argument("--json-out", default="")
    build_case_catalog.add_argument("--json-out-catalog", default="")
    build_case_catalog.add_argument("--products-csv-out", default="")
    build_case_catalog.add_argument("--modules-csv-out", default="")
    build_case_catalog.add_argument("--matrix-csv-out", default="")
    build_case_catalog.add_argument("--duplicates-csv-out", default="")
    build_case_catalog.add_argument("--matrix-md-out", default="")
    build_case_catalog.add_argument("--resolution-out", default="")
    build_case_catalog.add_argument("--summary-out", default="")
    build_case_catalog.set_defaults(func=forward_build_case_catalog)

    generate_suite = subparsers.add_parser("generate-suite", help="Generate testCases.csv and deviceInfo_generated.json from one zip")
    generate_suite.add_argument("--test-catalog-json", default="")
    generate_suite.add_argument("--package-zip", default="")
    generate_suite.add_argument("--web-config", default="")
    generate_suite.add_argument("--profile", choices=["parameter", "auto", "base", "changed", "voice-reg", "multi-wke"], default="auto")
    generate_suite.add_argument("--override", action="append", default=[])
    generate_suite.add_argument("--out-dir", default="")
    generate_suite.add_argument("-p", "--port", default=DEFAULT_PORT)
    generate_suite.add_argument("--ctrl-port", default=DEFAULT_CTRL_PORT)
    generate_suite.set_defaults(func=forward_generate_suite)

    validate = subparsers.add_parser("validate", help="Run config-only or device-side validation on an existing suite")
    validate.add_argument("--suite-dir", required=True)
    validate.add_argument("--package-zip", default="")
    validate.add_argument("--web-config", default="")
    validate.add_argument("--base-script", default="")
    validate.add_argument("--config-only", action="store_true")
    validate.add_argument("-f", "--file", default="deviceInfo_generated.json")
    validate.add_argument("-r", "--run-times", type=int, default=0)
    validate.add_argument("-l", "--label", default="mars-belt")
    validate.add_argument("-p", "--port", default=DEFAULT_PORT)
    validate.add_argument("--ctrl-port", default=DEFAULT_CTRL_PORT)
    validate.add_argument("--protocol-port", default="")
    validate.add_argument("--pretest", action="store_true")
    validate.add_argument("--skip-pretest", action="store_true")
    validate.add_argument("--update-audio-skills", action="store_true")
    validate.set_defaults(func=forward_validate)

    ensure_audio = subparsers.add_parser("ensure-audio-skills", help="确保 mars-belt/tools/audio 下存在 listenai-play / listenai-laid-installer")
    ensure_audio.add_argument("--update-audio-skills", action="store_true", help="即使本地 tools/audio 里已存在 audio skills，也执行 git pull --ff-only")
    ensure_audio.add_argument("--ensure-laid", action="store_true", help="完成 repo 校验后再执行 listenai-play ensure-laid")
    ensure_audio.add_argument("--force-laid", action="store_true", help="配合 --ensure-laid 使用，强制刷新 laid")
    ensure_audio.set_defaults(func=run_ensure_audio_skills)

    set_volume = subparsers.add_parser("set-volume", help="Set the Windows playback volume to a target percentage")
    set_volume.add_argument("--percent", type=int, default=50)
    set_volume.set_defaults(func=lambda args: set_playback_volume(args.percent))

    burn = subparsers.add_parser("burn", help="Burn one package zip or explicit fw.bin")
    burn.add_argument("--package-zip", default="")
    burn.add_argument("--firmware-bin", default="")
    burn.add_argument("--ctrl-port", default=DEFAULT_CTRL_PORT)
    burn.add_argument("--burn-port", default=DEFAULT_PORT)
    burn.add_argument("--ctrl-baud", type=int, default=115200)
    burn.add_argument("--baud", type=int, default=460800)
    burn.add_argument("--runtime-log-port", default="")
    burn.add_argument("--runtime-log-baud", type=int, default=115200)
    burn.add_argument("--max-retry", type=int, default=3)
    burn.add_argument("--cmd-delay-ms", type=int, default=300)
    burn.add_argument("--burn-mode-wait-ms", type=int, default=2000)
    burn.add_argument("--pre-burn-reboot", action="store_true", help="Send reboot on ctrl port before the 4-step burn entry flow")
    burn.add_argument("--boot-wait-seconds", type=int, default=5)
    burn.add_argument("--skip-loglevel", action="store_true")
    burn.add_argument("--keep-extracted", action="store_true")
    burn.set_defaults(func=run_burn)

    backup = subparsers.add_parser("backup-state", help=argparse.SUPPRESS)
    backup.add_argument("--root", default="")
    backup.add_argument("--memory-root", default="")
    backup.add_argument("--interval-seconds", type=int, default=300)
    backup.add_argument("--loop", action="store_true")
    backup.set_defaults(func=run_backup_state)

    merge_report = subparsers.add_parser("merge-report", help="把多份验证结果汇总为一个多 sheet 工作簿")
    merge_report.add_argument("--output", required=True)
    merge_report.add_argument("--summary-json", default="")
    merge_report.add_argument("--config-json", default="")
    merge_report.add_argument("--xlsx", action="append", default=[], help="格式: 标签=路径，或直接传入 xlsx 路径")
    merge_report.set_defaults(func=run_merge_report)

    return parser


def main() -> int:
    parser = build_parser()
    args = parser.parse_args()
    try:
        return int(args.func(args))
    except SystemExit:
        raise
    except Exception as exc:
        append_error_entry(
            command_name=getattr(args, "command", "mars-belt"),
            cli_args=vars(args),
            phenomenon=f"{type(exc).__name__}: {exc}",
            work_dir=str(Path.cwd()),
        )
        raise


if __name__ == "__main__":
    raise SystemExit(main())
