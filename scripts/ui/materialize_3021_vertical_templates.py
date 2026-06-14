#!/usr/bin/env python3
from __future__ import annotations

import argparse
import json
import re
from copy import deepcopy
from pathlib import Path
from typing import Any, Dict, Iterable, List, Sequence, Set

from openpyxl import load_workbook

SHEET = "词条预处理"
RESERVED_VOICE_REG_CONTROLS = {
    "学习命令词",
    "删除命令词",
    "学习唤醒词",
    "删除唤醒词",
    "删除全部命令词",
    "退出学习",
    "退出删除",
}


def norm(text: Any) -> str:
    return re.sub(r"\s+", "", str(text or ""))


def safe_name(text: str) -> str:
    return re.sub(r"[^\w\u4e00-\u9fa5.-]+", "_", text)[:120]


def protocol(code: int, recv: bool = False) -> str:
    cmd = 0x82 if recv else 0x81
    chk = (0xA5 + 0xFA + 0x00 + cmd + code + 0x00) & 0xFF
    return f"A5 FA 00 {cmd:02X} {code:02X} 00 {chk:02X} FB"


def existing_words(ws) -> Set[str]:
    words: Set[str] = set()
    for row in ws.iter_rows(values_only=True):
        # Template columns: idx, word, extWord, type, reply, replyMode, sndProtocol, recProtocol.
        for value in row[1:3] if len(row) >= 3 else row:
            if value is None:
                continue
            text = str(value).strip()
            if text:
                words.add(norm(text))
    return words


def collect_target_words(job: Dict[str, Any]) -> List[str]:
    cfg = job.get("algoConfig") or {}
    words: List[str] = []
    for key in ("studyCommandWords", "runtimeCommandWords", "businessCommandWords"):
        for value in cfg.get(key) or []:
            text = str(value or "").strip()
            if text and text not in words:
                words.append(text)
    return words


def append_command_rows(wb, words: Sequence[str], start_code: int) -> List[Dict[str, Any]]:
    if SHEET not in wb.sheetnames:
        raise RuntimeError(f"sheet not found: {SHEET}")
    ws = wb[SHEET]
    present = existing_words(ws)
    appended: List[Dict[str, Any]] = []
    next_code = start_code
    next_idx = max([int(c.value) for c in ws["A"] if isinstance(c.value, int)] or [100]) + 1
    for word in words:
        if word in RESERVED_VOICE_REG_CONTROLS:
            raise RuntimeError(f"reserved voice registration control word cannot be materialized as normal command: {word}")
        if norm(word) in present:
            continue
        while next_code in {0x01, 0x02, 0x03, 0x04, 0x05, 0x06, 0x07, 0x08, 0x09, 0x0A, 0x30, 0x31, 0x32, 0x33, 0x40, 0x41, 0x42, 0x43, 0x44, 0x45, 0x46, 0x47, 0x48, 0x49, 0x57, 0x58, 0x61, 0x62, 0x70, 0x71, 0x72, 0x73}:
            next_code += 1
        if next_code > 0xEF:
            raise RuntimeError("not enough protocol command code space for generated rows")
        ws.append([next_idx, word, word, "命令词", "好的", "主", protocol(next_code, False), protocol(next_code, True)])
        appended.append({"word": word, "idx": next_idx, "code": f"0x{next_code:02X}", "sndProtocol": protocol(next_code, False), "recProtocol": protocol(next_code, True)})
        present.add(norm(word))
        next_idx += 1
        next_code += 1
    return appended


def materialize(plan: Dict[str, Any], repo_root: Path, out_dir: Path) -> tuple[Dict[str, Any], List[Dict[str, Any]]]:
    out_plan = deepcopy(plan)
    manifest: List[Dict[str, Any]] = []
    out_dir.mkdir(parents=True, exist_ok=True)
    for job in out_plan.get("jobs") or []:
        template = str(job.get("algoTemplate") or "")
        if not template:
            continue
        source = (repo_root / template).resolve()
        if not source.exists():
            raise RuntimeError(f"template not found: {template}")
        words = collect_target_words(job)
        if not words:
            manifest.append({"jobId": job.get("jobId"), "source": template, "generated": template, "appended": []})
            continue
        wb = load_workbook(source)
        # Keep generated protocol codes deterministic but separated by job.
        start_code = 0x74 + (abs(hash(str(job.get("jobId") or ""))) % 24)
        appended = append_command_rows(wb, words, start_code=start_code)
        if appended:
            generated = out_dir / f"{safe_name(str(job.get('jobId') or 'job'))}.xlsx"
            wb.save(generated)
            rel = generated.relative_to(repo_root) if generated.is_relative_to(repo_root) else generated
            job["algoTemplate"] = str(rel)
        manifest.append({"jobId": job.get("jobId"), "source": template, "generated": job.get("algoTemplate"), "targetWords": words, "appended": appended})
    return out_plan, manifest


def main() -> int:
    parser = argparse.ArgumentParser(description="Create per-job xlsx templates so UI voice-registration selections match the imported vertical commands.")
    parser.add_argument("--plan", required=True)
    parser.add_argument("--repo-root", default=".")
    parser.add_argument("--out-plan", required=True)
    parser.add_argument("--out-dir", required=True)
    parser.add_argument("--manifest", default="")
    args = parser.parse_args()
    repo_root = Path(args.repo_root).resolve()
    plan_path = Path(args.plan)
    plan = json.loads(plan_path.read_text(encoding="utf-8"))
    out_plan, manifest = materialize(plan, repo_root, (repo_root / args.out_dir).resolve())
    out_path = Path(args.out_plan)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    out_path.write_text(json.dumps(out_plan, ensure_ascii=False, indent=2), encoding="utf-8")
    if args.manifest:
        m = Path(args.manifest)
        m.parent.mkdir(parents=True, exist_ok=True)
        m.write_text(json.dumps({"sourcePlan": str(plan_path), "outPlan": str(out_path), "items": manifest}, ensure_ascii=False, indent=2), encoding="utf-8")
    summary = {"jobs": len(out_plan.get("jobs") or []), "templatesGenerated": sum(1 for item in manifest if item.get("appended")), "rowsAppended": sum(len(item.get("appended") or []) for item in manifest)}
    print(json.dumps(summary, ensure_ascii=False, indent=2))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
