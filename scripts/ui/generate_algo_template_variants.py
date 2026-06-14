#!/usr/bin/env python3
from __future__ import annotations

import argparse
import json
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any, Callable, Dict, Iterable, List, Sequence

from openpyxl import load_workbook

SHEET = "词条预处理"
HEADERS = ["语义标签", "语义(最小功能词)", "功能泛化词", "功能类型", "播报语句", "播报模式", "发送协议", "接收协议"]


@dataclass(frozen=True)
class Profile:
    id: str
    language: str
    filename: str
    purpose: str
    recommended_for: List[str]
    required_config: Dict[str, Any]
    covers: List[str]
    parameters: List[str]
    risk: str
    row_factory: Callable[[str], List[List[Any]]]
    compat_alias_for: str | None = None


def protocol(code: int, recv: bool = False) -> str:
    cmd = 0x82 if recv else 0x81
    chk = (0xA5 + 0xFA + 0x00 + cmd + code + 0x00) & 0xFF
    return f"A5 FA 00 {cmd:02X} {code:02X} 00 {chk:02X} FB"


def row(
    idx: int | None,
    word: str | None,
    ext: str | None,
    typ: str | None,
    reply: str | None,
    mode: str | None,
    code: int | None = None,
    *,
    snd: str | None = None,
    rec: str | None = None,
) -> List[Any]:
    if code is not None:
        snd = protocol(code, False)
        rec = protocol(code, True)
    return [idx, word, ext, typ, reply, mode, snd, rec]


def clone_rows(rows: Iterable[Sequence[Any]]) -> List[List[Any]]:
    return [list(values) for values in rows]


def dedupe_rows(rows: Iterable[Sequence[Any]]) -> List[List[Any]]:
    result: List[List[Any]] = []
    seen = set()
    for values in rows:
        values_list = list(values)
        # Child rows may intentionally share empty idx/word but different extWord.
        key = tuple("" if value is None else str(value) for value in values_list)
        if key in seen:
            continue
        seen.add(key)
        result.append(values_list)
    return result


@dataclass(frozen=True)
class LocaleText:
    wake_default: str
    wake_alt_1: str
    wake_alt_2: str
    wake_negative: str
    cmd_negative: str
    on: str
    on_ext: str
    on_child_ext: str
    off: str
    status: str
    reply_ok: str
    reply_on: str
    reply_off: str
    reply_status: str
    volume_up: str
    volume_down: str
    volume_max: str
    volume_min: str
    volume_mid: str
    exit_rec: str
    exit_reply: str
    welcome: str
    passive_broadcast: str
    passive_state: str
    rest: str
    switch_wake: str
    query_wake: str
    restore_wake: str
    switch_to_alt_1: str
    switch_to_alt_2: str
    learn_command: str
    learn_wakeup: str
    delete_command: str
    delete_wakeup: str
    delete_all: str
    exit_learning: str
    exit_deletion: str
    learn_command_reply: str
    learn_wakeup_reply: str
    delete_command_reply: str
    delete_wakeup_reply: str
    delete_all_reply: str
    valid_learn_target_1: str
    valid_learn_target_2: str
    short_boundary_hint: str
    long_boundary_hint: str


TEXT: Dict[str, LocaleText] = {
    "zh": LocaleText(
        wake_default="小聆小聆",
        wake_alt_1="小优小优",
        wake_alt_2="小乐小乐",
        wake_negative="[小李小李/小牛小牛]",
        cmd_negative="[打开电视/关闭电视/播放音乐]",
        on="打开风扇",
        on_ext="[开机/开启风扇/启动风扇]",
        on_child_ext="[帮我打开风扇/请打开风扇]",
        off="关闭风扇",
        status="查询状态",
        reply_ok="好的",
        reply_on="风扇已打开",
        reply_off="风扇已关闭",
        reply_status="当前状态正常",
        volume_up="增大音量",
        volume_down="减小音量",
        volume_max="最大音量",
        volume_min="最小音量",
        volume_mid="中等音量",
        exit_rec="退出识别",
        exit_reply="主人我休息了，有事再叫我",
        welcome="欢迎使用智能产品，我是小聆，快来唤醒我吧",
        passive_broadcast="协议播报已触发",
        passive_state="被动播报状态正常",
        rest="我先离开，有需要再叫我",
        switch_wake="切换唤醒词",
        query_wake="查询唤醒词",
        restore_wake="恢复默认唤醒词",
        switch_to_alt_1="切换到小优小优",
        switch_to_alt_2="切换到小乐小乐",
        learn_command="学习命令词",
        learn_wakeup="学习唤醒词",
        delete_command="删除命令词",
        delete_wakeup="删除唤醒词",
        delete_all="删除全部命令词",
        exit_learning="退出学习",
        exit_deletion="退出删除",
        learn_command_reply="请说要学习的命令词",
        learn_wakeup_reply="请说要学习的唤醒词",
        delete_command_reply="请说要删除的命令词",
        delete_wakeup_reply="请说要删除的唤醒词",
        delete_all_reply="已删除全部命令词",
        valid_learn_target_1="打开学习灯",
        valid_learn_target_2="关闭学习灯",
        short_boundary_hint="短词",
        long_boundary_hint="超长注册命令词样例",
    ),
    "en": LocaleText(
        wake_default="Hello My Dear",
        wake_alt_1="Smart Helper",
        wake_alt_2="Voice Partner",
        wake_negative="Hello My Car",
        cmd_negative="Open TV",
        on="Start Fan",
        on_ext="Turn On Fan",
        on_child_ext="Please Start Fan",
        off="Stop Fan",
        status="Query Status",
        reply_ok="Okay",
        reply_on="Fan started",
        reply_off="Fan stopped",
        reply_status="Current status is normal",
        volume_up="Volume Up",
        volume_down="Volume Down",
        volume_max="Set Volume To Max",
        volume_min="Set Volume To Min",
        volume_mid="Set Medium Volume",
        exit_rec="Exit Recognition",
        exit_reply="I will rest now",
        welcome="Welcome your smart product is ready",
        passive_broadcast="Protocol broadcast triggered",
        passive_state="Passive status is normal",
        rest="I will rest now call me if you need help",
        switch_wake="Switch Wake Word",
        query_wake="Query Wake Word",
        restore_wake="Restore Default Wake Word",
        switch_to_alt_1="Switch To Smart Helper",
        switch_to_alt_2="Switch To Voice Partner",
        learn_command="Learn Command",
        learn_wakeup="Learn Wake Word",
        delete_command="Delete Command",
        delete_wakeup="Delete Wake Word",
        delete_all="Delete All Commands",
        exit_learning="Exit Learning",
        exit_deletion="Exit Deletion",
        learn_command_reply="Please say the command to learn",
        learn_wakeup_reply="Please say the wake word to learn",
        delete_command_reply="Please say the command to delete",
        delete_wakeup_reply="Please say the wake word to delete",
        delete_all_reply="All commands deleted",
        valid_learn_target_1="Start Study Light",
        valid_learn_target_2="Stop Study Light",
        short_boundary_hint="Short Word",
        long_boundary_hint="Very Long Command Sample",
    ),
}


def common_tail(lang: str, start_idx: int = 90) -> List[List[Any]]:
    t = TEXT[lang]
    return [
        row(start_idx, None, None, "欢迎语", t.welcome, "主", 0x70),
        row(start_idx + 1, None, None, "播报语", t.passive_broadcast, "被", 0x71),
        row(start_idx + 2, None, None, "休息语", t.rest, "主", 0x72),
        row(start_idx + 3, None, None, "心跳协议", None, None, 0x73),
    ]


def base_core_rows(lang: str) -> List[List[Any]]:
    t = TEXT[lang]
    rows = [
        row(1, t.wake_default, None, "唤醒词", "我在/在呢/我来啦" if lang == "zh" else "Yes/Hello/What can I do for you", "主", 0x01),
        row(2, t.on, t.on_ext, "命令词", t.reply_on, "主", 0x02),
        row(None, None, t.on_child_ext, "命令词", None, None, None),
        row(3, t.off, t.off, "命令词", t.reply_off, "主", 0x03),
        row(4, t.status, t.status, "命令词", t.reply_status, "主", 0x04),
        row(5, t.volume_up, t.volume_up, "增大音量", "增大音量" if lang == "zh" else "Volume increased", "主", 0x05),
        row(6, t.volume_down, t.volume_down, "减小音量", "减小音量" if lang == "zh" else "Volume decreased", "主", 0x06),
        row(7, t.volume_max, t.volume_max, "最大音量", "最大音量" if lang == "zh" else "Maximum volume", "主", 0x07),
        row(8, t.volume_min, t.volume_min, "最小音量", "最小音量" if lang == "zh" else "Minimum volume", "主", 0x08),
        row(9, t.volume_mid, t.volume_mid, "中等音量", "中等音量" if lang == "zh" else "Medium volume", "主", 0x09),
        row(10, t.exit_rec, t.exit_rec, "退出识别", t.exit_reply, "主", 0x0A),
        row(11, "唤醒词负性词" if lang == "zh" else "Wake Negative Words", t.wake_negative, "唤醒词负性词", None, None, None),
        row(12, "命令词负性词" if lang == "zh" else "Command Negative Words", t.cmd_negative, "命令词负性词", None, None, None),
    ]
    return rows + common_tail(lang, 13)


def protocol_rows(lang: str) -> List[List[Any]]:
    t = TEXT[lang]
    rows = base_core_rows(lang) + [
        row(30, "查询设备状态" if lang == "zh" else "Query Device Status", "查询设备状态" if lang == "zh" else "Query Device Status", "命令词", t.reply_status, "被", 0x30),
        row(31, None, None, "播报语", t.passive_state, "被", 0x31),
        row(32, None, None, "播报语", t.passive_broadcast, "被", 0x32),
        row(33, None, None, "心跳协议", None, None, 0x33),
    ]
    return dedupe_rows(rows)


def multi_common_rows(lang: str) -> List[List[Any]]:
    t = TEXT[lang]
    return [
        row(40, t.wake_alt_1, None, "唤醒词", ("小优在" if lang == "zh" else "Smart helper here"), "主", 0x40),
        row(41, t.wake_alt_2, None, "唤醒词", ("小乐在" if lang == "zh" else "Voice partner here"), "主", 0x41),
        row(42, t.switch_wake, t.switch_wake, "命令词", ("已切换唤醒词" if lang == "zh" else "Wake word switched"), "主", 0x42),
        row(43, t.query_wake, t.query_wake, "命令词", ("当前唤醒词" if lang == "zh" else "Current wake word"), "主", 0x43),
        row(44, t.restore_wake, t.restore_wake, "命令词", ("已恢复默认唤醒词" if lang == "zh" else "Default wake word restored"), "主", 0x44),
    ]


def multi_loop_rows(lang: str) -> List[List[Any]]:
    return dedupe_rows(base_core_rows(lang) + multi_common_rows(lang))


def multi_specified_rows(lang: str) -> List[List[Any]]:
    t = TEXT[lang]
    extra = [
        row(45, t.switch_to_alt_1, t.switch_to_alt_1, "命令词", ("已切换到候选唤醒词一" if lang == "zh" else "Switched to candidate one"), "主", 0x45),
        row(46, t.switch_to_alt_2, t.switch_to_alt_2, "命令词", ("已切换到候选唤醒词二" if lang == "zh" else "Switched to candidate two"), "主", 0x46),
    ]
    return dedupe_rows(base_core_rows(lang) + multi_common_rows(lang) + extra)


def multi_protocol_rows(lang: str) -> List[List[Any]]:
    t = TEXT[lang]
    extra = [
        row(47, t.switch_to_alt_1, t.switch_to_alt_1, "命令词", ("等待协议确认" if lang == "zh" else "Waiting for protocol confirm"), "被", 0x47),
        row(48, t.switch_to_alt_2, t.switch_to_alt_2, "命令词", ("等待协议确认" if lang == "zh" else "Waiting for protocol confirm"), "被", 0x48),
        row(49, None, None, "播报语", ("协议切换成功" if lang == "zh" else "Protocol switch succeeded"), "被", 0x49),
    ]
    return dedupe_rows(base_core_rows(lang) + multi_common_rows(lang) + extra)


def voice_reg_common_rows(lang: str) -> List[List[Any]]:
    t = TEXT[lang]
    # Voice-registration control words are generated by the UI registration
    # config as special commands. Importing them here as normal protocol rows
    # makes runtime hit the protocol command instead of the registration state.
    return [
        row(57, t.valid_learn_target_1, t.valid_learn_target_1, "命令词", t.reply_ok, "主", 0x57),
        row(58, t.valid_learn_target_2, t.valid_learn_target_2, "命令词", t.reply_ok, "主", 0x58),
    ]


def voice_reg_specific_rows(lang: str) -> List[List[Any]]:
    return dedupe_rows(base_core_rows(lang) + voice_reg_common_rows(lang))


def voice_reg_continuous_rows(lang: str) -> List[List[Any]]:
    # contLearn uses the same UI control words as specificLearn; the difference is
    # releaseRegist.registMode and runtime state-machine expectations, not duplicate
    # command rows in the import template.
    return dedupe_rows(base_core_rows(lang) + voice_reg_common_rows(lang))


def voice_reg_boundary_delete_rows(lang: str) -> List[List[Any]]:
    t = TEXT[lang]
    extra = [
        row(61, t.short_boundary_hint, t.short_boundary_hint, "命令词", ("边界测试短词" if lang == "zh" else "Short boundary sample"), "主", 0x61),
        row(62, t.long_boundary_hint, t.long_boundary_hint, "命令词", ("边界测试长词" if lang == "zh" else "Long boundary sample"), "主", 0x62),
    ]
    return dedupe_rows(base_core_rows(lang) + voice_reg_common_rows(lang) + extra)


def depth_tuning_rows(lang: str) -> List[List[Any]]:
    t = TEXT[lang]
    rows = base_core_rows(lang) + [
        row(65, t.wake_alt_1, None, "唤醒词", ("备用唤醒命中" if lang == "zh" else "Backup wake hit"), "主", 0x65),
        row(66, "打开窗帘" if lang == "zh" else "Open Curtain", "[打开一号窗帘/开启窗帘]" if lang == "zh" else "Open Curtain", "命令词", t.reply_ok, "主", 0x66),
        row(67, "关闭窗帘" if lang == "zh" else "Close Curtain", "[关闭一号窗帘/合上窗帘]" if lang == "zh" else "Close Curtain", "命令词", t.reply_ok, "主", 0x67),
    ]
    return dedupe_rows(rows)


def full_stateful_rows(lang: str) -> List[List[Any]]:
    return dedupe_rows(
        base_core_rows(lang)
        + multi_common_rows(lang)
        + [r for r in multi_specified_rows(lang) if r[0] in {45, 46}]
        + voice_reg_common_rows(lang)
        + [r for r in protocol_rows(lang) if r[0] in {30, 31, 32, 33}]
    )


BASE_PARAMS = [
    "timeout",
    "volLevel",
    "defaultVol",
    "volMaxOverflow",
    "volMinOverflow",
    "uportBaud",
    "logLevel",
    "volSave",
    "releaseAlgoList[*].word",
    "releaseAlgoList[*].extWord",
    "releaseAlgoList[*].children[*].extWord",
    "releaseAlgoList[*].type",
    "releaseAlgoList[*].reply",
    "releaseAlgoList[*].replyMode",
    "releaseAlgoList[*].sndProtocol",
    "releaseAlgoList[*].recProtocol",
]
PROTOCOL_PARAMS = ["releaseAlgoList[*].replyMode", "releaseAlgoList[*].sndProtocol", "releaseAlgoList[*].recProtocol", "protocolConfig"]
MULTI_PARAMS = [
    "multiWkeEnable",
    "multiWkeMode",
    "wakeWordSave",
    "releaseMultiWke.common[*].condition",
    "releaseMultiWke.common[*].reply",
    "releaseMultiWke.wkelist[*].condition",
    "releaseMultiWke.wkelist[*].reply",
    "releaseMultiWke.wkelist[*].sndProtocol",
    "releaseMultiWke.wkelist[*].recProtocol",
    "releaseMultiWke.wkelist[*].isDefault",
    "releaseMultiWke.wkelist[*].isFrozen",
]
VOICE_REG_PARAMS = [
    "voiceRegEnable",
    "releaseRegist.registMode",
    "releaseRegist.wakeupRepeatCount",
    "releaseRegist.commandRepeatCount",
    "releaseRegist.wakeupRetryCount",
    "releaseRegist.commandRetryCount",
    "releaseRegist.wakeupRegistMaxLimit",
    "releaseRegist.commandRegistMaxLimit",
    "releaseRegist.wakeupWordsMinLimit",
    "releaseRegist.wakeupWordsMaxLimit",
    "releaseRegist.commandWordsMinLimit",
    "releaseRegist.commandWordsMaxLimit",
    "releaseRegist.wakeupSensitivity",
    "releaseRegist.commandSensitivity",
    "releaseRegist.reply",
    "releaseRegist.replyMode",
    "releaseRegist.sndProtocol",
    "releaseRegist.recProtocol",
    "releaseRegistConfig.*.triggers.*.stages.*.(condition|reply|delReply)",
]
DEPTH_PARAMS = [
    "sensitivity",
    "releaseDepthList[*].pinyin",
    "releaseDepthList[*].decEnable",
    "releaseDepthList[*].decThreshold",
    "releaseDepthList[*].e2eEnable",
    "releaseDepthList[*].e2eThreshold",
    "releaseDepthList[*].embeddedEnable",
    "releaseDepthList[*].embeddedThreshold",
    "releaseDepthList[*].asrFreeEnable",
    "releaseDepthList[*].asrFreeThreshold",
]


def make_profiles() -> List[Profile]:
    profile_specs = [
        {
            "id": "base_core",
            "purpose": "基础功能：默认唤醒、业务命令、音量边界、退出识别、负性词、欢迎/播报/休息/心跳协议。",
            "recommended_for": ["基础中值包", "左右边界包", "无专项功能的垂类烟测"],
            "required_config": {"voiceRegEnable": False, "multiWkeEnable": False, "replyMode": "主+被", "protocol": "snd+rec"},
            "covers": ["基础唤醒", "业务命令", "音量调节/边界", "主动播报", "被动播报", "心跳协议", "负性词"],
            "parameters": BASE_PARAMS,
            "risk": "低：推荐作为所有包的基础模板。",
            "factory": base_core_rows,
        },
        {
            "id": "protocol_active_passive",
            "purpose": "协议专项：主动识别发送协议、被动接收协议触发播报、被动命令回复、心跳协议。",
            "recommended_for": ["协议主动/被动专项", "recProtocol/sndProtocol 回归", "replyMode 主/被切换验证"],
            "required_config": {"replyMode": "主+被", "protocol": "snd+rec", "uportBaud": "按包配置"},
            "covers": ["主动协议", "被动协议", "被动播报", "被动命令", "心跳协议"],
            "parameters": BASE_PARAMS + PROTOCOL_PARAMS,
            "risk": "低：协议口波特率必须和包配置一致。",
            "factory": protocol_rows,
        },
        {
            "id": "multi_wakeup_loop",
            "purpose": "多唤醒循环切换：默认唤醒词、两个候选唤醒词、切换/查询/恢复公共指令。",
            "recommended_for": ["multiWkeMode=loop", "wakeWordSave 与循环切换联动"],
            "required_config": {"multiWkeEnable": True, "multiWkeMode": "loop", "wakeWordSave": "true/false 分包覆盖", "extraWakeWords": 2},
            "covers": ["多唤醒候选词", "循环切换", "查询唤醒词", "恢复默认", "掉电保持联动"],
            "parameters": BASE_PARAMS + MULTI_PARAMS,
            "risk": "中：需要产品支持多唤醒且候选词不能与默认唤醒冲突。",
            "factory": multi_loop_rows,
        },
        {
            "id": "multi_wakeup_specified",
            "purpose": "多唤醒指定切换：候选唤醒词、指定切换到候选词、查询和恢复默认。",
            "recommended_for": ["multiWkeMode=specified", "指定切换正反例", "默认/冻结唤醒词验证"],
            "required_config": {"multiWkeEnable": True, "multiWkeMode": "specified", "extraWakeWords": 2},
            "covers": ["指定切换", "候选唤醒词", "默认唤醒词", "冻结唤醒词", "查询/恢复"],
            "parameters": BASE_PARAMS + MULTI_PARAMS,
            "risk": "中：必须通过 UI 当前多唤醒表格配置候选词属性。",
            "factory": multi_specified_rows,
        },
        {
            "id": "multi_wakeup_protocol",
            "purpose": "多唤醒协议切换：协议确认型候选切换、协议触发播报、查询/恢复公共链路。",
            "recommended_for": ["multiWkeMode=protocol", "协议切换候选唤醒词"],
            "required_config": {"multiWkeEnable": True, "multiWkeMode": "protocol", "releaseMultiWke.wkelist[*].recProtocol": "必填"},
            "covers": ["协议切换唤醒词", "多唤醒 snd/recProtocol", "被动播报", "查询/恢复"],
            "parameters": BASE_PARAMS + MULTI_PARAMS + PROTOCOL_PARAMS,
            "risk": "中：协议口和波特率不一致会造成假失败。",
            "factory": multi_protocol_rows,
        },
        {
            "id": "voice_reg_specific",
            "purpose": "语音注册指定学习：基础宿主动作、可学习目标命令；学习/删除/退出控制词由 UI 语音注册配置生成 special 词条。",
            "recommended_for": ["voiceRegEnable=true", "releaseRegist.registMode=specificLearn", "repeat/retry 正反例"],
            "required_config": {"voiceRegEnable": True, "releaseRegist.registMode": "specificLearn"},
            "covers": ["指定学习", "学习命令词", "学习唤醒词", "删除命令词", "删除唤醒词", "退出学习/删除", "全部删除"],
            "parameters": BASE_PARAMS + VOICE_REG_PARAMS,
            "risk": "中：模板内不能重复导入普通协议控制词。",
            "factory": voice_reg_specific_rows,
        },
        {
            "id": "voice_reg_continuous",
            "purpose": "语音注册连续学习：基础宿主动作、可学习目标命令；连续学习入口由 UI 语音注册配置生成 special 词条。",
            "recommended_for": ["releaseRegist.registMode=contLearn", "连续学习状态机回归"],
            "required_config": {"voiceRegEnable": True, "releaseRegist.registMode": "contLearn"},
            "covers": ["连续学习", "自动下一条", "两步删除命令词", "模板满处理", "重试耗尽"],
            "parameters": BASE_PARAMS + VOICE_REG_PARAMS,
            "risk": "中：学习态需等待提示播报结束和算法重建完成。",
            "factory": voice_reg_continuous_rows,
        },
        {
            "id": "voice_reg_boundary_delete",
            "purpose": "语音注册边界/删除：字数上下限、重试次数、模板上限、删除闭环；删除入口由 UI special 词条提供。",
            "recommended_for": ["语音注册边界包", "删除闭环包", "retryCount/repeatCount/maxLimit 边界"],
            "required_config": {"voiceRegEnable": True, "releaseRegist.*Limit": "按边界包配置", "releaseRegist.*RetryCount": "1/2/3 分包覆盖"},
            "covers": ["字数上下限", "重试次数", "模板上限", "删除闭环", "失败后不生效"],
            "parameters": BASE_PARAMS + VOICE_REG_PARAMS,
            "risk": "中：字数边界语料运行时合成，模板只提供宿主动作。",
            "factory": voice_reg_boundary_delete_rows,
        },
        {
            "id": "depth_tuning",
            "purpose": "深度调优：提供唤醒词、命令词、子泛化词基础数据，用于 pinyin/DEC/E2E/embedded/ASRFree 调优。",
            "recommended_for": ["灵敏度/深度调优专项", "releaseDepthList 阈值边界"],
            "required_config": {"sensitivity": "low/mid/high 分包覆盖", "depthThreshold": "按推荐边界配置"},
            "covers": ["唤醒词调优", "命令词调优", "子泛化词", "pinyin", "DEC/E2E/ASRFree/embedded"],
            "parameters": BASE_PARAMS + DEPTH_PARAMS,
            "risk": "中：阈值变化需要固定音频集统计，不宜仅凭一次识别判定。",
            "factory": depth_tuning_rows,
        },
        {
            "id": "full_feature_stateful",
            "purpose": "综合状态包：基础功能 + 协议主动/被动 + 多唤醒指定切换 + 语音注册宿主动作。",
            "recommended_for": ["全功能保持开启包", "功能耦合冒烟", "报告前综合回归"],
            "required_config": {"multiWkeEnable": True, "voiceRegEnable": True, "wakeWordSave": True, "volSave": True},
            "covers": ["基础功能", "协议", "多唤醒", "语音注册", "掉电保持"],
            "parameters": BASE_PARAMS + PROTOCOL_PARAMS + MULTI_PARAMS + VOICE_REG_PARAMS,
            "risk": "高：词量较多；语音注册控制词必须由 UI special 配置生成，3021 上如出现内存超限应拆分专项模板定位。",
            "factory": full_stateful_rows,
        },
    ]
    profiles: List[Profile] = []
    for lang in ["zh", "en"]:
        for spec in profile_specs:
            profiles.append(
                Profile(
                    id=f"{lang}_{spec['id']}",
                    language=lang,
                    filename=f"algo_{lang}_{spec['id']}.xlsx",
                    purpose=spec["purpose"],
                    recommended_for=list(spec["recommended_for"]),
                    required_config=dict(spec["required_config"]),
                    covers=list(spec["covers"]),
                    parameters=list(dict.fromkeys(spec["parameters"])),
                    risk=spec["risk"],
                    row_factory=spec["factory"],
                )
            )
    # Backward-compatible aliases. Existing docs/scripts may still reference these names.
    alias_map = {
        "zh_basic": "zh_base_core",
        "zh_multi_wakeup": "zh_multi_wakeup_specified",
        "zh_voice_register": "zh_voice_reg_specific",
        "en_basic": "en_base_core",
        "en_multi_wakeup": "en_multi_wakeup_specified",
        "en_voice_register": "en_voice_reg_specific",
    }
    by_id = {profile.id: profile for profile in profiles}
    for alias_id, target_id in alias_map.items():
        target = by_id[target_id]
        profiles.append(
            Profile(
                id=alias_id,
                language=target.language,
                filename=f"algo_{alias_id}.xlsx",
                purpose=f"兼容旧文件名，内容等同于 {target.id}: {target.purpose}",
                recommended_for=target.recommended_for,
                required_config=target.required_config,
                covers=target.covers,
                parameters=target.parameters,
                risk=target.risk,
                row_factory=target.row_factory,
                compat_alias_for=target.id,
            )
        )
    return profiles


PARAMETER_REQUIREMENTS = [
    {
        "group": "基础配置",
        "parameters": ["timeout", "volLevel", "defaultVol", "volMaxOverflow", "volMinOverflow", "uportBaud", "logLevel", "volSave"],
        "template_types": ["base_core", "protocol_active_passive"],
        "why": "需要唤醒、命令词、音量调节、退出识别、主动/被动播报数据支撑设备侧验证。",
    },
    {
        "group": "算法词条",
        "parameters": [
            "releaseAlgoList[*].word",
            "releaseAlgoList[*].extWord",
            "releaseAlgoList[*].children[*].extWord",
            "releaseAlgoList[*].type",
            "releaseAlgoList[*].reply",
            "releaseAlgoList[*].replyMode",
            "releaseAlgoList[*].sndProtocol",
            "releaseAlgoList[*].recProtocol",
        ],
        "template_types": ["base_core", "protocol_active_passive", "depth_tuning"],
        "why": "模板必须提供不同功能类型、父/子泛化词、回复语、主/被播报和协议字段。",
    },
    {
        "group": "多唤醒",
        "parameters": MULTI_PARAMS,
        "template_types": ["multi_wakeup_loop", "multi_wakeup_specified", "multi_wakeup_protocol"],
        "why": "loop/specified/protocol 三种模式的切换触发方式不同，必须分模板分包验证。",
    },
    {
        "group": "语音注册",
        "parameters": VOICE_REG_PARAMS,
        "template_types": ["voice_reg_specific", "voice_reg_continuous", "voice_reg_boundary_delete"],
        "why": "specificLearn 与 contLearn 状态机不同；边界、重试、模板上限和删除闭环需要独立数据。",
    },
    {
        "group": "深度调优",
        "parameters": DEPTH_PARAMS,
        "template_types": ["depth_tuning"],
        "why": "阈值、拼音和使能项需要同时存在唤醒词、命令词、子泛化词，便于固定音频集统计。",
    },
    {
        "group": "综合状态",
        "parameters": ["wakeWordSave", "volSave", "multiWkeEnable", "voiceRegEnable"],
        "template_types": ["full_feature_stateful"],
        "why": "状态保持和能力耦合需要同包冒烟，但容量异常时必须拆回专项模板定位。",
    },
]


def clone_template(base: Path, out: Path, rows: Sequence[Sequence[Any]]) -> None:
    wb = load_workbook(base)
    if SHEET not in wb.sheetnames:
        raise RuntimeError(f"missing sheet {SHEET}: {base}")
    ws = wb[SHEET]
    if ws.max_row > 1:
        ws.delete_rows(2, ws.max_row - 1)
    for col, header in enumerate(HEADERS, start=1):
        ws.cell(row=1, column=col, value=header)
    for values in rows:
        ws.append(list(values))
    out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out)


def render_matrix_md(manifest: Dict[str, Any]) -> str:
    lines = [
        "# 固件打包算法模板覆盖矩阵",
        "",
        "生成来源：`scripts/ui/generate_algo_template_variants.py`。",
        "",
        "## 使用原则",
        "",
        "- 模板按参数能力和测试类型选择，不按“中文/英文基础三件套”粗略选择。",
        "- 正式 UI 打包优先使用平台 UI 当前下载的最新模板作为底板；本目录生成的模板是 fallback 和测试数据参考。",
        "- UI 页面下拉、芯片、语言、垂类、SDK 版本必须运行时从 UI 当前页面确认，不能从本 manifest 固化。",
        "- 语音注册负例不得使用语音注册控制词，避免误触发真实学习/删除功能。",
        "- 语音注册算法模板的 `词条预处理` 不得导入 `学习命令词/删除命令词/学习唤醒词/删除唤醒词/删除全部命令词/退出学习/退出删除` 等普通协议命令；这些控制词必须由 UI 语音注册配置生成 special 控制词。否则运行时会命中普通协议命令并发送协议帧，无法进入学习态。",
        "- 综合模板词量较大；容量或编译失败时拆分为专项模板复测。",
        "",
        "## 模板清单",
        "",
        "| 模板 | 语言 | 用途 | 适用测试 | 覆盖能力 | 风险 |",
        "| --- | --- | --- | --- | --- | --- |",
    ]
    for item in manifest["templates"]:
        if item.get("compatAliasFor"):
            continue
        lines.append(
            "| "
            + " | ".join(
                [
                    f"`{item['file']}`",
                    item["language"],
                    item["purpose"],
                    "；".join(item["recommendedFor"]),
                    "；".join(item["covers"]),
                    item["risk"],
                ]
            )
            + " |"
        )
    lines.extend(["", "## 参数到模板映射", "", "| 配置组 | 参数 | 应使用模板类型 | 数据目的 |", "| --- | --- | --- | --- |"])
    for req in manifest["parameterRequirements"]:
        lines.append(
            "| "
            + " | ".join(
                [
                    req["group"],
                    "<br>".join(f"`{p}`" for p in req["parameters"]),
                    "<br>".join(f"`{p}`" for p in req["templateTypes"]),
                    req["why"],
                ]
            )
            + " |"
        )
    lines.extend(["", "## 兼容旧文件名", "", "| 旧模板名 | 等价 profile |", "| --- | --- |"])
    for item in manifest["templates"]:
        if item.get("compatAliasFor"):
            lines.append(f"| `{item['file']}` | `{item['compatAliasFor']}` |")
    return "\n".join(lines).rstrip() + "\n"


def build(out_dir: Path, zh_base: Path, en_base: Path) -> Dict[str, Any]:
    manifest: Dict[str, Any] = {
        "generatedBy": "scripts/ui/generate_algo_template_variants.py",
        "schemaVersion": 2,
        "dataSourcePrinciple": "Generated templates are fallback test data. Real UI packaging must confirm current UI options and prefer the latest UI-downloaded template as base.",
        "voiceRegControlRule": "语音注册控制词不得在算法模板词条预处理中作为普通协议命令重复导入；学习/删除/退出控制入口必须由 UI 语音注册配置生成 special_type=语音注册控制相关 词条。",
        "parameterRequirements": [
            {
                "group": item["group"],
                "parameters": item["parameters"],
                "templateTypes": item["template_types"],
                "why": item["why"],
            }
            for item in PARAMETER_REQUIREMENTS
        ],
        "templates": [],
    }
    for profile in make_profiles():
        base = zh_base if profile.language == "zh" else en_base
        out = out_dir / profile.filename
        rows = profile.row_factory(profile.language)
        clone_template(base, out, rows)
        manifest["templates"].append(
            {
                "id": profile.id,
                "file": str(out),
                "language": profile.language,
                "purpose": profile.purpose,
                "recommendedFor": profile.recommended_for,
                "requiredConfig": profile.required_config,
                "covers": profile.covers,
                "parameters": profile.parameters,
                "risk": profile.risk,
                "rows": len(rows),
                "sourceBase": str(base),
                **({"compatAliasFor": profile.compat_alias_for} if profile.compat_alias_for else {}),
            }
        )
    manifest_path = out_dir / "template_manifest.json"
    matrix_path = out_dir / "template_requirement_matrix.md"
    manifest_path.write_text(json.dumps(manifest, ensure_ascii=False, indent=2), encoding="utf-8")
    matrix_path.write_text(render_matrix_md(manifest), encoding="utf-8")
    return manifest


def main() -> int:
    parser = argparse.ArgumentParser(description="Generate parameter-driven Chinese/English algorithm templates for UI firmware packaging tests.")
    parser.add_argument("--zh-base", default="assets/templates/聆思科技_命令词播报词协议配置表V1.0_中文模板.xlsx")
    parser.add_argument("--en-base", default="assets/templates/聆思科技_算法配置英文模板.xlsx")
    parser.add_argument("--out-dir", default="assets/templates")
    args = parser.parse_args()
    manifest = build(Path(args.out_dir), Path(args.zh_base), Path(args.en_base))
    print(json.dumps({"templates": len(manifest["templates"]), "matrix": str(Path(args.out_dir) / "template_requirement_matrix.md")}, ensure_ascii=False, indent=2))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
