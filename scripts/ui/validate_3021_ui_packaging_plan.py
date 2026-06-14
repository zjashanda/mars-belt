#!/usr/bin/env python3
from __future__ import annotations

import argparse
import json
import re
from collections import Counter, defaultdict
from pathlib import Path
from typing import Any, Dict, Iterable, List, Tuple

RESERVED_VOICE_REG_CONTROLS = {
    "学习命令词",
    "删除命令词",
    "学习唤醒词",
    "删除唤醒词",
    "删除全部命令词",
    "退出学习",
    "退出删除",
}


def norm(text: Any) -> str:
    return re.sub(r"\s+", "", str(text or ""))


def load_plan(path: Path) -> Dict[str, Any]:
    data = json.loads(path.read_text(encoding="utf-8"))
    if not isinstance(data, dict) or not isinstance(data.get("jobs"), list):
        raise SystemExit(f"plan must be a JSON object with jobs[]: {path}")
    return data


def read_xlsx_cells(path: Path) -> List[str]:
    try:
        from openpyxl import load_workbook
    except Exception as exc:  # pragma: no cover - environment issue
        raise SystemExit(f"openpyxl is required to validate xlsx templates: {exc}") from exc
    wb = load_workbook(path, data_only=True, read_only=True)
    cells: List[str] = []
    for ws in wb.worksheets:
        for row in ws.iter_rows(values_only=True):
            for value in row:
                if value is None:
                    continue
                text = str(value).strip()
                if text:
                    cells.append(text)
    return cells


def vertical_name(job: Dict[str, Any]) -> str:
    version = str(((job.get("product") or {}).get("versionLabel")) or "")
    return version.split("-")[0] if version else "UNKNOWN"


def feature_state(job: Dict[str, Any], key: str) -> str:
    return str((job.get("feature") or {}).get(key) or "Unsupported")


def expected_min_jobs(jobs: List[Dict[str, Any]]) -> Dict[str, int]:
    by_vertical: Dict[str, List[Dict[str, Any]]] = defaultdict(list)
    for job in jobs:
        by_vertical[vertical_name(job)].append(job)
    expected: Dict[str, int] = {}
    for name, items in by_vertical.items():
        any_voice = any(feature_state(item, "voice_regist") == "Optional" for item in items)
        any_multi = any(feature_state(item, "multi_wakeup") == "Optional" for item in items)
        # Default minimal rule used by this skill: three vectors for ordinary or
        # multi-only verticals; add one off-negative isolation vector when voice
        # registration is enabled because voice control words must be isolated.
        expected[name] = 4 if any_voice else 3 if any_multi else 3
    return expected


def validate(plan: Dict[str, Any], root: Path, allow_v1: bool = False) -> Tuple[List[Dict[str, Any]], Dict[str, Any]]:
    jobs: List[Dict[str, Any]] = plan["jobs"]
    findings: List[Dict[str, Any]] = []
    template_cache: Dict[Path, List[str]] = {}

    def add(level: str, code: str, job_id: str, message: str, **extra: Any) -> None:
        item = {"level": level, "code": code, "jobId": job_id, "message": message}
        item.update(extra)
        findings.append(item)

    ids = [str(j.get("jobId") or "") for j in jobs]
    for job_id, count in Counter(ids).items():
        if not job_id:
            add("FAIL", "PLAN-JOBID-MISSING", "", "jobId is required")
        elif count > 1:
            add("FAIL", "PLAN-JOBID-DUP", job_id, f"duplicate jobId count={count}")

    product_names_by_vertical: Dict[str, set[str]] = defaultdict(set)
    profiles_by_vertical: Dict[str, set[str]] = defaultdict(set)
    product_labels_by_vertical: Dict[str, set[str]] = defaultdict(set)

    for job in jobs:
        job_id = str(job.get("jobId") or "")
        product = job.get("product") or {}
        cfg = job.get("algoConfig") or {}
        feature = job.get("feature") or {}
        vertical = vertical_name(job)
        product_name = str(job.get("productName") or "")
        profile = str(job.get("profile") or "")
        template = str(job.get("algoTemplate") or "")
        template_path = (root / template).resolve()

        product_names_by_vertical[vertical].add(product_name)
        profiles_by_vertical[vertical].add(profile)
        product_labels_by_vertical[vertical].add(str(product.get("productLabel") or product.get("productValue") or ""))

        if "打包测试" not in product_name:
            add("FAIL", "PLAN-PRODUCT-NAME", job_id, "productName must include 打包测试", productName=product_name)
        if not (str(product.get("moduleBoard") or "").startswith("CSK3021") or str(product.get("moduleMark") or "") == "CSK3021"):
            add("FAIL", "PLAN-MODULE", job_id, "job is not scoped to 3021", product=product)
        if not product.get("versionLabel") or not product.get("productLabel"):
            add("FAIL", "PLAN-PRODUCT-FIELDS", job_id, "productLabel/versionLabel are required", product=product)
        if "V1.0" in str(product.get("versionLabel") or "") and not allow_v1:
            add("FAIL", "PLAN-LEGACY-V1", job_id, "V1.0 legacy version is not allowed in current UI-only packaging unless explicitly allowed", versionLabel=product.get("versionLabel"))
        if not template:
            add("FAIL", "PLAN-TEMPLATE-MISSING", job_id, "algoTemplate is required")
            continue
        if not template_path.exists():
            add("FAIL", "PLAN-TEMPLATE-NOT-FOUND", job_id, f"template not found: {template}")
            continue

        cells = template_cache.get(template_path)
        if cells is None:
            cells = read_xlsx_cells(template_path)
            template_cache[template_path] = cells
        normalized_cells = {norm(cell) for cell in cells}
        reserved_hits = sorted(RESERVED_VOICE_REG_CONTROLS & set(cells))
        if reserved_hits:
            add("FAIL", "VOICE-REG-TEMPLATE-RESERVED", job_id, "voice registration control words must not be imported as normal algorithm rows", template=template, hits=reserved_hits)

        if cfg.get("voiceRegEnable"):
            if feature.get("voice_regist") != "Optional":
                add("FAIL", "VOICE-REG-FEATURE-GATE", job_id, "voiceRegEnable=true but feature.voice_regist is not Optional", feature=feature)
            if cfg.get("registMode") not in {"specificLearn", "contLearn"}:
                add("FAIL", "VOICE-REG-MODE", job_id, "voiceRegEnable requires registMode specificLearn or contLearn", algoConfig=cfg)
            learn_words = [str(x) for x in (cfg.get("studyCommandWords") or []) if str(x).strip()]
            if not learn_words:
                add("FAIL", "VOICE-REG-LEARN-WORDS", job_id, "voice registration package must declare studyCommandWords")
            bad = sorted(set(learn_words) & RESERVED_VOICE_REG_CONTROLS)
            if bad:
                add("FAIL", "VOICE-REG-LEARN-WORDS-RESERVED", job_id, "studyCommandWords must not use registration control words", words=bad)
            missing = [w for w in learn_words if norm(w) not in normalized_cells]
            if missing:
                add("FAIL", "VOICE-REG-LEARN-WORDS-NOT-IN-TEMPLATE", job_id, "studyCommandWords must be selectable from imported algorithm command rows", missing=missing, template=template)

        if cfg.get("multiWkeEnable") and feature.get("multi_wakeup") != "Optional":
            add("FAIL", "MULTI-WAKE-FEATURE-GATE", job_id, "multiWkeEnable=true but feature.multi_wakeup is not Optional", feature=feature)
        if cfg.get("multiWkeEnable") and cfg.get("multiWkeMode") not in {"specified", "loop", "protocol"}:
            add("FAIL", "MULTI-WAKE-MODE", job_id, "multiWkeEnable requires multiWkeMode specified/loop/protocol", algoConfig=cfg)

    for vertical, names in sorted(product_names_by_vertical.items()):
        if len(names) != 1:
            add("FAIL", "PLAN-PRODUCT-GROUPING", vertical, "one representative product container per vertical is required", productNames=sorted(names))
    for vertical, labels in sorted(product_labels_by_vertical.items()):
        if len(labels) != 1:
            add("FAIL", "PLAN-REPRESENTATIVE-PRODUCT", vertical, "one representative product category per vertical is required", productLabels=sorted(labels))
    expected = expected_min_jobs(jobs)
    for vertical, want in sorted(expected.items()):
        got = len([j for j in jobs if vertical_name(j) == vertical])
        if got < want:
            add("FAIL", "PLAN-MINIMAL-COVERAGE-COUNT", vertical, f"vertical has too few package vectors: got={got}, expected>={want}")

    summary = {
        "jobs": len(jobs),
        "verticals": len(product_names_by_vertical),
        "fail": sum(1 for f in findings if f["level"] == "FAIL"),
        "warn": sum(1 for f in findings if f["level"] == "WARN"),
        "expectedMinJobsByVertical": expected,
        "productsByVertical": {k: sorted(v) for k, v in sorted(product_names_by_vertical.items())},
        "profilesByVertical": {k: sorted(v) for k, v in sorted(profiles_by_vertical.items())},
    }
    return findings, summary


def write_md(path: Path, findings: List[Dict[str, Any]], summary: Dict[str, Any]) -> None:
    lines = [
        "# 3021 UI Packaging Plan Preflight",
        "",
        f"- Jobs: {summary['jobs']}",
        f"- Verticals: {summary['verticals']}",
        f"- FAIL: {summary['fail']}",
        f"- WARN: {summary['warn']}",
        "",
        "## Products By Vertical",
        "",
    ]
    for vertical, names in summary["productsByVertical"].items():
        lines.append(f"- {vertical}: {', '.join(names)}")
    lines += ["", "## Profiles By Vertical", ""]
    for vertical, profiles in summary["profilesByVertical"].items():
        lines.append(f"- {vertical}: {', '.join(profiles)}")
    if findings:
        lines += ["", "## Findings", ""]
        for item in findings:
            lines.append(f"- {item['level']} `{item['code']}` `{item['jobId']}`: {item['message']}")
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text("\n".join(lines) + "\n", encoding="utf-8")


def main() -> int:
    parser = argparse.ArgumentParser(description="Validate a 3021 UI-only packaging plan before submitting browser automation.")
    parser.add_argument("--plan", required=True)
    parser.add_argument("--repo-root", default=".")
    parser.add_argument("--out-json", default="")
    parser.add_argument("--out-md", default="")
    parser.add_argument("--allow-v1", action="store_true")
    args = parser.parse_args()

    root = Path(args.repo_root).resolve()
    plan_path = Path(args.plan)
    plan = load_plan(plan_path)
    findings, summary = validate(plan, root, allow_v1=args.allow_v1)
    payload = {"summary": summary, "findings": findings}
    print(json.dumps(payload, ensure_ascii=False, indent=2))
    if args.out_json:
        out = Path(args.out_json)
        out.parent.mkdir(parents=True, exist_ok=True)
        out.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    if args.out_md:
        write_md(Path(args.out_md), findings, summary)
    return 1 if summary["fail"] else 0


if __name__ == "__main__":
    raise SystemExit(main())
