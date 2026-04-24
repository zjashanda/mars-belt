#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
轻量化语音功能测试脚本
仅依赖: pyserial (pip install pyserial)

功能:
  1. 串口日志读取 + 正则匹配 (唤醒词/识别词/协议等)
  2. 播放音频唤醒设备，播报命令词
  3. ASR 识别结果比对 (拼音自动转汉字)
  4. 命令词与协议对比
  5. 同义词判定 (同协议词/absorb词不算串扰)
  6. 测试结果 CSV 输出
  7. 原始串口日志保存
  8. 正则表达式自动发现 (首次运行)
  9. 设备重启检测 & 自动恢复日志等级
  10. 测试后 CrossTalk 自动分析 & spell2zh 修复
"""

import json
import os
import sys
import re
import time
import datetime
import threading
import argparse
import subprocess
import platform
import asyncio
import hashlib
import base64
import shlex
import shutil
import csv as csv_mod
from collections import defaultdict, deque

from listenai_audio_skill_bootstrap import resolve_listenai_play_script

try:
    import serial
except ImportError:
    print("缺少 pyserial，请执行: pip install pyserial")
    sys.exit(1)


class AbortTestRun(RuntimeError):
    """Raised when the device enters an unrecoverable test state."""


# ══════════════════════════════════════════════════════════════
#  日志
# ══════════════════════════════════════════════════════════════
class Logger:
    """轻量日志，同时输出到控制台和文件"""

    LEVELS = {"DEBUG": 0, "INFO": 1, "WARN": 2, "ERROR": 3}

    def __init__(self, log_dir, level="INFO"):
        self.min_level = self.LEVELS.get(level, 1)
        os.makedirs(log_dir, exist_ok=True)
        ts = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
        self._fp = open(os.path.join(log_dir, f"test_{ts}.log"), "a", encoding="utf-8")

    def _log(self, level, msg):
        now = datetime.datetime.now().strftime("%H:%M:%S.%f")[:-3]
        line = f"[{now}][{level}] {msg}"
        if self.LEVELS.get(level, 0) >= self.min_level:
            print(line)
        self._fp.write(line + "\n")
        self._fp.flush()

    def debug(self, msg): self._log("DEBUG", msg)
    def info(self, msg):  self._log("INFO", msg)
    def warn(self, msg):  self._log("WARN", msg)
    def error(self, msg): self._log("ERROR", msg)

    def close(self):
        self._fp.close()


SAFE_RESULT_PATH_LIMIT = 240
ANSI_ESCAPE_RE = re.compile(r"\x1B\[[0-?]*[ -/]*[@-~]")
ILLEGAL_XLSX_RE = re.compile(r"[\x00-\x08\x0B-\x0C\x0E-\x1F]")


def sanitize_log_text(value):
    return ANSI_ESCAPE_RE.sub("", str(value or ""))


def sanitize_xlsx_text(value):
    return ILLEGAL_XLSX_RE.sub("", sanitize_log_text(value))


def normalize_hex_protocol(value):
    text = str(value or "").strip()
    if not text:
        return ""
    parts = [part for part in text.split() if part]
    if not parts:
        return ""
    if not all(re.fullmatch(r"[0-9A-Fa-f]{2}", part) for part in parts):
        return ""
    return " ".join(part.upper() for part in parts)


def is_framed_uart_protocol(value):
    normalized = normalize_hex_protocol(value)
    if not normalized:
        return False
    parts = normalized.split()
    return len(parts) >= 4 and parts[0] == "A5" and parts[1] == "FA" and parts[-1] == "FB"


def _safe_result_token(text, fallback):
    token = re.sub(r'[\\/:*?"<>|]+', "_", str(text or "").strip())
    token = re.sub(r"\s+", "_", token)
    token = token.strip("._ ")
    return token or fallback


def build_result_dir_name(base_dir, timestamp, project, label):
    project_token = _safe_result_token(project, "project")
    label_token = _safe_result_token(label, "run")
    candidate = f"{timestamp}_{project_token}_{label_token}"
    probe = os.path.join(base_dir, "result", candidate, "test_YYYYMMDD_HHMMSS.log")
    if len(probe) <= SAFE_RESULT_PATH_LIMIT:
        return candidate

    digest = hashlib.sha1(f"{project_token}|{label_token}".encode("utf-8")).hexdigest()[:10]
    short_candidate = f"{timestamp}_{project_token[:24]}_{label_token[:24]}_{digest}"
    short_probe = os.path.join(base_dir, "result", short_candidate, "test_YYYYMMDD_HHMMSS.log")
    if len(short_probe) <= SAFE_RESULT_PATH_LIMIT:
        return short_candidate
    return f"{timestamp}_{label_token[:32]}_{digest}"


# ══════════════════════════════════════════════════════════════
#  串口读取线程
# ══════════════════════════════════════════════════════════════
class SerialReader(threading.Thread):
    """后台线程持续读取串口，按正则提取关键字段"""

    def __init__(self, port, baudrate, regex_map, log, serial_log_dir=None):
        super().__init__(daemon=True)
        self.port = port
        self.baudrate = baudrate
        self.regex_map = regex_map
        self.log = log
        self.ser = None
        self.stop_flag = False
        self._lock = threading.Lock()
        # 正则匹配结果 (最后一个值)
        self.regex_result = {k: None for k in regex_map}
        # 正则匹配全部结果 (累积列表, clear 时重置)
        self.regex_result_all = {k: [] for k in regex_map}
        # 接收到的协议消息列表
        self.recv_msg_list = []
        # 最近原始行 (用于 set_log_level 等检查)
        self.recent_lines = deque(maxlen=500)
        # 重启检测
        self.reboot_detected = False
        self.reboot_count = 0
        self.reboot_reasons = []
        # 协议日志可能被其他线程日志插断，按 TX/RX 分片重组。
        self._protocol_partial = {
            "sendMsg": {"bytes": [], "expected": None, "carry": ""},
            "recvMsg": {"bytes": [], "expected": None, "carry": ""},
        }
        # 原始串口日志文件
        self._serial_log_fp = None
        if serial_log_dir:
            log_path = os.path.join(serial_log_dir, "serial_raw.log")
            self._serial_log_fp = open(log_path, "a", encoding="utf-8")

    def connect(self):
        try:
            self.ser = serial.Serial(self.port, self.baudrate, timeout=1)
            self.log.info(f"串口 {self.port} 连接成功")
            return True
        except Exception as e:
            self.log.error(f"串口 {self.port} 连接失败: {e}")
            return False

    def run(self):
        if not self.ser or not self.ser.is_open:
            return
        while not self.stop_flag:
            try:
                raw = self.ser.readline()
                if not raw:
                    continue
                text = raw.decode("utf-8", errors="ignore").strip()
                if not text:
                    continue
                # 写入原始串口日志 (所有行，包括 lega_ota_write)
                if self._serial_log_fp:
                    ts = datetime.datetime.now().strftime("%H:%M:%S.%f")[:-3]
                    self._serial_log_fp.write(f"[{ts}] {text}\n")
                    self._serial_log_fp.flush()
                # 过滤 OTA 写入行 (不参与正则匹配)
                if "lega_ota_write" in text:
                    continue
                with self._lock:
                    self.recent_lines.append(text)
                self._match_regex(text)
            except Exception as e:
                if "拒绝访问" in str(e) or "Access is denied" in str(e):
                    self.log.error(f"串口 {self.port} 断开!")
                    self.stop_flag = True
                else:
                    self.log.debug(f"串口读取异常: {e}")

    def _match_regex(self, text):
        with self._lock:
            matched_tags = set()
            for tag, pattern in self.regex_map.items():
                if not pattern:
                    continue
                m = re.match(pattern, text)
                if m:
                    try:
                        val = m.group(1).strip()
                        if tag in {"sendMsg", "recvMsg"} and not self._protocol_value_complete(text, val):
                            continue
                        self.regex_result[tag] = val
                        self.regex_result_all[tag].append(val)
                        matched_tags.add(tag)
                        if tag == "recvMsg":
                            self.recv_msg_list.append(val)
                        if tag == "rebootReason":
                            self.reboot_detected = True
                            self.reboot_count += 1
                            self.reboot_reasons.append(val)
                            self.log.warn(f"  [重启检测] Boot Reason: {val}")
                        else:
                            self.log.debug(f"  正则[{tag}]: {val}")
                    except IndexError:
                        pass
            self._match_protocol_fragments_locked(text, matched_tags)

    def _protocol_value_complete(self, text, value):
        normalized = str(value or "").strip()
        if not normalized:
            return False
        if not re.fullmatch(r"[0-9A-Fa-f]{2}(?: [0-9A-Fa-f]{2})*", normalized):
            return False
        marker_match = re.search(r"\[(TX|RX)\]\s*\[(\d+)\]", text)
        if not marker_match:
            return True
        expected = int(marker_match.group(2))
        parts = [part for part in normalized.split() if part]
        return len(parts) == expected

    def _reset_protocol_partial_locked(self, tag=None):
        if tag:
            self._protocol_partial[tag] = {"bytes": [], "expected": None, "carry": ""}
            return
        for key in self._protocol_partial:
            self._protocol_partial[key] = {"bytes": [], "expected": None, "carry": ""}

    def _extract_protocol_payload_fragment(self, text):
        payload = str(text or "")
        if "->" in payload:
            payload = payload.split("->", 1)[1]
        if ":" in payload:
            payload = payload.split(":", 1)[1]
        payload = payload.strip()
        if "[" in payload:
            payload = payload.split("[", 1)[0].strip()
        return payload

    def _append_protocol_match_locked(self, tag, parts):
        normalized = " ".join(str(part or "").upper() for part in parts if str(part or "").strip())
        if not normalized:
            return
        self.regex_result[tag] = normalized
        self.regex_result_all[tag].append(normalized)
        if tag == "recvMsg":
            self.recv_msg_list.append(normalized)
        self.log.debug(f"  协议[{tag}]: {normalized}")

    def _consume_protocol_payload_locked(self, tag, payload):
        state = self._protocol_partial[tag]
        carry = state.get("carry", "")
        for token in re.findall(r"[0-9A-Fa-f]{1,2}", payload or ""):
            upper = token.upper()
            idx = 0
            while idx < len(upper):
                if carry:
                    need = 2 - len(carry)
                    carry += upper[idx:idx + need]
                    idx += need
                    if len(carry) == 2:
                        state["bytes"].append(carry)
                        carry = ""
                    continue
                remain = len(upper) - idx
                if remain >= 2:
                    state["bytes"].append(upper[idx:idx + 2])
                    idx += 2
                else:
                    carry = upper[idx]
                    idx += 1
        state["carry"] = carry

    def _try_flush_protocol_locked(self, tag):
        state = self._protocol_partial[tag]
        expected = state.get("expected")
        if expected:
            if len(state["bytes"]) >= expected and not state.get("carry"):
                self._append_protocol_match_locked(tag, state["bytes"][:expected])
                self._reset_protocol_partial_locked(tag)
            return
        if state["bytes"] and not state.get("carry"):
            self._append_protocol_match_locked(tag, state["bytes"])
            self._reset_protocol_partial_locked(tag)

    def _match_protocol_fragments_locked(self, text, matched_tags):
        marker_match = re.search(r"\[(TX|RX)\]\s*\[(\d+)\](.*)", text)
        if marker_match:
            tag = "sendMsg" if marker_match.group(1) == "TX" else "recvMsg"
            if tag in matched_tags:
                self._reset_protocol_partial_locked(tag)
                return
            payload = self._extract_protocol_payload_fragment(marker_match.group(3) or "")
            self._reset_protocol_partial_locked(tag)
            self._protocol_partial[tag]["expected"] = int(marker_match.group(2))
            self._consume_protocol_payload_locked(tag, payload)
            self._try_flush_protocol_locked(tag)
            return

        stripped = self._extract_protocol_payload_fragment(text)
        if not stripped or not re.fullmatch(r"[0-9A-Fa-f ]+", stripped):
            return

        for tag in ("sendMsg", "recvMsg"):
            state = self._protocol_partial[tag]
            if not state["bytes"] and not state.get("carry"):
                continue
            self._consume_protocol_payload_locked(tag, stripped)
            self._try_flush_protocol_locked(tag)

    def get(self, key):
        with self._lock:
            return self.regex_result.get(key)

    def get_all(self, key):
        """获取某个 key 从上次 clear 到现在的所有匹配值"""
        with self._lock:
            return list(self.regex_result_all.get(key, []))

    def get_recv_list(self):
        with self._lock:
            return list(self.recv_msg_list)

    def get_recent_lines(self):
        with self._lock:
            return list(self.recent_lines)

    def clear(self):
        with self._lock:
            self.regex_result = {k: None for k in self.regex_map}
            self.regex_result_all = {k: [] for k in self.regex_map}
            self.recv_msg_list.clear()
            self._reset_protocol_partial_locked()

    def is_rebooted(self):
        with self._lock:
            return self.reboot_detected

    def clear_reboot_flag(self):
        with self._lock:
            self.reboot_detected = False

    def get_reboot_count(self):
        with self._lock:
            return self.reboot_count

    def write(self, cmd):
        if self.ser and self.ser.is_open:
            self.ser.write(f"{cmd}\n".encode("utf-8"))

    def close(self):
        self.stop_flag = True
        if self.ser and self.ser.is_open:
            self.ser.close()
            self.log.info(f"串口 {self.port} 已关闭")
        if self._serial_log_fp:
            self._serial_log_fp.close()
            self._serial_log_fp = None


# ══════════════════════════════════════════════════════════════
#  音频播放 (跨平台)
# ══════════════════════════════════════════════════════════════
IS_WINDOWS = platform.system() == "Windows"
IS_LINUX = platform.system() == "Linux"
LISTENAI_PLAY_SCRIPT = os.environ.get("LISTENAI_PLAY_SCRIPT", "").strip()
LEGACY_LISTENAI_PLAY_SCRIPT = os.path.expanduser(
    "~/.codex/skills/listenai-play/scripts/listenai_play.py"
)


def _boolish(value, default=False):
    if isinstance(value, bool):
        return value
    if value is None:
        return default
    text = str(value).strip().lower()
    if text in {"1", "true", "yes", "on"}:
        return True
    if text in {"0", "false", "no", "off"}:
        return False
    return default


def _update_audio_skills_requested():
    return _boolish(os.environ.get("MARS_BELT_UPDATE_AUDIO_SKILLS"), False)


def _listenai_play_script_path(log=None):
    if LISTENAI_PLAY_SCRIPT:
        path = os.path.abspath(os.path.expanduser(LISTENAI_PLAY_SCRIPT))
        if os.path.isfile(path):
            return path
        if log:
            log.warn(f"环境变量 LISTENAI_PLAY_SCRIPT 指向的脚本不存在: {path}")

    try:
        path = resolve_listenai_play_script(
            update=_update_audio_skills_requested(),
            log=log,
        )
        if path and os.path.isfile(path):
            return path
    except Exception as exc:
        if log:
            log.warn(f"本地 tools/audio 依赖自恢复失败，尝试兼容旧路径: {exc}")

    legacy = os.path.abspath(os.path.expanduser(LEGACY_LISTENAI_PLAY_SCRIPT))
    return legacy if os.path.isfile(legacy) else ""


def _run_listenai_play(args, log=None):
    script_path = _listenai_play_script_path(log=log)
    if not script_path:
        raise RuntimeError(
            "未找到 listenai-play 脚本，且 mars-belt/tools/audio 下的依赖自恢复未成功"
        )
    cmd = [sys.executable or shutil.which("python3") or "python3", script_path] + list(args)
    return subprocess.run(
        cmd,
        stdout=subprocess.PIPE,
        stderr=subprocess.PIPE,
        text=True,
        encoding="utf-8",
        errors="replace",
    )


def _scan_listenai_render_devices(log=None):
    result = _run_listenai_play(["scan", "--json", "--direction", "Render"], log=log)
    if result.returncode != 0:
        details = (result.stderr or result.stdout or "scan failed").strip()
        raise RuntimeError(details)
    payload = json.loads((result.stdout or "[]").strip() or "[]")
    if isinstance(payload, dict):
        payload = [payload]
    return [item for item in payload if isinstance(item, dict)]


def _probe_listenai_render_device(device_key, log=None):
    cmd = ["probe"]
    if device_key:
        cmd.extend(["--device-key", device_key])
    result = _run_listenai_play(cmd, log=log)
    if result.returncode != 0:
        details = (result.stderr or result.stdout or "probe failed").strip()
        raise RuntimeError(details)
    return (result.stdout or "").strip()


def _play_via_listenai_device(filepath, device_key, log):
    cmd = ["play", "--audio-file", filepath, "--repeat", "1"]
    if device_key:
        cmd.extend(["--device-key", device_key])
    result = _run_listenai_play(cmd, log=log)
    if result.returncode == 0:
        return True
    details = (result.stderr or result.stdout or "play failed").strip()
    log.warn(f"指定 ListenAI 声卡播报失败: {details}")
    return False


def _format_listenai_devices(devices):
    if not devices:
        return "未发现 ListenAI Audio 渲染声卡"
    parts = []
    for item in devices:
        key = str(item.get("device_key") or "").strip()
        name = str(item.get("name") or item.get("backend_target") or "unknown").strip()
        parts.append(f"{key} ({name})")
    return "; ".join(parts)


def _ensure_audio_card_config(config):
    device_list = config.setdefault("deviceListInfo", {})
    audio_card = device_list.get("audioCard")
    if not isinstance(audio_card, dict):
        audio_card = {}
        device_list["audioCard"] = audio_card
    audio_card["deviceKey"] = str(audio_card.get("deviceKey") or "").strip()
    audio_card["useDefault"] = _boolish(audio_card.get("useDefault"), False)
    audio_card["fallbackToDefault"] = _boolish(audio_card.get("fallbackToDefault"), True)
    audio_card["name"] = str(audio_card.get("name") or "").strip()
    audio_card["backendTarget"] = str(audio_card.get("backendTarget") or "").strip()
    audio_card["lastError"] = str(audio_card.get("lastError") or "").strip()
    return audio_card


def _resolve_audio_card_selection(config, log):
    audio_card = _ensure_audio_card_config(config)
    explicit_key = audio_card.get("deviceKey", "")
    use_default = audio_card.get("useDefault", False)
    fallback_to_default = audio_card.get("fallbackToDefault", True)

    devices = []
    if _listenai_play_script_path(log=log):
        try:
            devices = _scan_listenai_render_devices(log=log)
        except Exception as exc:
            if explicit_key or use_default:
                raise RuntimeError(f"ListenAI 声卡扫描失败: {exc}") from exc
            log.warn(f"ListenAI 声卡扫描失败，暂按默认声卡继续: {exc}")
    elif explicit_key or use_default:
        raise RuntimeError("当前配置启用了声卡绑定，但本机缺少 listenai-play skill 脚本")

    device_count = len(devices)
    device_summary = _format_listenai_devices(devices)

    def use_default_target(reason, persist_error=""):
        if device_count > 1:
            raise RuntimeError(
                f"{reason}；当前存在多个 ListenAI Audio 渲染声卡，禁止使用默认声卡: {device_summary}"
            )
        audio_card["useDefault"] = True
        audio_card["name"] = "default"
        audio_card["backendTarget"] = "default"
        audio_card["lastError"] = persist_error
        log.warn(reason)
        log.info("音频输出已切换为默认声卡")
        return {
            "deviceKey": "",
            "useDefault": True,
            "name": "default",
            "backendTarget": "default",
            "allowDefaultFallback": True,
        }

    if use_default:
        return use_default_target("配置要求使用默认声卡")

    if explicit_key:
        match = None
        for item in devices:
            if str(item.get("device_key") or "").strip().upper() == explicit_key.upper():
                match = item
                break
        if not match:
            reason = f"未扫描到指定声卡 {explicit_key}"
            if fallback_to_default:
                return use_default_target(reason, persist_error=reason)
            raise RuntimeError(f"{reason}；当前设备: {device_summary}")
        try:
            _probe_listenai_render_device(explicit_key, log=log)
        except Exception as exc:
            reason = f"指定声卡 {explicit_key} 探测失败: {exc}"
            if fallback_to_default:
                return use_default_target(reason, persist_error=reason)
            raise RuntimeError(reason) from exc
        audio_card["useDefault"] = False
        audio_card["deviceKey"] = str(match.get("device_key") or explicit_key)
        audio_card["name"] = str(match.get("name") or "")
        audio_card["backendTarget"] = str(match.get("backend_target") or "")
        audio_card["lastError"] = ""
        log.info(
            f"音频输出绑定指定声卡: {audio_card['deviceKey']} -> "
            f"{audio_card['backendTarget']} ({audio_card['name']})"
        )
        return {
            "deviceKey": audio_card["deviceKey"],
            "useDefault": False,
            "name": audio_card["name"],
            "backendTarget": audio_card["backendTarget"],
            "allowDefaultFallback": fallback_to_default and device_count <= 1,
        }

    if device_count > 1:
        raise RuntimeError(
            f"检测到多个 ListenAI Audio 渲染声卡，必须先在配置中指定 deviceListInfo.audioCard.deviceKey: {device_summary}"
        )

    if device_count == 1:
        match = devices[0]
        auto_key = str(match.get("device_key") or "").strip()
        try:
            _probe_listenai_render_device(auto_key, log=log)
        except Exception as exc:
            reason = f"唯一 ListenAI 声卡 {auto_key} 探测失败: {exc}"
            if fallback_to_default:
                return use_default_target(reason, persist_error=reason)
            raise RuntimeError(reason) from exc
        audio_card["useDefault"] = False
        audio_card["deviceKey"] = auto_key
        audio_card["name"] = str(match.get("name") or "")
        audio_card["backendTarget"] = str(match.get("backend_target") or "")
        audio_card["lastError"] = ""
        log.info(
            f"音频输出自动绑定唯一声卡: {audio_card['deviceKey']} -> "
            f"{audio_card['backendTarget']} ({audio_card['name']})"
        )
        return {
            "deviceKey": audio_card["deviceKey"],
            "useDefault": False,
            "name": audio_card["name"],
            "backendTarget": audio_card["backendTarget"],
            "allowDefaultFallback": fallback_to_default,
        }

    audio_card["useDefault"] = True
    audio_card["name"] = "default"
    audio_card["backendTarget"] = "default"
    audio_card["lastError"] = ""
    log.info("未发现 ListenAI Audio 渲染声卡，音频输出使用默认声卡")
    return {
        "deviceKey": "",
        "useDefault": True,
        "name": "default",
        "backendTarget": "default",
        "allowDefaultFallback": True,
    }


def _play_ffplay(filepath, env):
    """通过 ffplay 播放 (Windows/Linux/macOS 通用)"""
    cmd = ["ffplay", "-nodisp", "-autoexit", "-loglevel", "error", filepath]
    result = subprocess.run(cmd, env=env, stdout=subprocess.PIPE, stderr=subprocess.PIPE)
    return result.returncode == 0


def _play_powershell(filepath):
    """Windows 备选: 通过 PowerShell MediaPlayer 播放 MP3"""
    ps_script = f'''
Add-Type -AssemblyName PresentationCore
$player = New-Object System.Windows.Media.MediaPlayer
$player.Open([uri]"{filepath}")
Start-Sleep -Milliseconds 500
$player.Play()
while ($player.NaturalDuration.HasTimeSpan -eq $false) {{ Start-Sleep -Milliseconds 100 }}
$dur = $player.NaturalDuration.TimeSpan.TotalMilliseconds
Start-Sleep -Milliseconds ($dur + 200)
$player.Stop()
$player.Close()
'''
    result = subprocess.run(
        ["powershell", "-NoProfile", "-Command", ps_script],
        stdout=subprocess.PIPE, stderr=subprocess.PIPE
    )
    return result.returncode == 0


def play_audio(filepath, log, audio_target=None):
    """
    跨平台播放音频文件
    优先使用 ffplay (需装 ffmpeg)，Windows 下自动回退 PowerShell MediaPlayer
    """
    if not os.path.isfile(filepath):
        log.error(f"音频不存在: {filepath}")
        return False

    device_key = ""
    allow_default_fallback = True
    if isinstance(audio_target, dict):
        device_key = str(audio_target.get("deviceKey") or "").strip()
        allow_default_fallback = _boolish(audio_target.get("allowDefaultFallback"), True)
    elif audio_target:
        device_key = str(audio_target).strip()

    if device_key:
        try:
            if _play_via_listenai_device(filepath, device_key, log):
                return True
            if not allow_default_fallback:
                return False
            log.warn("指定声卡播报失败，改用默认声卡重试")
        except Exception as exc:
            log.warn(f"指定声卡播报异常: {exc}")
            if not allow_default_fallback:
                return False
            log.warn("切换默认声卡重试")

    env = os.environ.copy()
    if IS_LINUX:
        pulse = "unix:/mnt/wslg/runtime-dir/pulse/native"
        if os.path.exists("/mnt/wslg"):
            env["PULSE_SERVER"] = pulse

    # 方式1: ffplay
    try:
        if _play_ffplay(filepath, env):
            return True
        log.warn("ffplay 播放返回非零，尝试备选方案...")
    except FileNotFoundError:
        log.debug("ffplay 不可用，尝试备选方案...")

    # 方式2: Windows PowerShell
    if IS_WINDOWS:
        try:
            if _play_powershell(filepath):
                return True
            log.warn("PowerShell 播放失败")
        except Exception as e:
            log.warn(f"PowerShell 播放异常: {e}")

    # 方式3: Linux ffplay/mpv/paplay/aplay
    if IS_LINUX:
        for player in ["ffplay -nodisp -loglevel quiet", "mpv --no-video", "paplay", "aplay"]:
            try:
                parts = player.split() + [filepath]
                result = subprocess.run(parts, env=env, stdout=subprocess.PIPE, stderr=subprocess.PIPE)
                if result.returncode == 0:
                    return True
            except FileNotFoundError:
                continue

    log.error(f"所有播放方式均失败，请安装 ffmpeg (ffplay)")
    return False


# ══════════════════════════════════════════════════════════════
#  TTS 语音合成 (讯飞 AIUI WebSocket API)
#  仅在音频缺失时触发, 需额外安装: pip install websockets
# ══════════════════════════════════════════════════════════════
async def _tts_synthesize_one(text, filepath, app_id, api_key,
                              vcn="x4_yezi", speed="50", pitch="50", volume="100"):
    """异步调用讯飞 AIUI WebSocket 合成单条语音, 保存为 MP3"""
    try:
        import websockets
    except ImportError:
        raise ImportError("TTS 合成需要 websockets, 请执行: pip install websockets")

    base_url = "ws://wsapi.xfyun.cn/v1/aiui"
    cur_time = str(int(time.time()))

    param = json.dumps({
        "auth_id": "894c985bf8b1111c6728db79d3479aeg",
        "data_type": "text",
        "speed": speed, "pitch": pitch, "volume": volume,
        "ent": "xtts", "vcn": vcn,
        "aue": "lame", "sfl": "1",
        "scene": "IFLYTEK.tts",
        "sample_rate": "16000",
        "vad_info": "end",
        "ver_type": "monitor",
        "result_level": "plain"
    })

    param_b64 = base64.b64encode(param.encode("utf-8")).decode()
    checksum = hashlib.md5(
        (api_key + cur_time + param_b64).encode("utf-8")
    ).hexdigest()
    conn_url = (f"{base_url}?appid={app_id}&checksum={checksum}"
                f"&param={param_b64}&curtime={cur_time}&signtype=md5")

    async with websockets.connect(conn_url, origin="*", close_timeout=30) as ws:
        await ws.send(text)
        await ws.recv()
        with open(filepath, "wb") as f:
            while True:
                resp = await ws.recv()
                data = json.loads(resp)
                content = data.get("data", {}).get("content", "")
                if content:
                    f.write(base64.b64decode(content))
                if data.get("data", {}).get("is_finish"):
                    break


def tts_generate(text, filepath, tts_cfg):
    """同步包装: 合成单条 MP3"""
    asyncio.run(_tts_synthesize_one(
        text, filepath,
        app_id=tts_cfg.get("app_id", ""),
        api_key=tts_cfg.get("api_key", ""),
        vcn=tts_cfg.get("vcn", "x4_yezi"),
        speed=tts_cfg.get("speed", "50"),
        pitch=tts_cfg.get("pitch", "50"),
        volume=tts_cfg.get("volume", "100"),
    ))


# ══════════════════════════════════════════════════════════════
#  正则自动发现 — 候选模式
# ══════════════════════════════════════════════════════════════
REGEX_CANDIDATES = {
    "wakeKw": [
        r'.*ncmThreshold.*keyword\\":\\"(.*?)\\".*intentStr.*',
        r'.*keyword\\":\\"(.*?)\\".*intentStr.*',
        r'.*keyword":"(.*?)".*intentStr.*',
        r'.*keyword["\s:=]+([^"\\,]+).*intent.*',
    ],
    "asrKw": [
        r'.*ncmThreshold.*keyword.*intentStr\\":\\"(.*?)\\".*',
        r'.*intentStr\\":\\"(.*?)\\".*',
        r'.*intentStr":"(.*?)".*',
        r'.*intent[Ss]tr["\s:=]+([^"\\,]+).*',
    ],
    "sendMsg": [
        r'.*send msg:: (.*)',
        r'.*send msg: (.*)',
        r'.*sendMsg:: (.*)',
        r'.*send_msg[:\s]+(.*)',
        r'.*\[TX\].*->\s*:\s*([0-9A-Fa-f ]{8,})',
        r'.*\[TX\].*?:\s*([0-9A-Fa-f ]{8,})',
        r'.*\[TX\].*?([0-9A-Fa-f]{2}(?: [0-9A-Fa-f]{2}){3,})',
    ],
    "recvMsg": [
        r'.*receive msg:: (.*)',
        r'.*receive msg: (.*)',
        r'.*recv msg:: (.*)',
        r'.*recvMsg:: (.*)',
        r'.*\[RX\].*?:\s*([0-9A-Fa-f ]{8,})',
    ],
    "playId": [
        r'.*play id : (\d+)',
        r'.*play id: (\d+)',
        r'.*play id :(\d+)',
        r'.*playId[:\s]+(\d+)',
    ],
    "volume": [
        r'.*volume\] set scale_vol : (.*)',
        r'.*set scale_vol : (.*)',
        r'.*scale_vol\s*:\s*(.*)',
        r'.*set vol:\s*\d+\s*->\s*(\d+)',
        r'.*app set (?:min|max) vol:\s*(\d+)',
        r'.*mini player set vol\s*:\s*(\d+)',
    ],
}



def normalize_pinyin_text(text):
    """Normalize space/underscore separated pinyin to a stable hyphen form."""
    if text is None:
        return ""
    normalized = str(text).strip().lower()
    if not normalized:
        return ""
    normalized = normalized.replace("_", "-")
    normalized = re.sub(r"\s+", "-", normalized)
    normalized = re.sub(r"-+", "-", normalized)
    return normalized.strip("-")

# ══════════════════════════════════════════════════════════════
#  语音测试主类
# ══════════════════════════════════════════════════════════════
class VoiceTest:
    def __init__(self, config, args):
        self.config = config
        self.config_file_path = os.path.abspath(args.file)
        self.work_dir = os.getcwd()
        self.project = config.get("projectInfo", "")
        self.wakeup_word = config.get("wakeupWord", "")
        self.word_list = config.get("wordList", [])
        self.spell2zh = config.get("spell2zh", {})
        self.spell2zh_normalized = {}
        self.zh2pinyin = {}
        for pinyin, zh in self.spell2zh.items():
            normalized = normalize_pinyin_text(pinyin)
            if normalized and normalized not in self.spell2zh_normalized:
                self.spell2zh_normalized[normalized] = zh
            if zh and normalized and zh not in self.zh2pinyin:
                self.zh2pinyin[zh] = normalized
        self.absorb = config.get("absorb", {})
        self.kw2proto = config.get("kw2protocol", {})
        self.tts_cfg = config.get("ttsConfig", {})
        self.command_random = config.get("commandRandom", 0)
        self.run_times = args.runTimes
        self.wav_dir = os.path.join(self.work_dir, "wavSource")

        # 串口配置 (命令行 --port 可覆盖配置文件)
        csk = config.get("deviceListInfo", {}).get("cskApLog", {})
        self.ser_port = getattr(args, 'port', '') or csk.get("port", "")
        self.ser_baud = csk.get("baudRate", 115200)
        self.regex_map = dict(csk.get("regex", {}))

        # 上电配置 (pretest)
        pretest_cfg = config.get("pretestConfig", {})
        self.pretest_enabled = getattr(args, 'pretest', False) or pretest_cfg.get("enabled", False)
        self.skip_pretest = getattr(args, 'skip_pretest', False)
        self.ctrl_port = getattr(args, 'ctrl_port', '') or pretest_cfg.get("ctrlPort", "")
        self.ctrl_baud = pretest_cfg.get("ctrlBaudRate", 115200)
        self.power_on_cmds = pretest_cfg.get("powerOnCmds", [
            "uut-switch2.off", "uut-switch1.off", "uut-switch1.on"
        ])
        self.audio_enable_cmds = pretest_cfg.get("audioEnableCmds", [
            "uut-pa.on", "pa-enable.set 0 17 0 1"
        ])
        self.pretest_cmd_delay = pretest_cfg.get("cmdDelay", 0.3)
        boot_wait = pretest_cfg.get("bootWait", 5)
        try:
            boot_wait = float(boot_wait)
        except Exception:
            boot_wait = 5.0
        self.pretest_boot_wait = max(boot_wait, 10.0)

        # 输出目录
        ts = datetime.datetime.now().strftime("%Y-%m-%d_%H_%M_%S")
        result_name = build_result_dir_name(self.work_dir, ts, self.project, args.label)
        self.result_dir = os.path.join(self.work_dir, "result", result_name)
        os.makedirs(self.result_dir, exist_ok=True)

        # 测试用例先复制到结果目录 (配置文件在测试结束后复制修改后的版本)
        test_cases_csv = os.path.join(self.work_dir, "testCases.csv")
        if os.path.isfile(test_cases_csv):
            shutil.copy2(test_cases_csv, self.result_dir)

        self.log = Logger(self.result_dir)
        self.reader = None
        self.audio_target = {}
        try:
            self.audio_target = _resolve_audio_card_selection(self.config, self.log)
            self._save_config(f"声卡配置已保存到配置文件: {self.config_file_path}")
        except Exception as exc:
            self.log.error(f"初始化播报声卡失败: {exc}")
            raise

        # 重启统计 (从 reader 提取，reader.close 前保存)
        self._reboot_count = 0
        self._reboot_reasons = []
        self.audio_retry_limit = 5
        self.pretest_power_on_retry_limit = 5
        self.regex_retry_limit = 5
        self.wakeup_retry_limit = 10
        self.command_retry_limit = 3
        self.protocol_retry_limit = 5
        self.unexpected_reboot_limit = 3
        self.abort_reason = ""
        self.exit_code = 1

        # 启动信息
        self.log.info(f"平台: {platform.system()} {platform.release()}")
        self.log.info(f"项目: {self.project}")
        self.log.info(f"串口: {self.ser_port} @ {self.ser_baud}")
        if self.pretest_enabled and self.ctrl_port:
            self.log.info(f"上电控制串口: {self.ctrl_port} @ {self.ctrl_baud}")
            self.log.info(f"上电命令: {self.power_on_cmds}")
        audio_key = str(self.audio_target.get("deviceKey") or "").strip()
        if audio_key:
            self.log.info(
                f"播报声卡: {audio_key} -> "
                f"{self.audio_target.get('backendTarget')} ({self.audio_target.get('name')})"
            )
        else:
            self.log.info("播报声卡: default")
        self.log.info(f"唤醒词: {self.wakeup_word}")
        self.log.info(f"命令词数: {len(self.word_list)}")
        self.log.info(f"音频目录: {self.wav_dir}")

        # 构建同义词组
        self.synonym_groups = self._build_synonym_groups()

    # ── 同义词组构建 ──────────────────────────────────────────
    def _build_synonym_groups(self):
        proto_to_words = defaultdict(set)
        for word, proto in self.kw2proto.items():
            proto_to_words[proto].add(word)
        groups = {}
        for proto, words in proto_to_words.items():
            for w in words:
                groups[w] = set(words)
        for word, absorb_list in self.absorb.items():
            if word not in groups:
                groups[word] = {word}
            if isinstance(absorb_list, list):
                groups[word].update(absorb_list)
            elif isinstance(absorb_list, str):
                groups[word].add(absorb_list)
        return groups

    def is_synonym(self, word_a, word_b):
        if word_a == word_b:
            return True
        group_a = self.synonym_groups.get(word_a, set())
        group_b = self.synonym_groups.get(word_b, set())
        return word_b in group_a or word_a in group_b

    # ── 拼音转汉字 ────────────────────────────────────────────
    def pinyin_to_zh(self, text):
        if not text:
            return ""
        if text in self.spell2zh:
            return self.spell2zh[text]
        normalized = normalize_pinyin_text(text)
        return self.spell2zh_normalized.get(normalized, text)

    # ── 音频检查 & TTS 合成 ─────────────────────────────────────
    def check_and_generate_audio(self):
        os.makedirs(self.wav_dir, exist_ok=True)
        all_words = [self.wakeup_word] + list(self.word_list)
        missing = []
        for word in all_words:
            path = os.path.join(self.wav_dir, f"{word}.mp3")
            if not os.path.isfile(path) or os.path.getsize(path) == 0:
                missing.append(word)
        if not missing:
            self.log.info(f"音频检查通过: 全部 {len(all_words)} 条音频就绪")
            return True
        self.log.warn(f"发现 {len(missing)} 条音频缺失或为空, 需要 TTS 合成")
        if not self.tts_cfg.get("app_id") or not self.tts_cfg.get("api_key"):
            self.log.error("缺少 TTS 配置, 无法自动合成!")
            self.log.error(f"缺失列表: {missing}")
            return False
        success, fail = 0, 0
        for i, word in enumerate(missing):
            path = os.path.join(self.wav_dir, f"{word}.mp3")
            self.log.info(f"  TTS [{i+1}/{len(missing)}]: {word}")
            generated = False
            for attempt in range(1, self.audio_retry_limit + 1):
                try:
                    tts_generate(word, path, self.tts_cfg)
                    if os.path.isfile(path) and os.path.getsize(path) > 0:
                        self.log.info(f"    合成成功: {word} (第{attempt}次)")
                        success += 1
                        generated = True
                        break
                    self.log.error(f"    合成异常: {word} (文件为空, 第{attempt}次)")
                except ImportError as e:
                    self.log.error(str(e))
                    return False
                except Exception as e:
                    self.log.error(f"    合成失败: {word} | {e} (第{attempt}次)")
                if attempt < self.audio_retry_limit:
                    time.sleep(1)
            if not generated:
                fail += 1
            time.sleep(0.3)
        self.log.info(f"TTS 合成完成: 成功={success}, 失败={fail}")
        return fail == 0

    # ── 音频播放 ──────────────────────────────────────────────
    def play(self, name):
        path = os.path.join(self.wav_dir, f"{name}.mp3")
        self.log.info(f"播放音频: {name}")
        return play_audio(path, self.log, self.audio_target)

    # ── 串口初始化 ────────────────────────────────────────────
    def init_serial(self):
        self.reader = SerialReader(
            self.ser_port, self.ser_baud, self.regex_map, self.log,
            serial_log_dir=self.result_dir
        )
        if not self.reader.connect():
            return False
        self.reader.start()
        time.sleep(1)
        return self.reader.ser and self.reader.ser.is_open

    def reconnect_serial(self):
        if self.reader:
            try:
                self.reader.close()
            except Exception as e:
                self.log.debug(f"关闭旧串口 reader 异常: {e}")
        self.reader = SerialReader(
            self.ser_port, self.ser_baud, self.regex_map, self.log,
            serial_log_dir=self.result_dir
        )
        if not self.reader.connect():
            return False
        self.reader.start()
        time.sleep(1)
        return self.reader.ser and self.reader.ser.is_open

    # ── 设备上电 (pretest) ────────────────────────────────────
    def _has_boot_signal(self):
        recent = self.reader.get_recent_lines()
        if not recent:
            return False
        for line in recent[-80:]:
            lowered = line.lower()
            if (
                "Boot Reason" in line
                or "boot" in lowered
                or "config version:" in lowered
                or "root:/$" in line
                or "wkword:" in lowered
                or "voice:" in lowered
                or "[I]" in line
                or "[D]" in line
            ):
                return True
        return False

    def _wait_for_boot_signal_window(self, wait_seconds):
        deadline = time.time() + max(float(wait_seconds or 0), 0.5)
        while time.time() < deadline:
            if self._has_boot_signal():
                return True
            time.sleep(0.5)
        return self._has_boot_signal()

    def _wait_ctrl_port_ready(self, attempts=20, delay_seconds=0.5):
        for _ in range(max(int(attempts), 1)):
            if self.ctrl_port and os.path.exists(self.ctrl_port):
                return True
            time.sleep(max(float(delay_seconds), 0.1))
        return bool(self.ctrl_port and os.path.exists(self.ctrl_port))

    def _send_ctrl_commands(self, commands, command_delay=None, description="控制命令"):
        if not commands:
            raise RuntimeError(f"{description} 为空")
        burn_root = os.path.abspath(os.path.join(os.path.dirname(__file__), "..", "burn"))
        ctrl_script = os.path.join(burn_root, "sudo_ctrl.py")
        delay_seconds = self.pretest_cmd_delay if command_delay is None else command_delay
        delay_ms = int(float(delay_seconds or 0) * 1000)
        if platform.system() == "Linux" and os.path.isfile(ctrl_script):
            if not self._wait_ctrl_port_ready():
                raise RuntimeError(f"控制串口未就绪: {self.ctrl_port}")
            cmd_args = [
                "python3",
                ctrl_script,
                "send",
                self.ctrl_port,
                str(self.ctrl_baud),
                "--delay-ms",
                str(delay_ms),
                *list(commands),
            ]
            cmd_str = " ".join(shlex.quote(arg) for arg in cmd_args)
            self.log.info(
                f"  通过 sg dialout 发送{description}: {list(commands)} "
                f"(delay={delay_ms}ms)"
            )
            completed = subprocess.run(
                ["sg", "dialout", "-c", cmd_str],
                cwd=burn_root,
                stdout=subprocess.PIPE,
                stderr=subprocess.STDOUT,
                text=True,
                encoding="utf-8",
                errors="ignore",
                check=False,
            )
            output = (completed.stdout or "").strip()
            if output:
                self.log.debug(output)
            if completed.returncode != 0:
                raise RuntimeError(f"sg dialout {description}失败: {output or completed.returncode}")
            return

        ctrl_ser = serial.Serial(self.ctrl_port, self.ctrl_baud, timeout=1)
        try:
            self.log.info(f"控制串口 {self.ctrl_port} 连接成功")
            for cmd in commands:
                self.log.info(f"  发送命令: {cmd}")
                ctrl_ser.write(f"{cmd}\r\n".encode("utf-8"))
                time.sleep(delay_seconds)
        finally:
            ctrl_ser.close()
            self.log.info(f"控制串口 {self.ctrl_port} 已关闭")

    def _send_pretest_power_commands(self):
        self._send_ctrl_commands(self.power_on_cmds, description="上电命令")

    def _send_audio_enable_commands(self):
        self._send_ctrl_commands(self.audio_enable_cmds, command_delay=0.2, description="音频使能命令")

    def _attempt_audio_enable_recovery(self):
        if not self.ctrl_port:
            self.log.warn("未配置控制串口，跳过音频使能补救")
            return False
        try:
            self.log.warn("检测到播报已执行但未捕获唤醒/识别日志，尝试补发音频使能命令")
            self._send_audio_enable_commands()
            time.sleep(1)
            return True
        except Exception as exc:
            self.log.warn(f"音频使能补救失败: {exc}")
            return False

    def _set_runtime_loglevel_via_helper(self, level=4):
        burn_root = os.path.abspath(os.path.join(os.path.dirname(__file__), "..", "burn"))
        ctrl_script = os.path.join(burn_root, "sudo_ctrl.py")
        if platform.system() == "Linux" and os.path.isfile(ctrl_script):
            if not self.ser_port or not self._wait_ctrl_port_ready(attempts=20, delay_seconds=0.5):
                raise RuntimeError(f"日志串口未就绪: {self.ser_port}")
            cmd_args = ["python3", ctrl_script, "setloglevel", self.ser_port, str(int(self.ser_baud or 115200))]
            cmd_str = " ".join(shlex.quote(arg) for arg in cmd_args)
            self.log.info(f"通过 sg dialout 设置日志等级 loglevel {level}")
            completed = subprocess.run(
                ["sg", "dialout", "-c", cmd_str],
                cwd=burn_root,
                stdout=subprocess.PIPE,
                stderr=subprocess.STDOUT,
                text=True,
                encoding="utf-8",
                errors="ignore",
                check=False,
            )
            output = (completed.stdout or "").strip()
            if output:
                self.log.debug(output)
            if completed.returncode != 0:
                raise RuntimeError(f"sg dialout 设置 loglevel 失败: {output or completed.returncode}")
            return
        self.reader.write(f"loglevel {level}")

    def pretest_power_on(self):
        """通过控制串口发送上电命令，在日志串口已连接后执行，确保捕获完整启动日志"""
        if self.skip_pretest:
            self.log.debug("skip_pretest 已启用，跳过 pretest_power_on")
            return True
        if not self.pretest_enabled:
            self.log.debug("上电配置未启用，跳过 pretest")
            return True
        if not self.ctrl_port:
            self.log.warn("上电配置已启用但未指定控制串口 (ctrlPort)，跳过 pretest")
            return True

        self.log.info(f"{'='*50}")
        self.log.info("=== 设备上电 (pretest) ===")
        self.log.info(f"控制串口: {self.ctrl_port} @ {self.ctrl_baud}")

        for attempt in range(1, self.pretest_power_on_retry_limit + 1):
            try:
                self.log.info(f"控制串口 {self.ctrl_port} 第{attempt}次执行上电序列")
                self.reader.clear()
                self._send_pretest_power_commands()

                self.log.info(f"上电命令发送完毕，等待设备启动 ({self.pretest_boot_wait}s)...")
                if self._wait_for_boot_signal_window(self.pretest_boot_wait):
                    self.log.info("设备上电完成，已获取到启动日志")
                    if self.audio_enable_cmds:
                        self.log.info("上电后补发音频使能命令")
                        self._send_audio_enable_commands()
                        time.sleep(0.8)
                    return True
                self.log.warn("首次等待未捕获到启动日志，尝试重连日志串口后复查")
                if self.reconnect_serial():
                    if self._wait_for_boot_signal_window(6):
                        self.log.info("设备上电完成，重连日志串口后已获取到启动日志")
                        if self.audio_enable_cmds:
                            self.log.info("重连后补发音频使能命令")
                            self._send_audio_enable_commands()
                            time.sleep(0.8)
                        return True
                self.log.warn("重连后仍未捕获到启动日志，尝试盲发 loglevel 4 后复查")
                try:
                    self._set_runtime_loglevel_via_helper(4)
                    if self.reconnect_serial() and self._wait_for_boot_signal_window(6):
                        self.log.info("设备上电完成，盲发 loglevel 4 后已获取到启动日志")
                        if self.audio_enable_cmds:
                            self.log.info("盲发 loglevel 后补发音频使能命令")
                            self._send_audio_enable_commands()
                            time.sleep(0.8)
                        return True
                except Exception as force_exc:
                    self.log.warn(f"盲发 loglevel 4 失败: {force_exc}")
                self.log.warn(f"第{attempt}次上电后未获取到启动日志")
            except Exception as e:
                self.log.error(f"第{attempt}次上电异常: {e}")
            if attempt < self.pretest_power_on_retry_limit:
                time.sleep(1)

        self.log.error("设备上电失败: 连续 5 次未获取到启动日志")
        return False

    # ── 设置日志等级 ──────────────────────────────────────────
    def set_log_level(self, level=4, retries=5):
        if self.skip_pretest:
            self.log.debug("skip_pretest 已启用，跳过 loglevel [D] 检测")
            return True
        for i in range(retries):
            self.log.info(f"设置日志等级 loglevel {level} (第{i+1}次)")
            recent_before = len(self.reader.get_recent_lines())
            self.reader.write(f"loglevel {level}")
            deadline = time.time() + 2.5
            while time.time() < deadline:
                recent_lines = self.reader.get_recent_lines()[recent_before:]
                saw_debug = any("[D]" in line for line in recent_lines)
                saw_echo = any(f"loglevel {level}" in line for line in recent_lines)
                saw_prompt = any("root:/$" in line or "root:/" in line for line in recent_lines)
                if saw_debug or (saw_echo and saw_prompt):
                    self.log.info("日志等级设置成功")
                    return True
                time.sleep(0.2)
            self.log.warn(f"第{i+1}次验证失败，未检测到 [D] 日志")
        self.log.error("日志等级设置失败")
        return False

    # ── 唤醒设备 ──────────────────────────────────────────────
    def wakeup(self, target_wakeup=None, max_retry=None):
        target_wakeup = target_wakeup or self.wakeup_word
        max_retry = max_retry or self.wakeup_retry_limit
        for i in range(max_retry):
            self.log.info(f"  指定唤醒第{i+1}/{max_retry}次 [{target_wakeup}]...")
            self.reader.clear()
            self.play(target_wakeup)
            deadline = time.time() + 1.5
            while time.time() < deadline:
                wake_all = self.reader.get_all("wakeKw")
                wake_unique = list(dict.fromkeys(wake_all))
                if len(wake_unique) > 1:
                    self.log.warn(f"  唤醒期间检测到多个不同结果: {[self.pinyin_to_zh(w) for w in wake_unique]}")
                wake_raw = self.reader.get("wakeKw")
                if wake_raw and self.pinyin_to_zh(wake_raw) == target_wakeup:
                    self.log.info("  唤醒成功")
                    return True
                time.sleep(0.1)
        self.log.error(f"唤醒失败! (已尝试{max_retry}次)")
        return False

    # ══════════════════════════════════════════════════════════
    #  正则自动发现
    # ══════════════════════════════════════════════════════════
    def auto_discover_regex(self):
        """播放唤醒词+命令词，从串口原始日志中推导正则表达式。"""
        max_retries_per_word = self.regex_retry_limit
        max_words = 5
        reboot_limit = self.unexpected_reboot_limit

        self.log.info("=" * 50)
        candidate_cmds = []
        for word in self.word_list:
            if word != self.wakeup_word and self.kw2proto.get(word):
                candidate_cmds.append((word, self.kw2proto[word]))
                if len(candidate_cmds) >= max_words:
                    break
        if not candidate_cmds:
            self.log.error("wordList 中没有带协议的命令词，无法进行正则发现!")
            return "DEVICE_FAIL"

        wake_pinyin = self.zh2pinyin.get(self.wakeup_word, "")
        all_discovered = {}
        all_lines_total = []
        recovery_attempted = False

        while True:
            current_discovered = {}
            current_lines_total = []
            if recovery_attempted:
                self.log.info("=== 正则自动发现重试（音频使能补救后） ===")
            else:
                self.log.info("=== 正则自动发现开始 ===")

            for cmd_idx, (test_cmd, test_proto) in enumerate(candidate_cmds, start=1):
                self.log.info(f"\n--- 使用命令词 [{test_cmd}] ({cmd_idx}/{len(candidate_cmds)}) ---")

                for attempt in range(1, max_retries_per_word + 1):
                    self.log.info(f"  第 {attempt}/{max_retries_per_word} 次尝试")
                    if self.reader.get_reboot_count() > reboot_limit:
                        self.log.error(f"设备意外重启超过 {reboot_limit} 次，停止正则发现")
                        return "DEVICE_FAIL"

                    if self.reader.is_rebooted():
                        reason = self.reader.get("rebootReason") or "unknown"
                        self.log.warn(f"  检测到设备重启 (Boot Reason: {reason})，等待恢复...")
                        self.reader.clear_reboot_flag()
                        time.sleep(5)
                        if not self.set_log_level(4):
                            return "DEVICE_FAIL"
                        continue

                    if not self.set_log_level(4):
                        return "DEVICE_FAIL"

                    self.reader.clear()
                    self.log.info(f"  播放唤醒词 [{self.wakeup_word}]...")
                    self.play(self.wakeup_word)
                    time.sleep(3)
                    if self.reader.is_rebooted():
                        self.log.warn("  播放唤醒词后设备重启!")
                        self.reader.clear_reboot_flag()
                        time.sleep(5)
                        continue
                    wake_lines = self.reader.get_recent_lines()

                    self.log.info(f"  播放命令词 [{test_cmd}] (协议={test_proto})...")
                    self.play(test_cmd)
                    time.sleep(3)
                    if self.reader.is_rebooted():
                        self.log.warn("  播放命令词后设备重启!")
                        self.reader.clear_reboot_flag()
                        time.sleep(5)
                        continue
                    cmd_lines = self.reader.get_recent_lines()

                    seen = set()
                    all_lines = []
                    for line in wake_lines + cmd_lines:
                        if line not in seen:
                            seen.add(line)
                            all_lines.append(line)
                    current_lines_total = all_lines
                    self.log.info(f"  采集到 {len(all_lines)} 行日志")
                    if len(all_lines) < 3:
                        self.log.warn("  日志行数过少，设备可能未响应")
                        time.sleep(2)
                        continue

                    for tag, candidates in REGEX_CANDIDATES.items():
                        if tag not in current_discovered:
                            result = self._try_candidates(tag, candidates, all_lines, wake_pinyin, test_proto, test_cmd)
                            if result:
                                current_discovered[tag] = result

                    has_wake = "wakeKw" in current_discovered
                    has_asr = "asrKw" in current_discovered
                    has_send = "sendMsg" in current_discovered
                    self.log.info(f"  wakeKw={'已找到' if has_wake else '未找到'}，asrKw={'已找到' if has_asr else '未找到'}，sendMsg={'已找到' if has_send else '未找到'}")
                    # 预检测只要求设备能稳定唤醒并识别命令词。
                    # sendMsg / playId 若缺失，允许进入正式 case，再由协议/播报相关用例单独判定。
                    if has_wake and has_asr:
                        status_text = "，sendMsg 已找到" if has_send else "，sendMsg 暂未找到"
                        self.log.info(f"  [预检测] wakeKw + asrKw 已获取到{status_text}，退出重试循环")
                        break

                if "wakeKw" in current_discovered and "asrKw" in current_discovered:
                    break

            all_discovered = current_discovered
            all_lines_total = current_lines_total
            if "wakeKw" in all_discovered and "asrKw" in all_discovered:
                break
            if recovery_attempted:
                break
            if not self._attempt_audio_enable_recovery():
                break
            recovery_attempted = True

        discovery_log = os.path.join(self.result_dir, "regex_discovery_raw.log")
        with open(discovery_log, "w", encoding="utf-8") as f:
            for line in all_lines_total:
                f.write(line + "\n")

        if self.reader.get_reboot_count() > reboot_limit:
            self.log.error(f"设备意外重启超过 {reboot_limit} 次，设备初始化失败!")
            return "DEVICE_FAIL"

        has_wake = "wakeKw" in all_discovered
        has_asr = "asrKw" in all_discovered
        has_send = "sendMsg" in all_discovered
        if not has_wake or not has_asr:
            missing = []
            if not has_wake: missing.append("wakeKw")
            if not has_asr: missing.append("asrKw")
            self.log.error(f"正则自动发现失败: {missing} 未全部获取到 (最多重试 {max_retries_per_word} 次)")
            return "DEVICE_FAIL"

        self.log.info(f"\n成功发现 {len(all_discovered)} 个正则模式:")
        for key, value in all_discovered.items():
            self.log.info(f"  {key}: {value}")
        missing_keys = [key for key in ["sendMsg", "playId"] if key not in all_discovered]
        if missing_keys:
            self.log.warn(f"以下正则未发现 (将在测试结果中体现): {missing_keys}")

        self.regex_map.update(all_discovered)
        self._save_regex_to_config()
        return True

    def _try_candidates(self, tag, candidates, lines, wake_pinyin, test_proto, test_cmd=None):
        """对一个 regex key 尝试所有候选模式，返回第一个验证通过的模式"""
        # 构建命令词拼音 (用于 asrKw 验证)
        cmd_pinyin = self.zh2pinyin.get(test_cmd, "") if test_cmd else ""
        for pattern in candidates:
            matched_values = []
            for line in lines:
                m = re.match(pattern, line)
                if m:
                    try:
                        matched_values.append(m.group(1).strip())
                    except IndexError:
                        pass

            if not matched_values:
                continue

            # 验证匹配结果的合理性
            if tag == "wakeKw":
                # wakeKw 应能匹配到唤醒词 (中文或拼音)
                for val in matched_values:
                    converted = self.pinyin_to_zh(val)

                    normalized_val = normalize_pinyin_text(val)
                    if converted == self.wakeup_word or val == self.wakeup_word:
                        self.log.info(f"  [{tag}] 验证通过: {pattern}")
                        self.log.info(f"    匹配值: {val} → {converted}")
                        return pattern
                    # 容忍拼音包含关系 (如 wake_pinyin 是子串)
                    if wake_pinyin and wake_pinyin in normalized_val:
                        self.log.info(f"  [{tag}] 验证通过(拼音包含): {pattern}")
                        return pattern

            elif tag == "asrKw":
                # asrKw 应能匹配到命令词的拼音/中文 (优先用命令词验证，因为 asrKw 是识别结果)
                for val in matched_values:
                    converted = self.pinyin_to_zh(val)

                    normalized_val = normalize_pinyin_text(val)
                    # 验证命令词匹配
                    if test_cmd and (converted == test_cmd or val == test_cmd):
                        self.log.info(f"  [{tag}] 验证通过(命令词匹配): {pattern}")
                        self.log.info(f"    匹配值: {val} → {converted}")
                        return pattern
                    if cmd_pinyin and cmd_pinyin in normalized_val:
                        self.log.info(f"  [{tag}] 验证通过(命令词拼音匹配): {pattern}")
                        return pattern
                    # 也接受唤醒词匹配 (有些设备唤醒后 intentStr 也返回唤醒词)
                    if converted == self.wakeup_word or val == self.wakeup_word:
                        self.log.info(f"  [{tag}] 验证通过(唤醒词匹配): {pattern}")
                        return pattern
                    if wake_pinyin and wake_pinyin in normalized_val:
                        self.log.info(f"  [{tag}] 验证通过(唤醒词拼音匹配): {pattern}")
                        return pattern

            elif tag == "sendMsg" and test_proto:
                # sendMsg 应含有期望的协议 hex
                for val in matched_values:
                    if val.strip().upper() == test_proto.strip().upper():
                        self.log.info(f"  [{tag}] 验证通过: {pattern}")
                        self.log.info(f"    匹配值: {val}")
                        return pattern

            elif tag == "recvMsg":
                # recvMsg 有任何 hex 格式匹配即可
                for val in matched_values:
                    if re.search(r'[0-9A-Fa-f]{2}(\s[0-9A-Fa-f]{2}){2,}', val):
                        self.log.info(f"  [{tag}] 验证通过: {pattern}")
                        return pattern

            elif tag == "playId":
                # playId 有任何数字匹配即可
                if matched_values:
                    self.log.info(f"  [{tag}] 验证通过: {pattern}")
                    return pattern

            elif tag == "volume":
                if matched_values:
                    self.log.info(f"  [{tag}] 验证通过: {pattern}")
                    return pattern

        return None

    def _save_config(self, success_msg="配置已保存"):
        try:
            with open(self.config_file_path, "w", encoding="utf-8") as f:
                json.dump(self.config, f, ensure_ascii=False, indent=2)
            if success_msg:
                self.log.info(success_msg)
            return True
        except Exception as e:
            self.log.error(f"保存配置文件失败: {e}")
            return False

    def _save_regex_to_config(self):
        """将当前 regex_map 写回配置文件"""
        csk = self.config.setdefault("deviceListInfo", {}).setdefault("cskApLog", {})
        csk["regex"] = dict(self.regex_map)
        self._save_config(f"正则已保存到配置文件: {self.config_file_path}")

    # ── 单条命令测试 ──────────────────────────────────────────
    def _run_single_command_attempt(
        self,
        command,
        run_idx,
        *,
        expected_proto="",
        timeout_seconds=3.0,
        do_wakeup=True,
    ):
        """执行一次命令词测试尝试。"""
        self.log.info(f"命令词 [{command}] 开始测试")
        self.reader.clear()
        expected_proto = expected_proto or self.kw2proto.get(command, "")

        if do_wakeup and not self.wakeup():
            wake_status = str(getattr(self, "_last_wakeup_status", "") or "").strip()
            wake_detail = str(getattr(self, "_last_wakeup_detail", "") or "").strip()
            wake_verdict = "Skip(人工)" if wake_status == "command-window-expired" else "WakeupFail"
            self.log.error(f"  [{command}] 唤醒失败，跳过当前命令词")
            row = {
                "测试次数": run_idx, "唤醒词": self.wakeup_word, "命令词": command,
                "期望协议": expected_proto, "识别原始结果": "", "识别结果": "",
                "播报ID": "", "实际发送协议": "", "协议收发一致": "",
                "协议比对": "", "识别判定": wake_verdict, "设备响应列表": wake_detail,
                "重启次数": self.reader.get_reboot_count(),
            }
            if wake_detail:
                row["识别结果"] = wake_detail
            return wake_verdict, row

        self.reader.clear()
        t_start = time.time()
        self.play(command)
        response_ms = None
        deadline = t_start + max(float(timeout_seconds), 1.5)
        while time.time() < deadline:
            if self.reader.get_all("asrKw") and response_ms is None:
                response_ms = int((time.time() - t_start) * 1000)
            time.sleep(0.1)

        asr_all = self.reader.get_all("asrKw")
        asr_raw = self.reader.get("asrKw")
        send_msg_all = self.reader.get_all("sendMsg")
        play_id = self.reader.get("playId") or ""
        recv_list = [
            value
            for value in (normalize_hex_protocol(item) for item in self.reader.get_recv_list())
            if value
        ]

        asr_unique = list(dict.fromkeys(asr_all))
        send_unique = self._split_protocol_values("|".join(send_msg_all))
        framed_send = [value for value in send_unique if is_framed_uart_protocol(value)]
        send_candidates = framed_send or send_unique
        expected_proto_norm = normalize_hex_protocol(expected_proto)
        if expected_proto_norm and expected_proto_norm in send_candidates:
            send_msg = expected_proto_norm
        else:
            send_msg = send_candidates[-1] if send_candidates else ""
        asr_zh_list = [self.pinyin_to_zh(item) for item in asr_unique]
        asr_zh = self.pinyin_to_zh(asr_raw) if asr_raw else ""
        asr_raw_display = "|".join(asr_unique) if asr_unique else ""
        asr_zh_display = "|".join(asr_zh_list) if asr_zh_list else ""
        send_msg_display = send_msg or ("|".join(send_candidates) if send_candidates else "")

        if not asr_all:
            verdict = "UnAsr"
            self.log.error(f"  [{command}] 未识别")
        elif len(asr_unique) > 1:
            verdict = "CrossTalk"
            self.log.warn(f"  [{command}] 多结果串扰! 识别到 {len(asr_unique)} 个不同结果:")
            for i, (raw, zh) in enumerate(zip(asr_unique, asr_zh_list)):
                self.log.warn(f"    结果{i+1}: {raw} → {zh}")
        elif asr_zh == command:
            verdict = "OK"
            self.log.info(f"  [{command}] 识别正确 → {asr_zh}")
        elif self.is_synonym(asr_zh, command):
            verdict = "OK"
            self.log.info(f"  [{command}] 同义词匹配 → {asr_zh}")
        else:
            verdict = "CrossTalk"
            self.log.warn(f"  [{command}] 串扰! 识别为: {asr_zh} (原始: {asr_raw})")

        if len(send_candidates) > 1:
            if expected_proto_norm and expected_proto_norm in send_candidates:
                self.log.warn(
                    f"  [{command}] 检测到附加发送协议，已按命令协议收口: {send_candidates}"
                )
            else:
                self.log.warn(f"  [{command}] 检测到多个发送协议: {send_candidates}")
                if verdict == "OK" and expected_proto:
                    verdict = "ConfigFail"

        send_recv_match = "1" if send_msg and send_msg in recv_list else "0"
        if expected_proto:
            if not send_msg:
                proto_verdict = "协议缺失"
                if verdict == "OK":
                    verdict = "ConfigFail"
                self.log.warn(f"  [{command}] 未捕获发送协议，期望=[{expected_proto}]")
            elif send_msg == expected_proto_norm:
                proto_verdict = "协议一致"
            else:
                proto_verdict = "协议不一致"
                if verdict == "OK":
                    verdict = "ConfigFail"
                self.log.warn(f"  协议不一致: 期望=[{expected_proto}] 实际=[{send_msg}]")
        else:
            proto_verdict = ""

        reboot_cnt = self.reader.get_reboot_count()
        row = {
            "测试次数": run_idx,
            "唤醒词": self.wakeup_word,
            "命令词": command,
            "期望协议": expected_proto,
            "识别原始结果": asr_raw_display,
            "识别结果": asr_zh_display,
            "播报ID": play_id,
            "实际发送协议": send_msg_display,
            "协议收发一致": send_recv_match,
            "协议比对": proto_verdict,
            "识别判定": verdict,
            "响应时间(ms)": response_ms if response_ms is not None else "",
            "设备响应列表": "|".join(recv_list),
            "重启次数": reboot_cnt,
        }

        rt_str = f" 响应{response_ms}ms" if response_ms is not None else ""
        if len(asr_unique) > 1:
            self.log.info(f"  第{run_idx}条结果: {verdict} (识别到{len(asr_unique)}个不同结果){rt_str}")
        else:
            self.log.info(f"  第{run_idx}条结果: {verdict}{rt_str}")
        return verdict, row

    @staticmethod
    def _split_protocol_values(raw_value):
        values = []
        for chunk in str(raw_value or "").split("|"):
            value = normalize_hex_protocol(chunk)
            if value and value not in values:
                values.append(value)
        return values

    def _is_protocol_retry_candidate(self, verdict, row, expected_proto=""):
        expected = str(expected_proto or row.get("期望协议", "") or "").strip()
        if verdict != "ConfigFail" or not expected:
            return False
        proto_verdict = str(row.get("协议比对", "") or "").strip()
        if proto_verdict in {"协议缺失", "协议不一致"}:
            return True
        return len(self._split_protocol_values(row.get("实际发送协议", ""))) > 1

    def _finalize_protocol_retry_row(self, row, expected_proto="", attempts_used=0, retry_limit=0, captured_protocols=None):
        updated = dict(row)
        merged_protocols = []
        for value in (captured_protocols or []):
            if value and value not in merged_protocols:
                merged_protocols.append(value)
        for value in self._split_protocol_values(updated.get("实际发送协议", "")):
            if value and value not in merged_protocols:
                merged_protocols.append(value)

        if merged_protocols:
            updated["实际发送协议"] = "|".join(merged_protocols)

        expected = str(expected_proto or updated.get("期望协议", "") or "").strip()
        retry_desc = f"协议日志异常；已重试{attempts_used}/{retry_limit}次仍未稳定"
        if expected:
            retry_desc = f"{retry_desc}；期望协议={expected}"
        if merged_protocols:
            retry_desc = f"{retry_desc}；保留断开协议={updated['实际发送协议']}"
        else:
            retry_desc = f"{retry_desc}；多次重试仍未捕获到发送协议"

        existing = str(updated.get("设备响应列表", "") or "").strip()
        updated["设备响应列表"] = f"{existing}|{retry_desc}".strip("|") if existing else retry_desc
        updated["协议比对"] = "协议打印异常"
        return updated

    def test_one(self, command, run_idx):
        """唤醒 → 播报命令词 → 读取识别结果 → 比对。
        单条命令识别失败默认最多重试 3 次；若命中协议日志异常，则放宽到 5 次。
        """
        final_verdict = "UnAsr"
        final_row = {
            "测试次数": run_idx,
            "唤醒词": self.wakeup_word,
            "命令词": command,
            "期望协议": self.kw2proto.get(command, ""),
            "识别原始结果": "",
            "识别结果": "",
            "播报ID": "",
            "实际发送协议": "",
            "协议收发一致": "",
            "协议比对": "",
            "识别判定": "UnAsr",
            "响应时间(ms)": "",
            "设备响应列表": "",
            "重启次数": self.reader.get_reboot_count(),
        }
        base_retry_limit = max(1, self.command_retry_limit)
        effective_retry_limit = base_retry_limit
        protocol_retry_limit = max(base_retry_limit, getattr(self, "protocol_retry_limit", base_retry_limit))
        had_protocol_issue = False
        captured_protocols = []
        attempts_used = 0
        attempt = 1
        while attempt <= effective_retry_limit:
            attempts_used = attempt
            verdict, row = self._run_single_command_attempt(command, run_idx)
            protocol_issue = self._is_protocol_retry_candidate(verdict, row, row.get("期望协议", ""))
            if protocol_issue:
                had_protocol_issue = True
                effective_retry_limit = max(effective_retry_limit, protocol_retry_limit)
                for value in self._split_protocol_values(row.get("实际发送协议", "")):
                    if value not in captured_protocols:
                        captured_protocols.append(value)
            detail = row.get("设备响应列表", "")
            row["设备响应列表"] = f"attempt={attempt}/{effective_retry_limit}|{detail}".strip("|")
            final_verdict = verdict
            final_row = row
            if verdict == "OK":
                break
            if verdict == "WakeupFail":
                break
            if self.reader.get_reboot_count() > self.unexpected_reboot_limit:
                self.abort_reason = f"设备意外重启超过 {self.unexpected_reboot_limit} 次，停止测试"
                break
            if attempt < effective_retry_limit:
                if protocol_issue and effective_retry_limit > base_retry_limit:
                    self.log.warn(
                        f"  [{command}] 第 {attempt} 次失败({verdict})，检测到协议日志异常，准备继续重试..."
                    )
                else:
                    self.log.warn(f"  [{command}] 第 {attempt} 次失败({verdict})，准备重试...")
                time.sleep(0.6)
            attempt += 1
        if final_verdict != "OK" and had_protocol_issue:
            final_row = self._finalize_protocol_retry_row(
                final_row,
                final_row.get("期望协议", ""),
                attempts_used=attempts_used,
                retry_limit=effective_retry_limit,
                captured_protocols=captured_protocols,
            )
        return final_verdict, final_row

    # ── CSV 测试用例加载 ────────────────────────────────────────
    def load_test_cases(self):
        """读取 testCases.csv, 返回测试用例列表; 不存在则返回 None"""
        csv_path = os.path.join(os.getcwd(), "testCases.csv")
        if not os.path.isfile(csv_path):
            self.log.warn(f"测试用例文件不存在: {csv_path}, 回退到 wordList 模式")
            return None
        cases = []
        with open(csv_path, "r", encoding="utf-8-sig") as f:
            reader = csv_mod.DictReader(f)
            for row in reader:
                cases.append(row)
        self.log.info(f"加载测试用例: {len(cases)} 条 (来自 {csv_path})")
        return cases

    # ── 测试类型分发表 ────────────────────────────────────────
    # 自动执行的测试类型 → 执行模式
    EXEC_MODES = {
        "唤醒识别": "wakeup",
        "唤醒稳定性": "wakeup_stability",
        "误唤醒测试": "false_wakeup",
        "命令词识别": "single",
        "协议校验": "single",
        "同义词验证": "single",
        "相似命令词区分": "single",
        "命令词拒识": "rejection",
        "开关配对": "pair",
        "开关反复切换": "scenario",
        "场景组合": "scenario",
        "重复稳定性": "repeat",
    }
    # 需人工参与的测试类型 (自动标记为 Skip)
    MANUAL_TYPES = {"播报验证", "超时退出", "快速交互", "重启恢复", "功能验证"}

    # ── 唤醒测试 (仅唤醒, 不发命令词) ──────────────────────────
    def test_wakeup_only(self, tc):
        """唤醒识别: 使用与正式命令链路一致的唤醒标准"""
        target_wakeup = tc.get("命令词", "") or tc.get("唤醒词", "") or self.wakeup_word
        wake_ok = self.wakeup(target_wakeup, max_retry=self.wakeup_retry_limit)

        wake_all = self.reader.get_all("wakeKw")
        wake_unique = list(dict.fromkeys(wake_all))
        wake_raw = self.reader.get("wakeKw")
        play_id = self.reader.get("playId") or ""

        wake_raw_display = "|".join(wake_unique) if wake_unique else ""
        wake_zh_list = [self.pinyin_to_zh(w) for w in wake_unique]
        wake_zh_display = "|".join(wake_zh_list) if wake_zh_list else ""
        wake_zh = self.pinyin_to_zh(wake_raw) if wake_raw else ""

        if not wake_ok or not wake_all:
            verdict = "WakeupFail"
            self.log.error(f"  唤醒识别: 无响应")
        elif len(wake_unique) > 1:
            verdict = "CrossTalk"
            self.log.warn(f"  唤醒识别: 多结果串扰! {wake_zh_list}")
        elif wake_zh == target_wakeup:
            verdict = "OK"
            self.log.info(f"  唤醒识别: 成功 → {wake_zh}")
        else:
            verdict = "CrossTalk"
            self.log.warn(f"  唤醒识别: 识别为 {wake_zh} (原始: {wake_raw})")

        return verdict, {
            "用例编号": tc.get("用例编号", ""), "功能模块": tc.get("功能模块", ""),
            "测试类型": tc.get("测试类型", ""), "测试次数": "",
            "唤醒词": target_wakeup, "命令词": target_wakeup,
            "期望协议": "", "识别原始结果": wake_raw_display,
            "识别结果": wake_zh_display, "播报ID": play_id,
            "实际发送协议": "", "协议收发一致": "", "协议比对": "",
            "识别判定": verdict, "设备响应列表": "", "重启次数": self.reader.get_reboot_count(),
        }

    # ── 唤醒稳定性测试 ────────────────────────────────────────
    def test_wakeup_stability(self, tc, count=3):
        """连续唤醒 N 次, 统计成功率"""
        target_wakeup = tc.get("命令词", "") or tc.get("唤醒词", "") or self.wakeup_word
        success = 0
        details = []
        for i in range(count):
            self.reader.clear()
            self.play(target_wakeup)
            time.sleep(1.5)
            wake_raw = self.reader.get("wakeKw")
            if wake_raw and self.pinyin_to_zh(wake_raw) == target_wakeup:
                success += 1
                details.append(f"第{i+1}次:成功")
            else:
                details.append(f"第{i+1}次:失败")
            time.sleep(3)

        verdict = "OK" if success == count else "WakeupFail"
        self.log.info(f"  唤醒稳定性: {success}/{count} ({', '.join(details)})")
        return verdict, {
            "用例编号": tc.get("用例编号", ""), "功能模块": tc.get("功能模块", ""),
            "测试类型": "唤醒稳定性", "测试次数": f"{success}/{count}",
            "唤醒词": target_wakeup, "命令词": target_wakeup,
            "期望协议": "", "识别原始结果": "",
            "识别结果": f"{success}/{count}成功", "播报ID": "",
            "实际发送协议": "", "协议收发一致": "", "协议比对": "",
            "识别判定": verdict, "设备响应列表": ", ".join(details),
            "重启次数": self.reader.get_reboot_count(),
        }

    # ── 误唤醒测试 (自动化) ────────────────────────────────────
    def test_false_wakeup(self, tc):
        """误唤醒测试: 播放近似音, 验证设备不被唤醒 (wakeKw 无匹配 = OK)"""
        word = tc.get("命令词", "")
        self.log.info(f"  误唤醒测试: 播放近似音 [{word}]")
        self.reader.clear()
        self.play(word)
        time.sleep(2)

        wake_all = self.reader.get_all("wakeKw")

        if not wake_all:
            verdict = "OK"
            self.log.info(f"  [{word}] 未被唤醒 (正确)")
        else:
            wake_zh = self.pinyin_to_zh(wake_all[0])
            verdict = "FalseWakeup"
            self.log.warn(f"  [{word}] 被误唤醒! wakeKw={wake_zh}")

        return verdict, {
            "用例编号": tc.get("用例编号", ""), "功能模块": tc.get("功能模块", ""),
            "测试类型": "误唤醒测试", "测试次数": "",
            "唤醒词": self.wakeup_word, "命令词": word,
            "期望协议": "", "识别原始结果": "|".join(wake_all) if wake_all else "",
            "识别结果": self.pinyin_to_zh(wake_all[0]) if wake_all else "",
            "播报ID": "", "实际发送协议": "",
            "协议收发一致": "", "协议比对": "",
            "识别判定": verdict, "设备响应列表": "",
            "重启次数": self.reader.get_reboot_count(),
        }

    # ── 命令词拒识测试 (自动化) ──────────────────────────────
    def test_rejection(self, tc):
        """命令词拒识: 唤醒 → 播放无效命令 → 验证无 asrKw/sendMsg 匹配"""
        word = tc.get("命令词", "")
        self.log.info(f"  命令词拒识: 播放无效命令 [{word}]")
        self.reader.clear()

        # 先唤醒
        if not self.wakeup():
            self.log.error(f"  [{word}] 唤醒失败, 跳过拒识测试")
            return "WakeupFail", {
                "用例编号": tc.get("用例编号", ""), "功能模块": tc.get("功能模块", ""),
                "测试类型": "命令词拒识", "测试次数": "",
                "唤醒词": self.wakeup_word, "命令词": word,
                "期望协议": "", "识别原始结果": "", "识别结果": "",
                "播报ID": "", "实际发送协议": "",
                "协议收发一致": "", "协议比对": "",
                "识别判定": "WakeupFail", "设备响应列表": "",
                "重启次数": self.reader.get_reboot_count(),
            }

        # 播放无效命令
        self.reader.clear()
        self.play(word)
        time.sleep(3)

        asr_all = self.reader.get_all("asrKw")
        send_all = self.reader.get_all("sendMsg")

        if not asr_all and not send_all:
            verdict = "OK"
            self.log.info(f"  [{word}] 正确拒识 (无响应)")
        elif asr_all:
            asr_zh = self.pinyin_to_zh(asr_all[-1])
            verdict = "FalseAccept"
            self.log.warn(f"  [{word}] 被错误识别为 {asr_zh}!")
        else:
            verdict = "FalseAccept"
            self.log.warn(f"  [{word}] 有协议输出 {send_all}!")

        asr_display = "|".join(asr_all) if asr_all else ""
        send_display = "|".join(send_all) if send_all else ""

        return verdict, {
            "用例编号": tc.get("用例编号", ""), "功能模块": tc.get("功能模块", ""),
            "测试类型": "命令词拒识", "测试次数": "",
            "唤醒词": self.wakeup_word, "命令词": word,
            "期望协议": "", "识别原始结果": asr_display,
            "识别结果": self.pinyin_to_zh(asr_all[-1]) if asr_all else "",
            "播报ID": "", "实际发送协议": send_display,
            "协议收发一致": "", "协议比对": "",
            "识别判定": verdict, "设备响应列表": "",
            "重启次数": self.reader.get_reboot_count(),
        }

    # ── 开关配对测试 ────────────────────────────────────────────
    def test_pair(self, tc):
        """开关配对: 打开→关闭 (或关闭→打开), 两步分别验证"""
        cmd_text = tc.get("命令词", "")
        proto_text = tc.get("期望协议", "")
        cmds = [c.strip() for c in cmd_text.split("→") if c.strip()]
        protos = [p.strip() for p in proto_text.split("/") if p.strip()]

        rows = []
        all_ok = True
        for i, cmd in enumerate(cmds):
            self._check_reboot(rows)
            self.log.info(f"  配对步骤 {i+1}/{len(cmds)}: {cmd}")
            verdict, row = self.test_one(cmd, f"配对{i+1}")
            row["用例编号"] = tc.get("用例编号", "")
            row["功能模块"] = tc.get("功能模块", "")
            row["测试类型"] = "开关配对"
            if i < len(protos):
                row["期望协议"] = protos[i]
            if verdict != "OK":
                all_ok = False
            rows.append(row)
            time.sleep(1)

        return "OK" if all_ok else "Fail", rows

    # ── 场景组合测试 ────────────────────────────────────────────
    def test_scenario(self, tc):
        """场景组合: 从测试步骤中提取命令词序列, 依次执行"""
        steps_text = tc.get("测试步骤", "")
        # 提取 "唤醒→XXX" 中的 XXX
        cmds = re.findall(r'唤醒→(\S+)', steps_text)
        if not cmds:
            self.log.warn(f"  场景组合: 无法从步骤中提取命令词: {steps_text}")
            return "Skip", []

        rows = []
        all_ok = True
        for i, cmd in enumerate(cmds):
            self._check_reboot(rows)
            self.log.info(f"  场景步骤 {i+1}/{len(cmds)}: {cmd}")
            verdict, row = self.test_one(cmd, f"场景{i+1}")
            row["用例编号"] = tc.get("用例编号", "")
            row["功能模块"] = tc.get("功能模块", "")
            row["测试类型"] = "场景组合"
            if verdict != "OK":
                all_ok = False
            rows.append(row)
            time.sleep(1)

        return "OK" if all_ok else "Fail", rows

    # ── 重复稳定性测试 ────────────────────────────────────────
    def test_repeat(self, tc, count=3):
        """同一命令词重复 N 次"""
        cmd = tc.get("命令词", "")
        rows = []
        ok_count = 0
        for i in range(count):
            self._check_reboot(rows)
            self.log.info(f"  重复 {i+1}/{count}: {cmd}")
            verdict, row = self.test_one(cmd, f"稳定{i+1}/{count}")
            row["用例编号"] = tc.get("用例编号", "")
            row["功能模块"] = tc.get("功能模块", "")
            row["测试类型"] = "重复稳定性"
            if verdict == "OK":
                ok_count += 1
            rows.append(row)
            time.sleep(1)

        overall = "OK" if ok_count == count else f"{ok_count}/{count}"
        self.log.info(f"  重复稳定性: {ok_count}/{count} 通过")
        return overall, rows

    # ── 执行单条测试用例 (CSV 驱动) ──────────────────────────
    def execute_test_case(self, tc, idx):
        """根据测试类型分发执行, 返回 (verdict, [row_dicts])"""
        test_type = tc.get("测试类型", "")
        exec_mode = self.EXEC_MODES.get(test_type)

        if exec_mode is None:
            # 人工测试类型, 标记 Skip
            return "Skip", [{
                "用例编号": tc.get("用例编号", ""), "功能模块": tc.get("功能模块", ""),
                "测试类型": test_type, "测试次数": "",
                "唤醒词": self.wakeup_word, "命令词": tc.get("命令词", ""),
                "期望协议": tc.get("期望协议", ""), "识别原始结果": "",
                "识别结果": "", "播报ID": "", "实际发送协议": "",
                "协议收发一致": "", "协议比对": "",
                "识别判定": "Skip(人工)", "设备响应列表": "",
                "重启次数": self.reader.get_reboot_count(),
            }]

        if exec_mode == "wakeup":
            verdict, row = self.test_wakeup_only(tc)
            return verdict, [row]

        elif exec_mode == "wakeup_stability":
            verdict, row = self.test_wakeup_stability(tc)
            return verdict, [row]

        elif exec_mode == "false_wakeup":
            verdict, row = self.test_false_wakeup(tc)
            return verdict, [row]

        elif exec_mode == "rejection":
            verdict, row = self.test_rejection(tc)
            return verdict, [row]

        elif exec_mode == "single":
            cmd = tc.get("命令词", "")
            if not cmd or cmd == self.wakeup_word:
                return "Skip", []
            verdict, row = self.test_one(cmd, idx)
            row["用例编号"] = tc.get("用例编号", "")
            row["功能模块"] = tc.get("功能模块", "")
            row["测试类型"] = test_type
            return verdict, [row]

        elif exec_mode == "pair":
            return self.test_pair(tc)

        elif exec_mode == "scenario":
            return self.test_scenario(tc)

        elif exec_mode == "repeat":
            return self.test_repeat(tc)

        return "Skip", []

    # ── 重启检测处理 ────────────────────────────────────────────
    def _check_reboot(self, first_round_rows):
        """检查设备是否重启，如重启则记录并恢复日志等级"""
        if not self.reader.is_rebooted():
            return
        reason = self.reader.get("rebootReason") or "unknown"
        self.log.warn(f"{'!'*50}")
        self.log.warn(f"检测到设备重启! Boot Reason: {reason}")
        self.log.warn(f"{'!'*50}")
        first_round_rows.append({
            "测试次数": "", "唤醒词": self.wakeup_word, "命令词": "[设备重启]",
            "期望协议": "", "识别原始结果": "", "识别结果": f"Boot Reason: {reason}",
            "播报ID": "", "实际发送协议": "", "协议收发一致": "",
            "协议比对": "", "识别判定": "Reboot", "设备响应列表": "",
            "重启次数": self.reader.get_reboot_count(),
        })
        self.log.info("等待设备重启完成...")
        time.sleep(5)
        self.set_log_level(4)
        self.reader.clear_reboot_flag()
        if self.reader.get_reboot_count() > self.unexpected_reboot_limit:
            self.abort_reason = f"设备意外重启超过 {self.unexpected_reboot_limit} 次，停止测试"
            self.log.error(self.abort_reason)
            raise AbortTestRun(self.abort_reason)

    # ══════════════════════════════════════════════════════════
    #  CSV 驱动测试模式
    # ══════════════════════════════════════════════════════════
    def _run_csv_mode(self, test_cases):
        """根据 testCases.csv 驱动测试, 按测试类型分发执行"""
        # 筛选自动测试用例
        auto_cases = []
        manual_cases = []
        for tc in test_cases:
            test_type = tc.get("测试类型", "")
            method = tc.get("测试方法", "")
            if test_type in self.MANUAL_TYPES or method in ("人工",):
                manual_cases.append(tc)
            else:
                auto_cases.append(tc)

        # 应用 -r 限制
        total_auto = len(auto_cases)
        if self.run_times:
            auto_cases = auto_cases[:self.run_times]
        test_count = len(auto_cases)

        self.log.info(f"===== CSV 驱动测试 [{self.project}] =====")
        self.log.info(f"  自动用例: {test_count}/{total_auto} 条 (人工用例: {len(manual_cases)} 条跳过)")

        # 按测试类型统计
        type_counts = defaultdict(int)
        for tc in auto_cases:
            type_counts[tc.get("测试类型", "")] += 1
        for tt, cnt in type_counts.items():
            self.log.info(f"    {tt}: {cnt} 条")

        # ═══════════════ 首轮执行 ═══════════════
        first_round_rows = []
        user_abort = False

        for idx, tc in enumerate(auto_cases):
            tc_id = tc.get("用例编号", "")
            test_type = tc.get("测试类型", "")
            cmd = tc.get("命令词", "")

            self.log.info(f"{'='*50}")
            self.log.info(f"[{tc_id}] ({idx+1}/{test_count}) {test_type}: {cmd}")

            try:
                self._check_reboot(first_round_rows)
                verdict, rows = self.execute_test_case(tc, idx)
                first_round_rows.extend(rows)
            except AbortTestRun as exc:
                self.abort_reason = str(exc)
                self.log.error(self.abort_reason)
                break
            except KeyboardInterrupt:
                self.log.info("用户中断测试 (Ctrl+C)")
                user_abort = True
                break
            except Exception as e:
                self.log.error(f"测试异常: {e}")

        # 添加人工用例的 Skip 记录
        for tc in manual_cases:
            first_round_rows.append({
                "用例编号": tc.get("用例编号", ""), "功能模块": tc.get("功能模块", ""),
                "测试类型": tc.get("测试类型", ""), "测试次数": "",
                "唤醒词": self.wakeup_word, "命令词": tc.get("命令词", ""),
                "期望协议": tc.get("期望协议", ""), "识别原始结果": "",
                "识别结果": "", "播报ID": "", "实际发送协议": "",
                "协议收发一致": "", "协议比对": "",
                "识别判定": "Skip(人工)", "设备响应列表": "",
                "重启次数": "",
            })

        # ═══════════════ 首轮统计 ═══════════════
        active_rows = [r for r in first_round_rows if r.get("识别判定") not in ("Reboot", "Skip(人工)", None)]
        first_ok = sum(1 for r in active_rows if r.get("识别判定") == "OK")
        first_fail_rows = [r for r in active_rows if r.get("识别判定") != "OK"]
        first_ct = [r for r in first_fail_rows if r.get("识别判定") == "CrossTalk"]
        first_unasr = [r for r in first_fail_rows if r.get("识别判定") == "UnAsr"]
        first_wkfail = [r for r in first_fail_rows if r.get("识别判定") == "WakeupFail"]

        self.log.info(f"\n===== 首轮测试完成 =====")
        self.log.info(f"  OK: {first_ok}  CrossTalk: {len(first_ct)}  UnAsr: {len(first_unasr)}  WakeupFail: {len(first_wkfail)}")
        self.log.info(f"  跳过(人工): {len(manual_cases)} 条")

        # ═══════════════ 失败重测 (所有自动测试类型) ═══════════════
        retry_round_rows = []
        generic_retry_params = {
            "wakeupWord",
            "releaseAlgoList[*].word",
            "releaseAlgoList[*].extWord",
            "releaseAlgoList[*].children[*].extWord",
            "releaseAlgoList[*].sndProtocol",
            "releaseAlgoList[*].replyMode",
            "releaseAlgoList[*].reply",
            "releaseAlgoList[*].recProtocol",
            "releaseAlgoList[*].type",
        }
        # 从配对/场景等复合用例中提取单个命令词
        failed_cmds = []
        _seen = set()
        for r in first_fail_rows:
            tt = r.get("测试类型", "")
            if tt in self.MANUAL_TYPES:
                continue
            raw_param = str(r.get("原始参数", "")).strip()
            if raw_param not in generic_retry_params:
                continue
            cmd = r.get("命令词", "")
            if cmd and cmd not in _seen:
                _seen.add(cmd)
                failed_cmds.append(cmd)

        if failed_cmds and not user_abort:
            self.log.info(f"{'='*50}")
            self.log.info(f"===== 失败用例重测: 共 {len(failed_cmds)} 条 =====")
            for ri, cmd in enumerate(failed_cmds):
                self.log.info(f"{'='*50}")
                self.log.info(f"[重测] 第 {ri+1}/{len(failed_cmds)} 条: {cmd}")
                try:
                    self._check_reboot(retry_round_rows)
                    verdict, row = self.test_one(cmd, ri)
                    row["测试类型"] = "失败重测"
                    retry_round_rows.append(row)
                except AbortTestRun as exc:
                    self.abort_reason = str(exc)
                    self.log.error(self.abort_reason)
                    break
                except KeyboardInterrupt:
                    self.log.info("用户中断重测 (Ctrl+C)")
                    break
                except Exception as e:
                    self.log.error(f"重测异常: {e}")

            retry_ok = sum(1 for r in retry_round_rows if r.get("识别判定") == "OK")
            retry_ct = sum(1 for r in retry_round_rows if r.get("识别判定") == "CrossTalk")
            retry_unasr = sum(1 for r in retry_round_rows if r.get("识别判定") == "UnAsr")
            self.log.info(f"===== 重测完成: OK={retry_ok} CrossTalk={retry_ct} UnAsr={retry_unasr} =====")

        return first_round_rows, retry_round_rows, user_abort

    # ══════════════════════════════════════════════════════════
    #  wordList 模式 (CSV 不存在时回退)
    # ══════════════════════════════════════════════════════════
    def _run_wordlist_mode(self):
        """传统 wordList 模式: 遍历命令词逐个测试"""
        cmd_word_list = [w for w in self.word_list if w != self.wakeup_word]
        total = len(cmd_word_list)
        test_count = self.run_times if self.run_times else total
        self.log.info(f"===== wordList 模式 [{self.project}] 测试 {test_count}/{total} 条命令词 =====")

        first_round_rows = []
        idx = 0
        user_abort = False
        while True:
            if self.command_random:
                import random
                cmd = random.choice(cmd_word_list)
            else:
                if idx >= total:
                    break
                cmd = cmd_word_list[idx]

            self.log.info(f"{'='*50}")
            self.log.info(f"[首轮] 第 {idx+1}/{test_count} 条: {cmd}")

            try:
                self._check_reboot(first_round_rows)
                verdict, row = self.test_one(cmd, idx)
                first_round_rows.append(row)
            except AbortTestRun as exc:
                self.abort_reason = str(exc)
                self.log.error(self.abort_reason)
                break
            except KeyboardInterrupt:
                self.log.info("用户中断测试 (Ctrl+C)")
                user_abort = True
                break
            except Exception as e:
                self.log.error(f"测试异常: {e}")

            idx += 1
            if self.run_times and idx >= self.run_times:
                break

        # 首轮统计
        first_ok = sum(1 for r in first_round_rows if r.get("识别判定") == "OK")
        first_fail_rows = [r for r in first_round_rows if r.get("识别判定") not in ("OK", "Reboot", None)]
        first_ct = [r for r in first_fail_rows if r.get("识别判定") == "CrossTalk"]
        first_unasr = [r for r in first_fail_rows if r.get("识别判定") == "UnAsr"]
        first_wkfail = [r for r in first_fail_rows if r.get("识别判定") == "WakeupFail"]

        self.log.info(f"\n===== 首轮测试完成 =====")
        self.log.info(f"  OK: {first_ok}  CrossTalk: {len(first_ct)}  UnAsr: {len(first_unasr)}  WakeupFail: {len(first_wkfail)}")

        # 失败重测
        retry_round_rows = []
        failed_cmds = [r["命令词"] for r in first_fail_rows if r.get("命令词")]

        if failed_cmds and not user_abort:
            self.log.info(f"{'='*50}")
            self.log.info(f"===== 失败用例重测: 共 {len(failed_cmds)} 条 =====")
            for ri, cmd in enumerate(failed_cmds):
                self.log.info(f"{'='*50}")
                self.log.info(f"[重测] 第 {ri+1}/{len(failed_cmds)} 条: {cmd}")
                try:
                    self._check_reboot(retry_round_rows)
                    verdict, row = self.test_one(cmd, ri)
                    retry_round_rows.append(row)
                except AbortTestRun as exc:
                    self.abort_reason = str(exc)
                    self.log.error(self.abort_reason)
                    break
                except KeyboardInterrupt:
                    self.log.info("用户中断重测 (Ctrl+C)")
                    break
                except Exception as e:
                    self.log.error(f"重测异常: {e}")

            retry_ok = sum(1 for r in retry_round_rows if r.get("识别判定") == "OK")
            retry_ct = sum(1 for r in retry_round_rows if r.get("识别判定") == "CrossTalk")
            retry_unasr = sum(1 for r in retry_round_rows if r.get("识别判定") == "UnAsr")
            self.log.info(f"===== 重测完成: OK={retry_ok} CrossTalk={retry_ct} UnAsr={retry_unasr} =====")

        return first_round_rows, retry_round_rows, user_abort

    # ── 主测试流程 ────────────────────────────────────────────
    def run(self):
        # 1. 测试前检查音频
        if not self.check_and_generate_audio():
            self.log.error("音频准备失败, 退出!")
            self.log.close()
            return

        # 2. 初始化日志串口 (先连接，确保捕获上电后全部日志)
        if not self.init_serial():
            self.log.error("串口初始化失败，退出!")
            self.log.close()
            return

        # 3. 设备上电 (pretest: 通过控制串口上电，日志串口已在监听)
        if not self.pretest_power_on():
            self.log.error("!" * 50)
            self.log.error("设备上电失败! 控制串口无法通信，测试终止!")
            self.log.error("!" * 50)
            print(f"\n{'!'*50}")
            print(f"  设备上电失败! 控制串口无法通信，测试终止!")
            print(f"  请检查控制串口连接后重试")
            print(f"{'!'*50}")
            self.reader.close()
            self.log.close()
            return

        if not self.set_log_level(4):
            self.log.error("!" * 50)
            self.log.error("日志等级设置失败! 设备无法正常响应，测试终止!")
            self.log.error("!" * 50)
            print(f"\n{'!'*50}")
            print(f"  日志等级设置失败! 设备无法正常响应，测试终止!")
            print(f"  请检查设备状态后重试")
            print(f"{'!'*50}")
            self.reader.close()
            self.log.close()
            return
        # 5. 检查正则表达式，缺失则自动发现 (含设备健康检测)
        ESSENTIAL_KEYS = ["wakeKw", "asrKw"]
        missing = [k for k in ESSENTIAL_KEYS if not self.regex_map.get(k)]
        if missing:
            self.log.info(f"检测到缺失正则: {missing}")
            self.log.info("启动正则自动发现 (含设备健康检测)...")
            discover_result = self.auto_discover_regex()
            if discover_result == "DEVICE_FAIL":
                self.log.error("!" * 50)
                self.log.error("设备初始化失败! 测试终止!")
                self.log.error("可能原因: 设备持续重启 / 无串口日志 / 无唤醒识别响应")
                self.log.error("!" * 50)
                print(f"\n{'!'*50}")
                print(f"  设备初始化失败! 测试终止!")
                print(f"  请检查设备状态后重试")
                print(f"{'!'*50}")
                self.reader.close()
                self.log.close()
                return
            elif discover_result is True:
                self.reader.close()
                time.sleep(0.5)
                if not self.init_serial():
                    self.log.error("重新初始化串口失败!")
                    self.log.close()
                    return
            else:
                self.log.error("正则自动发现失败且未返回可用结果，测试终止!")
                self.reader.close()
                self.log.close()
                return

        # ═══════════════ 加载测试用例 ═══════════════
        test_cases = self.load_test_cases()

        # 为 CSV 中不在 wordList 的命令词生成音频 (误唤醒近似音、拒识无效命令等)
        if test_cases:
            extra_words = set()
            for tc in test_cases:
                cmd = tc.get("命令词", "")
                if cmd and cmd not in self.word_list and not cmd.startswith("("):
                    extra_words.add(cmd)
            for word in sorted(extra_words):
                mp3_path = os.path.join(self.wav_dir, f"{word}.mp3")
                if not os.path.isfile(mp3_path) or os.path.getsize(mp3_path) == 0:
                    self.log.info(f"生成额外音频: {word}")
                    try:
                        tts_generate(word, mp3_path, self.tts_cfg)
                    except Exception as e:
                        self.log.error(f"额外音频合成失败: {word} | {e}")
                        self.reader.close()
                        self.log.close()
                        return

        if test_cases:
            first_round_rows, retry_round_rows, user_abort = self._run_csv_mode(test_cases)
        else:
            first_round_rows, retry_round_rows, user_abort = self._run_wordlist_mode()

        self.log.info("===== 全部测试结束 =====")

        # 保存重启统计 (在 reader.close 前)
        self._reboot_count = self.reader.get_reboot_count()
        self._reboot_reasons = list(self.reader.reboot_reasons)
        self.reader.close()

        retry_ok_commands = {
            str(row.get("命令词", "")).strip()
            for row in retry_round_rows
            if row.get("识别判定") == "OK" and str(row.get("命令词", "")).strip()
        }
        unresolved_rows = []
        for row in first_round_rows:
            verdict = row.get("识别判定")
            if verdict in ("OK", "Skip(人工)", "Skip", None, ""):
                continue
            command = str(row.get("命令词", "")).strip()
            if command and command in retry_ok_commands:
                continue
            unresolved_rows.append(row)
        for row in retry_round_rows:
            verdict = row.get("识别判定")
            if verdict not in ("OK", "Skip(人工)", "Skip", None, ""):
                unresolved_rows.append(row)
        self.exit_code = 0 if not unresolved_rows and not self.abort_reason and not user_abort else 1

        # 自动分析 CrossTalk 并修复拼音映射
        analysis_data = self.post_analysis(first_round_rows, retry_round_rows)

        # 复制修改后的配置文件到结果目录 (包含自动发现的正则和修复的拼音映射)
        config_backup = os.path.join(self.result_dir, os.path.basename(self.config_file_path))
        try:
            with open(config_backup, "w", encoding="utf-8") as f:
                json.dump(self.config, f, ensure_ascii=False, indent=2)
            self.log.info(f"修改后的配置文件已复制到结果目录: {config_backup}")
        except Exception as e:
            self.log.error(f"复制配置文件失败: {e}")

        # 保存 xlsx 结果文件
        self._save_xlsx(first_round_rows, retry_round_rows, analysis_data)

        if self.abort_reason:
            self.log.error(f"测试异常终止: {self.abort_reason}")
        elif self.exit_code == 0:
            self.log.info("测试结论: PASS")
        else:
            self.log.error(f"测试结论: FAIL, unresolved={len(unresolved_rows)}")

        self.log.close()

    # ══════════════════════════════════════════════════════════
    #  测试后分析 — CrossTalk 自动诊断 & spell2zh 修复
    # ══════════════════════════════════════════════════════════
    def post_analysis(self, first_round_rows, retry_round_rows):
        """分析所有 CrossTalk，区分拼音映射误判和真实串扰，修复 spell2zh
        返回: analysis_data dict 供 xlsx 写入
        """
        self.log.info("=" * 50)
        self.log.info("=== 测试后自动分析开始 ===")

        # 合并所有行分析 (首轮 + 重测)
        all_rows = first_round_rows + retry_round_rows
        ct_rows = [r for r in all_rows if r.get("识别判定") == "CrossTalk"]

        false_ct = []
        true_ct = []
        spell2zh_fixes = {}

        for r in ct_rows:
            command = r["命令词"]
            asr_raw = r["识别原始结果"]
            asr_zh = r["识别结果"]
            send_msg = r["实际发送协议"]
            proto_verdict = r["协议比对"]
            expected_proto = r["期望协议"]

            is_pinyin = bool(re.match(r'^[a-z]+\d?\s', asr_raw)) if asr_raw else False
            proto_ok = (proto_verdict == "协议一致") or (
                expected_proto and send_msg and
                send_msg.strip().upper() == expected_proto.strip().upper()
            )

            if is_pinyin and proto_ok:
                false_ct.append({
                    "命令词": command, "ASR原始": asr_raw, "ASR转换": asr_zh,
                    "分析结论": "拼音映射缺失 (协议正确，设备实际识别正确)"
                })
                if asr_raw and asr_raw not in self.spell2zh:
                    spell2zh_fixes[asr_raw] = command
                    self.log.info(f"  [修复] 添加拼音映射: '{asr_raw}' → '{command}'")
            elif is_pinyin and not proto_ok and not expected_proto:
                false_ct.append({
                    "命令词": command, "ASR原始": asr_raw, "ASR转换": asr_zh,
                    "分析结论": "拼音映射缺失 (无期望协议，需人工确认)"
                })
                if asr_raw and asr_raw not in self.spell2zh:
                    spell2zh_fixes[asr_raw] = command
            else:
                true_ct.append({
                    "命令词": command, "ASR原始": asr_raw, "ASR转换": asr_zh,
                    "期望协议": expected_proto, "实际协议": send_msg,
                    "分析结论": "真实串扰"
                })
                self.log.warn(f"  [真实串扰] {command} → 识别为: {asr_zh}")

        # 修复 spell2zh 并写回配置
        if spell2zh_fixes:
            self.spell2zh.update(spell2zh_fixes)
            self.config["spell2zh"] = self.spell2zh
            self._save_config(f"已将 {len(spell2zh_fixes)} 条拼音映射修复写回配置文件")

        # ── 协议/播报ID 全部缺失检测 ──
        test_rows = [
            r for r in all_rows
            if r.get("识别判定") not in ("Reboot", None, "Skip(人工)", "Skip")
            and r.get("测试类型") != "固件版本校验"
        ]
        analyzer_rows = [
            r for r in test_rows
            if str(r.get("__suite_mode") or "device").strip().lower() == "device"
            and str(r.get("命令词") or "").strip()
        ]
        has_expected_proto = [r for r in analyzer_rows if r.get("期望协议")]
        device_warnings = []

        if analyzer_rows:
            # 检测协议缺失: 有期望协议的用例中，实际发送协议全部为空
            if has_expected_proto:
                rows_with_send = [r for r in has_expected_proto if r.get("实际发送协议")]
                if not rows_with_send:
                    warn_msg = f"所有有期望协议的命令词 ({len(has_expected_proto)} 条) 实际发送协议均为空，sendMsg 正则可能未匹配或设备未发送协议"
                    device_warnings.append(warn_msg)
                    self.log.error(f"  [异常] {warn_msg}")

            # 检测播报ID缺失: 所有用例播报ID全部为空
            rows_with_playid = [r for r in analyzer_rows if r.get("播报ID")]
            if not rows_with_playid:
                warn_msg = f"所有命令词 ({len(analyzer_rows)} 条) 播报ID均为空，playId 正则可能未匹配或设备未播报"
                device_warnings.append(warn_msg)
                self.log.error(f"  [异常] {warn_msg}")

            # 检测识别结果全部缺失
            rows_with_asr = [r for r in analyzer_rows if r.get("识别原始结果")]
            if not rows_with_asr:
                warn_msg = f"所有命令词 ({len(analyzer_rows)} 条) ASR识别结果均为空，asrKw 正则可能未匹配或设备未识别"
                device_warnings.append(warn_msg)
                self.log.error(f"  [异常] {warn_msg}")

        # 统计打印
        first_total = sum(1 for r in first_round_rows if r.get("识别判定") not in ("Reboot", None))
        first_ok = sum(1 for r in first_round_rows if r.get("识别判定") == "OK")
        first_wkfail = sum(1 for r in first_round_rows if r.get("识别判定") == "WakeupFail")
        retry_total = sum(1 for r in retry_round_rows if r.get("识别判定") not in ("Reboot", None))
        retry_ok = sum(1 for r in retry_round_rows if r.get("识别判定") == "OK")

        print(f"\n{'='*50}")
        print(f"  测试汇总")
        print(f"{'='*50}")
        print(f"  首轮: {first_ok}/{first_total} 通过 ({first_ok*100//first_total if first_total else 0}%)")
        if first_wkfail:
            print(f"  未唤醒: {first_wkfail}")
        if retry_total:
            print(f"  重测: {retry_ok}/{retry_total} 通过 ({retry_ok*100//retry_total if retry_total else 0}%)")
        if false_ct or true_ct:
            print(f"  拼音误判: {len(false_ct)} (已自动修复)")
            print(f"  真实串扰: {len(true_ct)}")
        if self._reboot_count:
            print(f"  设备重启: {self._reboot_count} 次")
        if device_warnings:
            print(f"  --- 设备异常警告 ---")
            for w in device_warnings:
                print(f"  [!] {w}")
        print(f"{'='*50}")

        return {
            "false_ct": false_ct,
            "true_ct": true_ct,
            "spell2zh_fixes": spell2zh_fixes,
            "first_total": first_total,
            "first_ok": first_ok,
            "first_wkfail": first_wkfail,
            "retry_total": retry_total,
            "retry_ok": retry_ok,
            "device_warnings": device_warnings,
        }

    # ══════════════════════════════════════════════════════════
    #  结果保存 — xlsx 多 sheet 输出
    # ══════════════════════════════════════════════════════════
    def _save_xlsx(self, first_round_rows, retry_round_rows, analysis_data):
        """将所有测试结果保存为 xlsx 文件，分多个 sheet"""
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        except ImportError:
            self.log.error("缺少 openpyxl，无法保存 xlsx! 请执行: pip install openpyxl")
            self.log.info("回退到 CSV 保存...")
            self._save_csv_fallback(first_round_rows, retry_round_rows)
            return

        xlsx_path = os.path.join(self.result_dir, "testResult.xlsx")
        wb = Workbook()

        # 样式定义
        header_font = Font(bold=True, size=11)
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font_white = Font(bold=True, size=11, color="FFFFFF")
        ok_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        ct_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        unasr_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
        wkfail_fill = PatternFill(start_color="F4B084", end_color="F4B084", fill_type="solid")
        reboot_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
        skip_fill = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )

        COLUMNS = ["用例编号", "功能模块", "测试类型", "测试次数", "唤醒词", "命令词",
                   "期望协议", "识别原始结果", "识别结果", "播报ID", "实际发送协议",
                   "协议收发一致", "协议比对", "识别判定", "响应时间(ms)",
                   "设备响应列表", "重启次数"]

        def safe_excel_value(value):
            return sanitize_xlsx_text(value)

        def write_sheet(ws, rows_data, sheet_title=None):
            """写入一个 sheet 的表头和数据行"""
            # 表头
            for col_idx, col_name in enumerate(COLUMNS, 1):
                cell = ws.cell(row=1, column=col_idx, value=col_name)
                cell.font = header_font_white
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal='center')
                cell.border = thin_border

            # 数据行
            for row_idx, row_data in enumerate(rows_data, 2):
                for col_idx, col_name in enumerate(COLUMNS, 1):
                    val = row_data.get(col_name, "")
                    cell = ws.cell(row=row_idx, column=col_idx, value=safe_excel_value(val))
                    cell.border = thin_border
                    cell.alignment = Alignment(horizontal='center')

                # 根据识别判定着色
                verdict = row_data.get("识别判定", "")
                fill = None
                if verdict == "OK":
                    fill = ok_fill
                elif verdict == "CrossTalk":
                    fill = ct_fill
                elif verdict == "UnAsr":
                    fill = unasr_fill
                elif verdict == "WakeupFail":
                    fill = wkfail_fill
                elif verdict == "FalseWakeup":
                    fill = wkfail_fill
                elif verdict == "FalseAccept":
                    fill = ct_fill
                elif verdict == "Reboot":
                    fill = reboot_fill
                elif "Skip" in verdict:
                    fill = skip_fill
                if fill:
                    for col_idx in range(1, len(COLUMNS) + 1):
                        ws.cell(row=row_idx, column=col_idx).fill = fill

            # 自动列宽
            for col_idx, col_name in enumerate(COLUMNS, 1):
                max_len = len(col_name) * 2  # 中文字符占2
                for row_idx in range(2, len(rows_data) + 2):
                    cell_val = str(ws.cell(row=row_idx, column=col_idx).value or "")
                    max_len = max(max_len, len(cell_val) + 2)
                ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = min(max_len, 40)

        # ── Sheet 1: 首轮测试结果 ──
        ws1 = wb.active
        ws1.title = "首轮测试结果"
        write_sheet(ws1, first_round_rows)

        # ── Sheet 2: 失败重测结果 ──
        ws2 = wb.create_sheet("失败重测结果")
        if retry_round_rows:
            write_sheet(ws2, retry_round_rows)
        else:
            ws2.cell(row=1, column=1, value="无失败用例需要重测")
            ws2.cell(row=1, column=1).font = Font(italic=True, color="808080")

        # ── Sheet 3: 分析报告 ──
        ws3 = wb.create_sheet("分析报告")
        self._write_analysis_sheet(ws3, first_round_rows, retry_round_rows,
                                   analysis_data, header_font, thin_border,
                                   ok_fill, ct_fill, unasr_fill)

        wb.save(xlsx_path)
        self.log.info(f"测试结果已保存: {xlsx_path}")
        print(f"\n  结果文件: {xlsx_path}")

    def _write_analysis_sheet(self, ws, first_rows, retry_rows, analysis,
                              header_font, thin_border, ok_fill, ct_fill, unasr_fill):
        """写入分析报告 sheet"""
        from openpyxl.styles import Font, PatternFill, Alignment

        row = 1
        title_font = Font(bold=True, size=14)
        section_font = Font(bold=True, size=12)
        bold_font = Font(bold=True)

        ws.cell(row=row, column=1, value="测试结果分析报告").font = title_font
        row += 1
        ws.cell(row=row, column=1,
                value=sanitize_xlsx_text(f"生成时间: {datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"))
        row += 2

        # ── 汇总统计 ──
        ws.cell(row=row, column=1, value="一、测试统计").font = section_font
        row += 1

        stats = [
            ("项目", "数值"),
            ("首轮测试总数", analysis.get("first_total", 0)),
            ("首轮通过", analysis.get("first_ok", 0)),
            ("首轮通过率", f"{analysis['first_ok']*100//analysis['first_total']}%" if analysis.get("first_total") else "N/A"),
            ("首轮未唤醒", analysis.get("first_wkfail", 0)),
            ("重测总数", analysis.get("retry_total", 0)),
            ("重测通过", analysis.get("retry_ok", 0)),
            ("设备重启次数", self._reboot_count),
        ]
        for i, (label, val) in enumerate(stats):
            ws.cell(row=row + i, column=1, value=label).font = bold_font if i == 0 else Font()
            ws.cell(row=row + i, column=2, value=sanitize_xlsx_text(val)).font = bold_font if i == 0 else Font()
            ws.cell(row=row + i, column=1).border = thin_border
            ws.cell(row=row + i, column=2).border = thin_border
        row += len(stats) + 1

        # ── 拼音映射误判 ──
        false_ct = analysis.get("false_ct", [])
        true_ct = analysis.get("true_ct", [])
        fixes = analysis.get("spell2zh_fixes", {})

        ws.cell(row=row, column=1, value="二、CrossTalk 分析").font = section_font
        row += 1
        ws.cell(row=row, column=1, value=f"拼音映射误判: {len(false_ct)} 条 (已自动修复)")
        row += 1
        ws.cell(row=row, column=1, value=f"真实串扰: {len(true_ct)} 条")
        row += 2

        if false_ct:
            ws.cell(row=row, column=1, value="拼音映射误判明细:").font = bold_font
            row += 1
            for col_idx, h in enumerate(["命令词", "ASR原始", "ASR转换", "分析结论"], 1):
                ws.cell(row=row, column=col_idx, value=h).font = bold_font
                ws.cell(row=row, column=col_idx).border = thin_border
            row += 1
            for item in false_ct:
                for col_idx, k in enumerate(["命令词", "ASR原始", "ASR转换", "分析结论"], 1):
                    ws.cell(row=row, column=col_idx, value=sanitize_xlsx_text(item.get(k, "")))
                    ws.cell(row=row, column=col_idx).border = thin_border
                    ws.cell(row=row, column=col_idx).fill = unasr_fill
                row += 1
            row += 1

        if true_ct:
            ws.cell(row=row, column=1, value="真实串扰明细:").font = bold_font
            row += 1
            for col_idx, h in enumerate(["命令词", "ASR原始", "ASR转换", "期望协议", "实际协议", "分析结论"], 1):
                ws.cell(row=row, column=col_idx, value=h).font = bold_font
                ws.cell(row=row, column=col_idx).border = thin_border
            row += 1
            for item in true_ct:
                for col_idx, k in enumerate(["命令词", "ASR原始", "ASR转换", "期望协议", "实际协议", "分析结论"], 1):
                    ws.cell(row=row, column=col_idx, value=sanitize_xlsx_text(item.get(k, "")))
                    ws.cell(row=row, column=col_idx).border = thin_border
                    ws.cell(row=row, column=col_idx).fill = ct_fill
                row += 1
            row += 1

        if fixes:
            ws.cell(row=row, column=1, value="三、自动修复的拼音映射").font = section_font
            row += 1
            ws.cell(row=row, column=1, value="拼音").font = bold_font
            ws.cell(row=row, column=2, value="对应命令词").font = bold_font
            ws.cell(row=row, column=1).border = thin_border
            ws.cell(row=row, column=2).border = thin_border
            row += 1
            for pinyin, word in fixes.items():
                ws.cell(row=row, column=1, value=sanitize_xlsx_text(pinyin)).border = thin_border
                ws.cell(row=row, column=2, value=sanitize_xlsx_text(word)).border = thin_border
                row += 1
            row += 1

        if self._reboot_reasons:
            ws.cell(row=row, column=1, value="四、设备重启记录").font = section_font
            row += 1
            for i, reason in enumerate(self._reboot_reasons, 1):
                ws.cell(row=row, column=1, value=f"第{i}次")
                ws.cell(row=row, column=2, value=sanitize_xlsx_text(reason))
                row += 1
            row += 1

        # ── 设备异常警告 ──
        device_warnings = analysis.get("device_warnings", [])
        if device_warnings:
            warn_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
            section_num = "四" if not self._reboot_reasons else "五"
            ws.cell(row=row, column=1, value=f"{section_num}、设备异常警告").font = section_font
            row += 1
            for w in device_warnings:
                cell = ws.cell(row=row, column=1, value=sanitize_xlsx_text(w))
                cell.fill = warn_fill
                cell.border = thin_border
                # 合并到多列显示完整内容
                ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
                row += 1
            row += 1

        # 设置列宽
        ws.column_dimensions['A'].width = 22
        ws.column_dimensions['B'].width = 30
        ws.column_dimensions['C'].width = 20
        ws.column_dimensions['D'].width = 30
        ws.column_dimensions['E'].width = 20
        ws.column_dimensions['F'].width = 35

    def _save_csv_fallback(self, first_round_rows, retry_round_rows):
        """openpyxl 不可用时回退到 CSV"""
        COLUMNS = ["用例编号", "功能模块", "测试类型", "测试次数", "唤醒词", "命令词",
                   "期望协议", "识别原始结果", "识别结果", "播报ID", "实际发送协议",
                   "协议收发一致", "协议比对", "识别判定", "响应时间(ms)",
                   "设备响应列表", "重启次数"]
        # 首轮
        csv_path = os.path.join(self.result_dir, "testResultSummary.csv")
        with open(csv_path, "w", encoding="utf-8-sig") as f:
            f.write(",".join(COLUMNS) + "\n")
            for row in first_round_rows:
                f.write(",".join(str(row.get(c, "")) for c in COLUMNS) + "\n")
        # 重测
        if retry_round_rows:
            csv2 = os.path.join(self.result_dir, "retryResult.csv")
            with open(csv2, "w", encoding="utf-8-sig") as f:
                f.write(",".join(COLUMNS) + "\n")
                for row in retry_round_rows:
                    f.write(",".join(str(row.get(c, "")) for c in COLUMNS) + "\n")
        self.log.info(f"结果已保存为 CSV: {csv_path}")


# ══════════════════════════════════════════════════════════════
#  入口
# ══════════════════════════════════════════════════════════════
if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="轻量化语音功能测试")
    parser.add_argument("-f", "--file", default="deviceInfo_generated.json",
                        help="配置文件路径 (JSON)")
    parser.add_argument("-r", "--runTimes", type=int, default=10,
                        help="测试次数, 0=全部命令词各测一遍, 默认10")
    parser.add_argument("-l", "--label", default="语音功能测试",
                        help="测试标签 (结果目录名称后缀)")
    parser.add_argument("-p", "--port", type=str, default="COM14",
                        help="日志串口端口 (覆盖配置文件, 如 COM3 或 /dev/ttyACM1)")
    parser.add_argument("--ctrl-port", type=str, default="",
                        help="控制串口端口 (上电用, 覆盖配置文件, 如 COM5 或 /dev/ttyACM0)")
    parser.add_argument("--pretest", action="store_true", default=False,
                        help="启用设备上电 (pretest), 覆盖配置文件中的 enabled 设置")
    parser.add_argument("--skip-pretest", action="store_true", default=False,
                        help="跳过 pretest 的 loglevel/[D] 检测，直接运行测试")
    parser.add_argument("--update-audio-skills", action="store_true", default=False,
                        help="即使本地 tools 里已存在 audio skills，也执行 git pull --ff-only")
    args = parser.parse_args()

    if not os.path.isfile(args.file):
        print(f"配置文件不存在: {args.file}")
        sys.exit(1)

    with open(args.file, "r", encoding="utf-8") as f:
        config = json.load(f)

    if args.update_audio_skills:
        os.environ["MARS_BELT_UPDATE_AUDIO_SKILLS"] = "1"

    test = VoiceTest(config, args)
    test.run()
