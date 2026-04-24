import argparse
import csv
import json
import os
import re
import shutil
import subprocess
import sys
import time
import traceback
import zipfile
from copy import deepcopy
from datetime import datetime, timedelta, time as dtime
from pathlib import Path
from types import SimpleNamespace
from typing import Any, Dict, List, Optional, Sequence

from listenai_advanced_combo_trials import build_feature_toggle
from listenai_advanced_combo_trials import build_specific_voice_reg_payload
from listenai_advanced_combo_trials import package_release_with_algo_unified
from listenai_auto_package import ListenAIClient
from listenai_auto_package import require_ok
from listenai_executable_case_suite import first_version, load_web_config
from listenai_profile_suite import build_default_specified_multi_wke_payload, build_profile_payload, export_suite
from listenai_packaging_rules import build_short_release_comment_from_selected, build_weekly_product_name
from listenai_parameter_catalog import build_catalog_payload
from listenai_resolve_and_package import family_of_version
from listenai_shared_product_flow import ensure_shared_product, package_release_for_existing_product
from listenai_test_case_catalog import build_test_catalog
from listenai_task_support import (
    BURN_LOG_ROOT,
    RUNTIME_ROOT,
    TASKS_ROOT,
    ensure_runtime_dir,
    ensure_task_dir,
    load_global_tts_config,
    load_global_audio_card_config,
    resolve_user_path,
    resolve_listenai_token,
    runtime_dir_for_task,
)
from listenai_custom_package import retry_sleep_seconds

ROOT = Path(__file__).resolve().parent.parent
VOICE = ROOT / "py/listenai_voice_test_lite.py"
CLI = ROOT / "mars_belt.py"
RESROOT = TASKS_ROOT
PKGROOT = TASKS_ROOT
BURNROOT = ROOT / "burn"
IS_WINDOWS = os.name == "nt"
DEFAULT_TRACE_PORT = "COM14" if IS_WINDOWS else "/dev/ttyACM0"
DEFAULT_PROTOCOL_PORT = "COM13" if IS_WINDOWS else "/dev/ttyACM1"
DEFAULT_CTRL_PORT = "COM15" if IS_WINDOWS else "/dev/ttyACM4"
DEFAULT_BURN_PORT = "COM14" if IS_WINDOWS else "/dev/ttyACM0"

CASE_ID = "\u7528\u4f8b\u7f16\u53f7"
MOD = "\u529f\u80fd\u6a21\u5757"
TEST_TYPE = "\u6d4b\u8bd5\u7c7b\u578b"
COMMAND = "\u547d\u4ee4\u8bcd"
FUNC_TYPE = "\u529f\u80fd\u7c7b\u578b"
EXPECT_PROTO = "\u671f\u671b\u534f\u8bae"
EXPECT_REPLY = "\u671f\u671b\u64ad\u62a5"
REPLY_MODE = "\u64ad\u62a5\u6a21\u5f0f"
PRIORITY = "\u4f18\u5148\u7ea7"
METHOD = "\u6d4b\u8bd5\u65b9\u6cd5"
PRECOND = "\u524d\u7f6e\u6761\u4ef6"
STEPS = "\u6d4b\u8bd5\u6b65\u9aa4"
EXPECTED = "\u9884\u671f\u7ed3\u679c"
EXECUTOR = "\u6267\u884c\u5668"
PACK_ARGS = "\u6253\u5305\u53c2\u6570"
CFG_ASSERT = "\u914d\u7f6e\u65ad\u8a00"
RUN_ASSERT = "\u8fd0\u884c\u65ad\u8a00"
RAW_PARAM = "\u539f\u59cb\u53c2\u6570"
INPUT_VALUE = "\u8f93\u5165\u503c"
LINKAGE = "\u8054\u52a8\u8bf4\u660e"
LITE = "voiceTestLite\u9002\u914d"
VERDICT_COL = "\u8bc6\u522b\u5224\u5b9a"

HEADERS = [
    CASE_ID,
    MOD,
    TEST_TYPE,
    COMMAND,
    FUNC_TYPE,
    EXPECT_PROTO,
    EXPECT_REPLY,
    REPLY_MODE,
    PRIORITY,
    METHOD,
    PRECOND,
    STEPS,
    EXPECTED,
    EXECUTOR,
    PACK_ARGS,
    CFG_ASSERT,
    RUN_ASSERT,
    RAW_PARAM,
    INPUT_VALUE,
    LINKAGE,
    LITE,
]


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description="ListenAI validation runner")
    parser.add_argument("--product", default="\u53d6\u6696\u5668")
    parser.add_argument("--module", default="3021")
    parser.add_argument("--language", default="\u4e2d\u6587")
    parser.add_argument("--version", default="\u901a\u7528\u5782\u7c7b")
    parser.add_argument("--scene", default="\u7eaf\u79bb\u7ebf")
    parser.add_argument("--source-release-id", default="")
    parser.add_argument("--refresh-live", action="store_true")
    parser.add_argument("--timeout-sec", type=int, default=1800)
    parser.add_argument("--log-port", default=DEFAULT_TRACE_PORT)
    parser.add_argument("--protocol-port", default=DEFAULT_PROTOCOL_PORT)
    parser.add_argument("--uart0-port", default="")
    parser.add_argument("--uart1-port", default="")
    parser.add_argument("--ctrl-port", default=DEFAULT_CTRL_PORT)
    parser.add_argument("--burn-port", default=DEFAULT_BURN_PORT)
    parser.add_argument("--task-dir", default="", help="复用指定测试模式目录，便于在同一批产物上续跑")
    parser.add_argument("--variants", default="", help="仅执行指定变体，逗号分隔，如 pkg-01-mid-stable,pkg-03-right-boundary,pkg-04-full-save-on")
    parser.add_argument("--update-audio-skills", action="store_true", help="即使本地 tools/audio 里已存在 audio skills，也执行 git pull --ff-only")
    parser.add_argument("--skip-device", action="store_true")
    return parser


def log(text: str, payload: Optional[Dict[str, Any]] = None) -> None:
    print(f"\n[{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}] {text}")
    if payload:
        print(json.dumps(payload, ensure_ascii=False, indent=2))
    sys.stdout.flush()


def mkdir(path: Path) -> Path:
    path.mkdir(parents=True, exist_ok=True)
    return path


def jload(path: Path) -> Dict[str, Any]:
    return json.loads(path.read_text(encoding="utf-8"))


def jsave(path: Path, payload: Dict[str, Any]) -> None:
    mkdir(path.parent)
    path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")


def week_stamp() -> str:
    today = datetime.now().date()
    monday = today - timedelta(days=today.weekday())
    return monday.strftime("%m%d")


def sdk_tag(version_label: str) -> str:
    matches = re.findall(r"(A\d+(?:\.\d+)*)", str(version_label or ""), re.I)
    return matches[-1].upper() if matches else "SDK"


def product_name(module: str, product_label: str, version_label: str) -> str:
    return build_weekly_product_name(module, product_label, version_label)


def language_tag(language: str) -> str:
    if language == "\u4e2d\u6587":
        return "zh"
    if language == "\u82f1\u6587":
        return "en"
    return re.sub(r"[^0-9A-Za-z]+", "", language.lower()) or "lang"


def family_tag(version_label: str) -> str:
    return family_of_version(version_label or "") or "custom"


def run_id_for(selected: Dict[str, Any]) -> str:
    match = re.search(r"(\d{4})", str(selected.get("moduleBoard") or ""))
    module = match.group(1) if match else str(selected.get("moduleBoard") or "module")
    return f"{module}_{selected.get('productLabel')}_{language_tag(str(selected.get('language') or ''))}_{family_tag(str(selected.get('versionLabel') or ''))}_{week_stamp()}"


def tts_fallback() -> Dict[str, str]:
    direct = load_global_tts_config()
    if direct:
        return direct
    candidates = [
        ROOT / "deviceInfo_generated.json",
        ROOT / "result/listenai_executable_suite_heater_3021_zh_generic/deviceInfo_generated.json",
    ]
    for path in candidates:
        if not path.exists():
            continue
        try:
            tts = (jload(path).get("ttsConfig") or {})
        except Exception:
            continue
        if tts.get("app_id") and tts.get("api_key"):
            return {key: str(tts.get(key, "")) for key in ["app_id", "api_key", "vcn", "speed", "pitch", "volume"]}
    return {}


def split_choices(text: str) -> List[str]:
    return [item.strip() for item in str(text or "").replace("|", "/").split("/") if item.strip()]


def hexp(text: Any) -> str:
    return " ".join(item for item in re.split(r"[\s,]+", str(text or "").strip().upper()) if item)


def row(**kwargs: Any) -> Dict[str, Any]:
    payload = {key: "" for key in HEADERS}
    payload.update(kwargs)
    return payload


def decode_cli_text(text: Any) -> str:
    value = str(text or "")
    if "\\u" in value or "\\x" in value:
        try:
            return value.encode("ascii").decode("unicode_escape")
        except Exception:
            return value
    return value


def write_suite(suite_dir: Path, rows: Sequence[Dict[str, Any]], device_info: Dict[str, Any]) -> None:
    mkdir(suite_dir)
    with (suite_dir / "testCases.csv").open("w", encoding="utf-8-sig", newline="") as fp:
        writer = csv.DictWriter(fp, fieldnames=HEADERS)
        writer.writeheader()
        for item in rows:
            writer.writerow({key: item.get(key, "") for key in HEADERS})
    (suite_dir / "executable_cases.json").write_text(
        json.dumps({"generatedAt": datetime.now().isoformat(timespec="seconds"), "rows": list(rows)}, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )
    (suite_dir / "deviceInfo_generated.json").write_text(json.dumps(device_info, ensure_ascii=False, indent=2), encoding="utf-8")


def load_base_rows(base_suite_dir: Path) -> List[Dict[str, Any]]:
    payload = base_suite_dir / "executable_cases.json"
    if not payload.exists():
        return []
    return [dict(item) for item in (jload(payload).get("rows") or [])]


def seed_audio(device_info: Dict[str, Any], suite_dir: Path) -> Dict[str, int]:
    words: List[str] = []
    wakeup = str(device_info.get("wakeupWord") or "").strip()
    if wakeup:
        words.append(wakeup)
    for item in device_info.get("wordList") or []:
        word = str(item or "").strip()
        if word and word not in words:
            words.append(word)
    target_dir = mkdir(suite_dir / "wavSource")
    copied = 0
    existing = 0
    for word in words:
        target = target_dir / f"{word}.mp3"
        if target.exists() and target.stat().st_size > 0:
            existing += 1
            continue
        for source_dir in [ROOT / "wavSource", ROOT / "result/listenai_executable_suite_heater_3021_zh_generic/wavSource"]:
            source = source_dir / f"{word}.mp3"
            if source.exists() and source.stat().st_size > 0:
                shutil.copy2(source, target)
                copied += 1
                break
    return {"copied": copied, "existing": existing}


def mk_device(config: Dict[str, Any], project_name: str, log_port: str, ctrl_port: str, tts: Dict[str, str]) -> Dict[str, Any]:
    context = build_voice_context(config)
    device_info = context.get("device_info") or build_device_info_template(config, context)
    device_info["projectInfo"] = project_name
    tts_cfg = device_info.setdefault("ttsConfig", {})
    if not str(tts_cfg.get("app_id") or "").strip() and tts:
        tts_cfg.update(tts)
    device_list = device_info.setdefault("deviceListInfo", {})
    csk = device_list.setdefault("cskApLog", {})
    csk["port"] = log_port
    csk["baudRate"] = int(csk.get("baudRate") or 115200)
    regex = csk.setdefault("regex", {})
    for key, value in {
        "rebootReason": r".*Boot Reason: (.*)",
        "endLine": "",
        "wakeKw": "",
        "asrKw": "",
        "sendMsg": "",
        "recvMsg": "",
        "playId": "",
        "volume": "",
    }.items():
        regex.setdefault(key, value)
    audio_defaults = load_global_audio_card_config()
    audio_card = device_list.setdefault("audioCard", {})
    audio_card["deviceKey"] = str(audio_card.get("deviceKey") or audio_defaults.get("deviceKey") or "")
    audio_card["useDefault"] = as_bool(
        audio_card.get("useDefault"),
        as_bool(audio_defaults.get("useDefault"), False),
    )
    audio_card["fallbackToDefault"] = as_bool(
        audio_card.get("fallbackToDefault"),
        as_bool(audio_defaults.get("fallbackToDefault"), True),
    )
    if "fallbackToDefault" not in audio_card and "fallbackToDefault" in audio_defaults:
        audio_card["fallbackToDefault"] = as_bool(audio_defaults.get("fallbackToDefault"), True)
    audio_card["name"] = str(audio_card.get("name") or audio_defaults.get("name") or "")
    audio_card["backendTarget"] = str(audio_card.get("backendTarget") or audio_defaults.get("backendTarget") or "")
    audio_card["lastError"] = str(audio_card.get("lastError") or audio_defaults.get("lastError") or "")
    pretest = device_info.setdefault("pretestConfig", {})
    pretest["enabled"] = bool(ctrl_port)
    pretest["ctrlPort"] = ctrl_port
    pretest["ctrlBaudRate"] = 115200
    pretest.setdefault("powerOnCmds", ["uut-switch2.off", "uut-switch1.off", "uut-switch1.on"])
    pretest.setdefault("cmdDelay", 0.3)
    pretest.setdefault("bootWait", 5)
    return device_info


def pick_row(rows: Sequence[Dict[str, Any]], parameter: str, value: str) -> Dict[str, Any]:
    for item in rows:
        if item.get(RAW_PARAM) == parameter and str(item.get(INPUT_VALUE, "")) == str(value):
            return deepcopy(item)
    raise RuntimeError(f"missing base case {parameter}={value}")


def rows_words(config: Dict[str, Any]) -> List[Dict[str, Any]]:
    version = (config.get("_ver_list") or [{}])[0]
    result: List[Dict[str, Any]] = []
    for idx, item in enumerate(version.get("asr_cmds") or [], start=1):
        if not isinstance(item, dict):
            continue
        intent = str(item.get("intent") or "").strip()
        item_type = str(item.get("type") or "").strip()
        if not intent:
            continue
        snd = hexp(item.get("snd_protocol"))
        rec = hexp(item.get("rec_protocol"))
        reply = (split_choices(item.get("reply")) or [""])[0]
        reply_mode = str(item.get("reply_mode") or "").strip()
        raw_param = "releaseAlgoList[*].sndProtocol" if snd else "releaseAlgoList[*].word"
        input_value = snd or intent
        if item_type == "\u5524\u9192\u8bcd":
            raw_param = "releaseAlgoList[*].word"
            input_value = intent
        if item_type in {"\u5524\u9192\u8bcd\u8d1f\u6027\u8bcd", "\u547d\u4ee4\u8bcd\u8d1f\u6027\u8bcd"}:
            result.append(
                row(
                    **{
                        CASE_ID: f"WORDS-{idx:03d}",
                        MOD: "\u8bcd\u8868\u9a8c\u8bc1",
                        TEST_TYPE: "negative",
                        COMMAND: intent,
                        FUNC_TYPE: "structure",
                        PRIORITY: "P1",
                        METHOD: "manual",
                        PRECOND: "flash target firmware",
                        STEPS: f"play {intent}",
                        EXPECTED: "should not trigger",
                        EXECUTOR: "manual",
                        PACK_ARGS: "{}",
                        RUN_ASSERT: f"negative={intent}",
                        RAW_PARAM: "negativeWord",
                        INPUT_VALUE: intent,
                        LINKAGE: "manual check",
                        LITE: "partial",
                    }
                )
            )
            continue
        config_assert = f"_ver_list[0].asr_cmds[*].intent contains {intent}"
        if raw_param == "releaseAlgoList[*].sndProtocol":
            config_assert = f"_ver_list[0].asr_cmds[*].snd_protocol contains {snd}"
        run_assert = f"word={intent};snd={snd};reply={reply}"
        if rec:
            run_assert += f";rec={rec}"
        result.append(
            row(
                **{
                    CASE_ID: f"WORDS-{idx:03d}",
                    MOD: "\u8bcd\u8868\u9a8c\u8bc1",
                    TEST_TYPE: "wake" if item_type == "\u5524\u9192\u8bcd" else "command",
                    COMMAND: intent,
                    FUNC_TYPE: "structure",
                    EXPECT_PROTO: snd,
                    EXPECT_REPLY: reply,
                    REPLY_MODE: reply_mode,
                    PRIORITY: "P1" if item_type == "\u5524\u9192\u8bcd" else "P0",
                    METHOD: "auto+manual",
                    PRECOND: "flash target firmware",
                    STEPS: f"play {intent}",
                    EXPECTED: "recognition and protocol match config",
                    EXECUTOR: "device",
                    PACK_ARGS: "{}",
                    CFG_ASSERT: config_assert,
                    RUN_ASSERT: run_assert,
                    RAW_PARAM: raw_param,
                    INPUT_VALUE: input_value,
                    LINKAGE: item_type,
                    LITE: "yes",
                }
            )
        )
    return result


def voice_reg_learn_words(config: Dict[str, Any]) -> List[str]:
    version = (config.get("_ver_list") or [{}])[0]
    study = ((version.get("firmware") or {}).get("study_config") or {})
    words: List[str] = []
    for item in study.get("reg_commands") or []:
        if not isinstance(item, dict):
            continue
        word = str(item.get("word") or item.get("condition") or "").strip()
        if word and word not in words:
            words.append(word)
    return words


def rows_voice_reg(base_rows: Sequence[Dict[str, Any]], config: Dict[str, Any]) -> List[Dict[str, Any]]:
    try:
        main = pick_row(base_rows, "voiceRegEnable", "true")
    except Exception:
        main = row(**{CASE_ID: "VOICE-001", RAW_PARAM: "voiceRegEnable", INPUT_VALUE: "true", LITE: "partial"})
    learn_words = voice_reg_learn_words(config)
    learn_word = learn_words[0] if learn_words else ""
    pack = {"voiceRegEnable": True}
    if learn_words:
        pack["studyRegCommands"] = learn_words
    cfg_asserts = ["firmware.study_config.enable eq true"]
    if learn_word:
        cfg_asserts.append(f"firmware.study_config.reg_commands[*].word contains {learn_word}")
    main.update(
        {
            CASE_ID: "VOICE-001",
            MOD: "\u8bed\u97f3\u6ce8\u518c",
            TEST_TYPE: "voice-reg",
            COMMAND: "\u5b66\u4e60\u547d\u4ee4\u8bcd",
            FUNC_TYPE: "direct",
            EXPECT_REPLY: "\u8bf7\u5728\u5b89\u9759\u73af\u5883\u4e0b\uff0c\u8bf4\u51fa\u4f60\u60f3\u8981\u5b66\u4e60\u7684\u547d\u4ee4\u8bcd",
            REPLY_MODE: "main",
            PRIORITY: "P1",
            METHOD: "auto+manual",
            PRECOND: "flash voice-reg firmware",
            STEPS: "enter learning flow and verify no reboot",
            EXPECTED: "learning flow available and no reboot",
            EXECUTOR: "semi-auto",
            PACK_ARGS: json.dumps(pack, ensure_ascii=False),
            CFG_ASSERT: " ; ".join(cfg_asserts),
            RUN_ASSERT: f"learn_word={learn_word or 'NONE'};check no reboot",
            LINKAGE: f"studyRegCommands={json.dumps(learn_words, ensure_ascii=False)}",
        }
    )
    smoke = rows_words(config)
    if smoke:
        extra = deepcopy(smoke[0])
        extra[CASE_ID] = "VOICE-002"
        extra[MOD] = "\u8bed\u97f3\u6ce8\u518c\u70df\u6d4b"
        extra[LINKAGE] = "verify base recognition still works"
        return [main, extra]
    return [main]


def build_target_args(cli_args: argparse.Namespace, out_dir: Path) -> SimpleNamespace:
    catalog_cache_dir = RUNTIME_ROOT / "catalog"
    reference_cache_dir = RUNTIME_ROOT / "reference_cache"
    return SimpleNamespace(
        token=os.environ.get("LISTENAI_TOKEN", "").strip(),
        refresh_live=bool(cli_args.refresh_live),
        product=cli_args.product,
        module=cli_args.module,
        language=cli_args.language,
        version=cli_args.version,
        scene=cli_args.scene,
        source_release_id=cli_args.source_release_id,
        md_out=str(out_dir / "parameter_catalog.md"),
        json_out=str(out_dir / "parameter_catalog.json"),
        dict_tree_in=str(reference_cache_dir / "dict_tree.json"),
        config_details_in=str(reference_cache_dir / "config_details.json"),
        config_sensitivity_in=str(reference_cache_dir / "config_sensitivity.json"),
        config_recommended_in=str(reference_cache_dir / "config_recommended.json"),
        release_detail_in="",
        json_out_catalog=str(catalog_cache_dir / "listenai_product_options.json"),
        products_csv_out=str(catalog_cache_dir / "listenai_product_catalog.csv"),
        modules_csv_out=str(catalog_cache_dir / "listenai_module_catalog.csv"),
        matrix_csv_out=str(catalog_cache_dir / "listenai_product_options_matrix.csv"),
        duplicates_csv_out=str(catalog_cache_dir / "listenai_version_defid_duplicates.csv"),
        matrix_md_out=str(catalog_cache_dir / "listenai_product_options_matrix.md"),
        resolution_out=str(out_dir / "resolved_product.json"),
        summary_out=str(out_dir / "selected_package_summary.json"),
    )


def ensure_target_metadata(cli_args: argparse.Namespace, work_dir: Path) -> Dict[str, Any]:
    target_args = build_target_args(cli_args, work_dir / "catalog")
    catalog = build_catalog_payload(target_args)
    test_catalog = build_test_catalog(catalog)
    catalog_dir = mkdir(work_dir / "catalog")
    (catalog_dir / "parameter_catalog.json").write_text(json.dumps(catalog, ensure_ascii=False, indent=2), encoding="utf-8")
    (catalog_dir / "test_case_catalog.json").write_text(json.dumps(test_catalog, ensure_ascii=False, indent=2), encoding="utf-8")
    return {"catalog": catalog, "testCatalog": test_catalog}


def build_base_rows(test_catalog: Dict[str, Any], base_zip: Path, base_suite_dir: Path, selected: Dict[str, Any], cli_args: argparse.Namespace, tts: Dict[str, str]) -> List[Dict[str, Any]]:
    config = load_web_config(str(base_zip), "")
    suite_payload = build_profile_payload(
        web_config=config or {},
        profile="base",
        metadata={
            "scalars": {
                "product": selected.get("productLabel", ""),
                "module": selected.get("moduleBoard", ""),
                "language": selected.get("language", ""),
                "version": selected.get("versionLabel", ""),
                "scene": selected.get("sceneLabel", ""),
                "defId": selected.get("defId", ""),
                "generatedAt": datetime.now().isoformat(timespec="seconds"),
            },
            "appliedOverrides": {},
            "learnWords": [],
            "finalRelease": {},
            "comments": "基础配置",
        },
        selected_meta=selected,
    )
    seed_audio(suite_payload.get("deviceInfo") or {}, base_suite_dir)
    export_suite(base_suite_dir, suite_payload)
    return [dict(item) for item in (suite_payload.get("rows") or [])]


def run(cmd: List[str], cwd: Optional[Path] = None) -> int:
    log("\u6267\u884c\u547d\u4ee4", {"command": " ".join(cmd), "cwd": str(cwd or ROOT)})
    return subprocess.run(cmd, cwd=str(cwd or ROOT), check=False).returncode


def latest_dir(path: Path) -> Optional[Path]:
    if not path.exists():
        return None
    dirs = [item for item in path.iterdir() if item.is_dir()]
    if not dirs:
        return None
    return sorted(dirs, key=lambda item: item.stat().st_mtime, reverse=True)[0]


def existing_file(value: Any) -> Optional[Path]:
    text = str(value or "").strip()
    if not text:
        return None
    path = Path(text)
    return path if path.is_file() else None


def existing_dir(value: Any) -> Optional[Path]:
    text = str(value or "").strip()
    if not text:
        return None
    path = Path(text)
    return path if path.is_dir() else None


def parameter_map(catalog: Dict[str, Any]) -> Dict[str, Dict[str, Any]]:
    return {str(item.get("key") or ""): dict(item) for item in catalog.get("parameters") or [] if item.get("key")}


def parameter_default(catalog: Dict[str, Any], key: str, fallback: Any = None) -> Any:
    entry = parameter_map(catalog).get(key) or {}
    value = entry.get("default")
    return fallback if value in (None, "") else value


def parameter_enum_entries(catalog: Dict[str, Any], key: str) -> List[Dict[str, Any]]:
    entry = parameter_map(catalog).get(key) or {}
    return [dict(item) for item in entry.get("enum_values") or [] if item.get("value") not in (None, "")]


def parameter_enum_values(catalog: Dict[str, Any], key: str) -> List[Any]:
    return [item.get("value") for item in parameter_enum_entries(catalog, key)]


def parameter_enum_values_with_fallback(catalog: Dict[str, Any], key: str, fallback: Sequence[Any]) -> List[Any]:
    values = parameter_enum_values(catalog, key)
    if values:
        return values
    return list(fallback)


def speaker_is_language_compatible(language: Any, value: Any, label: Any = "") -> bool:
    lang = str(language or "").strip()
    value_text = str(value or "").strip().lower()
    label_text = str(label or "").strip()
    english_like = any(token in value_text for token in ["_en", "enus", "enuk"]) or any(token in label_text for token in ["英语", "英文"])
    foreign_non_target = any(token in value_text for token in ["jajp", "kokr"]) or any(token in label_text for token in ["日语", "韩语"])
    if lang == "中文":
        return not english_like and not foreign_non_target
    if lang == "英文":
        return english_like and not foreign_non_target
    return True


def language_filtered_vcn_values(catalog: Dict[str, Any], language: Any) -> List[Any]:
    entries = parameter_enum_entries(catalog, "vcn")
    filtered = [item.get("value") for item in entries if speaker_is_language_compatible(language, item.get("value"), item.get("label"))]
    if filtered:
        return filtered
    return [item.get("value") for item in entries]


def pick_low_mid_high(values: Sequence[Any], fallback: Any = None) -> tuple[Any, Any, Any]:
    unique: List[Any] = []
    for value in values:
        if value not in unique:
            unique.append(value)
    if not unique:
        return fallback, fallback, fallback
    low = unique[0]
    high = unique[-1]
    mid = unique[len(unique) // 2]
    return low, mid, high


def as_bool(value: Any, default: bool = False) -> bool:
    if isinstance(value, bool):
        return value
    text = str(value or "").strip().lower()
    if text in {"1", "true", "yes", "on"}:
        return True
    if text in {"0", "false", "no", "off"}:
        return False
    return default


def compact_overrides(values: Dict[str, Any]) -> Dict[str, Any]:
    return {key: value for key, value in values.items() if value not in (None, "")}


def scalar_text(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def effective_uart_port_map(cli_args: argparse.Namespace) -> Dict[str, str]:
    # 当前设备本地连线约定：
    # - UART1(默认日志串口) -> --log-port
    # - UART0(默认协议串口) -> --protocol-port
    uart0 = str(getattr(cli_args, "uart0_port", "") or getattr(cli_args, "protocol_port", "") or DEFAULT_PROTOCOL_PORT).strip()
    uart1 = str(getattr(cli_args, "uart1_port", "") or getattr(cli_args, "log_port", "") or DEFAULT_TRACE_PORT).strip()
    return {"0": uart0, "1": uart1}


def resolve_runtime_serials(config: Dict[str, Any], cli_args: argparse.Namespace) -> Dict[str, Any]:
    ver = first_version(config or {})
    firmware = ver.get("firmware") or {}
    uart_cfg = firmware.get("uart_config") or {}
    port_map = effective_uart_port_map(cli_args)
    trace_uart = scalar_text(uart_cfg.get("trace_uart"))
    uport_uart = scalar_text(uart_cfg.get("uport_uart"))
    runtime_trace_uart = trace_uart
    if trace_uart == "0":
        # 当前设备/固件上 traceUart=0 运行态未把启动日志切到 UART0，日志仍固定出现在 UART1。
        runtime_trace_uart = "1"
    # The bench wiring is stable and the CLI ports are the source of truth.
    # Firmware UART metadata is still recorded, but it must not override the
    # physical ports explicitly passed by the operator for this test station.
    log_port = str(getattr(cli_args, "log_port", "") or port_map.get(runtime_trace_uart) or port_map.get("1") or DEFAULT_TRACE_PORT).strip()
    protocol_port = str(getattr(cli_args, "protocol_port", "") or port_map.get(uport_uart) or port_map.get("0") or DEFAULT_PROTOCOL_PORT).strip()
    log_baud = int(uart_cfg.get("trace_baud") or 115200)
    protocol_baud = int(uart_cfg.get("uport_baud") or 9600)
    return {
        "traceUart": trace_uart,
        "runtimeTraceUart": runtime_trace_uart,
        "uportUart": uport_uart,
        "logPort": log_port,
        "protocolPort": protocol_port,
        "logBaud": log_baud,
        "protocolBaud": protocol_baud,
    }


def sum_cfg(result_dir: Optional[Path]) -> Dict[str, Any]:
    payload = result_dir / "suite_config_assert_result.json" if result_dir else None
    if not payload or not payload.exists():
        return {"status": "missing"}
    summary = (jload(payload).get("summary") or {})
    return {"status": "ok", "resultDir": str(result_dir), "total": summary.get("total"), "counters": summary.get("counters") or {}}


def sum_dev(result_dir: Optional[Path]) -> Dict[str, Any]:
    summary_csv = result_dir / "testResultSummary.csv" if result_dir else None
    if summary_csv and summary_csv.exists():
        counters: Dict[str, int] = {}
        rows = list(csv.DictReader(summary_csv.open("r", encoding="utf-8-sig", newline="")))
        for item in rows:
            verdict = str(item.get(VERDICT_COL) or "UNKNOWN").strip() or "UNKNOWN"
            counters[verdict] = counters.get(verdict, 0) + 1
        return {"status": "ok", "resultDir": str(result_dir), "total": len(rows), "counters": counters}

    xlsx_path = result_dir / "testResult.xlsx" if result_dir else None
    if xlsx_path and xlsx_path.exists():
        try:
            summary = load_device_xlsx_summary(xlsx_path)
            summary["resultDir"] = str(result_dir)
            return summary
        except Exception as exc:
            return {"status": "xlsx_parse_failed", "resultDir": str(result_dir), "error": f"{type(exc).__name__}: {exc}"}

    tail = ""
    logs = sorted((result_dir or Path(".")).glob("test_*.log")) if result_dir and result_dir.exists() else []
    if logs:
        tail = "\n".join(logs[-1].read_text(encoding="utf-8", errors="ignore").splitlines()[-20:])
    return {"status": "missing_csv", "resultDir": str(result_dir) if result_dir else "", "logTail": tail}


def param_txt(selected: Dict[str, Any], shared_name: str, spec: Dict[str, Any], summary: Dict[str, Any], final_release: Dict[str, Any]) -> str:
    applied = summary.get("appliedOverrides") or spec.get("overrides") or {}
    learn_words = summary.get("resolvedVoiceRegLearnCommands") or []
    return "\n".join(
        [
            f"\u751f\u6210\u65f6\u95f4: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}",
            f"\u5171\u4eab\u4ea7\u54c1\u540d: {shared_name}",
            f"variantId: {spec['id']}",
            f"variantTitle: {spec['title']}",
            f"product: {selected.get('productLabel')}",
            f"module: {selected.get('moduleBoard')}",
            f"language: {selected.get('language')}",
            f"scene: {selected.get('sceneLabel')}",
            f"version: {selected.get('versionLabel')}",
            f"defId: {selected.get('defId')}",
            f"releaseId: {summary.get('releaseId')}",
            f"releaseVersion: {summary.get('releaseVersion')}",
            f"comments: {spec.get('comments')}",
            "",
            "[appliedOverrides]",
            json.dumps(applied, ensure_ascii=False, indent=2),
            "",
            "[voiceRegLearnWords]",
            json.dumps(learn_words, ensure_ascii=False, indent=2),
            "",
            "[finalRelease]",
            json.dumps({key: final_release.get(key) for key in ["timeout", "volLevel", "defaultVol", "voiceRegEnable", "multiWkeEnable", "multiWkeMode", "algoViewMode", "comments", "status"]}, ensure_ascii=False, indent=2),
            "",
        ]
    )


def inject(zip_path: Path, name: str, text: str) -> None:
    tmp = Path(str(zip_path) + ".tmp")
    with zipfile.ZipFile(zip_path, "r") as src, zipfile.ZipFile(tmp, "w", compression=zipfile.ZIP_DEFLATED) as dst:
        for info in src.infolist():
            if info.filename != name:
                dst.writestr(info, src.read(info.filename))
        dst.writestr(name, text.encode("utf-8"))
    tmp.replace(zip_path)


def sdk_hits(zip_path: Path) -> List[str]:
    try:
        with zipfile.ZipFile(zip_path) as zf:
            return [name for name in zf.namelist() if any(key in name.lower() for key in ["burn", "flash", "download", "program", "uart"])][:100]
    except Exception:
        return []


def latest_burn_log() -> str:
    burn_log = BURN_LOG_ROOT
    if not burn_log.exists():
        return ""
    logs = sorted(burn_log.glob("*.log"), key=lambda item: item.stat().st_mtime, reverse=True)
    if not logs:
        return ""
    return "\n".join(logs[0].read_text(encoding="utf-8", errors="ignore").splitlines()[-30:])


def burn_package(zip_path: Path, cli_args: argparse.Namespace, runtime_serial: Dict[str, Any]) -> int:
    return run(
        [
            sys.executable,
            str(CLI),
            "burn",
            "--package-zip",
            str(zip_path),
            "--ctrl-port",
            cli_args.ctrl_port,
            "--burn-port",
            cli_args.burn_port,
            "--runtime-log-port",
            str(runtime_serial.get("logPort") or cli_args.log_port),
            "--runtime-log-baud",
            str(runtime_serial.get("logBaud") or 115200),
        ],
        cwd=ROOT,
    )


def maybe_pause(manifest_path: Path, state: Dict[str, Any]) -> None:
    if os.environ.get("MARS_BELT_DISABLE_PAUSE", "").strip().lower() in {"1", "true", "yes", "on"}:
        return
    now = datetime.now()
    start = datetime.combine(now.date(), dtime(12, 30))
    end = datetime.combine(now.date(), dtime(13, 30))
    if start <= now < end:
        jsave(manifest_path, state)
        if CLI.exists():
            subprocess.run([sys.executable, str(CLI), "backup-state"], cwd=str(ROOT), check=False)
        seconds = max(int((end - now).total_seconds()), 0)
        log("\u5348\u95f4\u6682\u505c\u4e2d", {"sleepSeconds": seconds, "resumeAt": end.strftime("%Y-%m-%d %H:%M:%S")})
        time.sleep(seconds)


def build_variant_specs(selected: Dict[str, Any], catalog: Dict[str, Any]) -> List[Dict[str, Any]]:
    feature_map = dict(catalog.get("featureMap") or {})
    voice_supported = str(feature_map.get("voice_regist") or "") == "Optional"
    multi_supported = str(feature_map.get("multi_wakeup") or "") == "Optional"

    timeout_low, timeout_mid, timeout_high = 1, 30, 60
    vol_level_low, vol_level_mid, vol_level_high = 2, 5, 10
    default_vol_low, default_vol_mid, default_vol_high = 1, 3, 10
    baud_low, baud_mid, baud_high = pick_low_mid_high(
        parameter_enum_values_with_fallback(catalog, "uportBaud", [2400, 4800, 9600, 19200, 38400, 57600, 115200, 921600]),
        9600,
    )
    log_level_low, log_level_mid, log_level_high = pick_low_mid_high(
        parameter_enum_values_with_fallback(catalog, "logLevel", [0, 1, 2, 3, 4, 5]),
        4,
    )
    compress_low, compress_mid, compress_high = pick_low_mid_high(
        parameter_enum_values_with_fallback(catalog, "compress", [1, 2, 3]),
        2,
    )

    language = str(selected.get("language") or "")
    vcn_default = parameter_default(catalog, "vcn", "")
    vcn_values = language_filtered_vcn_values(catalog, language)
    if vcn_default and vcn_default not in vcn_values:
        vcn_default = vcn_values[len(vcn_values) // 2] if vcn_values else vcn_default
    vcn_low, vcn_mid, vcn_high = pick_low_mid_high(vcn_values, vcn_default)

    needs_vcn_override = bool(vcn_values) and (not vcn_default or vcn_default not in vcn_values)
    stable_mid_core_overrides = compact_overrides({
        "timeout": timeout_mid,
        "volLevel": vol_level_mid,
        "defaultVol": default_vol_mid,
        "uportBaud": baud_mid,
        "logLevel": log_level_mid,
        "volSave": False,
        "vcn": vcn_mid if needs_vcn_override else None,
        "speed": 50,
        "vol": 50,
        "compress": compress_mid,
        "paConfigEnable": False,
    })
    base_mid_overrides = compact_overrides({
        **deepcopy(stable_mid_core_overrides),
        "volMaxOverflow": "中值最大音量播报",
        "volMinOverflow": "中值最小音量播报",
    })

    specs = [
        {
            "id": "pkg-01-mid-stable",
            "title": "PKG-01 基础中值稳定包",
            "kind": "changed",
            "pack": "base",
            "profile": "changed",
            "overrides": deepcopy(base_mid_overrides),
            "extraPhrases": ["基础中值稳定包"],
            "featurePoints": [
                "基础数值中值",
                "字符串参数一次性入包",
                "volSave=false",
                "paConfigEnable=false",
                "合法语种发音人保持默认或最小修正",
            ],
        },
        {
            "id": "pkg-02-left-boundary",
            "title": "PKG-02 基础左边界包",
            "kind": "changed",
            "pack": "base",
            "profile": "changed",
            "overrides": compact_overrides({
                "timeout": timeout_low,
                "volLevel": vol_level_low,
                "defaultVol": default_vol_low,
                "uportBaud": baud_low,
                "logLevel": log_level_low,
                "speed": 1,
                "vol": 1,
                "compress": compress_low,
                "paConfigEnable": False,
            }),
            "extraPhrases": ["基础左边界包"],
            "featurePoints": [
                "唤醒时长左边界",
                "音量档位左边界",
                "默认音量左边界",
                "低波特率",
                "日志级别左边界",
            ],
        },
        {
            "id": "pkg-03-right-boundary",
            "title": "PKG-03 基础右边界包",
            "kind": "changed",
            "pack": "base",
            "profile": "changed",
            "overrides": compact_overrides({
                "timeout": timeout_high,
                "volLevel": vol_level_high,
                "defaultVol": default_vol_high,
                "uportBaud": baud_high,
                "logLevel": log_level_high,
                "volSave": True,
                "speed": 100,
                "vol": 100,
                "compress": compress_high,
                "paConfigEnable": True,
            }),
            "extraPhrases": ["基础右边界包"],
            "featurePoints": [
                "唤醒时长右边界",
                "音量档位右边界",
                "默认音量右边界",
                "高波特率",
                "日志级别右边界",
                "volSave=true",
                "paConfigEnable=true",
            ],
        },
    ]

    if voice_supported and multi_supported:
        specs.append(
            {
                "id": "pkg-04-full-save-on",
                "title": "PKG-04 全功能保持开启包",
                "kind": "full-feature",
                "pack": "algo",
                "profile": "changed",
                "overrides": compact_overrides({
                    **deepcopy(stable_mid_core_overrides),
                    "voiceRegEnable": True,
                    **build_specific_voice_reg_payload(),
                    **build_default_specified_multi_wke_payload(),
                    "wakeWordSave": True,
                }),
                "extraPhrases": ["全功能保持开启包"],
                "featurePoints": [
                    "基础中值稳定基线",
                    "语音注册 specificLearn",
                    "多唤醒 specified",
                    "新增两个唤醒词",
                    "wakeWordSave=true",
                ],
            }
        )
        specs.append(
            {
                "id": "pkg-05-multi-save-off",
                "title": "PKG-05 多唤醒保持关闭隔离包",
                "kind": "changed",
                "pack": "algo",
                "profile": "changed",
                "overrides": compact_overrides({
                    **deepcopy(stable_mid_core_overrides),
                    **build_default_specified_multi_wke_payload(),
                    "wakeWordSave": False,
                }),
                "extraPhrases": ["多唤醒保持关闭隔离包"],
                "featurePoints": [
                    "基础中值稳定基线",
                    "多唤醒 specified",
                    "新增两个唤醒词",
                    "wakeWordSave=false",
                    "隔离唤醒词掉电保持关闭",
                ],
            }
        )
    elif multi_supported:
        specs.append(
            {
                "id": "pkg-04-multi-save-on",
                "title": "PKG-04 多唤醒保持开启包",
                "kind": "changed",
                "pack": "algo",
                "profile": "changed",
                "overrides": compact_overrides({
                    **deepcopy(stable_mid_core_overrides),
                    **build_default_specified_multi_wke_payload(),
                    "wakeWordSave": True,
                }),
                "extraPhrases": ["多唤醒保持开启包"],
                "featurePoints": [
                    "基础中值稳定基线",
                    "多唤醒 specified",
                    "新增两个唤醒词",
                    "wakeWordSave=true",
                ],
            }
        )
        specs.append(
            {
                "id": "pkg-05-multi-save-off",
                "title": "PKG-05 多唤醒保持关闭隔离包",
                "kind": "changed",
                "pack": "algo",
                "profile": "changed",
                "overrides": compact_overrides({
                    **deepcopy(stable_mid_core_overrides),
                    **build_default_specified_multi_wke_payload(),
                    "wakeWordSave": False,
                }),
                "extraPhrases": ["多唤醒保持关闭隔离包"],
                "featurePoints": [
                    "基础中值稳定基线",
                    "多唤醒 specified",
                    "新增两个唤醒词",
                    "wakeWordSave=false",
                    "隔离唤醒词掉电保持关闭",
                ],
            }
        )
    elif voice_supported:
        specs.append(
            {
                "id": "pkg-04-voice-reg-only",
                "title": "PKG-04 语音注册专项包",
                "kind": "changed",
                "pack": "algo",
                "profile": "changed",
                "overrides": compact_overrides({
                    **deepcopy(stable_mid_core_overrides),
                    "voiceRegEnable": True,
                    **build_specific_voice_reg_payload(),
                }),
                "extraPhrases": ["语音注册专项包"],
                "featurePoints": [
                    "基础中值稳定基线",
                    "语音注册 specificLearn",
                ],
            }
        )

    for spec in specs:
        spec["comments"] = build_short_release_comment_from_selected(
            selected,
            spec.get("overrides") or {},
            extra_phrases=spec.get("extraPhrases") or [],
        )
    return specs


def write_feature_matrix(task_dir: Path, selected: Dict[str, Any], specs: Sequence[Dict[str, Any]]) -> None:
    lines = [
        f"# 功能点验证矩阵 - {selected.get('productLabel')}",
        "",
        f"- 目标：`{selected.get('productPath')}` | `{selected.get('moduleBoard')}` | `{selected.get('language')}` | `{selected.get('versionLabel')}`",
        "",
        "| 变体 | 目标 | 关键覆盖点 |",
        "| --- | --- | --- |",
    ]
    for spec in specs:
        lines.append(
            f"| `{spec['id']}` | `{spec.get('pack')}/{spec.get('profile')}` | {'；'.join(spec.get('featurePoints') or [])} |"
        )
    (task_dir / "feature_matrix.md").write_text("\n".join(lines).rstrip() + "\n", encoding="utf-8")


def copy_variant_artifacts(task_dir: Path, variant_id: str, suite_dir: Path) -> None:
    latest = latest_dir(suite_dir / "result")
    if latest is None:
        return
    mapping = {
        "serial_raw.log": f"{variant_id}_serial_raw.log",
        "testResult.xlsx": f"{variant_id}_testResult.xlsx",
        "suite_config_assert_result.json": f"{variant_id}_suite_config_assert_result.json",
        "suite_config_assert_result.csv": f"{variant_id}_suite_config_assert_result.csv",
    }
    for source_name, target_name in mapping.items():
        source = latest / source_name
        if source.exists():
            shutil.copy2(source, task_dir / target_name)
    test_logs = sorted(latest.glob("test_*.log"), key=lambda item: item.stat().st_mtime, reverse=True)
    if test_logs:
        shutil.copy2(test_logs[0], task_dir / f"{variant_id}_test_tool.log")
    burn_logs = sorted(BURN_LOG_ROOT.glob("*.log"), key=lambda item: item.stat().st_mtime, reverse=True)
    if burn_logs:
        shutil.copy2(burn_logs[0], task_dir / f"{variant_id}_burn.log")


def short_text(value: Any, limit: int = 140) -> str:
    text = re.sub(r"\s+", " ", str(value or "")).strip()
    if len(text) <= limit:
        return text
    return text[: max(limit - 3, 0)] + "..."


def xlsx_text(row_item: Sequence[Any], header_index: Dict[str, int], name: str) -> str:
    idx = header_index.get(name)
    if idx is None or idx >= len(row_item):
        return ""
    value = row_item[idx]
    if value is None:
        return ""
    return str(value).strip()


def device_case_match_keys(row_item: Sequence[Any], header_index: Dict[str, int]) -> List[tuple]:
    wake = xlsx_text(row_item, header_index, "唤醒词")
    command = xlsx_text(row_item, header_index, COMMAND)
    expect_proto = hexp(xlsx_text(row_item, header_index, EXPECT_PROTO))
    keys: List[tuple] = []
    if wake or command or expect_proto:
        keys.append((wake, command, expect_proto))
        keys.append((wake, command, ""))
        return keys
    return [("case", xlsx_text(row_item, header_index, CASE_ID))]


def load_device_xlsx_summary(xlsx_path: Path) -> Dict[str, Any]:
    from openpyxl import load_workbook

    wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    try:
        ws = wb["首轮测试结果"] if "首轮测试结果" in wb.sheetnames else wb[wb.sheetnames[0]]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return {"status": "ok_from_xlsx", "total": 0, "counters": {}, "finalFailures": []}
        headers = [str(cell) if cell is not None else "" for cell in rows[0]]
        header_index = {name: idx for idx, name in enumerate(headers) if name}
        verdict_idx = header_index.get(VERDICT_COL)
        detail_candidates = [
            header_index.get("设备响应列表"),
            header_index.get("识别结果"),
            header_index.get("实际发送协议"),
            header_index.get("识别原始结果"),
        ]
        counters: Dict[str, int] = {}
        failures: List[Dict[str, Any]] = []
        total = 0
        for row_item in rows[1:]:
            if not any(value is not None and str(value).strip() for value in row_item):
                continue
            verdict = "UNKNOWN"
            if verdict_idx is not None and verdict_idx < len(row_item) and row_item[verdict_idx] is not None:
                verdict = str(row_item[verdict_idx]).strip() or "UNKNOWN"
            counters[verdict] = counters.get(verdict, 0) + 1
            total += 1
            if verdict == "OK":
                continue
            case_id = xlsx_text(row_item, header_index, CASE_ID) or "-"
            detail = ""
            detail_fragments: List[str] = []
            for idx in detail_candidates:
                if idx is None or idx >= len(row_item):
                    continue
                value = row_item[idx]
                if value is None:
                    continue
                text = str(value).strip()
                if text:
                    detail_fragments.append(text)
                if text and not detail:
                    detail = short_text(text)
            failures.append(
                {
                    "caseId": case_id,
                    "verdict": verdict,
                    "detail": detail,
                    "rawDetail": " | ".join(detail_fragments),
                    "matchKeys": device_case_match_keys(row_item, header_index),
                }
            )

        retry_success_keys = set()
        if "失败重测结果" in wb.sheetnames:
            retry_ws = wb["失败重测结果"]
            retry_rows = list(retry_ws.iter_rows(values_only=True))
            if retry_rows:
                retry_headers = [str(cell) if cell is not None else "" for cell in retry_rows[0]]
                retry_header_index = {name: idx for idx, name in enumerate(retry_headers) if name}
                retry_verdict_idx = retry_header_index.get(VERDICT_COL)
                for row_item in retry_rows[1:]:
                    if not any(value is not None and str(value).strip() for value in row_item):
                        continue
                    verdict = "UNKNOWN"
                    if retry_verdict_idx is not None and retry_verdict_idx < len(row_item) and row_item[retry_verdict_idx] is not None:
                        verdict = str(row_item[retry_verdict_idx]).strip() or "UNKNOWN"
                    if verdict == "OK":
                        retry_success_keys.update(device_case_match_keys(row_item, retry_header_index))

        retry_resolvable = {"WakeupFail", "ProtoFail", "AsrFail", "Fail", "UNKNOWN", "ConfigFail"}
        final_counters = dict(counters)
        final_failures: List[Dict[str, Any]] = []
        for item in failures:
            retry_matched = any(key in retry_success_keys for key in (item.get("matchKeys") or []))
            # 最终报告应以“失败重测后的最终结论”为准。
            # 设备侧 xlsx 的首轮 ConfigFail 往往只是命令链路/协议抓取瞬时异常，
            # 失败重测同 case 已恢复为 OK 时，不应继续算作最终失败。
            if retry_matched and item["verdict"] in retry_resolvable:
                final_counters[item["verdict"]] = max(final_counters.get(item["verdict"], 0) - 1, 0)
                final_counters["OK"] = final_counters.get("OK", 0) + 1
                continue
            final_failures.append(item)
        final_counters = {key: value for key, value in final_counters.items() if value}
        return {"status": "ok_from_xlsx", "total": total, "counters": final_counters, "finalFailures": final_failures}
    finally:
        wb.close()


def device_result_has_issue(device_result: Dict[str, Any], exit_code: int) -> bool:
    if exit_code != 0:
        return True
    if not isinstance(device_result, dict):
        return True
    final_failures = list(device_result.get("finalFailures") or [])
    actionable_failures = [
        item
        for item in final_failures
        if scalar_text(item.get("verdict")) not in {"Skip", "Skip(人工)"}
    ]
    if actionable_failures:
        return True
    counters = dict(device_result.get("counters") or {})
    for key, value in counters.items():
        if key in {"OK", "Skip", "Skip(人工)"}:
            continue
        try:
            if int(value) > 0:
                return True
        except Exception:
            return True
    return False


def ordered_counter_parts(counters: Dict[str, Any]) -> List[str]:
    preferred = ["OK", "Skip", "Reboot", "ConfigFail", "WakeupFail", "ProtoFail", "AsrFail", "Fail", "UNKNOWN"]
    parts: List[str] = []
    seen = set()
    for key in preferred:
        if key in counters:
            parts.append(f"{key}={counters[key]}")
            seen.add(key)
    for key in sorted(counters):
        if key in seen:
            continue
        parts.append(f"{key}={counters[key]}")
    return parts


def result_summary_text(result: Dict[str, Any]) -> str:
    if not isinstance(result, dict) or not result:
        return "-"
    parts = ordered_counter_parts(dict(result.get("counters") or {}))
    total = result.get("total")
    if total not in (None, ""):
        parts.append(f"total={total}")
    status = scalar_text(result.get("status"))
    if status and status not in {"ok", "ok_from_xlsx"}:
        parts.append(f"status={status}")
    exit_code = result.get("exitCode")
    if exit_code not in (None, ""):
        parts.append(f"exit={exit_code}")
    if not parts:
        tail = short_text(result.get("logTail") or result.get("error") or "", limit=180)
        if tail:
            parts.append(tail)
    return "；".join(parts) or (status or "-")


def _named_value(text: Any, *keys: str) -> str:
    source = str(text or "")
    for key in keys:
        match = re.search(rf"{re.escape(key)}=([^；;| ]+)", source)
        if match:
            return match.group(1).strip()
    return ""


def _contains_any(text: str, parts: Sequence[str]) -> bool:
    return any(part and part in text for part in parts)


def summarize_failure_entry(case_id: str, verdict: str, detail: str, raw_detail: str = "") -> str:
    full_text = " | ".join(part for part in [str(detail or "").strip(), str(raw_detail or "").strip()] if part).strip()
    case_id = scalar_text(case_id) or "-"
    verdict = scalar_text(verdict) or "UNKNOWN"

    if verdict == "Reboot" or _contains_any(full_text.lower(), ["unexpected-reboot", "自动重启", "self reboot"]):
        if case_id.startswith("VOICE-"):
            summary = "执行语音注册链路时设备发生异常重启，按规则直接判 FAIL。"
        else:
            summary = "执行过程中设备发生异常重启，按规则直接判 FAIL。"
        return f"{case_id} [{verdict}] {summary}"

    if case_id == "CORE-DEFAULTVOL-001":
        target_default = _named_value(full_text, "expected_default", "defaultVol")
        inferred_default = _named_value(full_text, "inferred_default")
        running_volume = _named_value(full_text, "running_volume", "volume")
        summary = (
            f"默认音量设为 {target_default or '?'} 档后，上电推断默认档为 {inferred_default or '?'}，"
            f"运行音量={running_volume or '?'}，与预期不符。"
        )
        return f"{case_id} [{verdict}] {summary}"

    if case_id == "CORE-VOLUME-001" and _contains_any(full_text, ["协议不一致", "未捕获发送协议"]):
        summary = "音量调节命令触发后回传协议与期望不一致，音量边界行为异常。"
        return f"{case_id} [{verdict}] {summary}"

    if case_id == "CORE-VOLUME-001":
        cfg_match = re.search(r"配置档位=(\d+)", full_text)
        actual_match = re.search(r"实际检测=(\d+)", full_text)
        cfg_value = cfg_match.group(1) if cfg_match else ""
        actual_value = actual_match.group(1) if actual_match else ""
        if _contains_any(full_text, ["no volume values captured", "volume=None", "档位数不匹配"]):
            summary = (
                f"配置音量档位为 {cfg_value or '?'} 档，但执行增减音量后未采集到有效音量变化；"
                f"实际检测档位={actual_value or '?'}，与配置不符。"
            )
        else:
            summary = "音量档位边界验证未通过，设备侧实际档位行为与配置不一致。"
        return f"{case_id} [{verdict}] {summary}"

    if case_id == "CHG-VOLSAVE-001":
        expected_save = _named_value(full_text, "expected_save")
        expected_value = _named_value(full_text, "expected_boot_index", "expected")
        observed_value = _named_value(full_text, "observed_after_down_index", "observed")
        boot_value = _named_value(full_text, "boot_running_index", "boot-running-config")
        mode_text = "开启" if expected_save == "True" else "关闭" if expected_save == "False" else "配置"
        summary = (
            f"音量掉电保存{mode_text}后，重启恢复档位或减一档结果与预期不一致；"
            f"期望重启档位={expected_value or '?'}，实测重启档位={boot_value or '?'}，减档后实测={observed_value or '?'}。"
        )
        if "assist-save-window=refresh=missing；save=missing" in full_text:
            summary += " 保存窗口日志缺失仅作辅助观察，不单独作为失败依据。"
        return f"{case_id} [{verdict}] {summary}"

    if case_id == "CHG-WAKEWORDSAVE-001":
        expected_save = _named_value(full_text, "wakeWordSave")
        if expected_save == "False" or "after-reboot-inactive-" in full_text:
            summary = "关闭唤醒词掉电保存后，重启后非默认唤醒词仍可继续唤醒，未恢复到默认态。"
        else:
            summary = "开启唤醒词掉电保存后，重启后未保持切换后的目标唤醒词。"
        return f"{case_id} [{verdict}] {summary}"

    if case_id == "MWK-001":
        if "query-current=ConfigFail" in full_text:
            summary = "默认态下“查询唤醒词”未按预期识别，或查询后当前生效词发生异常变化。"
        elif "should_work=False" in full_text and "accepted=" in full_text:
            summary = "默认态下非当前唤醒词仍可继续唤醒，唤醒词隔离失败。"
        else:
            summary = "多唤醒默认态验证未通过。"
        return f"{case_id} [{verdict}] {summary}"

    if case_id in {"MWK-002", "MWK-004", "MWK-006"}:
        if "should_work=False" in full_text and "accepted=" in full_text:
            summary = "切换或恢复默认后，非当前目标唤醒词仍可继续唤醒，指定切换行为异常。"
        else:
            summary = "多唤醒切换/恢复链路未达到预期。"
        return f"{case_id} [{verdict}] {summary}"

    if case_id == "MWK-003":
        if "query-current=ConfigFail" in full_text:
            summary = "查询当前唤醒词命令未按预期识别，或查询后当前生效词发生异常变化。"
        else:
            summary = "查询当前唤醒词场景未通过，当前生效词校验异常。"
        return f"{case_id} [{verdict}] {summary}"

    if case_id == "MWK-005":
        if "restore-default-entry=" in full_text or "scenario=restore_default" in full_text:
            summary = "恢复默认唤醒词后，默认词未重新生效，或其他候选词未正确失效。"
        else:
            summary = "指定切换输入无效目标后，当前唤醒词未保持不变。"
        return f"{case_id} [{verdict}] {summary}"

    if case_id == "MWK-SAVE-001":
        if _contains_any(full_text, ["after-reboot-inactive-", "should_work=False"]) and "accepted=" in full_text:
            summary = "切换唤醒词后断电重启，非当前唤醒词仍可继续唤醒，掉电保持后的隔离行为异常。"
        else:
            summary = "切换唤醒词后断电重启，重启后的当前生效词与 wakeWordSave 配置预期不一致。"
        return f"{case_id} [{verdict}] {summary}"

    if case_id == "VOICE-006":
        if _contains_any(full_text, ["unexpected-learning-success", "reg success"]):
            summary = "唤醒词失败重试耗尽场景下，设备却仍把失败语料当成学习成功。"
        else:
            summary = "唤醒词失败重试耗尽后，失败语料未按预期被拒绝。"
        return f"{case_id} [{verdict}] {summary}"

    if case_id == "VOICE-007":
        summary = "已支持的功能命令词未被正确拦截，命令词冲突保护异常。"
        return f"{case_id} [{verdict}] {summary}"

    if case_id == "VOICE-009":
        summary = "学习/删除/退出类保留词未被正确拦截，命令词保留词冲突保护异常。"
        return f"{case_id} [{verdict}] {summary}"

    if case_id == "VOICE-011":
        summary = "删除命令词后，学习命令词未被彻底移除或删除链路引发异常。"
        return f"{case_id} [{verdict}] {summary}"

    if case_id == "VOICE-012":
        summary = "删除命令词未完成或主动退出后，学习命令词未按预期保持。"
        return f"{case_id} [{verdict}] {summary}"

    if case_id == "VOICE-013":
        summary = "删除唤醒词后，学习唤醒词未被彻底移除或删除链路引发异常。"
        return f"{case_id} [{verdict}] {summary}"

    if case_id == "VOICE-014":
        summary = "删除唤醒词未完成或主动退出后，学习唤醒词未按预期保持。"
        return f"{case_id} [{verdict}] {summary}"

    concise = short_text(detail or raw_detail or "", limit=180)
    return f"{case_id} [{verdict}] {concise}".strip()


def summarize_runtime_issue(status: str, device_result: Optional[Dict[str, Any]] = None) -> List[str]:
    if not isinstance(device_result, dict):
        return []
    tail = short_text(device_result.get("logTail") or device_result.get("error") or "", limit=260)
    if not tail:
        return []
    if status == "device_burn_failed":
        return [f"烧录失败：{tail}"]
    if "wakeKw=已找到" in tail and "asrKw=已找到" in tail and "sendMsg=未找到" in tail:
        return ["设备已能识别唤醒和命令词，但未捕获到协议发送(sendMsg)，结果表未生成。"]
    if "未获取启动日志" in tail:
        return ["设备执行阶段未获取到启动日志，结果表未生成。"]
    if "missing_csv" in tail:
        return ["设备执行阶段未生成结果表，请结合结果目录中的日志继续定位。"]
    return [tail]


def extract_failed_config_cases(csv_path: Path, limit: int = 8) -> List[str]:
    if not csv_path.exists():
        return []
    try:
        rows = list(csv.DictReader(csv_path.open("r", encoding="utf-8-sig", newline="")))
    except Exception as exc:
        return [f"配置断言结果读取失败: {type(exc).__name__}: {exc}"]
    failures: List[str] = []
    for row_item in rows:
        verdict = scalar_text(row_item.get("执行结果"))
        if verdict == "OK" or verdict.startswith("Skip"):
            continue
        case_id = scalar_text(row_item.get("用例编号")) or "-"
        detail = short_text(row_item.get("结果详情") or row_item.get("配置断言") or "", limit=220)
        failures.append(summarize_failure_entry(case_id, verdict or "UNKNOWN", detail, row_item.get("结果详情") or ""))
        if len(failures) >= limit:
            break
    return failures


def extract_failed_device_cases(xlsx_path: Path, limit: int = 8) -> List[str]:
    if not xlsx_path.exists():
        return []
    try:
        summary = load_device_xlsx_summary(xlsx_path)
        failures: List[str] = []
        for item in summary.get("finalFailures") or []:
            failures.append(
                summarize_failure_entry(
                    scalar_text(item.get("caseId")) or "-",
                    scalar_text(item.get("verdict")) or "UNKNOWN",
                    scalar_text(item.get("detail")) or "",
                    scalar_text(item.get("rawDetail")) or "",
                )
            )
            if len(failures) >= limit:
                break
        return failures
    except Exception as exc:
        return [f"设备结果读取失败: {type(exc).__name__}: {exc}"]


def variant_status_text(status: str) -> str:
    mapping = {
        "planned": "未执行",
        "packaged": "已打包",
        "device_skipped": "设备跳过",
        "done": "通过",
        "config_failed": "配置断言失败",
        "pack_failed": "打包失败",
        "device_burn_failed": "烧录失败",
        "device_done_with_issue": "设备存在失败项",
    }
    return mapping.get(status, status or "-")


def variant_status_class(status: str) -> str:
    if status == "done":
        return "status-ok"
    if status in {"planned", "packaged", "device_skipped"}:
        return "status-warn"
    return "status-fail"


def overrides_summary_text(values: Dict[str, Any], limit: int = 220) -> str:
    parts: List[str] = []
    for key, value in (values or {}).items():
        if key == "comments":
            continue
        if isinstance(value, dict):
            if key == "releaseMultiWakeup":
                mode = scalar_text(value.get("mode"))
                word_count = len(value.get("wakeupWords") or [])
                parts.append(f"{key}={{mode={mode}, wakeupWords={word_count}}}")
            elif key == "releaseRegist":
                parts.append(f"{key}={{registMode={scalar_text(value.get('registMode'))}}}")
            else:
                parts.append(f"{key}=<complex>")
            continue
        if isinstance(value, list):
            parts.append(f"{key}=[{len(value)}]")
            continue
        parts.append(f"{key}={value}")
    return short_text("；".join(parts), limit=limit) or "-"


def find_variant_failure_details(task_dir: Path, variant_id: str, status: str, device_result: Optional[Dict[str, Any]] = None) -> List[str]:
    details = extract_failed_device_cases(task_dir / f"{variant_id}_testResult.xlsx")
    if details:
        return details
    details = extract_failed_config_cases(task_dir / f"{variant_id}_config_suite_config_assert_result.csv")
    if details:
        return details
    if status == "device_burn_failed":
        burn_log = task_dir / f"{variant_id}_burn.log"
        if burn_log.exists():
            tail = short_text("\n".join(burn_log.read_text(encoding="utf-8", errors="ignore").splitlines()[-10:]), limit=260)
            return [tail] if tail else []
    details = summarize_runtime_issue(status, device_result)
    if details:
        return details
    return []


def refresh_variant_human_fields(task_dir: Path, variants: Sequence[Dict[str, Any]]) -> None:
    for item in variants:
        if not isinstance(item, dict):
            continue
        variant_id = scalar_text(item.get("id"))
        status = scalar_text(item.get("status"))
        is_target = variant_id.startswith("cv-") or scalar_text(item.get("kind")) == "targeted-retest"

        if status == "done":
            item["humanResult"] = "打包成功；目标功能点验证通过。" if is_target else "打包成功；组合包执行完成。"
            if is_target:
                existing = scalar_text(item.get("humanEvidence"))
                item["humanEvidence"] = existing if existing and ("通过" in existing or "PASS" in existing.upper()) else "目标用例通过。"
            else:
                item["humanEvidence"] = "组合包已执行完成，可用于边界/中值覆盖，不直接作为单点归因结论。"
            item["humanFailures"] = []
            continue

        if status not in {"pack_failed", "config_failed", "device_burn_failed", "device_done_with_issue"}:
            continue

        failures = find_variant_failure_details(task_dir, variant_id, status, item.get("deviceResult"))
        if not failures:
            failures = summarize_runtime_issue(status, item.get("deviceResult"))
        if not failures:
            failures = [result_summary_text(item.get("deviceResult") or item.get("configResult") or {})]
        failures = [scalar_text(text) for text in failures if scalar_text(text)]
        item["humanFailures"] = failures[:3]
        item["humanEvidence"] = failures[0] if failures else "-"

        if status == "pack_failed":
            item["humanResult"] = "打包失败。"
        elif status == "config_failed":
            item["humanResult"] = "打包成功；配置断言失败。"
        elif status == "device_burn_failed":
            item["humanResult"] = "打包成功；烧录失败。"
        else:
            item["humanResult"] = "打包成功；目标功能点验证失败。" if is_target else "打包成功；组合包存在失败项。"


def safe_dir_fragment(value: Any, fallback: str = "result") -> str:
    text = re.sub(r"\s+", "", str(value or "")).strip()
    text = re.sub(r'[\\/:*?"<>|]+', "-", text)
    text = text.strip(".-")
    return text or fallback


def config_csv_path_for_variant(task_dir: Path, item: Dict[str, Any]) -> Optional[Path]:
    variant_id = scalar_text(item.get("id"))
    direct = task_dir / f"{variant_id}_config_suite_config_assert_result.csv"
    if direct.exists():
        return direct
    result_dir = existing_dir((item.get("configResult") or {}).get("resultDir"))
    if result_dir:
        csv_path = result_dir / "suite_config_assert_result.csv"
        if csv_path.exists():
            return csv_path
    suite_dir = existing_dir(item.get("suiteDir"))
    latest = latest_dir(suite_dir / "result") if suite_dir else None
    if latest:
        csv_path = latest / "suite_config_assert_result.csv"
        if csv_path.exists():
            return csv_path
    return None


def device_xlsx_path_for_variant(task_dir: Path, item: Dict[str, Any]) -> Optional[Path]:
    variant_id = scalar_text(item.get("id"))
    direct = task_dir / f"{variant_id}_testResult.xlsx"
    if direct.exists():
        return direct
    result_dir = existing_dir((item.get("deviceResult") or {}).get("resultDir"))
    if result_dir:
        xlsx_path = result_dir / "testResult.xlsx"
        if xlsx_path.exists():
            return xlsx_path
    suite_dir = existing_dir(item.get("suiteDir"))
    latest = latest_dir(suite_dir / "result") if suite_dir else None
    if latest:
        xlsx_path = latest / "testResult.xlsx"
        if xlsx_path.exists():
            return xlsx_path
    return None


def human_param_lines(values: Dict[str, Any]) -> List[str]:
    label_map = {
        "timeout": "唤醒时长",
        "volLevel": "音量档位数",
        "defaultVol": "初始化默认音量",
        "volMaxOverflow": "最大音量上溢播报语",
        "volMinOverflow": "最小音量下溢播报语",
        "uportBaud": "协议串口波特率",
        "logLevel": "日志级别",
        "wakeWordSave": "唤醒词掉电保存",
        "volSave": "音量掉电保存",
        "vcn": "合成发音人",
        "speed": "合成语速",
        "vol": "合成音量",
        "compress": "播报音压缩比",
        "paConfigEnable": "功放配置开关",
        "voiceRegEnable": "语音注册",
        "multiWkeEnable": "多唤醒开关",
        "multiWkeMode": "多唤醒模式",
    }
    ordered_keys = [
        "timeout",
        "volLevel",
        "defaultVol",
        "uportBaud",
        "logLevel",
        "volSave",
        "wakeWordSave",
        "vcn",
        "speed",
        "vol",
        "compress",
        "volMaxOverflow",
        "volMinOverflow",
        "paConfigEnable",
        "voiceRegEnable",
        "multiWkeEnable",
        "multiWkeMode",
        "releaseRegist",
        "releaseRegistConfig",
        "releaseMultiWke",
        "releaseAlgoList",
    ]

    def render_scalar(key: str, value: Any) -> str:
        label = label_map.get(key, key)
        if isinstance(value, bool):
            return f"{label}={'true' if value else 'false'}"
        return f"{label}={value}"

    lines: List[str] = []
    seen = set()
    for key in ordered_keys + sorted((values or {}).keys()):
        if key in seen or key == "comments" or key not in (values or {}):
            continue
        seen.add(key)
        value = values.get(key)
        if value in (None, ""):
            continue
        if key == "releaseMultiWke" and isinstance(value, dict):
            words = [scalar_text(item.get("word")) for item in (value.get("wkelist") or value.get("switch_list") or []) if scalar_text(item.get("word"))]
            mode = scalar_text(value.get("mode"))
            parts = [f"多唤醒模式={mode}"] if mode else []
            if words:
                parts.append(f"多唤醒候选词={'/'.join(words)}")
            lines.extend(parts or ["多唤醒配置=<complex>"])
            continue
        if key == "releaseAlgoList" and isinstance(value, list):
            wake_words = [scalar_text(item.get("word")) for item in value if scalar_text(item.get("type")) == "唤醒词" and scalar_text(item.get("word"))]
            if wake_words:
                lines.append(f"算法唤醒词={'/'.join(wake_words)}")
            else:
                lines.append(f"算法词条数量={len(value)}")
            continue
        if key == "releaseRegist" and isinstance(value, dict):
            cmd_words = [scalar_text(item.get("word")) for item in (value.get("regCommands") or value.get("reg_commands") or []) if scalar_text(item.get("word"))]
            wake_words = [scalar_text(item.get("word")) for item in (value.get("regWakewords") or value.get("reg_wakewords") or []) if scalar_text(item.get("word"))]
            if cmd_words:
                lines.append(f"语音注册命令词={'/'.join(cmd_words)}")
            if wake_words:
                lines.append(f"语音注册唤醒词={'/'.join(wake_words)}")
            if not cmd_words and not wake_words:
                lines.append("语音注册配置=<complex>")
            continue
        if key == "releaseRegistConfig" and isinstance(value, dict):
            summary_parts = []
            for sub_key in ["asrStudyRepeatCount", "asrStudyRetryCount", "wakeupStudyRepeatCount", "wakeupStudyRetryCount"]:
                if sub_key in value and value.get(sub_key) not in (None, ""):
                    summary_parts.append(f"{sub_key}={value.get(sub_key)}")
            lines.append("语音注册参数=" + ("；".join(summary_parts) if summary_parts else "<complex>"))
            continue
        if isinstance(value, dict):
            lines.append(f"{label_map.get(key, key)}=<complex>")
            continue
        if isinstance(value, list):
            lines.append(f"{label_map.get(key, key)}=[{len(value)}]")
            continue
        lines.append(render_scalar(key, value))
    return lines or ["保持默认配置"]


def derive_bundle_dir_name(task_dir: Path, item: Dict[str, Any]) -> str:
    bundle_name = scalar_text(item.get("resultBundleDirName"))
    if bundle_name:
        return bundle_name
    firmware = existing_file(item.get("downloadedFirmwarePath"))
    if firmware and task_dir not in firmware.parents:
        candidate = firmware.parent.name
        if candidate:
            return safe_dir_fragment(candidate, scalar_text(item.get("id")) or "result")
    variant_id = scalar_text(item.get("id")) or "variant"
    title = scalar_text(item.get("title")) or variant_id
    return safe_dir_fragment(f"{variant_id}_{title}", variant_id)


def bundle_variant_results(task_dir: Path, item: Dict[str, Any]) -> Dict[str, str]:
    bundle_root = task_dir / "result"
    bundle_root.mkdir(parents=True, exist_ok=True)
    dir_name = derive_bundle_dir_name(task_dir, item)
    bundle_dir = bundle_root / dir_name
    if bundle_dir.exists():
        shutil.rmtree(bundle_dir)
    bundle_dir.mkdir(parents=True, exist_ok=True)

    variant_id = scalar_text(item.get("id"))
    firmware = existing_file(item.get("downloadedFirmwarePath"))
    parameter_txt = existing_file(item.get("parameterTxtPath"))
    config_result_dir = existing_dir((item.get("configResult") or {}).get("resultDir"))
    device_result_dir = existing_dir((item.get("deviceResult") or {}).get("resultDir"))
    suite_dir = existing_dir(item.get("suiteDir"))

    for source in [firmware, parameter_txt]:
        if source and source.exists():
            shutil.copy2(source, bundle_dir / source.name)

    if config_result_dir:
        shutil.copytree(config_result_dir, bundle_dir / "config_result", dirs_exist_ok=True)
    config_csv = config_csv_path_for_variant(task_dir, item)
    if config_csv and config_csv.exists():
        shutil.copy2(config_csv, bundle_dir / config_csv.name)

    if device_result_dir:
        shutil.copytree(device_result_dir, bundle_dir / "device_result", dirs_exist_ok=True)
    device_xlsx = device_xlsx_path_for_variant(task_dir, item)
    if device_xlsx and device_xlsx.exists():
        shutil.copy2(device_xlsx, bundle_dir / device_xlsx.name)

    for suffix in ["_burn.log", "_serial_raw.log", "_test_tool.log", "_config_suite_config_assert_result.json"]:
        path = task_dir / f"{variant_id}{suffix}"
        if path.exists():
            shutil.copy2(path, bundle_dir / path.name)

    if suite_dir:
        for name in ["testCases.csv", "executable_cases.json", "README.md", "deviceInfo_generated.json"]:
            source = suite_dir / name
            if source.exists():
                target_dir = bundle_dir / "suite"
                target_dir.mkdir(parents=True, exist_ok=True)
                shutil.copy2(source, target_dir / name)

    summary_path = bundle_dir / "summary.txt"
    summary_lines = [
        f"执行包: {dir_name}",
        f"变体ID: {variant_id or '-'}",
        f"标题: {scalar_text(item.get('title')) or '-'}",
        f"状态: {variant_status_text(scalar_text(item.get('status')))}",
        "参数配置:",
    ]
    summary_lines.extend(f"- {line}" for line in human_param_lines(item.get("resolvedOverrides") or item.get("overrides") or {}))
    summary_lines.extend(
        [
            f"执行结果: {scalar_text(item.get('humanResult')) or (result_summary_text(item.get('configResult') or {}) + ' / ' + result_summary_text(item.get('deviceResult') or {}))}",
            f"关键说明: {scalar_text(item.get('humanEvidence')) or '; '.join(find_variant_failure_details(task_dir, variant_id, scalar_text(item.get('status')), item.get('deviceResult'))[:2] or ['无失败项'])}",
            f"固件路径: {str(firmware) if firmware else '-'}",
            f"参数文件: {str(parameter_txt) if parameter_txt else '-'}",
            f"配置结果目录: {str(config_result_dir) if config_result_dir else '-'}",
            f"设备结果目录: {str(device_result_dir) if device_result_dir else '-'}",
        ]
    )
    summary_path.write_text("\n".join(summary_lines).rstrip() + "\n", encoding="utf-8")

    item["resultBundleDirName"] = dir_name
    item["resultBundleDir"] = str(bundle_dir.resolve())
    return {"dirName": dir_name, "relativePath": f"result/{dir_name}", "path": str(bundle_dir.resolve())}


def load_case_ids_from_xlsx(xlsx_path: Path) -> List[str]:
    from openpyxl import load_workbook

    wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    try:
        ws = wb["首轮测试结果"] if "首轮测试结果" in wb.sheetnames else wb[wb.sheetnames[0]]
        rows = ws.iter_rows(values_only=True)
        headers = next(rows, None)
        if not headers:
            return []
        header_index = {str(value) if value is not None else "": idx for idx, value in enumerate(headers)}
        case_idx = header_index.get(CASE_ID)
        if case_idx is None:
            return []
        result: List[str] = []
        for row in rows:
            if case_idx >= len(row):
                continue
            value = row[case_idx]
            text = str(value).strip() if value is not None else ""
            if text:
                result.append(text)
        return result
    finally:
        wb.close()


def build_algo_summary(task_dir: Path, variants: Sequence[Dict[str, Any]], feature_map: Dict[str, Any]) -> List[Dict[str, str]]:
    specs: List[Dict[str, Any]] = [
        {"feature": "唤醒识别", "prefixes": ("CORE-WAKE-",), "enabled": True},
        {"feature": "命令词识别", "prefixes": ("CORE-CMD-",), "enabled": True},
        {"feature": "协议链路", "prefixes": ("CORE-CMD-", "CHG-SNDPROTO-", "CHG-RECPROTO-"), "enabled": True},
        {"feature": "语音注册", "prefixes": ("VOICE-",), "enabled": str(feature_map.get("voice_regist") or "") == "Optional"},
        {"feature": "多唤醒词", "prefixes": ("MWK-",), "enabled": str(feature_map.get("multi_wakeup") or "") == "Optional"},
    ]
    algo_issue_statuses = {"pack_failed", "config_failed", "device_burn_failed", "device_done_with_issue"}
    result: List[Dict[str, str]] = []
    for spec in specs:
        if not spec["enabled"]:
            continue
        matched_variants: List[str] = []
        failures: List[str] = []
        fallback_failures: List[str] = []
        for item in variants:
            overrides = dict(item.get("resolvedOverrides") or item.get("overrides") or {})
            variant_id = scalar_text(item.get("id"))
            variant_dir = derive_bundle_dir_name(task_dir, item)
            xlsx_path = device_xlsx_path_for_variant(task_dir, item)
            if not xlsx_path or not xlsx_path.exists():
                runtime_issue = summarize_runtime_issue(scalar_text(item.get("status")), item.get("deviceResult"))
                fallback_issue = scalar_text(item.get("humanEvidence")) or (runtime_issue[0] if runtime_issue else result_summary_text(item.get("deviceResult") or item.get("configResult") or {}))
                if spec["feature"] == "语音注册" and overrides.get("voiceRegEnable") and scalar_text(item.get("status")) in algo_issue_statuses:
                    matched_variants.append(variant_dir)
                    fallback_failures.append(fallback_issue)
                if spec["feature"] == "多唤醒词" and overrides.get("multiWkeEnable") and scalar_text(item.get("status")) in algo_issue_statuses:
                    matched_variants.append(variant_dir)
                    fallback_failures.append(fallback_issue)
                continue
            case_ids = load_case_ids_from_xlsx(xlsx_path)
            if not any(any(case_id.startswith(prefix) for prefix in spec["prefixes"]) for case_id in case_ids):
                continue
            matched_variants.append(variant_dir)
            summary = load_device_xlsx_summary(xlsx_path)
            for failure in summary.get("finalFailures") or []:
                case_id = scalar_text(failure.get("caseId"))
                verdict = scalar_text(failure.get("verdict")) or "UNKNOWN"
                if verdict.startswith("Skip"):
                    continue
                if any(case_id.startswith(prefix) for prefix in spec["prefixes"]):
                    failures.append(
                        summarize_failure_entry(
                            case_id,
                            verdict,
                            scalar_text(failure.get("detail")) or "",
                            scalar_text(failure.get("rawDetail")) or "",
                        )
                    )
        if failures:
            result.append(
                {
                    "feature": spec["feature"],
                    "coverage": " / ".join(dict.fromkeys(matched_variants)) or "已执行",
                    "status": "FAIL",
                    "summary": failures[0],
                }
            )
            continue
        if fallback_failures:
            result.append(
                {
                    "feature": spec["feature"],
                    "coverage": " / ".join(dict.fromkeys(matched_variants)) or "已执行",
                    "status": "FAIL",
                    "summary": fallback_failures[0],
                }
            )
            continue
        if matched_variants:
            result.append(
                {
                    "feature": spec["feature"],
                    "coverage": " / ".join(dict.fromkeys(matched_variants)),
                    "status": "PASS",
                    "summary": "相关自动用例执行通过，未发现遗留失败项。",
                }
            )
            continue
        result.append(
            {
                "feature": spec["feature"],
                "coverage": "-",
                "status": "BLOCK",
                "summary": "当前任务未纳入该算法专项执行。",
            }
        )
    return result


def generate_weekly_email_report(task_dir: Path, runtime_dir: Path, state: Dict[str, Any]) -> None:
    from html import escape

    selected = dict(state.get("selectedMeta") or {})
    variants = [dict(item) for item in (state.get("variants") or []) if isinstance(item, dict)]
    refresh_variant_human_fields(task_dir, variants)
    catalog_payload = {}
    catalog_path = existing_file((state.get("catalogPaths") or {}).get("parameterCatalog"))
    if catalog_path:
        catalog_payload = jload(catalog_path)
    feature_map = dict(catalog_payload.get("featureMap") or {})
    product_name = scalar_text(state.get("productName")) or build_weekly_product_name(str(selected.get("moduleMark") or ""), str(selected.get("productLabel") or ""), str(selected.get("versionLabel") or ""))
    manifest_path = runtime_dir / "manifest.json"

    bundle_map: Dict[str, Dict[str, str]] = {}
    for item in variants:
        variant_id = scalar_text(item.get("id"))
        if not variant_id:
            continue
        bundle_map[variant_id] = bundle_variant_results(task_dir, item)

    state["variants"] = variants
    algo_summary = build_algo_summary(task_dir, variants, feature_map)
    state["algoSummary"] = algo_summary
    state["updatedAt"] = datetime.now().isoformat(timespec="seconds")
    jsave(manifest_path, state)

    total_variants = len(variants)
    pass_variants = sum(1 for item in variants if str(item.get("status") or "") == "done")
    issue_statuses = {"pack_failed", "config_failed", "device_burn_failed", "device_done_with_issue"}
    issue_variants = sum(1 for item in variants if str(item.get("status") or "") in issue_statuses)
    pending_variants = sum(1 for item in variants if str(item.get("status") or "") in {"planned", "packaged"})
    pass_rate = int(pass_variants / total_variants * 100) if total_variants else 0
    if issue_variants:
        conclusion = "FAIL"
        result_text = f"存在异常包 {issue_variants}/{total_variants}"
    elif pending_variants:
        conclusion = "PARTIAL"
        result_text = f"存在未完成包 {pending_variants}/{total_variants}"
    else:
        conclusion = "PASS"
        result_text = f"全部包通过 {pass_variants}/{total_variants}"

    feature_rows = [
        ("voice_regist", scalar_text(feature_map.get("voice_regist")) or "-", "Unsupported 不纳入本轮" if scalar_text(feature_map.get("voice_regist")) == "Unsupported" else "按产品能力决定"),
        ("multi_wakeup", scalar_text(feature_map.get("multi_wakeup")) or "-", "Optional 纳入本轮专项" if scalar_text(feature_map.get("multi_wakeup")) == "Optional" else "不纳入专项"),
    ]
    feature_rows_html = "\n".join(
        "                <tr>"
        f"<td>{escape(name)}</td>"
        f"<td>{escape(value)}</td>"
        f"<td>{escape(action)}</td>"
        "</tr>"
        for name, value, action in feature_rows
    )

    feature_summary = [
        dict(item)
        for item in (state.get("featureSummary") or [])
        if isinstance(item, dict) and scalar_text(item.get("feature")) != "欢迎语 TTS 文案"
    ]
    feature_badge_class = {"PASS": "status-ok", "FAIL": "status-fail", "BLOCK": "status-warn"}
    overall_feature_summary = feature_summary + algo_summary
    if overall_feature_summary:
        pass_count = sum(1 for item in overall_feature_summary if scalar_text(item.get("status")).upper() == "PASS")
        fail_count = sum(1 for item in overall_feature_summary if scalar_text(item.get("status")).upper() == "FAIL")
        block_count = sum(1 for item in overall_feature_summary if scalar_text(item.get("status")).upper() == "BLOCK")
        total_features = len(overall_feature_summary)
        pass_rate = int(pass_count / total_features * 100) if total_features else pass_rate
        if fail_count:
            conclusion = "FAIL"
            result_text = f"功能点 PASS {pass_count} / FAIL {fail_count} / BLOCK {block_count}"
        elif block_count:
            conclusion = "PARTIAL"
            result_text = f"功能点 PASS {pass_count} / BLOCK {block_count}"
        else:
            conclusion = "PASS"
            result_text = f"功能点全部通过 {pass_count}/{total_features}"
    else:
        pass_count = fail_count = block_count = 0

    def rows_html(items: Sequence[Dict[str, Any]]) -> str:
        return "\n".join(
            "                <tr>\n"
            f"                    <td>{escape(scalar_text(item.get('feature')) or '-')}</td>\n"
            f"                    <td>{escape(scalar_text(item.get('coverage')) or '-')}</td>\n"
            f"                    <td><span class=\"{feature_badge_class.get(scalar_text(item.get('status')).upper(), 'status-warn')}\">{escape(scalar_text(item.get('status')).upper() or '-')}</span></td>\n"
            f"                    <td>{escape(scalar_text(item.get('summary')) or '-')}</td>\n"
            "                </tr>"
            for item in items
        )

    feature_section_html = ""
    if feature_summary:
        feature_section_html = (
            '<div class="section-title">功能点结果</div>\n'
            '<div class="attn-box ok-box">这里只展示本轮需要自动验证的可配置参数，已按新口径剔除页面 TTS 文案。</div>\n'
            "<table>\n"
            "                <thead>\n"
            "                    <tr><th style=\"width: 22%;\">功能点</th><th style=\"width: 22%;\">本轮覆盖</th><th style=\"width: 10%;\">结果</th><th style=\"width: 46%;\">说明</th></tr>\n"
            "                </thead>\n"
            "                <tbody>\n"
            f"{rows_html(feature_summary)}\n"
            "                </tbody>\n"
            "            </table>"
        )
    algo_section_html = ""
    if algo_summary:
        algo_section_html = (
            '<div class="section-title">算法功能结果</div>\n'
            '<div class="attn-box ok-box">算法侧单独展示唤醒、命令词、协议、语音注册、多唤醒，避免被参数结果淹没。</div>\n'
            "<table>\n"
            "                <thead>\n"
            "                    <tr><th style=\"width: 18%;\">算法功能</th><th style=\"width: 26%;\">覆盖包</th><th style=\"width: 10%;\">结果</th><th style=\"width: 46%;\">说明</th></tr>\n"
            "                </thead>\n"
            "                <tbody>\n"
            f"{rows_html(algo_summary)}\n"
            "                </tbody>\n"
            "            </table>"
        )

    control_findings = [dict(item) for item in (state.get("controlFindings") or []) if isinstance(item, dict)]
    control_md_path = runtime_dir / "control_variable_analysis.md"
    if control_findings:
        control_lines = [
            f"# 控制变量诊断 - {product_name}",
            "",
            "## 结论",
            f"- {scalar_text(state.get('controlConclusion')) or '无'}",
            "",
            "## 诊断明细",
            "",
        ]
        for item in control_findings:
            control_lines.extend(
                [
                    f"### {scalar_text(item.get('id')) or '-'}",
                    f"- 配置：{scalar_text(item.get('change')) or '-'}",
                    f"- 结果：{scalar_text(item.get('result')) or '-'}",
                    f"- 结论：{scalar_text(item.get('conclusion')) or '-'}",
                    "",
                ]
            )
        control_md_path.write_text("\n".join(control_lines).rstrip() + "\n", encoding="utf-8")
        control_rows_html = "\n".join(
            "                <tr>\n"
            f"                    <td>{escape(scalar_text(item.get('id')) or '-')}</td>\n"
            f"                    <td>{escape(scalar_text(item.get('change')) or '-')}</td>\n"
            f"                    <td>{escape(scalar_text(item.get('result')) or '-')}</td>\n"
            f"                    <td>{escape(scalar_text(item.get('conclusion')) or '-')}</td>\n"
            "                </tr>"
            for item in control_findings
        )
        control_section_html = (
            '<div class="section-title">控制变量诊断</div>\n'
            '<div class="attn-box ok-box">'
            f"<strong>最终归因：</strong>{escape(scalar_text(state.get('controlConclusion')) or '-')}"
            "</div>\n"
            "<table>\n"
            "                <thead>\n"
            "                    <tr><th style=\"width: 16%;\">定位包</th><th style=\"width: 34%;\">变更点</th><th style=\"width: 20%;\">设备结果</th><th style=\"width: 30%;\">结论</th></tr>\n"
            "                </thead>\n"
            "                <tbody>\n"
            f"{control_rows_html}\n"
            "                </tbody>\n"
            "            </table>"
        )
    else:
        control_section_html = ""

    def execution_stage_text(item: Dict[str, Any]) -> str:
        status = scalar_text(item.get("status"))
        if status == "pack_failed":
            return "打包"
        if status == "config_failed":
            return "打包 + 配置断言"
        if status == "device_burn_failed":
            return "打包 + 配置断言 + 烧录"
        if status in {"done", "device_done_with_issue"}:
            return "打包 + 配置断言 + 烧录 + 设备执行"
        if status == "device_skipped":
            return "打包 + 配置断言"
        if item.get("downloadedFirmwarePath"):
            return "打包"
        return "-"

    def device_exec_text(result: Dict[str, Any]) -> str:
        if not isinstance(result, dict) or not result:
            return "-"
        status = scalar_text(result.get("status"))
        if status == "skip_due_config_fail":
            return "未执行（配置断言失败）"
        if status == "skip_by_flag":
            return "未执行（人工跳过）"
        if status == "skip_no_flash_tool":
            return "未执行（无烧录工具）"
        return result_summary_text(result)

    def coverage_note(item: Dict[str, Any]) -> str:
        variant_id = scalar_text(item.get("id"))
        overrides = item.get("resolvedOverrides") or item.get("overrides") or {}
        try:
            timeout_value = int(overrides.get("timeout"))
        except Exception:
            timeout_value = 0
        if variant_id == "pkg-02-left-boundary" and timeout_value <= 1:
            return (
                "未执行普通命令词/全功能用例；timeout=1s 时自动链路等待唤醒确认后再播命令会超过命令窗口，"
                "该包只验证左边界配置、超时、唤醒、播报和串口观察项；命令词全功能由 pkg-01/pkg-03 覆盖。"
            )
        return scalar_text(item.get("coverageNote"))

    def execution_result_text(item: Dict[str, Any]) -> str:
        human_result = scalar_text(item.get("humanResult"))
        if human_result:
            return escape(human_result)
        return escape(result_summary_text(item.get("configResult") or {})) + "<br>" + escape(device_exec_text(item.get("deviceResult") or {}))

    def failure_text(item: Dict[str, Any]) -> str:
        variant_id = scalar_text(item.get("id"))
        human_failures = [scalar_text(x) for x in (item.get("humanFailures") or []) if scalar_text(x)]
        if human_failures:
            return "<br>".join(escape(text) for text in human_failures[:2])
        human_evidence = scalar_text(item.get("humanEvidence"))
        if human_evidence:
            return escape(human_evidence)
        failures = find_variant_failure_details(task_dir, variant_id, scalar_text(item.get("status")), item.get("deviceResult"))[:2] or ["无失败项"]
        note = coverage_note(item)
        if note and failures == ["无失败项"]:
            failures = [f"未测/裁剪：{note}"]
        return "<br>".join(escape(text) for text in failures)

    def result_dir_text(item: Dict[str, Any]) -> str:
        variant_id = scalar_text(item.get("id"))
        bundle = bundle_map.get(variant_id) or {}
        return escape(bundle.get("relativePath") or "-")

    execution_rows_html = "\n".join(
        "                <tr>\n"
        f"                    <td>{escape((bundle_map.get(scalar_text(item.get('id'))) or {}).get('dirName') or scalar_text(item.get('id')) or '-')}</td>\n"
        f"                    <td>{'<br>'.join(escape(line) for line in human_param_lines(item.get('resolvedOverrides') or item.get('overrides') or {}))}</td>\n"
        f"                    <td>{escape(execution_stage_text(item))}</td>\n"
        f"                    <td>{execution_result_text(item)}</td>\n"
        f"                    <td>{failure_text(item)}</td>\n"
        f"                    <td>{result_dir_text(item)}</td>\n"
        "                </tr>"
        for item in variants
    )
    execution_md_path = runtime_dir / "execution_results.md"
    execution_lines = [
        f"# 执行结果 - {product_name}",
        "",
        "## 主包/定位包执行结果",
        "",
    ]
    for item in variants:
        variant_id = scalar_text(item.get("id")) or "-"
        bundle = bundle_map.get(variant_id) or {}
        execution_lines.extend(
            [
                f"### {bundle.get('dirName') or variant_id}",
                "- 参数配置：",
            ]
        )
        execution_lines.extend(f"  - {line}" for line in human_param_lines(item.get("resolvedOverrides") or item.get("overrides") or {}))
        execution_lines.extend(
            [
                f"- 已执行阶段：{execution_stage_text(item)}",
                f"- 执行结果：{scalar_text(item.get('humanResult')) or (result_summary_text(item.get('configResult') or {}) + ' / ' + device_exec_text(item.get('deviceResult') or {}))}",
                f"- 异常/结论：{scalar_text(item.get('humanEvidence')) or '; '.join(find_variant_failure_details(task_dir, variant_id, scalar_text(item.get('status')), item.get('deviceResult'))[:2] or ['无失败项'])}",
                f"- 未测/裁剪：{coverage_note(item) or '无'}",
                f"- 结果目录：{bundle.get('relativePath') or '-'}",
                "",
            ]
        )
    execution_md_path.write_text("\n".join(execution_lines).rstrip() + "\n", encoding="utf-8")
    execution_section_html = (
        '<div class="section-title">执行结果</div>\n'
        '<div class="attn-box ok-box">执行包列直接对应附件中的目录名；每个目录内都包含固件、参数说明、日志和结果表。</div>\n'
        "<table>\n"
        "                <thead>\n"
        "                    <tr><th style=\"width: 14%;\">执行包</th><th style=\"width: 24%;\">当前配置参数</th><th style=\"width: 15%;\">已执行阶段</th><th style=\"width: 15%;\">执行结果</th><th style=\"width: 18%;\">Fail / Block 说明</th><th style=\"width: 14%;\">结果目录</th></tr>\n"
        "                </thead>\n"
        "                <tbody>\n"
        f"{execution_rows_html}\n"
        "                </tbody>\n"
        "            </table>"
    )

    variant_rows_html: List[str] = []
    issue_items_html: List[str] = []
    for item in variants:
        variant_id = scalar_text(item.get("id")) or "-"
        status = scalar_text(item.get("status"))
        bundle = bundle_map.get(variant_id) or {}
        failures = find_variant_failure_details(task_dir, variant_id, status, item.get("deviceResult"))
        failure_text = "<br>".join(escape(text) for text in failures) if failures else "-"
        variant_rows_html.append(
            "                <tr>\n"
            f"                    <td>{escape(bundle.get('dirName') or variant_id)}</td>\n"
            f"                    <td>{escape(variant_status_text(status))}</td>\n"
            f"                    <td>{escape('；'.join(human_param_lines(item.get('resolvedOverrides') or item.get('overrides') or {})))}</td>\n"
            f"                    <td>{escape(result_summary_text(item.get('configResult') or {}))}</td>\n"
            f"                    <td>{escape(result_summary_text(item.get('deviceResult') or {}))}</td>\n"
            f"                    <td>{failure_text}</td>\n"
            "                </tr>"
        )
        if status in issue_statuses:
            summary = scalar_text(item.get("humanResult")) or result_summary_text(item.get("deviceResult") or item.get("configResult") or {})
            issue_items_html.append(
                f"<li><strong>{escape(bundle.get('dirName') or variant_id)}</strong>：参数 {escape('；'.join(human_param_lines(item.get('resolvedOverrides') or item.get('overrides') or {})))}；"
                f"<span class=\"{variant_status_class(status)}\">{escape(variant_status_text(status))}</span>；{escape(summary)}"
                + (f"<br>{failure_text}" if failures else "")
                + "</li>"
            )

    if not issue_items_html:
        issue_html = '<div class="attn-box ok-box">本轮未发现新增异常包。</div>'
    else:
        issue_html = '<div class="attn-box"><strong>技术异常摘要</strong><ul>' + "".join(issue_items_html) + "</ul></div>"

    updated_at = scalar_text(state.get("updatedAt")).replace("T", " ") or datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    feature_matrix_path = task_dir / "feature_matrix.md"
    summary_md_path = runtime_dir / "summary.md"
    summary_lines = [
        f"# 周测汇总 - {product_name}",
        "",
        f"- 测试时间：`{updated_at}`",
        f"- 结果：`{result_text}`",
        "",
    ]
    coverage_notes = [
        (scalar_text(item.get("id")) or "-", coverage_note(item))
        for item in variants
        if coverage_note(item)
    ]
    if coverage_notes:
        summary_lines.extend(["## 未测/裁剪范围", ""])
        for variant_id, note in coverage_notes:
            summary_lines.append(f"- `{variant_id}`：{note}")
        summary_lines.append("")
    summary_lines.extend(["## 参数功能点结果", ""])
    for item in feature_summary:
        summary_lines.extend(
            [
                f"- {scalar_text(item.get('feature')) or '-'}：{scalar_text(item.get('status')) or '-'}",
                f"  - 覆盖：{scalar_text(item.get('coverage')) or '-'}",
                f"  - 说明：{scalar_text(item.get('summary')) or '-'}",
            ]
        )
    summary_lines.extend(["", "## 算法功能结果", ""])
    for item in algo_summary:
        summary_lines.extend(
            [
                f"- {scalar_text(item.get('feature')) or '-'}：{scalar_text(item.get('status')) or '-'}",
                f"  - 覆盖：{scalar_text(item.get('coverage')) or '-'}",
                f"  - 说明：{scalar_text(item.get('summary')) or '-'}",
            ]
        )
    summary_md_path.write_text("\n".join(summary_lines).rstrip() + "\n", encoding="utf-8")

    html_body = f"""<!DOCTYPE html>
<html>
<head>
    <meta charset="UTF-8">
    <style>
        body {{ font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif; background-color: #f5f7fa; margin: 0; padding: 20px; }}
        .container {{ max-width: 1080px; margin: 0 auto; background: #ffffff; border-radius: 12px; overflow: hidden; box-shadow: 0 4px 20px rgba(0,0,0,0.08); }}
        .header {{ background: linear-gradient(135deg, #0f766e 0%, #2563eb 100%); color: white; padding: 28px 30px; }}
        .header h1 {{ margin: 0; font-size: 22px; font-weight: 600; }}
        .header .subtitle {{ margin-top: 6px; opacity: 0.9; font-size: 13px; }}
        .content {{ padding: 28px 30px; }}
        .info-grid {{ display: grid; grid-template-columns: repeat(3, 1fr); gap: 12px; margin-bottom: 20px; }}
        .info-item {{ background: #f8f9fa; padding: 12px 14px; border-radius: 8px; border-left: 4px solid #2563eb; }}
        .info-item label {{ display: block; color: #6c757d; font-size: 11px; margin-bottom: 4px; text-transform: uppercase; letter-spacing: 0.5px; }}
        .info-item span {{ color: #1f2937; font-weight: 600; font-size: 13px; }}
        .section-title {{ color: #1f2937; font-size: 15px; font-weight: 700; margin: 20px 0 12px; padding-bottom: 8px; border-bottom: 2px solid #2563eb; display: inline-block; }}
        table {{ width: 100%; border-collapse: collapse; margin-top: 12px; table-layout: fixed; }}
        th {{ background: #2563eb; color: white; padding: 10px 12px; text-align: left; font-size: 12px; }}
        td {{ padding: 9px 12px; border-bottom: 1px solid #e5e7eb; font-size: 12px; color: #374151; vertical-align: top; word-break: break-word; }}
        td:first-child {{ font-weight: 600; color: #111827; }}
        .result-box {{ background: linear-gradient(135deg, #11998e 0%, #38ef7d 100%); color: white; padding: 22px; border-radius: 10px; text-align: center; margin-bottom: 20px; }}
        .result-box.fail {{ background: linear-gradient(135deg, #dc2626 0%, #f97316 100%); }}
        .result-box.partial {{ background: linear-gradient(135deg, #d97706 0%, #f59e0b 100%); }}
        .pass-rate {{ font-size: 40px; font-weight: 700; }}
        .result-text {{ font-size: 16px; margin-top: 4px; opacity: 0.95; }}
        .status-ok {{ color: #166534; font-weight: 600; background: #dcfce7; padding: 3px 9px; border-radius: 20px; font-size: 11px; }}
        .status-warn {{ color: #92400e; font-weight: 600; background: #fef3c7; padding: 3px 9px; border-radius: 20px; font-size: 11px; }}
        .status-fail {{ color: #b91c1c; font-weight: 600; background: #fee2e2; padding: 3px 9px; border-radius: 20px; font-size: 11px; }}
        .attn-box {{ background: #fff7ed; padding: 13px 15px; border-radius: 8px; border-left: 4px solid #ea580c; font-size: 12px; color: #4a5568; margin-top: 14px; }}
        .ok-box {{ background: #ecfdf5; border-left-color: #16a34a; }}
        .attn-box ul {{ margin: 8px 0 0 18px; padding: 0; }}
        .footer {{ background: #f8f9fa; padding: 18px; text-align: center; color: #6b7280; font-size: 11px; }}
    </style>
</head>
<body>
    <div class="container">
        <div class="header">
            <h1>周测固件汇总报告</h1>
            <div class="subtitle">{escape(product_name)}</div>
        </div>
        <div class="content">
            <div class="info-grid">
                <div class="info-item"><label>产品名称</label><span>{escape(product_name)}</span></div>
                <div class="info-item"><label>芯片型号</label><span>{escape(str(selected.get('moduleMark') or selected.get('moduleBoard') or '-'))}</span></div>
                <div class="info-item"><label>语言 / 场景</label><span>{escape(str(selected.get('language') or '-'))} / {escape(str(selected.get('sceneLabel') or '-'))}</span></div>
                <div class="info-item"><label>版本</label><span>{escape(str(selected.get('versionLabel') or '-'))}</span></div>
                <div class="info-item"><label>测试时间</label><span>{escape(updated_at)}</span></div>
                <div class="info-item"><label>目标范围</label><span>{escape(str(selected.get('productPath') or '-'))}</span></div>
            </div>
            <div class="result-box{' fail' if conclusion == 'FAIL' else ' partial' if conclusion == 'PARTIAL' else ''}">
                <div class="pass-rate">{pass_rate}%</div>
                <div class="result-text">{escape(result_text)}</div>
            </div>
            {feature_section_html}
            {algo_section_html}
            <div class="section-title">产品能力裁剪</div>
            <table>
                <thead>
                    <tr><th>能力点</th><th>平台状态</th><th>本轮处理</th></tr>
                </thead>
                <tbody>
{feature_rows_html}
                </tbody>
            </table>
            {control_section_html}
            {execution_section_html}
            <div class="section-title">详细技术矩阵（可忽略）</div>
            <table>
                <thead>
                    <tr><th style="width: 14%;">执行目录</th><th style="width: 10%;">状态</th><th style="width: 28%;">关键配置</th><th style="width: 16%;">配置断言</th><th style="width: 16%;">设备验证</th><th style="width: 16%;">异常摘要</th></tr>
                </thead>
                <tbody>
{chr(10).join(variant_rows_html)}
                </tbody>
            </table>
            {issue_html}
            <div class="section-title">附件说明</div>
            <div class="attn-box ok-box">
                <strong>result.zip</strong> 包含：
                <ul>
                    <li>周测汇总 HTML、feature matrix、manifest、summary</li>
                    <li>`result/` 目录下每个执行包的独立归档目录</li>
                    <li>每个目录内均包含固件 zip、参数说明、配置结果、设备结果、日志和 suite</li>
                </ul>
            </div>
        </div>
        <div class="footer">此邮件由自动测试系统发送 | 测试结论：<strong>{escape(conclusion)}</strong></div>
    </div>
</body>
</html>
"""

    html_path = task_dir / "test_report.html"
    html_path.write_text(html_body, encoding="utf-8")

    def compact_counter_text(counters: Dict[str, Any]) -> str:
        if not isinstance(counters, dict) or not counters:
            return "-"
        parts: List[str] = []
        for key, value in counters.items():
            if value in (None, "", 0):
                continue
            parts.append(f"{key} {value}")
        return " / ".join(parts) or "-"

    def compact_override_summary(values: Dict[str, Any]) -> str:
        values = dict(values or {})
        ordered = [
            ("timeout", "超时"),
            ("volLevel", "档位"),
            ("defaultVol", "默认音量"),
            ("volSave", "音量保存"),
            ("wakeWordSave", "唤醒词保存"),
            ("uportBaud", "波特率"),
            ("logLevel", "日志级别"),
            ("speed", "语速"),
            ("voiceRegEnable", "语音注册"),
            ("multiWkeEnable", "多唤醒"),
            ("multiWkeMode", "多唤醒模式"),
        ]
        parts: List[str] = []
        for key, label in ordered:
            if key not in values:
                continue
            value = values.get(key)
            if value in (None, ""):
                continue
            if isinstance(value, bool):
                rendered = "开" if value else "关"
            else:
                rendered = str(value)
            parts.append(f"{label} {rendered}")
        return " / ".join(parts[:6]) or short_text("；".join(human_param_lines(values)), limit=120) or "-"

    def compact_scope_summary(item: Dict[str, Any]) -> str:
        variant_id = scalar_text(item.get("id"))
        overrides = item.get("resolvedOverrides") or item.get("overrides") or {}
        scopes: List[str] = ["基础功能", "唤醒/命令词/协议"]
        if overrides.get("voiceRegEnable"):
            scopes.append("语音注册")
        if overrides.get("multiWkeEnable"):
            scopes.append("多唤醒")
        if variant_id == "pkg-02-left-boundary":
            scopes.insert(0, "左边界配置")
        elif variant_id == "pkg-03-right-boundary":
            scopes.insert(0, "右边界配置")
        elif variant_id == "pkg-01-mid-stable":
            scopes.insert(0, "中值稳定配置")
        return " / ".join(dict.fromkeys(scopes))

    def compact_variant_result(item: Dict[str, Any]) -> str:
        status = scalar_text(item.get("status"))
        config_text = compact_counter_text((item.get("configResult") or {}).get("counters") or {})
        device_text = compact_counter_text((item.get("deviceResult") or {}).get("counters") or {})
        if status == "done":
            return device_text if device_text != "-" else config_text or "通过"
        if status == "device_done_with_issue":
            if device_text != "-":
                return device_text
            return scalar_text(item.get("humanResult")) or "存在失败项"
        return scalar_text(item.get("humanResult")) or variant_status_text(status)

    def compact_coverage_note(item: Dict[str, Any]) -> str:
        return coverage_note(item)

    def compact_issue_blocks(item: Dict[str, Any]) -> List[Dict[str, str]]:
        variant_id = scalar_text(item.get("id"))
        failures = list((item.get("deviceResult") or {}).get("finalFailures") or [])
        case_ids = {scalar_text(entry.get("caseId")) for entry in failures if scalar_text(entry.get("caseId"))}
        blocks: List[Dict[str, str]] = []

        if variant_id == "pkg-02-left-boundary" and any(case_id.startswith("CORE-CMD-") for case_id in case_ids):
            return [
                {
                    "title": "左边界命令词阶段超时",
                    "actual": "执行路径：先播报唤醒词，待设备进入唤醒态后继续播报 7 条基础命令词。实际结果：命令阶段稳定回落到 `A5 FA 00 81 08 00 28 FB`，未进入命令识别窗口。",
                    "expected": "唤醒响应结束后仍应保持命令词识别状态，并对基础命令词给出协议或播报响应。",
                    "analysis": "当前包 `timeout=1s`，左边界窗口过短，命令词到达时设备大概率已退出唤醒态；更像时序边界配置问题，不是指定声卡播报链路异常。",
                }
            ]

        if variant_id == "pkg-04-full-save-on":
            if "VOICE-006" in case_ids:
                blocks.append(
                    {
                        "title": "失败语料被误学成功",
                        "actual": "执行路径：进入学习唤醒词流程，按失败场景连续重试直到次数耗尽。实际结果：设备仍把失败语料学习成功，后续可被当作已注册词使用。",
                        "expected": "失败重试耗尽后应明确学习失败，语料不能入库生效。",
                        "analysis": "失败分支没有把学习状态正确回滚，设备侧学习状态机可能仍提交了本应失败的语料。",
                    }
                )
            if "VOICE-007" in case_ids:
                blocks.append(
                    {
                        "title": "内置命令词冲突未被拒绝",
                        "actual": "执行路径：进入学习命令词流程，尝试学习已支持的内置命令词。实际结果：设备没有稳定给出明确拒绝，冲突词拦截表现异常。",
                        "expected": "已支持的内置命令词应在学习入口就被直接拒绝，不能继续进入成功路径。",
                        "analysis": "命令词冲突校验前置不足，固件没有在学习流程前半段拦住已占用词。",
                    }
                )
            if "VOICE-008" in case_ids:
                blocks.append(
                    {
                        "title": "默认唤醒词冲突判定不稳定",
                        "actual": "执行路径：进入学习唤醒词流程，尝试把默认唤醒词再次注册为新词。实际结果：冲突场景未稳定收敛到明确拒绝，需要结合实际唤醒效果才能判断状态。",
                        "expected": "默认唤醒词应被明确拒绝学习，且原默认唤醒词持续保持可用。",
                        "analysis": "当前更像冲突失败分支和自动判定边界都存在模糊地带，需要设备播报、唤醒状态和保留词状态联合确认。",
                    }
                )
            if "VOICE-009" in case_ids:
                blocks.append(
                    {
                        "title": "保留命令词冲突未被拒绝",
                        "actual": "执行路径：进入学习命令词流程，尝试学习保留命令词。实际结果：设备没有稳定明确拒绝，保留词冲突拦截异常。",
                        "expected": "保留命令词应被直接拒绝，不能继续进入学习成功路径。",
                        "analysis": "保留词名单校验未在学习前置流程生效，导致冲突词进入了后续学习分支。",
                    }
                )
            if "VOICE-010" in case_ids:
                blocks.append(
                    {
                        "title": "保留唤醒词冲突未被拒绝",
                        "actual": "执行路径：进入学习唤醒词流程，尝试学习保留唤醒词。实际结果：设备没有稳定给出冲突拒绝，异常路径与正常学习路径混杂。",
                        "expected": "保留唤醒词应被直接拒绝学习，且现有唤醒词状态保持不变。",
                        "analysis": "保留唤醒词冲突路径没有和普通学习失败路径彻底分离，导致判定和设备表现都不够清晰。",
                    }
                )
            if "VOICE-011" in case_ids:
                blocks.append(
                    {
                        "title": "删除命令词触发重启",
                        "actual": "执行路径：进入删除命令词流程并删除已注册命令词。实际结果：删除后抓到 boot 签名，设备发生异常重启。",
                        "expected": "删除完成后应保持正常运行，不应重启，也不应丢失当前会话状态。",
                        "analysis": "更像固件侧删除持久化或状态迁移异常，不是测试链路误判。",
                    }
                )
            if "MWK-CFG-001" in case_ids:
                blocks.append(
                    {
                        "title": "人工配置约束项",
                        "actual": "执行路径：自动化设备侧用例完成后，剩余冻结词/默认词约束只能通过页面或接口继续确认。实际结果：当前自动化不会把这类后台配置项判成设备通过或失败。",
                        "expected": "需页面或接口继续人工确认配置约束是否符合设计。",
                        "analysis": "这是验证边界，不是设备运行异常；应作为人工复核项保留。",
                    }
                )
            if blocks:
                return blocks

        if variant_id == "pkg-05-multi-save-off" and case_ids == {"MWK-CFG-001"}:
            return [
                {
                    "title": "仅剩人工配置确认项",
                    "actual": "执行路径：多唤醒设备侧自动化用例已全部完成。实际结果：当前只剩 `MWK-CFG-001` 需要到页面或接口核对冻结词/默认词约束。",
                    "expected": "完成后台配置确认后即可收口该包。",
                    "analysis": "属于页面/API 侧人工复核项，不是设备功能异常。",
                }
            ]

        concise = find_variant_failure_details(task_dir, variant_id, scalar_text(item.get("status")), item.get("deviceResult"))
        for text in concise[:3]:
            blocks.append(
                {
                    "title": "异常摘要",
                    "actual": f"执行路径：按该包既定 suite 自动执行。实际结果：{text}",
                    "expected": "详见附件中的测试结果与原始日志，按对应功能设计判断是否达标。",
                    "analysis": "当前摘要仅保留核心现象；精确归因需结合附件中的原始日志、结果表和串口记录继续定位。",
                }
            )
        return blocks

    total_variants = len(variants)
    variant_pass_count = sum(1 for item in variants if scalar_text(item.get("status")) == "done")
    variant_fail_count = sum(1 for item in variants if scalar_text(item.get("status")) in issue_statuses)
    variant_pending_count = sum(1 for item in variants if scalar_text(item.get("status")) in {"planned", "packaged", "device_skipped"})

    overview_cards_html = "\n".join(
        "                <div class=\"pkg-card "
        + ("pkg-ok" if scalar_text(item.get("status")) == "done" else "pkg-fail" if scalar_text(item.get("status")) in issue_statuses else "pkg-warn")
        + "\">\n"
        f"                    <div class=\"pkg-top\"><span class=\"pkg-id\">{escape(scalar_text(item.get('id')) or '-')}</span><span class=\"pkg-state\">{escape(variant_status_text(scalar_text(item.get('status'))))}</span></div>\n"
        f"                    <div class=\"pkg-title\">{escape(scalar_text(item.get('title')) or scalar_text(item.get('id')) or '-')}</div>\n"
        f"                    <div class=\"pkg-line\"><span>配置</span><strong>{escape(compact_override_summary(item.get('resolvedOverrides') or item.get('overrides') or {}))}</strong></div>\n"
        f"                    <div class=\"pkg-line\"><span>测试</span><strong>{escape(compact_scope_summary(item))}</strong></div>\n"
        f"                    <div class=\"pkg-line\"><span>结果</span><strong>{escape(compact_variant_result(item))}</strong></div>\n"
        + (
            f"                    <div class=\"pkg-line\"><span>未测</span><strong>{escape(compact_coverage_note(item))}</strong></div>\n"
            if compact_coverage_note(item)
            else ""
        )
        + "                </div>"
        for item in variants
    )

    issue_cards: List[str] = []
    for item in variants:
        note = compact_coverage_note(item)
        if not note:
            continue
        issue_cards.append(
            "                <div class=\"issue-card issue-ok\">\n"
            f"                    <div class=\"issue-head\">{escape(scalar_text(item.get('title')) or scalar_text(item.get('id')) or '-')} - 未测范围说明</div>\n"
            f"                    <div class=\"issue-meta\">{escape(compact_override_summary(item.get('resolvedOverrides') or item.get('overrides') or {}))}</div>\n"
            "                    <div class=\"issue-line\">\n"
            "                        <div class=\"issue-title\">范围裁剪说明</div>\n"
            f"                        <div class=\"issue-text\"><span>现象</span>{escape(note)}</div>\n"
            "                        <div class=\"issue-text\"><span>预期</span>未执行范围必须在邮件正文和附件 summary 中明确说明，不能写成全功能通过。</div>\n"
            "                        <div class=\"issue-text\"><span>分析</span>这是测试范围裁剪说明，不是设备功能失败。</div>\n"
            "                    </div>\n"
            "                </div>"
        )
    for item in variants:
        status = scalar_text(item.get("status"))
        if status not in issue_statuses:
            continue
        blocks = compact_issue_blocks(item)
        if not blocks:
            continue
        issue_lines_html = "".join(
            "                    <div class=\"issue-line\">\n"
            f"                        <div class=\"issue-title\">{escape(block.get('title') or '-')}</div>\n"
            f"                        <div class=\"issue-text\"><span>现象</span>{escape(block.get('actual') or '-')}</div>\n"
            f"                        <div class=\"issue-text\"><span>预期</span>{escape(block.get('expected') or '-')}</div>\n"
            f"                        <div class=\"issue-text\"><span>分析</span>{escape(block.get('analysis') or '-')}</div>\n"
            "                    </div>\n"
            for block in blocks
        )
        issue_cards.append(
            "                <div class=\"issue-card\">\n"
            f"                    <div class=\"issue-head\">{escape(scalar_text(item.get('title')) or scalar_text(item.get('id')) or '-')}</div>\n"
            f"                    <div class=\"issue-meta\">{escape(compact_override_summary(item.get('resolvedOverrides') or item.get('overrides') or {}))}</div>\n"
            f"{issue_lines_html}"
            "                </div>"
        )

    compact_issue_section_html = (
        "\n".join(issue_cards)
        if issue_cards
        else "                <div class=\"issue-card issue-ok\"><div class=\"issue-head\">本轮无新增核心异常</div><div class=\"issue-meta\">详细日志与结果表请直接查看附件。</div></div>"
    )

    mail_summary_body = f"""<!DOCTYPE html>
<html>
<head>
    <meta charset="UTF-8">
    <style>
        body {{ margin: 0; padding: 0; background: #08111f; color: #d8e3f0; font-family: 'Segoe UI', 'PingFang SC', 'Microsoft YaHei', sans-serif; }}
        .shell {{ padding: 28px 18px; background:
            radial-gradient(circle at top right, rgba(34,211,238,0.16), transparent 28%),
            radial-gradient(circle at top left, rgba(59,130,246,0.18), transparent 30%),
            #08111f; }}
        .panel {{ max-width: 980px; margin: 0 auto; background: rgba(10, 18, 34, 0.92); border: 1px solid rgba(148, 163, 184, 0.18); border-radius: 18px; overflow: hidden; box-shadow: 0 18px 60px rgba(2, 6, 23, 0.45); }}
        .hero {{ padding: 28px 30px 22px; border-bottom: 1px solid rgba(148, 163, 184, 0.12); background: linear-gradient(135deg, rgba(8,17,31,0.9), rgba(11,28,53,0.96)); }}
        .eyebrow {{ color: #67e8f9; font-size: 12px; letter-spacing: 1.4px; text-transform: uppercase; }}
        .hero h1 {{ margin: 10px 0 8px; font-size: 28px; color: #f8fafc; }}
        .hero p {{ margin: 0; color: #94a3b8; font-size: 13px; }}
        .hero-grid {{ display: grid; grid-template-columns: repeat(4, 1fr); gap: 12px; margin-top: 18px; }}
        .metric {{ background: rgba(15, 23, 42, 0.72); border: 1px solid rgba(148, 163, 184, 0.14); border-radius: 14px; padding: 14px 16px; }}
        .metric label {{ display: block; font-size: 11px; color: #94a3b8; text-transform: uppercase; letter-spacing: 0.8px; margin-bottom: 6px; }}
        .metric strong {{ font-size: 22px; color: #f8fafc; }}
        .metric span {{ display: block; margin-top: 6px; font-size: 12px; color: #cbd5e1; }}
        .content {{ padding: 24px 30px 30px; }}
        .section {{ margin-top: 24px; }}
        .section:first-child {{ margin-top: 0; }}
        .section-head {{ display: flex; align-items: center; justify-content: space-between; margin-bottom: 14px; }}
        .section-head h2 {{ margin: 0; font-size: 16px; color: #f8fafc; }}
        .section-head span {{ color: #67e8f9; font-size: 12px; }}
        .hint {{ color: #94a3b8; font-size: 12px; margin: 0 0 14px; }}
        .pkg-grid {{ display: grid; grid-template-columns: repeat(2, 1fr); gap: 14px; }}
        .pkg-card {{ border-radius: 14px; padding: 16px; border: 1px solid rgba(148, 163, 184, 0.14); background: rgba(15, 23, 42, 0.7); }}
        .pkg-ok {{ box-shadow: inset 0 0 0 1px rgba(34, 197, 94, 0.22); }}
        .pkg-fail {{ box-shadow: inset 0 0 0 1px rgba(248, 113, 113, 0.24); }}
        .pkg-warn {{ box-shadow: inset 0 0 0 1px rgba(250, 204, 21, 0.24); }}
        .pkg-top {{ display: flex; justify-content: space-between; align-items: center; margin-bottom: 8px; }}
        .pkg-id {{ color: #67e8f9; font-size: 12px; font-family: Consolas, monospace; }}
        .pkg-state {{ font-size: 11px; color: #cbd5e1; background: rgba(30, 41, 59, 0.9); border: 1px solid rgba(148, 163, 184, 0.16); border-radius: 999px; padding: 3px 10px; }}
        .pkg-title {{ color: #f8fafc; font-size: 15px; font-weight: 700; margin-bottom: 10px; }}
        .pkg-line {{ display: grid; grid-template-columns: 48px 1fr; gap: 10px; margin-top: 8px; font-size: 12px; }}
        .pkg-line span {{ color: #94a3b8; }}
        .pkg-line strong {{ color: #dbeafe; font-weight: 600; }}
        .issue-card {{ margin-top: 14px; border-radius: 14px; padding: 16px; border: 1px solid rgba(248, 113, 113, 0.24); background: linear-gradient(180deg, rgba(30, 41, 59, 0.82), rgba(15, 23, 42, 0.92)); }}
        .issue-ok {{ border-color: rgba(34, 197, 94, 0.22); }}
        .issue-head {{ color: #f8fafc; font-size: 15px; font-weight: 700; }}
        .issue-meta {{ color: #94a3b8; font-size: 12px; margin-top: 6px; }}
        .issue-line {{ margin-top: 14px; padding-top: 14px; border-top: 1px dashed rgba(148, 163, 184, 0.16); }}
        .issue-line:first-of-type {{ border-top: none; padding-top: 0; }}
        .issue-title {{ color: #fca5a5; font-size: 13px; font-weight: 700; margin-bottom: 8px; }}
        .issue-text {{ display: grid; grid-template-columns: 42px 1fr; gap: 10px; font-size: 12px; color: #dbe4ee; margin-top: 6px; }}
        .issue-text span {{ color: #67e8f9; }}
        .foot {{ margin-top: 22px; padding: 14px 16px; border-radius: 12px; background: rgba(15, 23, 42, 0.82); border: 1px solid rgba(148, 163, 184, 0.14); color: #cbd5e1; font-size: 12px; }}
        .foot strong {{ color: #f8fafc; }}
        @media (max-width: 860px) {{
            .hero-grid, .pkg-grid {{ grid-template-columns: 1fr; }}
        }}
    </style>
</head>
<body>
    <div class="shell">
        <div class="panel">
            <div class="hero">
                <div class="eyebrow">ListenAI Validation Mail Summary</div>
                <h1>Mars-Belt Test Result</h1>
                <p>{escape(product_name)} | {escape(str(selected.get('moduleMark') or selected.get('moduleBoard') or '-'))} | {escape(str(selected.get('language') or '-'))} | {escape(str(selected.get('versionLabel') or '-'))}</p>
                <div class="hero-grid">
                    <div class="metric"><label>包结果</label><strong>{variant_pass_count}/{total_variants}</strong><span>通过包数</span></div>
                    <div class="metric"><label>异常包</label><strong>{variant_fail_count}</strong><span>需重点关注</span></div>
                    <div class="metric"><label>未完成</label><strong>{variant_pending_count}</strong><span>本轮未收口包数</span></div>
                    <div class="metric"><label>功能点</label><strong>{escape(result_text)}</strong><span>{escape(updated_at)}</span></div>
                </div>
            </div>
            <div class="content">
                <div class="section">
                    <div class="section-head">
                        <h2>本轮打包与测试范围</h2>
                        <span>只保留关键配置与结果</span>
                    </div>
                    <p class="hint">详细用例、原始日志、Excel 结果保持在附件中；正文只展示“打包了什么、测了什么、结果如何”。</p>
                    <div class="pkg-grid">
{overview_cards_html}
                    </div>
                </div>
                <div class="section">
                    <div class="section-head">
                        <h2>需关注项</h2>
                        <span>现象 / 预期 / 分析</span>
                    </div>
{compact_issue_section_html}
                </div>
                <div class="foot">
                    <strong>附件说明</strong><br>
                    `result.zip` 内保留完整 `test_report.html`、`summary.md`、`manifest.json`、每包结果目录、Excel、串口日志与 suite。需要定位细节时直接看附件，不再在邮件正文展开。
                </div>
            </div>
        </div>
    </div>
</body>
</html>
"""

    mail_summary_path = task_dir / "mail_summary.html"
    mail_summary_path.write_text(mail_summary_body, encoding="utf-8")

    zip_path = task_dir / "result.zip"
    with zipfile.ZipFile(zip_path, "w", zipfile.ZIP_DEFLATED) as zipf:
        for source, arcname in [
            (html_path, "test_report.html"),
            (feature_matrix_path, "feature_matrix.md"),
            (summary_md_path, "weekly/summary.md"),
            (manifest_path, "weekly/manifest.json"),
            (control_md_path, "weekly/control_variable_analysis.md"),
            (execution_md_path, "weekly/execution_results.md"),
        ]:
            if source.exists():
                zipf.write(source, arcname)
        result_root = task_dir / "result"
        if result_root.exists():
            for item in sorted(result_root.rglob("*")):
                if item.is_dir():
                    continue
                zipf.write(item, str(item.relative_to(task_dir)))

    log("周测汇总邮件产物已生成", {"html": str(html_path), "mailSummary": str(mail_summary_path), "zip": str(zip_path)})


def init_state(manifest_path: Path, selected: Dict[str, Any], source_release_id: str, shared_name: str, specs: Sequence[Dict[str, Any]], catalog_paths: Dict[str, str]) -> Dict[str, Any]:
    if manifest_path.exists():
        return jload(manifest_path)
    return {
        "generatedAt": datetime.now().isoformat(timespec="seconds"),
        "updatedAt": datetime.now().isoformat(timespec="seconds"),
        "selectedMeta": selected,
        "sourceReleaseId": source_release_id,
        "productName": shared_name,
        "catalogPaths": catalog_paths,
        "sharedProduct": {"productName": shared_name, "productId": "", "productDetail": None},
        "variants": [
            {
                "id": spec["id"],
                "title": spec["title"],
                "kind": spec["kind"],
                "pack": spec["pack"],
                "comments": spec["comments"],
                "overrides": deepcopy(spec["overrides"]),
                "resolvedOverrides": {},
                "resolvedVoiceRegLearnCommands": [],
                "status": "planned",
                "packageAttemptCount": 0,
                "downloadedFirmwarePath": "",
                "downloadedSdkPath": "",
                "sdkToolHits": [],
                "parameterTxtPath": "",
                "suiteDir": "",
                "configResult": {},
                "deviceResult": {},
                "error": "",
            }
            for spec in specs
        ],
    }


def build_variant_rows(base_rows: Sequence[Dict[str, Any]], spec: Dict[str, Any], config: Dict[str, Any]) -> List[Dict[str, Any]]:
    _ = base_rows
    suite_payload = build_profile_payload(
        web_config=config or {},
        profile=str(spec.get("profile") or "base"),
        metadata={
            "scalars": {"generatedAt": datetime.now().isoformat(timespec="seconds")},
            "appliedOverrides": spec.get("overrides") or {},
            "learnWords": [],
            "finalRelease": {},
            "comments": spec.get("comments", ""),
        },
    )
    return [dict(item) for item in (suite_payload.get("rows") or [])]


def adapt_suite_payload_for_variant(suite_payload: Dict[str, Any], spec: Dict[str, Any]) -> Dict[str, Any]:
    overrides = spec.get("overrides") or {}
    try:
        timeout_value = int(overrides.get("timeout"))
    except Exception:
        timeout_value = 0
    if spec.get("id") == "pkg-02-left-boundary" and timeout_value and timeout_value <= 1:
        rows = [dict(item) for item in (suite_payload.get("rows") or [])]
        suite_payload["rows"] = [
            item for item in rows
            if not str(item.get("用例编号") or "").startswith("CORE-CMD-")
        ]
        metadata = suite_payload.setdefault("metadata", {})
        notes = metadata.setdefault("executionNotes", [])
        if isinstance(notes, list):
            notes.append(
                "pkg-02 timeout=1s: 未执行普通 CORE-CMD 命令词/全功能用例；"
                "该包只验证 timeout 左边界、配置边界、唤醒、播报和串口观察项，命令词全功能由 pkg-01/pkg-03 覆盖。"
            )
    return suite_payload


def main() -> int:
    args = build_parser().parse_args()
    if args.update_audio_skills:
        os.environ["MARS_BELT_UPDATE_AUDIO_SKILLS"] = "1"
    args.product = decode_cli_text(args.product)
    args.language = decode_cli_text(args.language)
    args.version = decode_cli_text(args.version)
    args.scene = decode_cli_text(args.scene)
    token = resolve_listenai_token()

    log("\u5f00\u59cb\u51c6\u5907\u76ee\u6807\u5143\u6570\u636e", {"product": args.product, "module": args.module, "language": args.language, "version": args.version, "scene": args.scene})
    prep_dir = ensure_runtime_dir("prep", "weekly", f"{args.module}_{args.product}_{week_stamp()}")
    metadata = ensure_target_metadata(args, prep_dir)
    catalog = metadata["catalog"]
    selected = dict(catalog["selected"])
    source_release_id = str(args.source_release_id or catalog.get("sourceReleaseId") or "")
    if not source_release_id:
        raise RuntimeError("sourceReleaseId missing for selected target")

    specs = build_variant_specs(selected, catalog)
    selected_variant_ids = {item.strip() for item in str(args.variants or "").split(",") if item.strip()}
    if selected_variant_ids:
        specs = [spec for spec in specs if spec["id"] in selected_variant_ids]
        if not specs:
            raise RuntimeError(f"no variants matched --variants={args.variants}")
    run_id = run_id_for(selected)
    shared_name = product_name(args.module, selected["productLabel"], selected["versionLabel"])
    if str(args.task_dir or "").strip():
        task_dir = Path(resolve_user_path(args.task_dir, ROOT))
        task_dir.mkdir(parents=True, exist_ok=True)
    else:
        task_dir = ensure_task_dir(RESROOT, shared_name, "测试模式")
    runtime_dir = runtime_dir_for_task(task_dir, "weekly")
    package_dir = task_dir
    manifest_path = runtime_dir / "manifest.json"
    catalog_paths = {
        "parameterCatalog": str((prep_dir / "catalog/parameter_catalog.json").resolve()),
        "testCaseCatalog": str((prep_dir / "catalog/test_case_catalog.json").resolve()),
    }
    state = init_state(manifest_path, selected, source_release_id, shared_name, specs, catalog_paths)
    state["runId"] = run_id
    state["updatedAt"] = datetime.now().isoformat(timespec="seconds")
    jsave(manifest_path, state)
    write_feature_matrix(task_dir, selected, specs)
    log("\u76ee\u6807\u5df2\u786e\u8ba4", {"selected": selected, "sourceReleaseId": source_release_id, "sharedProductName": shared_name, "runId": run_id})

    client = ListenAIClient(token=token, timeout=args.timeout_sec)
    feature_toggle = build_feature_toggle(dict(catalog.get("featureMap") or {}))
    tts = tts_fallback()
    shared_manifest = {"selectedMeta": selected, "sharedProduct": state.get("sharedProduct") or {"productName": shared_name, "productId": "", "productDetail": None}}
    product_detail = ensure_shared_product(client, shared_manifest)
    state["sharedProduct"] = shared_manifest["sharedProduct"]
    state["updatedAt"] = datetime.now().isoformat(timespec="seconds")
    jsave(manifest_path, state)
    log("\u5171\u4eab\u4ea7\u54c1\u5df2\u786e\u8ba4", {"productName": state["sharedProduct"].get("productName"), "productId": state["sharedProduct"].get("productId")})

    local_flash_tool = CLI.exists() and any((BURNROOT / name).exists() for name in ("Uart_Burn_Tool", "Uart_Burn_Tool.exe"))
    if local_flash_tool and not args.skip_device:
        run([sys.executable, "-X", "utf8", str(CLI), "set-volume", "--percent", "30"])

    for spec in specs:
        maybe_pause(manifest_path, state)
        current = next(item for item in state["variants"] if item["id"] == spec["id"])
        current.update({"title": spec["title"], "kind": spec["kind"], "pack": spec["pack"], "comments": spec["comments"], "overrides": deepcopy(spec["overrides"])})

        if current.get("status") in {"done", "device_skipped", "pack_failed"} and current.get("downloadedFirmwarePath"):
            log("\u8df3\u8fc7\u5df2\u5b8c\u6210\u53d8\u4f53", {"variant": spec["id"], "status": current.get("status")})
            continue

        try:
            firmware_zip: Optional[Path] = None
            suite_dir: Optional[Path] = None
            result = None
            reuse_existing = (
                current.get("status") in {"device_done_with_issue", "device_burn_failed", "packaged"}
                and str((current.get("configResult") or {}).get("status") or "") == "ok"
            )
            if reuse_existing:
                firmware_zip = existing_file(current.get("downloadedFirmwarePath"))
                suite_dir = existing_dir(current.get("suiteDir"))
                reuse_existing = bool(firmware_zip and suite_dir)
            if reuse_existing:
                current["status"] = "packaged"
                current["error"] = ""
                runtime_serial = resolve_runtime_serials(load_web_config(str(firmware_zip), "") or {}, args)
                current["runtimeSerial"] = deepcopy(runtime_serial)
                log("复用已有固件与 suite，直接重跑设备验证", {"variant": spec["id"], "zipPath": str(firmware_zip), "suiteDir": str(suite_dir)})
            else:
                last_error = ""
                for attempt in range(1, 3):
                    log("\u5f00\u59cb\u6253\u5305", {"variant": spec["id"], "attempt": attempt, "comments": spec["comments"]})
                    overrides = deepcopy(spec.get("overrides") or {})
                    overrides["comments"] = spec["comments"]
                    try:
                        if spec["pack"] == "algo":
                            summary = package_release_with_algo_unified(client, product_detail, source_release_id, args.timeout_sec, overrides, feature_toggle)
                        else:
                            summary = package_release_for_existing_product(client, product_detail, source_release_id, args.timeout_sec, overrides)
                        final_release = require_ok(client.get("/fw/release/detail", params={"id": summary["releaseId"]}), "release detail").get("data") or {}
                        result = {"summary": summary, "final": final_release, "attempt": attempt}
                        break
                    except Exception as exc:
                        last_error = f"{type(exc).__name__}: {exc}"
                        mkdir(runtime_dir / "attempt_logs")
                        (runtime_dir / "attempt_logs" / f"{spec['id']}_attempt_{attempt:02d}.txt").write_text(traceback.format_exc(), encoding="utf-8")
                        log("\u6253\u5305\u5931\u8d25", {"variant": spec["id"], "attempt": attempt, "error": last_error})
                        if attempt < 2:
                            time.sleep(retry_sleep_seconds(exc))
                if not result:
                    raise RuntimeError(f"{spec['id']} failed after 2 attempts: {last_error}")

                current["packageAttemptCount"] = result["attempt"]
                current["resolvedOverrides"] = deepcopy((result["summary"] or {}).get("appliedOverrides") or {})
                current["resolvedVoiceRegLearnCommands"] = deepcopy((result["summary"] or {}).get("resolvedVoiceRegLearnCommands") or [])
                release_version = str((result["summary"] or {}).get("releaseVersion") or "package")
                slug = re.sub(r"[^A-Za-z0-9._-]+", "-", release_version)[:32]
                firmware_zip = package_dir / f"{spec['id']}_{slug}.zip"
                client.download(str(result["summary"]["pkgUrl"]), str(firmware_zip))
                current["downloadedFirmwarePath"] = str(firmware_zip.resolve())
                current["status"] = "packaged"
                current["error"] = ""

                sdk_url = str((result["summary"] or {}).get("pkgSDKUrl") or "").strip()
                if sdk_url:
                    sdk_zip = package_dir / f"{spec['id']}_{slug}_SDK.zip"
                    client.download(sdk_url, str(sdk_zip))
                    current["downloadedSdkPath"] = str(sdk_zip.resolve())
                    current["sdkToolHits"] = sdk_hits(sdk_zip)
                    log("SDK \u4e0b\u8f7d\u5b8c\u6210", {"variant": spec["id"], "sdkPath": str(sdk_zip), "toolHits": current["sdkToolHits"][:20]})

                parameter_text = param_txt(selected, shared_name, spec, result["summary"], result["final"])
                parameter_path = package_dir / f"{spec['id']}_package_params.txt"
                parameter_path.write_text(parameter_text, encoding="utf-8")
                inject(firmware_zip, "validation_params.txt", parameter_text)
                current["parameterTxtPath"] = str(parameter_path.resolve())
                log("\u53c2\u6570\u8bf4\u660e\u5df2\u5199\u5165\u56fa\u4ef6", {"variant": spec["id"], "zipPath": str(firmware_zip)})

                config = load_web_config(str(firmware_zip), "")
                runtime_serial = resolve_runtime_serials(config or {}, args)
                current["runtimeSerial"] = deepcopy(runtime_serial)
                suite_dir = mkdir(runtime_dir / "suites" / spec["id"])
                suite_payload = build_profile_payload(
                    web_config=config or {},
                    profile=str(spec.get("profile") or "base"),
                    metadata={
                        "scalars": {
                            "product": selected.get("productLabel", ""),
                            "module": selected.get("moduleBoard", ""),
                            "language": selected.get("language", ""),
                            "version": selected.get("versionLabel", ""),
                            "scene": selected.get("sceneLabel", ""),
                            "defId": selected.get("defId", ""),
                            "generatedAt": datetime.now().isoformat(timespec="seconds"),
                        },
                        "appliedOverrides": current["resolvedOverrides"] or spec.get("overrides") or {},
                        "learnWords": current["resolvedVoiceRegLearnCommands"] or [],
                        "finalRelease": result["final"],
                        "variantId": spec["id"],
                        "variantTitle": spec["title"],
                        "comments": spec["comments"],
                    },
                    selected_meta=selected,
                )
                suite_payload = adapt_suite_payload_for_variant(suite_payload, spec)
                device_info = dict(suite_payload.get("deviceInfo") or {})
                seeded = seed_audio(device_info, suite_dir)
                export_suite(suite_dir, suite_payload)
                current["suiteDir"] = str(suite_dir.resolve())
                log("\u4e13\u5c5e suite \u5df2\u751f\u6210", {"variant": spec["id"], "suiteDir": str(suite_dir), "caseCount": len(suite_payload.get("rows") or []), "audioSeed": seeded})

                cfg_cmd = [
                    sys.executable,
                    "-X",
                    "utf8",
                    str(VOICE),
                    "--suite-dir",
                    str(suite_dir),
                    "--package-zip",
                    str(firmware_zip),
                    "--config-only",
                    "-l",
                    f"{spec['id']}_config",
                ]
                if args.update_audio_skills:
                    cfg_cmd.append("--update-audio-skills")
                cfg_exit = run(cfg_cmd)
                current["configResult"] = {"exitCode": cfg_exit, **sum_cfg(latest_dir(suite_dir / "result"))}
                copy_variant_artifacts(task_dir, f"{spec['id']}_config", suite_dir)
                state["updatedAt"] = datetime.now().isoformat(timespec="seconds")
                jsave(manifest_path, state)
                if cfg_exit != 0:
                    current["deviceResult"] = {"status": "skip_due_config_fail", "reason": "config-only validation failed"}
                    current["status"] = "config_failed"
                    state["updatedAt"] = datetime.now().isoformat(timespec="seconds")
                    jsave(manifest_path, state)
                    log("离线校验失败，跳过设备验证", {"variant": spec["id"], "exitCode": cfg_exit})
                    continue

            if args.skip_device:
                current["deviceResult"] = {"status": "skip_by_flag", "reason": "--skip-device"}
                current["status"] = "device_skipped"
                state["updatedAt"] = datetime.now().isoformat(timespec="seconds")
                jsave(manifest_path, state)
                continue
            if not local_flash_tool:
                current["deviceResult"] = {"status": "skip_no_flash_tool"}
                current["status"] = "device_skipped"
                state["updatedAt"] = datetime.now().isoformat(timespec="seconds")
                jsave(manifest_path, state)
                continue

            maybe_pause(manifest_path, state)
            runtime_serial = resolve_runtime_serials(load_web_config(str(firmware_zip), "") or {}, args)
            current["runtimeSerial"] = deepcopy(runtime_serial)
            burn_exit = burn_package(firmware_zip, args, runtime_serial)
            if burn_exit != 0:
                current["deviceResult"] = {"status": "burn_failed", "exitCode": burn_exit, "logTail": latest_burn_log()}
                current["status"] = "device_burn_failed"
                state["updatedAt"] = datetime.now().isoformat(timespec="seconds")
                jsave(manifest_path, state)
                log("\u70e7\u5f55\u5931\u8d25\uff0c\u8df3\u8fc7\u8bbe\u5907\u7528\u4f8b\u6267\u884c", {"variant": spec["id"], "exitCode": burn_exit})
                continue

            dev_cmd = [
                sys.executable,
                "-X",
                "utf8",
                str(VOICE),
                "--suite-dir",
                str(suite_dir),
                "--package-zip",
                str(firmware_zip),
                "-f",
                "deviceInfo_generated.json",
                "-r",
                "0",
                "-l",
                spec["id"],
                "-p",
                str(runtime_serial.get("logPort") or args.log_port),
                "--ctrl-port",
                args.ctrl_port,
                "--protocol-port",
                str(runtime_serial.get("protocolPort") or args.protocol_port),
                "--pretest",
            ]
            if args.update_audio_skills:
                dev_cmd.append("--update-audio-skills")
            dev_exit = run(dev_cmd)
            current["deviceResult"] = {"exitCode": dev_exit, **sum_dev(latest_dir(suite_dir / "result"))}
            current["status"] = "device_done_with_issue" if device_result_has_issue(current["deviceResult"], dev_exit) else "done"
            copy_variant_artifacts(task_dir, spec["id"], suite_dir)
            state["updatedAt"] = datetime.now().isoformat(timespec="seconds")
            jsave(manifest_path, state)
        except Exception as exc:
            current["status"] = "pack_failed"
            current["error"] = f"{type(exc).__name__}: {exc}"
            state["updatedAt"] = datetime.now().isoformat(timespec="seconds")
            jsave(manifest_path, state)
            log("\u53d8\u4f53\u6267\u884c\u5931\u8d25", {"variant": spec["id"], "error": current["error"]})

    lines = [
        f"# \u5468\u5ea6\u9a8c\u8bc1\u7ed3\u679c - {shared_name}",
        "",
        f"- \u76ee\u6807\u4ea7\u54c1\uff1a`{selected.get('productPath')}`",
        f"- \u76ee\u6807\u6a21\u7ec4\uff1a`{selected.get('moduleBoard')}`",
        f"- \u76ee\u6807\u8bed\u8a00\uff1a`{selected.get('language')}`",
        f"- \u76ee\u6807\u7248\u672c\uff1a`{selected.get('versionLabel')}`",
        f"- sourceReleaseId\uff1a`{source_release_id}`",
        "",
    ]
    for item in state.get("variants") or []:
        lines.extend(
            [
                f"## {item['id']}",
                f"- \u72b6\u6001\uff1a`{item.get('status')}`",
                f"- \u56fa\u4ef6\uff1a`{item.get('downloadedFirmwarePath', '')}`",
                f"- \u53c2\u6570\u8bf4\u660e\uff1a`{item.get('parameterTxtPath', '')}`",
                f"- Suite\uff1a`{item.get('suiteDir', '')}`",
                f"- 运行串口：`{json.dumps(item.get('runtimeSerial') or {}, ensure_ascii=False)}`",
                f"- \u751f\u6548\u8986\u76d6\u53c2\u6570\uff1a`{json.dumps(item.get('resolvedOverrides') or item.get('overrides') or {}, ensure_ascii=False)}`",
                f"- \u8bed\u97f3\u6ce8\u518c\u5b66\u4e60\u8bcd\uff1a`{json.dumps(item.get('resolvedVoiceRegLearnCommands') or [], ensure_ascii=False)}`",
                f"- \u79bb\u7ebf\u6821\u9a8c\uff1a`{json.dumps(item.get('configResult') or {}, ensure_ascii=False)}`",
                f"- \u8bbe\u5907\u7ed3\u679c\uff1a`{json.dumps(item.get('deviceResult') or {}, ensure_ascii=False)}`",
                f"- SDK \u5de5\u5177\u547d\u4e2d\uff1a`{json.dumps((item.get('sdkToolHits') or [])[:20], ensure_ascii=False)}`",
                f"- \u9519\u8bef\uff1a`{item.get('error', '')}`",
                "",
            ]
        )
    (runtime_dir / "summary.md").write_text("\n".join(lines).rstrip() + "\n", encoding="utf-8")
    try:
        generate_weekly_email_report(task_dir, runtime_dir, state)
    except Exception as exc:
        log("周测汇总邮件产物生成失败", {"error": f"{type(exc).__name__}: {exc}"})
    log("\u5468\u5ea6\u4efb\u52a1\u5b8c\u6210", {"taskDir": str(task_dir), "manifest": str(manifest_path), "summary": str(runtime_dir / "summary.md")})
    failed_statuses = {"pack_failed", "config_failed", "device_burn_failed", "device_done_with_issue"}
    has_failure = any(str(item.get("status") or "") in failed_statuses for item in state.get("variants") or [])
    return 1 if has_failure else 0


if __name__ == "__main__":
    raise SystemExit(main())
