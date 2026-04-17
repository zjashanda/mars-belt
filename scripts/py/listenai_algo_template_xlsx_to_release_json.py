#!/usr/bin/env python3
from __future__ import annotations

import argparse
import json
from pathlib import Path
from typing import Any, Dict, List, Optional

from openpyxl import load_workbook


DEFAULT_SHEET = "词条预处理"

HEADER_MAP = {
    "语义标签": "idx",
    "语义(最小功能词)": "word",
    "功能泛化词": "extWord",
    "功能类型": "type",
    "播报语句": "reply",
    "播报模式": "replyMode",
    "发送协议": "sndProtocol",
    "接收协议": "recProtocol",
}


def normalize_text(value: Any) -> Optional[str]:
    if value is None:
        return None
    text = str(value).strip()
    return text or None


def normalize_idx(value: Any, fallback: int) -> int:
    if value is None:
        return fallback
    if isinstance(value, int):
        return value
    if isinstance(value, float) and value.is_integer():
        return int(value)
    text = str(value).strip()
    try:
        return int(text)
    except Exception:
        return fallback


def make_item(release_id: str, row: Dict[str, Any], fallback_idx: int) -> Dict[str, Any]:
    return {
        "id": "",
        "releaseId": release_id,
        "pid": "0",
        "idx": normalize_idx(row.get("idx"), fallback_idx),
        "word": normalize_text(row.get("word")),
        "extWord": normalize_text(row.get("extWord")),
        "type": normalize_text(row.get("type")),
        "reply": normalize_text(row.get("reply")),
        "replyMode": normalize_text(row.get("replyMode")),
        "sndProtocol": normalize_text(row.get("sndProtocol")),
        "recProtocol": normalize_text(row.get("recProtocol")),
        "recoId": "",
        "recoExtWordStr": None,
        "asrFreeEnable": None,
        "relatedId": None,
        "relatedType": None,
        "pinyin": None,
        "deleteFlag": "NOT_DELETE",
        "createTime": None,
        "createUser": None,
        "updateTime": None,
        "updateUser": None,
        "children": [],
    }


def read_template(path: Path, sheet_name: str) -> List[Dict[str, Any]]:
    wb = load_workbook(path, data_only=True)
    if sheet_name not in wb.sheetnames:
        raise RuntimeError(f"sheet not found: {sheet_name}")
    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [normalize_text(cell) for cell in rows[0]]
    mapped: List[Dict[str, Any]] = []
    for raw in rows[1:]:
        record: Dict[str, Any] = {}
        has_data = False
        for index, value in enumerate(raw):
            header = headers[index] if index < len(headers) else None
            key = HEADER_MAP.get(header or "")
            if not key:
                continue
            record[key] = value
            if value not in (None, ""):
                has_data = True
        if not has_data:
            continue
        if not normalize_text(record.get("type")):
            continue
        mapped.append(record)
    return mapped


def main() -> int:
    parser = argparse.ArgumentParser(description="Convert an algorithm template xlsx into releaseAlgoList JSON.")
    parser.add_argument("--xlsx", required=True, help="Path to the xlsx template")
    parser.add_argument("--release-id", required=True, help="releaseId to stamp onto each item")
    parser.add_argument("--output", required=True, help="Output JSON file path")
    parser.add_argument("--sheet", default=DEFAULT_SHEET, help=f"Worksheet name, default: {DEFAULT_SHEET}")
    args = parser.parse_args()

    template_rows = read_template(Path(args.xlsx), args.sheet)
    if not template_rows:
        raise RuntimeError("no usable rows found in template")

    payload = [
        make_item(str(args.release_id), row, fallback_idx=index)
        for index, row in enumerate(template_rows, start=1)
    ]
    output_path = Path(args.output)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    output_path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    print(output_path)
    print(len(payload))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
